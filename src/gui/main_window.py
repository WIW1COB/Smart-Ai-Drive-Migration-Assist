"""Main window for the Migration Analysis Tool GUI"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import json
import logging
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

from src.utils.comparison_engine import compare_folders, cleanup_temp_dirs
from src.utils.credential_manager import CredentialManager
from src.gui.results_viewer import show_results_dialog
from src.rtc.connection import RTCConnection, get_rtc_connection
from src.config import settings

# Configure logging — write INFO+ to both console and a rotating log file
_log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "rtc_comparison.log")
_file_handler = logging.FileHandler(_log_file, mode='w', encoding='utf-8')
_file_handler.setLevel(logging.DEBUG)
_file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(name)s: %(message)s'))
_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.INFO)
_console_handler.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))

_root_logger = logging.getLogger()
_root_logger.setLevel(logging.DEBUG)
# Clear any existing handlers set by basicConfig before adding ours
_root_logger.handlers.clear()
_root_logger.addHandler(_file_handler)
_root_logger.addHandler(_console_handler)

logger = logging.getLogger(__name__)


class MigrationAnalysisGUI:
    """Main GUI window for Migration Analysis Tool"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Migration Analysis Report Generator")
        self.root.geometry("780x540")
        self.root.config(bg="#EAF3FB")
        
        # RTC Connection variables
        self.rtc_username = tk.StringVar()
        self.rtc_password = tk.StringVar()
        self.rtc_server_url = tk.StringVar(value=settings.RTC_SERVER_URL)
        self.rtc_connection = None
        self.keep_signed_in_var = tk.BooleanVar(value=False)
        self.cached_credentials_loaded = False
        self.save_credentials_on_success = False
        self.component_name = tk.StringVar(value="ALL")

        # GenMake / Cfg_DBFiles filter
        self.genmake_enabled_var = tk.BooleanVar(value=False)
        self.genmake_csv_path = tk.StringVar(value="")
        self._genmake_filter = None          # GenMakeFilter instance when loaded
        self._genmake_type_vars: dict = {}   # {type_str: BooleanVar}

        self.setup_ui()
        self._load_cached_credentials()
    
    def setup_ui(self):
        """Setup the main UI components"""
        # Header
        self.create_header()
        
        # Comparison Mode Selection
        self.create_mode_selection()
        
        # Component Name Selection
        self.create_component_selection()

        # GenMake CSV filter section
        self.create_genmake_section()

        # Input frames (all modes)
        self.create_folder_input_frame()       # Offline → Offline
        self.create_snapshot_input_frame()     # Online → Online
        self.create_hybrid_input_frame()       # Online → Offline
        self.create_interface_check_frame()    # Interface List Check
        
        # RTC Integration Section
        self.create_rtc_section()
        
        # Action buttons
        self.create_action_buttons()
        
        # Progress section
        self.create_progress_section()
        
        # Initially show offline_offline mode
        self.toggle_input_mode()
    
    def create_header(self):
        """Create the Bosch-style header"""
        header_frame = tk.Frame(self.root, bg="#EAF3FB")
        header_frame.pack(fill="x")
        
        top_strip = tk.Label(header_frame, bg="#E60000", height=1)
        top_strip.pack(fill="x")
        
        title_frame = tk.Frame(header_frame, bg="#003366")
        title_frame.pack(fill="x")
        title_frame.grid_columnconfigure(0, weight=1)
        title_frame.grid_columnconfigure(1, weight=2)
        title_frame.grid_columnconfigure(2, weight=1)
        
        title_label = tk.Label(
            title_frame,
            text="Migration Analysis Report Generator",
            font=("Segoe UI", 18, "bold"),
            bg="#003366",
            fg="white"
        )
        title_label.grid(row=0, column=1, padx=10, pady=10)
        
        auth_frame = tk.Frame(title_frame, bg="#003366")
        auth_frame.grid(row=0, column=2, sticky="e", padx=10, pady=8)
        
        self.auth_status_label = tk.Label(
            auth_frame,
            text="RTC: signed out",
            font=("Segoe UI", 8),
            bg="#003366",
            fg="#DDEBFF"
        )
        self.auth_status_label.pack(side=tk.LEFT, padx=(0, 8))
        
        self.logout_btn = tk.Button(
            auth_frame,
            text="Logout",
            command=self.on_logout,
            font=("Segoe UI", 9, "bold"),
            bg="#666666",
            fg="white",
            padx=10,
            pady=4,
            state="disabled"
        )
        self.logout_btn.pack(side=tk.LEFT)
    
    def create_mode_selection(self):
        """Create comparison mode selection"""
        mode_frame = tk.Frame(self.root, bg="#EAF3FB")
        mode_frame.pack(pady=(20, 10))
        
        tk.Label(
            mode_frame,
            text="Comparison Mode:",
            bg="#EAF3FB",
            font=("Segoe UI", 11, "bold")
        ).pack(side="left", padx=5)
        
        self.comparison_mode = tk.StringVar(value="offline_offline")
        
        # Mode 1: Offline → Offline (Folders/ZIPs)
        tk.Radiobutton(
            mode_frame,
            text="📁 Offline → Offline",
            variable=self.comparison_mode,
            value="offline_offline",
            bg="#EAF3FB",
            font=("Segoe UI", 10),
            command=self.toggle_input_mode
        ).pack(side="left", padx=10)
        
        # Mode 2: Online → Online (RTC URLs)
        tk.Radiobutton(
            mode_frame,
            text="🌐 Online → Online",
            variable=self.comparison_mode,
            value="online_online",
            bg="#EAF3FB",
            font=("Segoe UI", 10),
            command=self.toggle_input_mode
        ).pack(side="left", padx=10)
        
        # Mode 3: Online → Offline (RTC URL + Folder)
        tk.Radiobutton(
            mode_frame,
            text="🔄 Online → Offline",
            variable=self.comparison_mode,
            value="online_offline",
            bg="#EAF3FB",
            font=("Segoe UI", 10),
            command=self.toggle_input_mode
        ).pack(side="left", padx=10)
        
        # Mode 4: Interface List Check
        tk.Radiobutton(
            mode_frame,
            text="📝 Interface List Check",
            variable=self.comparison_mode,
            value="interface_check",
            bg="#EAF3FB",
            font=("Segoe UI", 10),
            command=self.toggle_input_mode
        ).pack(side="left", padx=10)
    
    def create_component_selection(self):
        """Create Component Name selection dropdown"""
        comp_frame = tk.Frame(self.root, bg="#EAF3FB")
        comp_frame.pack(pady=(0, 5))

        tk.Label(
            comp_frame,
            text="Component Name:",
            bg="#EAF3FB",
            font=("Segoe UI", 11, "bold")
        ).pack(side="left", padx=5)

        component_options = ["ALL", "DEM", "DCOM", "NET", "RTE", "VAR"]
        component_cb = ttk.Combobox(
            comp_frame,
            textvariable=self.component_name,
            values=component_options,
            state="readonly",
            width=10,
            font=("Segoe UI", 10)
        )
        component_cb.pack(side="left", padx=5)

        tk.Label(
            comp_frame,
            text="(DEM: filters to .c/.h/.proc/.bcfg/.cs/.xpt/.arxml/.txt/.dpp/.mk/.pdm only)",
            bg="#EAF3FB",
            font=("Segoe UI", 8),
            fg="#555555"
        ).pack(side="left", padx=5)

    def create_genmake_section(self):
        """Create the GenMake / Cfg_DBFiles compilation-filter section."""
        outer = tk.LabelFrame(
            self.root,
            text=" 📋 GenMake Compilation Filter (Cfg_DBFiles_GenMake.csv) ",
            bg="#EAF3FB",
            font=("Segoe UI", 9, "bold"),
            fg="#003366",
            pady=4,
        )
        outer.pack(fill="x", padx=10, pady=(0, 4))
        self._genmake_outer_frame = outer

        top_row = tk.Frame(outer, bg="#EAF3FB")
        top_row.pack(fill="x", padx=6, pady=2)

        tk.Checkbutton(
            top_row,
            text="Enable: compare only files listed in GenMake CSV",
            variable=self.genmake_enabled_var,
            bg="#EAF3FB",
            font=("Segoe UI", 10),
            command=self._on_genmake_toggle,
        ).pack(side="left")

        self._genmake_detail_frame = tk.Frame(outer, bg="#EAF3FB")

        csv_row = tk.Frame(self._genmake_detail_frame, bg="#EAF3FB")
        csv_row.pack(fill="x", padx=6, pady=2)

        tk.Label(
            csv_row, text="CSV file:", bg="#EAF3FB", font=("Segoe UI", 10)
        ).pack(side="left")

        self._genmake_path_entry = tk.Entry(
            csv_row, textvariable=self.genmake_csv_path, width=60, state="readonly"
        )
        self._genmake_path_entry.pack(side="left", padx=(4, 4))

        tk.Button(
            csv_row,
            text="Browse",
            bg="#007B3E",
            fg="white",
            font=("Segoe UI", 9),
            command=self._browse_genmake_csv,
        ).pack(side="left", padx=2)

        tk.Button(
            csv_row,
            text="Clear",
            bg="#666666",
            fg="white",
            font=("Segoe UI", 9),
            command=self._clear_genmake_csv,
        ).pack(side="left", padx=2)

        self._genmake_status_label = tk.Label(
            self._genmake_detail_frame,
            text="",
            bg="#EAF3FB",
            font=("Segoe UI", 8),
            fg="#555555",
        )
        self._genmake_status_label.pack(anchor="w", padx=6)

        # File-type multi-select (populated after CSV load)
        self._genmake_types_frame = tk.Frame(self._genmake_detail_frame, bg="#EAF3FB")
        self._genmake_types_frame.pack(fill="x", padx=6, pady=(2, 4))

    def _on_genmake_toggle(self):
        """Show/hide the GenMake detail controls based on the checkbox."""
        if self.genmake_enabled_var.get():
            self._genmake_detail_frame.pack(fill="x")
        else:
            self._genmake_detail_frame.pack_forget()

    def _browse_genmake_csv(self):
        """Open file dialog to pick a Cfg_DBFiles_GenMake.csv file."""
        path = filedialog.askopenfilename(
            title="Select Cfg_DBFiles_GenMake CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        self._load_genmake_csv(path)

    def _load_genmake_csv(self, path: str):
        """Parse the selected CSV and update the filter + type-selector UI."""
        from src.utils.genmake_filter import GenMakeFilter
        try:
            gf = GenMakeFilter.from_csv(path)
            self._genmake_filter = gf
            self.genmake_csv_path.set(path)
            self._genmake_status_label.config(
                text=f"✓ Loaded {len(gf)} entries  |  File types: {', '.join(gf.all_file_types) or 'none'}",
                fg="#007B3E",
            )
            self._rebuild_type_checkboxes(gf.all_file_types)
        except Exception as exc:
            self._genmake_filter = None
            self._genmake_status_label.config(
                text=f"✗ Failed to load: {exc}", fg="#B00020"
            )

    def _clear_genmake_csv(self):
        """Remove the loaded CSV filter."""
        self._genmake_filter = None
        self.genmake_csv_path.set("")
        self._genmake_status_label.config(text="", fg="#555555")
        for widget in self._genmake_types_frame.winfo_children():
            widget.destroy()
        self._genmake_type_vars.clear()

    def _rebuild_type_checkboxes(self, file_types: list):
        """Recreate the file-type checkbox grid after a CSV load."""
        for widget in self._genmake_types_frame.winfo_children():
            widget.destroy()
        self._genmake_type_vars.clear()

        if not file_types:
            return

        tk.Label(
            self._genmake_types_frame,
            text="Include file types (uncheck to exclude):",
            bg="#EAF3FB",
            font=("Segoe UI", 9),
        ).grid(row=0, column=0, columnspan=10, sticky="w")

        cols = 6
        for idx, ft in enumerate(file_types):
            var = tk.BooleanVar(value=True)
            self._genmake_type_vars[ft] = var
            tk.Checkbutton(
                self._genmake_types_frame,
                text=ft,
                variable=var,
                bg="#EAF3FB",
                font=("Segoe UI", 8),
            ).grid(row=1 + idx // cols, column=idx % cols, sticky="w", padx=4)

    def _get_active_genmake_filter(self):
        """Return the active GenMakeFilter (with selected types applied), or None."""
        if not self.genmake_enabled_var.get() or self._genmake_filter is None:
            return None
        # Apply type selection
        selected_types = {
            ft for ft, var in self._genmake_type_vars.items() if var.get()
        } or None  # None means all types
        self._genmake_filter.filter_file_types(selected_types)
        return self._genmake_filter

    def create_folder_input_frame(self):
        """Create offline → offline input section (Folders/ZIPs)"""
        self.folder_input_frame = tk.Frame(self.root, bg="#EAF3FB")
        self.folder_input_frame.pack(fill="x")
        
        # Info label
        tk.Label(
            self.folder_input_frame,
            text="📁 Offline → Offline: Compare local folders or ZIP files",
            bg="#EAF3FB",
            font=("Segoe UI", 9, "italic"),
            fg="#666666"
        ).pack(pady=(5, 10))
        
        tk.Label(
            self.folder_input_frame,
            text="Select Platform Folder (or ZIP file):",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(5, 5))
        
        self.folder1_entry = tk.Entry(self.folder_input_frame, width=85)
        self.folder1_entry.pack()
        
        tk.Button(
            self.folder_input_frame,
            text="Browse",
            bg="#007B3E",
            fg="white",
            command=lambda: self.browse_folder(self.folder1_entry)
        ).pack(pady=5)
        
        tk.Label(
            self.folder_input_frame,
            text="Select Project Folder (or ZIP file):",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(10, 5))
        
        self.folder2_entry = tk.Entry(self.folder_input_frame, width=85)
        self.folder2_entry.pack()
        
        tk.Button(
            self.folder_input_frame,
            text="Browse",
            bg="#007B3E",
            fg="white",
            command=lambda: self.browse_folder(self.folder2_entry)
        ).pack(pady=5)
    
    def create_snapshot_input_frame(self):
        """Create online → online input section (RTC URLs)"""
        self.snapshot_input_frame = tk.Frame(self.root, bg="#EAF3FB")
        
        # Info label
        tk.Label(
            self.snapshot_input_frame,
            text="🌐 Online → Online: Compare RTC snapshots/workspaces via URLs",
            bg="#EAF3FB",
            font=("Segoe UI", 9, "italic"),
            fg="#666666"
        ).pack(pady=(5, 10))
        
        tk.Label(
            self.snapshot_input_frame,
            text="Platform Snapshot/Workspace URL or UUID:",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(5, 5))
        
        self.snapshot1_entry = tk.Entry(self.snapshot_input_frame, width=85)
        self.snapshot1_entry.pack()
        
        tk.Label(
            self.snapshot_input_frame,
            text="📝 From RTC web: Copy snapshot URL (with id=_xxxxx) or just the UUID like _ojreQAAbEfG1br8X33nQcA",
            bg="#EAF3FB",
            font=("Segoe UI", 8),
            fg="gray"
        ).pack()
        
        tk.Label(
            self.snapshot_input_frame,
            text="Project Snapshot/Workspace URL or UUID:",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(10, 5))
        
        self.snapshot2_entry = tk.Entry(self.snapshot_input_frame, width=85)
        self.snapshot2_entry.pack()
        
        tk.Label(
            self.snapshot_input_frame,
            text="📝 From RTC web: Copy snapshot URL (with id=_xxxxx) or just the UUID like _i3S_vwAaEfG3rPS3zZLwKA",
            bg="#EAF3FB",
            font=("Segoe UI", 8),
            fg="gray"
        ).pack()
    
    def create_hybrid_input_frame(self):
        """Create online → offline input section (RTC Snapshot + Local Root Folder)"""
        self.hybrid_input_frame = tk.Frame(self.root, bg="#EAF3FB")
        
        # Info label
        tk.Label(
            self.hybrid_input_frame,
            text="🔄 Online → Offline: Fetch RTC snapshot, select component, compare with local folder hierarchy",
            bg="#EAF3FB",
            font=("Segoe UI", 9, "italic"),
            fg="#666666"
        ).pack(pady=(5, 10))
        
        tk.Label(
            self.hybrid_input_frame,
            text="Snapshot URL or UUID (Online):",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(5, 5))
        
        self.hybrid_url_entry = tk.Entry(self.hybrid_input_frame, width=85)
        self.hybrid_url_entry.pack()
        
        tk.Label(
            self.hybrid_input_frame,
            text="📝 From RTC web: Copy snapshot URL (with id=_xxxxx) or just the UUID",
            bg="#EAF3FB",
            font=("Segoe UI", 8),
            fg="gray"
        ).pack()
        
        tk.Label(
            self.hybrid_input_frame,
            text="Local Root Folder (component folders live here):",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(10, 5))
        
        self.hybrid_folder_entry = tk.Entry(self.hybrid_input_frame, width=85)
        self.hybrid_folder_entry.pack()
        
        tk.Button(
            self.hybrid_input_frame,
            text="Browse",
            bg="#007B3E",
            fg="white",
            command=lambda: self._browse_folder_only(self.hybrid_folder_entry)
        ).pack(pady=5)
        
        tk.Label(
            self.hybrid_input_frame,
            text=(
                "📂 Component name  rb.as.ms.fiatgen.cswpr  →  <root>/rb/as/ms/fiatgen/cswpr\n"
                "   Tool will prompt you to select which snapshot component(s) to compare."
            ),
            bg="#EAF3FB",
            font=("Segoe UI", 8),
            fg="#444444",
            justify="left"
        ).pack(pady=(2, 0))
    
    def create_interface_check_frame(self):
        """Create interface list check input section"""
        self.interface_check_frame = tk.Frame(self.root, bg="#EAF3FB")
        
        # Info label
        tk.Label(
            self.interface_check_frame,
            text="📝 Interface List Check: Analyze interfaces, switches, and dependencies",
            bg="#EAF3FB",
            font=("Segoe UI", 9, "italic"),
            fg="#666666"
        ).pack(pady=(5, 10))
        
        # Analysis mode selection
        self.interface_check_mode = tk.StringVar(value="single")
        
        mode_selector_frame = tk.Frame(self.interface_check_frame, bg="#EAF3FB")
        mode_selector_frame.pack(pady=10)
        
        tk.Label(
            mode_selector_frame,
            text="Analysis Mode:",
            bg="#EAF3FB",
            font=("Segoe UI", 10, "bold")
        ).pack(side="left", padx=5)
        
        tk.Radiobutton(
            mode_selector_frame,
            text="Option 1: Single Workspace",
            variable=self.interface_check_mode,
            value="single",
            bg="#EAF3FB",
            font=("Segoe UI", 10),
            command=self.toggle_interface_check_mode
        ).pack(side="left", padx=10)
        
        tk.Radiobutton(
            mode_selector_frame,
            text="Option 2: Compare Two Workspaces",
            variable=self.interface_check_mode,
            value="compare",
            bg="#EAF3FB",
            font=("Segoe UI", 10),
            command=self.toggle_interface_check_mode
        ).pack(side="left", padx=10)
        
        # Single workspace input
        self.single_workspace_frame = tk.Frame(self.interface_check_frame, bg="#EAF3FB")
        
        tk.Label(
            self.single_workspace_frame,
            text="Workspace Path (Platform or Project):",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(5, 5))
        
        self.single_workspace_entry = tk.Entry(self.single_workspace_frame, width=85)
        self.single_workspace_entry.pack()
        
        tk.Button(
            self.single_workspace_frame,
            text="Browse Folder",
            bg="#007B3E",
            fg="white",
            command=lambda: self.browse_folder(self.single_workspace_entry)
        ).pack(pady=5)
        
        # Dual workspace input (hidden by default)
        self.dual_workspace_frame = tk.Frame(self.interface_check_frame, bg="#EAF3FB")
        
        tk.Label(
            self.dual_workspace_frame,
            text="Platform Workspace Path:",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(5, 5))
        
        self.platform_workspace_entry = tk.Entry(self.dual_workspace_frame, width=85)
        self.platform_workspace_entry.pack()
        
        tk.Button(
            self.dual_workspace_frame,
            text="Browse Platform Folder",
            bg="#007B3E",
            fg="white",
            command=lambda: self.browse_folder(self.platform_workspace_entry)
        ).pack(pady=5)
        
        tk.Label(
            self.dual_workspace_frame,
            text="Project Workspace Path:",
            bg="#EAF3FB",
            font=("Segoe UI", 11)
        ).pack(pady=(10, 5))
        
        self.project_workspace_entry = tk.Entry(self.dual_workspace_frame, width=85)
        self.project_workspace_entry.pack()
        
        tk.Button(
            self.dual_workspace_frame,
            text="Browse Project Folder",
            bg="#007B3E",
            fg="white",
            command=lambda: self.browse_folder(self.project_workspace_entry)
        ).pack(pady=5)
        
        # Initially show single workspace mode
        self.toggle_interface_check_mode()
    
    def toggle_interface_check_mode(self):
        """Toggle between single and dual workspace analysis"""
        mode = self.interface_check_mode.get()
        
        # Hide both frames first
        self.single_workspace_frame.pack_forget()
        self.dual_workspace_frame.pack_forget()
        
        # Show the appropriate frame
        if mode == "single":
            self.single_workspace_frame.pack(fill="x", pady=10)
        else:
            self.dual_workspace_frame.pack(fill="x", pady=10)
    
    def create_rtc_section(self):
        """Create RTC integration section"""
        self.rtc_frame = tk.Frame(self.root, bg="#EAF3FB")
        self.rtc_frame.pack(pady=10)
        
        self.rtc_enabled_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            self.rtc_frame,
            text="Enable RTC/ALM WorkItem Integration",
            variable=self.rtc_enabled_var,
            bg="#EAF3FB",
            font=("Segoe UI", 10)
        ).grid(row=0, column=0, sticky="w", padx=5)
    
    def create_action_buttons(self):
        """Create action buttons"""
        button_frame = tk.Frame(self.root, bg="#EAF3FB")
        button_frame.pack(pady=20)
        
        # Main comparison button
        self.compare_btn = tk.Button(
            button_frame,
            text="Start Comparison",
            bg="#003366",
            fg="white",
            font=("Segoe UI", 12, "bold"),
            width=20,
            height=2,
            command=self.start_comparison
        )
        self.compare_btn.pack(side="left", padx=10)
        
        # Platform Dependency Analysis button
        self.platform_dep_btn = tk.Button(
            button_frame,
            text="🔍 Platform Dependency\n     Analysis",
            bg="#0066CC",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            width=18,
            height=2,
            command=self.open_platform_dependency_analysis
        )
        self.platform_dep_btn.pack(side="left", padx=10)
    
    def open_platform_dependency_analysis(self):
        """Open Platform Dependency Analysis tool for single folder analysis"""
        from src.gui.platform_dependency_viewer import PlatformDependencyViewer
        PlatformDependencyViewer(self.root)
    
    def create_progress_section(self):
        """Create progress display section"""
        self.progress_frame = tk.Frame(self.root, bg="#EAF3FB")
        self.progress_frame.pack(fill="both", expand=True, pady=10)
        
        self.progress_label = tk.Label(
            self.progress_frame,
            text="Ready to start comparison...",
            bg="#EAF3FB",
            font=("Segoe UI", 10)
        )
        self.progress_label.pack(pady=5)
        
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            length=600,
            mode='determinate'
        )
        self.progress_bar.pack(pady=5)
    
    def toggle_input_mode(self):
        """Toggle between different input modes"""
        mode = self.comparison_mode.get()
        
        # Hide all frames first
        self.folder_input_frame.pack_forget()
        self.snapshot_input_frame.pack_forget()
        self.hybrid_input_frame.pack_forget()
        self.interface_check_frame.pack_forget()
        
        # Show the appropriate frame
        if mode == "offline_offline":
            self.folder_input_frame.pack(fill="x", before=self.rtc_frame)
        elif mode == "online_online":
            self.snapshot_input_frame.pack(fill="x", before=self.rtc_frame)
        elif mode == "online_offline":
            self.hybrid_input_frame.pack(fill="x", before=self.rtc_frame)
        elif mode == "interface_check":
            self.interface_check_frame.pack(fill="x", before=self.rtc_frame)
    
    def _browse_folder_only(self, entry_field):
        """Browse for a folder only (no ZIP option, for root folder selection)."""
        folder_selected = filedialog.askdirectory(title="Select Local Root Folder")
        if folder_selected:
            entry_field.delete(0, tk.END)
            entry_field.insert(0, folder_selected)

    @staticmethod
    def _resolve_local_folder_for_component(root_folder, component_name):
        """
        Resolve a local folder path for a component by splitting on '.'.

        Example:
            root = C:/work/myproject
            component = rb.as.ms.fiatgen.cswpr
            result  = C:/work/myproject/rb/as/ms/fiatgen/cswpr

        Returns the resolved path string (regardless of whether it exists on disk).
        """
        parts = component_name.split('.')
        path = root_folder
        for part in parts:
            path = os.path.join(path, part)
        return path

    def browse_folder(self, entry_field):
        """Browse for folder or ZIP file"""
        choice = messagebox.askyesnocancel(
            "Select Input Type",
            "Do you want to select a FOLDER?\n\nYes = Select Folder\nNo = Select ZIP File\nCancel = Abort"
        )
        
        if choice is None:  # Cancel
            return
        elif choice:  # Yes - Select Folder
            folder_selected = filedialog.askdirectory()
            if folder_selected:
                entry_field.delete(0, tk.END)
                entry_field.insert(0, folder_selected)
        else:  # No - Select ZIP File
            zip_selected = filedialog.askopenfilename(
                title="Select ZIP File",
                filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
            )
            if zip_selected:
                entry_field.delete(0, tk.END)
                entry_field.insert(0, zip_selected)
    
    def start_comparison(self):
        """Start the comparison process"""
        mode = self.comparison_mode.get()
        
        if mode == "offline_offline":
            # Offline → Offline: Folders/ZIPs
            folder1 = self.folder1_entry.get().strip()
            folder2 = self.folder2_entry.get().strip()
            
            # Validate inputs
            if not folder1 or not folder2:
                messagebox.showerror(
                    "Missing Input",
                    "Please select both folders to compare."
                )
                return
            
            # Check if paths exist
            if not (os.path.exists(folder1) and os.path.exists(folder2)):
                messagebox.showerror(
                    "Invalid Path",
                    "One or both folder paths do not exist."
                )
                return
            
            # Show progress and start comparison in background thread
            self.run_folder_comparison(
                folder1, folder2,
                self.component_name.get(),
                genmake_filter=self._get_active_genmake_filter(),
            )
        
        elif mode == "online_online":
            # Online → Online: RTC Snapshots/URLs
            url1 = self.snapshot1_entry.get().strip()
            url2 = self.snapshot2_entry.get().strip()
            
            if not url1 or not url2:
                messagebox.showerror(
                    "Missing Input",
                    "Please enter both RTC snapshot URLs or UUIDs."
                )
                return
            
            # Check if different
            rtc_temp = RTCConnection(self.rtc_server_url.get(), "", "")
            uuid1 = rtc_temp.extract_snapshot_uuid(url1)
            uuid2 = rtc_temp.extract_snapshot_uuid(url2)
            
            if uuid1 == uuid2:
                messagebox.showerror(
                    "Invalid Input",
                    "⚠️ Both snapshot URLs point to the SAME snapshot!\n\n"
                    "Please select two different snapshots for comparison."
                )
                return
            
            # Request credentials if not already provided
            if not self.rtc_username.get() or not self.rtc_password.get():
                self.show_credential_dialog()
            else:
                # Start comparison with existing credentials
                self.start_snapshot_comparison(
                    url1, url2,
                    genmake_filter=self._get_active_genmake_filter(),
                )
        
        elif mode == "online_offline":
            # Online → Offline: RTC Snapshot + Local Root Folder
            url = self.hybrid_url_entry.get().strip()
            folder = self.hybrid_folder_entry.get().strip()
            
            if not url or not folder:
                messagebox.showerror(
                    "Missing Input",
                    "Please enter RTC snapshot URL/UUID and select a local root folder."
                )
                return
            
            if not os.path.isdir(folder):
                messagebox.showerror(
                    "Invalid Path",
                    "The local root folder path does not exist or is not a directory.\n\n"
                    "Please select a folder that contains the component sub-directories."
                )
                return
            
            # Check if RTC credentials are available
            if not self.rtc_username.get() or not self.rtc_password.get():
                # Try loading cached credentials first
                if not self.cached_credentials_loaded:
                    self._load_cached_credentials()
                
                # If still no credentials, show login dialog
                if not self.rtc_username.get() or not self.rtc_password.get():
                    self.show_signin_dialog(lambda: self.start_hybrid_comparison(url, folder))
                else:
                    # Credentials loaded from cache, proceed
                    self.start_hybrid_comparison(url, folder)
            else:
                # Credentials already available, proceed
                self.start_hybrid_comparison(url, folder)
        
        elif mode == "interface_check":
            # Interface List Check Mode
            check_mode = self.interface_check_mode.get()
            
            if check_mode == "single":
                # Single workspace analysis
                workspace_path = self.single_workspace_entry.get().strip()
                
                if not workspace_path:
                    messagebox.showerror(
                        "Missing Input",
                        "Please select a workspace path to analyze."
                    )
                    return
                
                if not os.path.exists(workspace_path):
                    messagebox.showerror(
                        "Invalid Path",
                        "The workspace path does not exist."
                    )
                    return
                
                # Start single workspace analysis
                self.run_interface_analysis_single(workspace_path)
            
            else:
                # Dual workspace comparison
                platform_path = self.platform_workspace_entry.get().strip()
                project_path = self.project_workspace_entry.get().strip()
                
                if not platform_path or not project_path:
                    messagebox.showerror(
                        "Missing Input",
                        "Please select both platform and project workspace paths."
                    )
                    return
                
                if not os.path.exists(platform_path) or not os.path.exists(project_path):
                    messagebox.showerror(
                        "Invalid Path",
                        "One or both workspace paths do not exist."
                    )
                    return
                
                # Start dual workspace comparison
                self.run_interface_analysis_compare(platform_path, project_path)
    
    def run_interface_analysis_single(self, workspace_path):
        """Run interface analysis on a single workspace"""
        self.compare_btn.config(state='disabled')
        self.progress_bar['value'] = 0
        self.progress_label.config(text="Starting interface analysis...")
        self.root.update()
        
        def analysis_thread():
            try:
                from src.utils.interface_list_analyzer import InterfaceListAnalyzer
                from src.utils.interface_list_reporter import InterfaceListReportGenerator
                
                # Update progress
                self.root.after(0, lambda: self.progress_label.config(
                    text="Analyzing workspace files..."
                ))
                self.root.after(0, lambda: self.progress_bar.config(value=20))
                
                # Run analysis
                analyzer = InterfaceListAnalyzer()
                analysis = analyzer.analyze_workspace(workspace_path)
                
                self.root.after(0, lambda: self.progress_bar.config(value=60))
                self.root.after(0, lambda: self.progress_label.config(
                    text="Generating reports..."
                ))
                
                # Generate reports
                workspace_name = os.path.basename(workspace_path)
                reporter = InterfaceListReportGenerator()
                html_path, excel_path = reporter.generate_single_workspace_report(
                    analysis, workspace_name
                )
                
                self.root.after(0, lambda: self.progress_bar.config(value=100))
                self.root.after(0, lambda: self.progress_label.config(
                    text="✓ Analysis complete!"
                ))
                
                # Show results
                self.root.after(0, lambda: self.show_interface_analysis_results(
                    html_path, excel_path, analysis
                ))
                
            except Exception as e:
                logger.error(f"Interface analysis failed: {e}", exc_info=True)
                self.root.after(0, lambda: messagebox.showerror(
                    "Analysis Failed",
                    f"An error occurred during analysis:\n{str(e)}"
                ))
            finally:
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
        
        thread = threading.Thread(target=analysis_thread, daemon=True)
        thread.start()
    
    def run_interface_analysis_compare(self, platform_path, project_path):
        """Run interface analysis comparing two workspaces"""
        self.compare_btn.config(state='disabled')
        self.progress_bar['value'] = 0
        self.progress_label.config(text="Starting workspace comparison...")
        self.root.update()
        
        def comparison_thread():
            try:
                from src.utils.interface_list_analyzer import InterfaceListAnalyzer
                from src.utils.interface_list_reporter import InterfaceListReportGenerator
                
                # Update progress
                self.root.after(0, lambda: self.progress_label.config(
                    text="Analyzing platform workspace..."
                ))
                self.root.after(0, lambda: self.progress_bar.config(value=15))
                
                # Run comparison
                analyzer = InterfaceListAnalyzer()
                comparison_data = analyzer.compare_workspaces(platform_path, project_path)
                
                self.root.after(0, lambda: self.progress_bar.config(value=60))
                self.root.after(0, lambda: self.progress_label.config(
                    text="Generating comparison reports..."
                ))
                
                # Generate reports
                platform_name = os.path.basename(platform_path)
                project_name = os.path.basename(project_path)
                reporter = InterfaceListReportGenerator()
                html_path, excel_path = reporter.generate_comparison_report(
                    comparison_data, platform_name, project_name
                )
                
                self.root.after(0, lambda: self.progress_bar.config(value=100))
                self.root.after(0, lambda: self.progress_label.config(
                    text="✓ Comparison complete!"
                ))
                
                # Show results
                self.root.after(0, lambda: self.show_interface_comparison_results(
                    html_path, excel_path, comparison_data
                ))
                
            except Exception as e:
                logger.error(f"Interface comparison failed: {e}", exc_info=True)
                self.root.after(0, lambda: messagebox.showerror(
                    "Comparison Failed",
                    f"An error occurred during comparison:\n{str(e)}"
                ))
            finally:
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
        
        thread = threading.Thread(target=comparison_thread, daemon=True)
        thread.start()
    
    def show_interface_analysis_results(self, html_path, excel_path, analysis):
        """Show interface analysis results dialog"""
        summary = f"""Interface Analysis Complete!

Workspace: {analysis.workspace_path}

Results:
\u2022 Files Analyzed: {analysis.analyzed_files} / {analysis.total_files}
\u2022 Interfaces Found: {len(analysis.interfaces)}
\u2022 Switches Found: {len(analysis.switches)}
\u2022 Dependencies Tracked: {len(analysis.dependencies)}
\u2022 Data Types Used: {len(analysis.data_types_used)}

Reports generated:
\U0001f4c4 HTML: {html_path}
\U0001f4ca Excel: {excel_path}
"""
        
        result = messagebox.askquestion(
            "Analysis Complete",
            summary + "\n\nDo you want to open the HTML report now?"
        )
        
        if result == 'yes':
            import webbrowser
            webbrowser.open(html_path)
    
    def show_interface_comparison_results(self, html_path, excel_path, comparison_data):
        """Show interface comparison results dialog"""
        platform = comparison_data['platform']
        project = comparison_data['project']
        differences = comparison_data['differences']
        
        summary = f"""Workspace Comparison Complete!

Platform: {platform.workspace_path}
   \u2022 Files: {platform.analyzed_files}
   \u2022 Interfaces: {len(platform.interfaces)}

Project: {project.workspace_path}
   \u2022 Files: {project.analyzed_files}
   \u2022 Interfaces: {len(project.interfaces)}

Differences Found:
   \u2022 Only in Platform: {len(differences['interfaces']['only_in_platform'])}
   \u2022 Only in Project: {len(differences['interfaces']['only_in_project'])}
   \u2022 Modified: {len(differences['interfaces']['modified'])}
   \u2022 Switch Changes: {len(differences['switches']['status_changed'])}

Reports generated:
\U0001f4c4 HTML: {html_path}
\U0001f4ca Excel: {excel_path}
"""
        
        result = messagebox.askquestion(
            "Comparison Complete",
            summary + "\n\nDo you want to open the HTML report now?"
        )
        
        if result == 'yes':
            import webbrowser
            webbrowser.open(html_path)
    
    def show_signin_dialog(self, callback=None):
        """Show dialog to input RTC credentials with optional callback
        
        Args:
            callback: Function to call after successful login (e.g., lambda: self.start_hybrid_comparison(url, folder))
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("RTC Login")
        dialog.geometry("480x360")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Header
        header = tk.Frame(dialog, bg='#003366', height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(
            header,
            text="🔐 RTC Login",
            font=('Segoe UI', 14, 'bold'),
            bg='#003366',
            fg='white'
        ).pack(pady=15)
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='white')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        tk.Label(
            main_frame,
            text="Enter RTC credentials:",
            font=('Segoe UI', 10),
            bg='white'
        ).pack(anchor=tk.W, pady=(0, 10))
        
        # Username
        tk.Label(main_frame, text="Username:", font=('Segoe UI', 10), bg='white').pack(anchor=tk.W)
        username_entry = ttk.Entry(main_frame, textvariable=self.rtc_username, width=40)
        username_entry.pack(fill=tk.X, pady=(0, 10))
        
        # Password
        tk.Label(main_frame, text="Password:", font=('Segoe UI', 10), bg='white').pack(anchor=tk.W)
        password_entry = ttk.Entry(main_frame, textvariable=self.rtc_password, show='●', width=40)
        password_entry.pack(fill=tk.X, pady=(0, 10))
        
        keep_signed_in_check = tk.Checkbutton(
            main_frame,
            text="Keep me signed in",
            variable=self.keep_signed_in_var,
            font=('Segoe UI', 10),
            bg='white',
            activebackground='white'
        )
        keep_signed_in_check.pack(anchor=tk.W, pady=(0, 6))
        
        cache_note = "Credentials are stored in Windows Credential Manager or encrypted local storage."
        tk.Label(
            main_frame,
            text=cache_note,
            font=('Segoe UI', 8),
            bg='white',
            fg='#666666',
            wraplength=420,
            justify=tk.LEFT
        ).pack(anchor=tk.W, pady=(0, 8))
        
        def on_ok():
            username = self.rtc_username.get().strip()
            password = self.rtc_password.get().strip()
            
            if not username or not password:
                messagebox.showerror('Error', 'Please enter both username and password.')
                return
                        
            dialog.destroy()
            
            # If callback provided (for hybrid mode), use it
            # Otherwise, use default Online→Online behavior
            if callback:
                callback()
            else:
                # Get snapshot URLs for Online→Online mode
                url1 = self.snapshot1_entry.get().strip()
                url2 = self.snapshot2_entry.get().strip()
                # Start comparison
                self.start_snapshot_comparison(
                    url1, url2,
                    genmake_filter=self._get_active_genmake_filter(),
                )
        
        def on_cancel():
            dialog.destroy()
        
        # Buttons
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(pady=(10, 0))
        
        tk.Button(
            button_frame,
            text="✓ Login",
            command=on_ok,
            font=('Segoe UI', 10, 'bold'),
            bg='#007B3E',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            button_frame,
            text="✕ Cancel",
            command=on_cancel,
            font=('Segoe UI', 10, 'bold'),
            bg='#666666',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            button_frame,
            text="Clear Cache",
            command=lambda: self.on_logout(show_message=True),
            font=('Segoe UI', 10, 'bold'),
            bg='#B00020',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        if self.rtc_username.get():
            password_entry.focus()
        else:
            username_entry.focus()
        dialog.bind('<Return>', lambda e: on_ok())
    
    def show_credential_dialog(self):
        """Legacy method - redirects to show_signin_dialog for backward compatibility"""
        self.show_signin_dialog()
        dialog.bind('<Escape>', lambda e: on_cancel())
    
    def _load_cached_credentials(self):
        """Load cached RTC credentials for online-online comparisons."""
        server_url = self.rtc_server_url.get()
        cached_credentials = CredentialManager.load_credentials(server_url)
        
        if cached_credentials:
            username, password = cached_credentials
            self.rtc_username.set(username)
            self.rtc_password.set(password)
            self.keep_signed_in_var.set(True)
            self.cached_credentials_loaded = True
            logger.info("Cached RTC credentials loaded for online-online mode")
        else:
            self.cached_credentials_loaded = False
        
        self._refresh_auth_status()
    
    def _refresh_auth_status(self):
        """Refresh the main window RTC auth display."""
        if not hasattr(self, 'auth_status_label') or not hasattr(self, 'logout_btn'):
            return
        
        if self.rtc_username.get() and self.rtc_password.get():
            username = self.rtc_username.get()
            label_text = f"RTC: {username}"
            self.auth_status_label.config(text=label_text)
            self.logout_btn.config(state="normal", bg="#B00020")
        else:
            self.auth_status_label.config(text="RTC: signed out")
            self.logout_btn.config(state="disabled", bg="#666666")
    
    def _update_cached_credentials_after_login(self, username, password, server_url):
        """Persist or clear RTC credentials based on the login checkbox."""
        # CRITICAL FIX: Always check current checkbox value at time of credential save
        # This ensures credentials are saved even if start_snapshot_comparison wasn't called yet
        should_save = self.keep_signed_in_var.get()
        
        if should_save:
            saved = CredentialManager.save_credentials(username, password, server_url)
            if saved:
                self.cached_credentials_loaded = True
                logger.info("✓ RTC credentials cached after successful login")
            else:
                self.cached_credentials_loaded = False
                logger.warning("✗ Failed to save credentials securely")
                self.root.after(
                    0,
                    lambda: messagebox.showwarning(
                        "Credential Cache Warning",
                        "Login succeeded, but credentials could not be saved securely.\n\n" +
                        "Possible causes:\n" +
                        "• keyring library not installed (pip install keyring)\n" +
                        "• cryptography library not installed (pip install cryptography)\n" +
                        "• File permission issues\n\n" +
                        "You will need to re-enter credentials on next login."
                    )
                )
        else:
            # User unchecked "Keep me signed in" - clear any existing cached credentials
            CredentialManager.clear_credentials(server_url)
            self.cached_credentials_loaded = False
            logger.info("Credentials not saved (Keep me signed in unchecked)")
        
        self.root.after(0, self._refresh_auth_status)
    
    def on_logout(self, show_message=True):
        """Clear online RTC login state and cached credentials."""
        server_url = self.rtc_server_url.get()
        CredentialManager.clear_credentials(server_url)
        self.rtc_username.set("")
        self.rtc_password.set("")
        self.keep_signed_in_var.set(False)
        self.cached_credentials_loaded = False
        self.rtc_connection = None
        self._refresh_auth_status()
        
        if show_message:
            messagebox.showinfo(
                "Logged Out",
                "RTC credentials have been cleared from this session and local secure storage."
            )
    
    def start_snapshot_comparison(self, url1, url2, genmake_filter=None):
        """Start RTC snapshot comparison in background thread"""
        # Disable button during processing
        self.save_credentials_on_success = self.keep_signed_in_var.get()
        self.compare_btn.config(state='disabled')
        self.progress_bar['value'] = 0
        self.progress_label.config(text="🌐 Connecting to RTC server...")
        self.root.update()

        # Run in background thread
        thread = threading.Thread(
            target=self._snapshot_comparison_thread,
            args=(url1, url2),
            kwargs={'genmake_filter': genmake_filter},
            daemon=True
        )
        thread.start()
    
    def _snapshot_comparison_thread(self, url1, url2, genmake_filter=None):
        """
        Background thread for snapshot comparison with comprehensive error handling

        Steps:
        1. Validate inputs
        2. Test RTC connection
        3. Fetch snapshot 1 components
        4. Fetch snapshot 2 components
        5. Compare snapshots (component-level)
        6. Optionally fetch file-level details
        7. Display results
        """
        try:
            logger.info("=" * 80)
            logger.info("SNAPSHOT COMPARISON STARTED")
            logger.info("=" * 80)
            
            # Get credentials
            username = self.rtc_username.get()
            password = self.rtc_password.get()
            server_url = self.rtc_server_url.get()
            
            # Validate credentials
            self.root.after(0, lambda: self._update_progress(5, "Validating credentials..."))
            if not username or not password:
                error_msg = "Username and password are required"
                logger.error(f"Validation failed: {error_msg}")
                self.root.after(0, lambda: messagebox.showerror('Validation Error', error_msg))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return
            
            logger.info(f"Server: {server_url}")
            logger.info(f"Username: {username}")
            
            # Test connection (with retries)
            self.root.after(0, lambda: self._update_progress(8, "🔗 Testing RTC connection (with retries)..."))
            rtc_conn = None
            error_msg = None
            
            for attempt in range(1, 4):  # 3 attempts
                logger.info(f"Connection attempt {attempt}/3...")
                rtc_conn, error_msg = get_rtc_connection(username, password, server_url)
                
                if rtc_conn:
                    logger.info("✓ RTC connection successful")
                    break
                
                if attempt < 3:
                    logger.warning(f"Attempt {attempt} failed: {error_msg}, retrying...")
                    import time
                    time.sleep(2)  # Wait 2 seconds before retry
            
            if not rtc_conn:
                detailed_error = self._format_connection_error(error_msg)
                logger.error(f"Connection failed after 3 attempts: {error_msg}")
                self.root.after(0, lambda: messagebox.showerror('Connection Error', detailed_error))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return
            
            self._update_cached_credentials_after_login(username, password, server_url)
            
            # Extract and validate UUIDs
            self.root.after(0, lambda: self._update_progress(10, "Validating snapshot URLs..."))
            logger.info("Extracting snapshot UUIDs...")
            
            uuid1 = rtc_conn.extract_snapshot_uuid(url1)
            uuid2 = rtc_conn.extract_snapshot_uuid(url2)
            
            logger.info(f"Snapshot 1 UUID: {uuid1}")
            logger.info(f"Snapshot 2 UUID: {uuid2}")
            
            # Validate UUIDs
            if not uuid1 or not uuid2:
                error_msg = "Could not extract snapshot UUIDs. Please verify the URLs/UUIDs are correct."
                logger.error(error_msg)
                self.root.after(0, lambda: messagebox.showerror('Invalid Input', error_msg))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return
            
            if uuid1 == uuid2:
                error_msg = "⚠️ Both snapshot UUIDs are identical!\n\nPlease select two different snapshots."
                logger.error(f"UUID validation error: {error_msg}")
                self.root.after(0, lambda: messagebox.showerror('Duplicate Snapshot', error_msg))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return
            
            # Fetch snapshots sequentially: Snapshot 1 first, then Snapshot 2.
            # Fetching in parallel caused Snapshot 1 to time out because Snapshot 2
            # immediately spawns 20 workers for its 422 baseline detail requests,
            # saturating the server before Snapshot 1's initial request completes.
            self.root.after(0, lambda: self._update_progress(20, "⬇️ Fetching Snapshot 1 components..."))
            logger.info("Fetching Snapshot 1 and Snapshot 2 components sequentially...")

            fetched_components = {}
            snapshot_fetches = {
                "Snapshot 1": uuid1,
                "Snapshot 2": uuid2
            }

            # Progress tracking for both snapshots
            import threading
            progress_lock = threading.Lock()
            snapshot_progress = {
                "Snapshot 1": {"current": 0, "total": 0, "message": "Starting..."},
                "Snapshot 2": {"current": 0, "total": 0, "message": "Starting..."}
            }
            
            def create_progress_callback(snap_name):
                """Create a progress callback for a specific snapshot"""
                def callback(current, total, message):
                    with progress_lock:
                        snapshot_progress[snap_name] = {
                            "current": current,
                            "total": total,
                            "message": message
                        }
                        # Calculate overall progress
                        snap1_prog = snapshot_progress["Snapshot 1"]
                        snap2_prog = snapshot_progress["Snapshot 2"]
                        
                        # Calculate percentage for each snapshot
                        snap1_pct = (snap1_prog["current"] / snap1_prog["total"] * 100) if snap1_prog["total"] > 0 else 0
                        snap2_pct = (snap2_prog["current"] / snap2_prog["total"] * 100) if snap2_prog["total"] > 0 else 0
                        
                        # Overall progress (20-50 range for fetching)
                        overall_pct = 20 + ((snap1_pct + snap2_pct) / 2) * 0.3  # 30% of progress bar for fetching
                        
                        # Create combined status message
                        status_msg = f"⬇️ Snap1: {snap1_prog['current']}/{snap1_prog['total']} | Snap2: {snap2_prog['current']}/{snap2_prog['total']}"
                        
                        # Update GUI from main thread
                        self.root.after(0, lambda p=overall_pct, m=status_msg: self._update_progress(p, m))
                return callback

            with ThreadPoolExecutor(max_workers=2) as executor:
                future_to_name = {
                    executor.submit(
                        rtc_conn.fetch_snapshot_components,
                        snapshot_uuid,
                        snapshot_name=snapshot_name,
                        progress_callback=create_progress_callback(snapshot_name)
                    ): snapshot_name
                    for snapshot_name, snapshot_uuid in snapshot_fetches.items()
                }

                for future in as_completed(future_to_name):
                    snapshot_name = future_to_name[future]
                    try:
                        result = future.result()
                        # Handle dict return format with name and components
                        if isinstance(result, dict):
                            components = result.get('components', [])
                            snap_name = result.get('name')
                        else:
                            # Fallback for old format
                            components = result if isinstance(result, list) else []
                            snap_name = None
                    except Exception as fetch_err:
                        logger.error(f"{snapshot_name}: fetch failed: {fetch_err}", exc_info=True)
                        components = []
                        snap_name = None

                    fetched_components[snapshot_name] = {
                        'components': components,
                        'name': snap_name
                    }
                    completed = len(fetched_components)
                    progress = 20 + (completed * 15)
                    self.root.after(
                        0,
                        lambda name=snapshot_name, count=len(components), value=progress:
                            self._update_progress(value, f"✓ {name}: fetched {count} components")
                    )

            snap1_components = fetched_components.get("Snapshot 1", {}).get('components', [])
            snap2_components = fetched_components.get("Snapshot 2", {}).get('components', [])
            snap1_actual_name = fetched_components.get("Snapshot 1", {}).get('name')
            snap2_actual_name = fetched_components.get("Snapshot 2", {}).get('name')

            if not snap1_components:
                error_msg = "❌ No components found in Snapshot 1\n\nPossible reasons:\n" \
                           "• Invalid snapshot UUID\n" \
                           "• Snapshot does not exist\n" \
                           "• Permission denied\n" \
                           "• Network connectivity issue"
                logger.error("Failed to fetch components from Snapshot 1")
                self.root.after(0, lambda: messagebox.showerror('Snapshot 1 Error', error_msg))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            if not snap2_components:
                error_msg = "❌ No components found in Snapshot 2\n\nPossible reasons:\n" \
                           "• Invalid snapshot UUID\n" \
                           "• Snapshot does not exist\n" \
                           "• Permission denied\n" \
                           "• Network connectivity issue"
                logger.error("Failed to fetch components from Snapshot 2")
                self.root.after(0, lambda: messagebox.showerror('Snapshot 2 Error', error_msg))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return
            
            # ===== COMPONENT SELECTION (Online Mode Specific) =====
            # Show component selection dialog to user (from main thread!)
            self.root.after(0, lambda: self._update_progress(60, "📋 Opening component selection dialog..."))
            logger.info("Showing component selection dialog for online mode...")
            
            from src.gui.dialogs import show_component_selection_dialog
            import threading
            
            # Use a threading event to coordinate dialog display
            dialog_event = threading.Event()
            selection_result = {'canceled': True, 'selected': []}
            
            def show_dialog_in_main_thread():
                nonlocal selection_result
                try:
                    selection_result = show_component_selection_dialog(snap1_components, snap2_components)
                    logger.info(f"Component selection result: {len(selection_result.get('selected', []))} components selected")
                except Exception as e:
                    logger.error(f"Error showing component selection dialog: {e}", exc_info=True)
                    selection_result = {'canceled': True, 'selected': []}
                finally:
                    dialog_event.set()  # Signal that dialog is done
            
            # Schedule dialog from main thread to ensure proper display
            self.root.after(0, show_dialog_in_main_thread)
            
            # Wait for dialog to complete (with timeout)
            if not dialog_event.wait(timeout=300):  # 5 minute timeout
                logger.warning("Component selection dialog timed out")
                self.root.after(0, lambda: messagebox.showwarning('Timeout', 'Component selection timed out.'))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return
            
            if selection_result['canceled']:
                logger.info("⚠️ User canceled component selection")
                self.root.after(0, lambda: messagebox.showinfo('Cancelled', 'Component selection cancelled. Comparison aborted.'))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return
            
            selected_comp_names = set(selection_result['selected'])
            component_mappings  = selection_result.get('component_mappings', {})  # {snap1_name: snap2_name}
            logger.info(f"✓ User selected {len(selected_comp_names)}/{len(selection_result['common'])} common components")
            if component_mappings:
                logger.info(f"✓ User mapped {len(component_mappings)} cross-snapshot component pair(s):")
                for s1n, s2n in component_mappings.items():
                    logger.info(f"    {s1n}  ↔  {s2n}")
            
            # Filter components to only selected ones
            if not selected_comp_names and not component_mappings:
                logger.info("No components selected. Comparison aborted.")
                self.root.after(0, lambda: messagebox.showwarning(
                    'No Components Selected',
                    'Please select at least one common component to compare or map at least one pair.'
                ))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            snap1_filtered = [c for c in snap1_components if c.get('name', str(c)) in selected_comp_names]
            snap2_filtered = [c for c in snap2_components if c.get('name', str(c)) in selected_comp_names]

            # ── Inject mapped pairs ────────────────────────────────────────
            # For each (snap1_name → snap2_name) user mapping, add the snap1 component
            # as-is, and add a renamed copy of the snap2 component so that
            # compare_snapshots pairs them by name naturally.
            if component_mappings:
                snap1_by_name = {c.get('name', str(c)): c for c in snap1_components}
                snap2_by_name = {c.get('name', str(c)): c for c in snap2_components}
                for s1_name, s2_name in component_mappings.items():
                    s1_comp = snap1_by_name.get(s1_name)
                    s2_comp = snap2_by_name.get(s2_name)
                    if s1_comp and s2_comp:
                        snap1_filtered.append(s1_comp)
                        # Rename the snap2 component so compare_snapshots pairs it with s1_name
                        renamed_s2 = dict(s2_comp)
                        renamed_s2['name'] = s1_name
                        renamed_s2['_original_name'] = s2_name   # preserved for reference
                        snap2_filtered.append(renamed_s2)
                        logger.info(f"  Injected mapped pair: {s1_name!r} (Snap1) ↔ {s2_name!r} (Snap2)")
                    else:
                        logger.warning(f"  Skipping mapped pair {s1_name!r}/{s2_name!r}: component not found")

            logger.info(f"Filtered: {len(snap1_filtered)} from Snapshot 1, {len(snap2_filtered)} from Snapshot 2")
            logger.info(f"Excluded: {len(selection_result['only_in_snap1'])} only in Snap1, "
                        f"{len(selection_result['only_in_snap2'])} only in Snap2, "
                        f"{len(component_mappings)} cross-mapped pair(s)")
            
            # Compare snapshots (only selected components) with progress tracking
            self.root.after(0, lambda: self._update_progress(70, f"🔍 Comparing {len(snap1_filtered)} selected components..."))
            logger.info(f"Comparing {len(snap1_filtered)} vs {len(snap2_filtered)} selected components (PARALLEL MODE)...")
            
            # Create progress callback for component comparison
            def comparison_progress_callback(current, total, message):
                # Calculate progress percentage (70-90% range for comparison phase)
                progress_pct = 70 + int((current / total) * 20)
                self.root.after(0, lambda p=progress_pct, m=message: self._update_progress(p, m))
            
            comparison_results = rtc_conn.compare_snapshots(
                snap1_filtered,
                snap2_filtered,
                progress_callback=comparison_progress_callback
            )

            # ── Apply GenMake filter to per-component file details ────────────────
            # Filter the 'details' dict inside each component's file_comparison so
            # only files listed in the CSV are shown in HTML reports and counted.
            #
            # Path mismatch handling
            # ─────────────────────
            # The SCM CLI often returns flat filenames (e.g. 'PostBuild.mk') with no
            # sub-directory prefix, while the CSV stores full workspace-relative paths
            # (e.g. 'rb\as\ms\ESP10E_MFA2\cswpr\cfg\PrePostBuild\PostBuild.mk').
            # A 1-component key will never match a ≥2-component suffix in the index,
            # so we also try prepending the component directory derived from its name
            # (dots → slashes) before concluding there is no match.
            if genmake_filter is not None:
                logger.info("Applying GenMake filter to component file details...")
                _genmake_excluded_total = 0
                for comp in comparison_results:
                    cname = comp.get('name', '')
                    # 'rb.as.ms.ESP10E_MFA2.cswpr' → 'rb/as/ms/ESP10E_MFA2/cswpr'
                    comp_dir = cname.replace('.', '/') if cname else ''

                    fcmp = comp.get('file_comparison')
                    if not fcmp or 'details' not in fcmp:
                        continue
                    original_details = fcmp['details']
                    filtered_details = {}
                    for fp, st in original_details.items():
                        # 1. Direct suffix-index match (path already has ≥2 components)
                        if genmake_filter.matches(fp):
                            filtered_details[fp] = st
                            continue
                        # 2. Flat-filename fallback: prefix with component dir and retry
                        #    e.g. 'PostBuild.mk' → 'rb/as/ms/ESP10E_MFA2/cswpr/PostBuild.mk'
                        if comp_dir:
                            if genmake_filter.matches(comp_dir + '/' + fp):
                                filtered_details[fp] = st
                                continue

                    excluded = len(original_details) - len(filtered_details)
                    _genmake_excluded_total += excluded
                    fcmp['details'] = filtered_details
                    # Recompute file-count summary from filtered details
                    fcmp['modified']  = sum(1 for s in filtered_details.values() if s == 'modified')
                    fcmp['added']     = sum(1 for s in filtered_details.values() if s == 'added')
                    fcmp['removed']   = sum(1 for s in filtered_details.values() if s == 'removed')
                    fcmp['unchanged'] = sum(1 for s in filtered_details.values() if s == 'unchanged')
                    logger.info(
                        f"  [{cname}] GenMake filter: {len(filtered_details)} kept, "
                        f"{excluded} excluded (comp_dir={comp_dir!r})"
                    )
                logger.info(
                    f"GenMake filter: excluded {_genmake_excluded_total} file entries "
                    f"across {len(comparison_results)} components"
                )

            # Calculate statistics
            different = sum(1 for r in comparison_results if r['status'] == 'Different')
            identical = sum(1 for r in comparison_results if r['status'] == 'Identical')
            added     = sum(1 for r in comparison_results if 'Added'   in r['status'])
            removed   = sum(1 for r in comparison_results if 'Removed' in r['status'])

            logger.info(f"Comparison results:")
            logger.info(f"  Different: {different}")
            logger.info(f"  Identical: {identical}")
            logger.info(f"  Added:     {added}")
            logger.info(f"  Removed:   {removed}")
            logger.info(f"  Total:     {len(comparison_results)}")
            
            # ── Fetch file content for inline HTML diffs ─────────────────────────
            # Strategy (per file):
            #   1. scm get file <baseline_uuid> -b -f <full_repo_path> -o <out>
            #      — mirrors results_viewer._get_baseline_file which is proven to work.
            #      Uses full_path from fi['path'] (e.g. 'src/foo.c'), NOT the
            #      filename-only key used in 'details' (e.g. 'foo.c').
            #   2. scm get file <item_id> <state_id> <out>
            #      — if item-id/state-id are populated in the folder structure.
            self.root.after(0, lambda: self._update_progress(88, "📥 Fetching modified file content for diffs..."))
            logger.info("Fetching file content for modified files...")

            from src.rtc.connection import RTCConnection as _RTCConn

            _BINARY_EXTS = {
                '.xls', '.xlsx', '.zip', '.exe', '.dll', '.so', '.a', '.o',
                '.png', '.jpg', '.jpeg', '.gif', '.bmp', '.pdf', '.doc', '.docx',
                '.ppt', '.pptx', '.bin', '.lib', '.obj', '.jar', '.class', '.pyc',
            }
            MAX_FILES_PER_COMP   = 200
            MAX_TOTAL_FETCH_JOBS = 1000
            TOTAL_BUDGET_SECS    = 900

            # Each task: (cname, fpath_key, full_path, item_id, state_id, baseline_uuid, snap_key)
            #   fpath_key   — key in details dict (just filename in flat SCM structure)
            #   full_path   — fi['path'] = full repo path (e.g. 'src/comp/foo.c')
            #   item_id     — from fi['uuid'] if populated
            #   state_id    — from fi['state-id'] if populated
            #   baseline_uuid — comp's baseline_uuid (for -b -f form)
            fetch_tasks = []
            for comp in comparison_results:
                if comp.get('status') != 'Different':
                    continue
                if len(fetch_tasks) >= MAX_TOTAL_FETCH_JOBS:
                    break

                cname  = comp.get('name', '')
                b1uuid = comp.get('baseline1_uuid', '')
                b2uuid = comp.get('baseline2_uuid', '')
                fcmp   = comp.get('file_comparison') or {}
                details = fcmp.get('details', {})

                struct1 = comp.get('folder_structure1') or {}
                struct2 = comp.get('folder_structure2') or {}
                files1_map = _RTCConn._get_all_files_from_structure(struct1)
                files2_map = _RTCConn._get_all_files_from_structure(struct2)

                # comp_dir for flat-filename fallback (same logic as details filter above)
                comp_dir_ft = cname.replace('.', '/') if cname else ''

                comp_jobs = 0
                for fpath_key, fstatus in sorted(details.items()):
                    if fstatus not in ('modified', 'added', 'removed'):
                        continue
                    if os.path.splitext(fpath_key.lower())[1] in _BINARY_EXTS:
                        continue
                    # GenMake filter — details dict is already filtered above, but guard
                    # here too in case of edge cases.  Use the same component-dir fallback
                    # so flat filenames (e.g. 'PostBuild.mk') aren't incorrectly skipped.
                    if genmake_filter is not None:
                        _passes = genmake_filter.matches(fpath_key)
                        if not _passes and comp_dir_ft:
                            _passes = genmake_filter.matches(comp_dir_ft + '/' + fpath_key)
                        if not _passes:
                            continue
                    if comp_jobs >= MAX_FILES_PER_COMP:
                        break
                    if len(fetch_tasks) >= MAX_TOTAL_FETCH_JOBS:
                        break

                    if fstatus in ('modified', 'removed') and b1uuid:
                        fi1 = files1_map.get(fpath_key, {})
                        full_path1 = fi1.get('path', fpath_key)   # full repo path
                        iid1 = fi1.get('uuid', '')
                        sid1 = fi1.get('state-id', '')
                        fetch_tasks.append((cname, fpath_key, full_path1, iid1, sid1, b1uuid, 'snap1'))

                    if fstatus in ('modified', 'added') and b2uuid:
                        fi2 = files2_map.get(fpath_key, {})
                        full_path2 = fi2.get('path', fpath_key)
                        iid2 = fi2.get('uuid', '')
                        sid2 = fi2.get('state-id', '')
                        fetch_tasks.append((cname, fpath_key, full_path2, iid2, sid2, b2uuid, 'snap2'))

                    comp_jobs += 1

            logger.info(f"Prepared {len(fetch_tasks)} file-content fetch tasks")
            if fetch_tasks:
                sample = fetch_tasks[0]
                logger.info(
                    f"Sample task: cname={sample[0]!r} key={sample[1]!r} "
                    f"full_path={sample[2]!r} item_id={sample[3]!r} "
                    f"state_id={sample[4]!r} baseline={sample[5]!r} snap={sample[6]!r}"
                )
            else:
                for comp in comparison_results:
                    if comp.get('status') == 'Different':
                        struct1 = comp.get('folder_structure1') or {}
                        fi_map = _RTCConn._get_all_files_from_structure(struct1)
                        if fi_map:
                            k = next(iter(fi_map))
                            logger.info(
                                f"Zero tasks — sample fi from {comp.get('name')!r}/{k!r}: {fi_map[k]}"
                            )
                        else:
                            logger.info(f"Zero tasks — folder_structure1 empty for {comp.get('name')!r}")
                        break

            file_contents_by_component = {}

            if fetch_tasks:
                def _fetch_one(task):
                    cname, fpath_key, full_path, item_id, state_id, baseline_uuid, snap_key = task
                    # Strategy 1: baseline UUID + full repo path (proven form)
                    content = rtc_conn.fetch_file_content_from_baseline(
                        baseline_uuid, full_path, cname)
                    # Strategy 2: item-id + state-id (if available and strategy 1 failed)
                    if content is None and item_id and state_id:
                        content = rtc_conn.fetch_file_content_by_item_state(
                            item_id, state_id, fpath_key)
                    return (cname, fpath_key, snap_key, content)

                import time as _time
                deadline = _time.time() + TOTAL_BUDGET_SECS
                total_file_tasks = len(fetch_tasks)
                completed_file_tasks = 0
                fetched_ok = 0

                with ThreadPoolExecutor(max_workers=10) as fetch_ex:
                    futures = {fetch_ex.submit(_fetch_one, t): t for t in fetch_tasks}
                    for fut in as_completed(futures):
                        if _time.time() > deadline:
                            logger.warning("File content fetch budget exceeded — stopping early")
                            break
                        completed_file_tasks += 1
                        try:
                            cname, fpath_key, snap_key, content = fut.result(timeout=90)
                            if content is not None:
                                file_contents_by_component.setdefault(
                                    cname, {}).setdefault(fpath_key, {})[snap_key] = content
                                fetched_ok += 1
                        except Exception as _fe:
                            logger.debug(f"Content fetch error: {_fe}")

                        # Update progress label with live counter every 5 files
                        if completed_file_tasks % 5 == 0 or completed_file_tasks == total_file_tasks:
                            _ct = completed_file_tasks
                            _tt = total_file_tasks
                            _ok = fetched_ok
                            self.root.after(0, lambda c=_ct, t=_tt, o=_ok:
                                self._update_progress(
                                    88,
                                    f"📥 Fetching file content: {c}/{t} fetched ({o} successful)"
                                )
                            )

                fetched_files = sum(len(v) for v in file_contents_by_component.values())
                logger.info(
                    f"✓ File content fetched for {fetched_files} files "
                    f"in {len(file_contents_by_component)} components"
                )

            # ── Fetch baseline info (name, comment, author, timestamp) ──────────
            # Used for the Changeset section in per-component HTML reports.
            self.root.after(0, lambda: self._update_progress(89, "📋 Fetching baseline/changeset metadata..."))
            logger.info("Fetching baseline metadata and changeset info for Different components...")

            baseline_info_cache = {}  # {baseline_uuid: {name, comment, author, timestamp}}
            changeset_by_component = {}  # {comp_name: {'baseline1': info, 'baseline2': info, 'changesets': []}}

            def _fetch_baseline_meta(bid):
                if bid in baseline_info_cache:
                    return baseline_info_cache[bid]
                info = rtc_conn.fetch_baseline_info(bid)
                baseline_info_cache[bid] = info
                return info

            _diff_comps = [c for c in comparison_results if c.get('status') == 'Different']
            _total_cs   = len(_diff_comps)
            _done_cs    = 0

            for comp in _diff_comps:
                cname  = comp.get('name', '')
                b1uuid = comp.get('baseline1_uuid', '')
                b2uuid = comp.get('baseline2_uuid', '')

                b1info = _fetch_baseline_meta(b1uuid) if b1uuid and b1uuid != 'N/A' else {}
                b2info = _fetch_baseline_meta(b2uuid) if b2uuid and b2uuid != 'N/A' else {}

                # Fetch changeset list from baseline2 (the newer one) via SCM CLI
                changesets = rtc_conn.fetch_baseline_changesets_scm(b2uuid, cname) if b2uuid and b2uuid != 'N/A' else []

                changeset_by_component[cname] = {
                    'baseline1': b1info,
                    'baseline2': b2info,
                    'changesets': changesets,
                }
                # If this component was compared via a cross-name mapping, record
                # the original Snapshot 2 component name for display in the report.
                if cname in component_mappings:
                    changeset_by_component[cname]['mapped_snap2_name'] = component_mappings[cname]

                _done_cs += 1
                _d = _done_cs
                _t = _total_cs
                _nc = len(changesets)
                self.root.after(0, lambda d=_d, t=_t, n=_nc, nm=cname:
                    self._update_progress(
                        89,
                        f"📋 Changeset metadata: {d}/{t} components "
                        f"({n} changeset{'s' if n != 1 else ''} for {nm})"
                    )
                )

            logger.info(
                f"✓ Baseline metadata fetched for {len(changeset_by_component)} components"
            )

            # Prepare results for viewer
            self.root.after(0, lambda: self._update_progress(90, "📊 Preparing results viewer..."))
            
            # Enable changeset fetching for modified components
            # IMPORTANT: Changeset fetching requires LSCM/SCM CLI to be installed and configured
            # Set to True to fetch changeset information (adds comprehensive change tracking to reports)
            # Set to False for faster comparisons without changeset details
            enable_changesets = True  # ENABLED: Fetch changeset information for modified files
            
            # Transform snapshot results into results_viewer format (with changeset fetching)
            if enable_changesets:
                logger.info("Changeset fetching ENABLED - fetching changesets for modified files...")
                self.root.after(0, lambda: self._update_progress(87, "🔍 Fetching changesets..."))
            else:
                logger.info("Changeset fetching DISABLED for faster performance")
            viewer_results = self._transform_snapshot_results_for_viewer(
                comparison_results,
                rtc_conn=rtc_conn,
                enable_changesets=enable_changesets
            )
            
            # Create timestamped output directory for this comparison run
            from datetime import datetime as _dt
            timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
            base_results_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                "Snapshot_Comparison_Results"
            )
            output_dir = os.path.join(base_results_dir, f"selected_components_{timestamp}")
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"Created output directory: {output_dir}")

            # Store comparison metadata for viewer
            comparison_metadata = {
                'snap1_url': url1,
                'snap2_url': url2,
                'snap1_name': snap1_actual_name,  # Actual snapshot name from API
                'snap2_name': snap2_actual_name,  # Actual snapshot name from API
                'snap1_components': snap1_filtered,
                'snap2_components': snap2_filtered,
                'selected_components': sorted(selected_comp_names),
                'total_common_components': len(selection_result['common']),
                'comparison_results': comparison_results,
                'file_contents_by_component': file_contents_by_component,
                'changeset_by_component': changeset_by_component,
                'rtc_server_url': server_url,
                'output_dir': output_dir,
            }
            
            # Display results using enhanced viewer
            self.root.after(0, lambda: self._display_snapshot_results_with_viewer(
                viewer_results, comparison_metadata
            ))
            
            self.root.after(0, lambda: self._update_progress(
                100,
                f"✅ Comparison complete: {len(comparison_results)} components analyzed "
                f"({different} different, {identical} identical, {added} added, {removed} removed)"
            ))
            
            logger.info("=" * 80)
            logger.info("SNAPSHOT COMPARISON COMPLETED SUCCESSFULLY")
            logger.info("=" * 80)
            
            self.root.after(2000, lambda: self.compare_btn.config(state='normal'))
            
        except Exception as e:
            logger.error(f"Snapshot comparison error: {e}", exc_info=True)
            error_msg = f"Snapshot comparison failed:\n\n{type(e).__name__}:\n{str(e)[:200]}"
            self.root.after(0, lambda: messagebox.showerror('Error', error_msg))
            self.root.after(0, lambda: self.compare_btn.config(state='normal'))
            self.root.after(0, lambda: self._update_progress(0, "Ready"))
    
    def _format_connection_error(self, error_msg):
        """Format connection error message with troubleshooting tips"""
        if not error_msg:
            return "Unknown connection error"
        
        tips = ""
        if "resolve host" in error_msg.lower():
            tips = "\n\nTroubleshooting:\n" \
                   "• Check network connection\n" \
                   "• Verify RTC server URL is reachable\n" \
                   "• May need VPN (if off-site)"
        elif "authentication" in error_msg.lower() or "401" in error_msg:
            tips = "\n\nTroubleshooting:\n" \
                   "• Verify username and password\n" \
                   "• Check credentials are correct\n" \
                   "• Account may need RTC permissions"
        elif "timeout" in error_msg.lower():
            tips = "\n\nTroubleshooting:\n" \
                   "• Server may be slow or offline\n" \
                   "• Try again in a moment\n" \
                   "• Check network connectivity"
        elif "ssl" in error_msg.lower():
            tips = "\n\nTroubleshooting:\n" \
                   "• SSL certificate issue (if self-signed)\n" \
                   "• May need certificate bypass\n" \
                   "• Contact IT support if persistent"
        
        return f"Cannot connect to RTC:\n{error_msg}{tips}"

    def start_hybrid_comparison(self, snapshot_url, local_folder):
        """Start hybrid comparison: RTC snapshot → Local folder"""
        self.save_credentials_on_success = self.keep_signed_in_var.get()
        self.compare_btn.config(state='disabled')
        self.progress_bar['value'] = 0
        self.progress_label.config(text="🔄 Starting hybrid comparison...")
        self.root.update()
        
        # Run in background thread
        thread = threading.Thread(
            target=self._hybrid_comparison_thread,
            args=(snapshot_url, local_folder),
            daemon=True
        )
        thread.start()
    
    def _hybrid_comparison_thread(self, snapshot_url, local_root_folder):
        """
        Background thread for Online → Offline comparison.

        Flow:
        1.  Connect to RTC and validate credentials.
        2.  Fetch all components from the provided snapshot.
        3.  Show component-selection dialog – each entry shows the resolved
            local folder path (component name split on '.').
        4.  For every selected component:
            a. Resolve the local folder:
               rb.as.ms.fiatgen.cswpr → <root>/rb/as/ms/fiatgen/cswpr
            b. Fetch the baseline file list from the online component via SCM CLI.
            c. Download each file's content to a temporary folder.
            d. Run compare_folders(temp_component_dir, local_resolved_dir).
            e. Collect the results.
        5.  Generate and open a combined HTML master report.
        """
        temp_dirs_to_cleanup = []

        try:
            logger.info("=" * 80)
            logger.info("HYBRID COMPARISON STARTED (Online → Offline)")
            logger.info(f"RTC Snapshot: {snapshot_url}")
            logger.info(f"Local Root Folder: {local_root_folder}")
            logger.info("=" * 80)

            # ── Credentials ────────────────────────────────────────────────
            username = self.rtc_username.get()
            password = self.rtc_password.get()
            server_url = self.rtc_server_url.get()

            self.root.after(0, lambda: self._update_progress(5, "Validating credentials..."))
            if not username or not password:
                self.root.after(0, lambda: messagebox.showerror(
                    'Validation Error', "Username and password are required"))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            # ── RTC connection ──────────────────────────────────────────────
            self.root.after(0, lambda: self._update_progress(10, "🔗 Testing RTC connection..."))
            rtc_conn, error_msg = get_rtc_connection(username, password, server_url)

            if not rtc_conn:
                detailed_error = self._format_connection_error(error_msg)
                logger.error(f"Connection failed: {error_msg}")
                self.root.after(0, lambda: messagebox.showerror(
                    'Connection Error', detailed_error))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            self._update_cached_credentials_after_login(username, password, server_url)

            # ── Extract snapshot UUID ───────────────────────────────────────
            self.root.after(0, lambda: self._update_progress(
                15, "Validating snapshot URL..."))
            snapshot_uuid = rtc_conn.extract_snapshot_uuid(snapshot_url)

            if not snapshot_uuid:
                self.root.after(0, lambda: messagebox.showerror(
                    'Invalid Input',
                    "Could not extract snapshot UUID.\n"
                    "Please verify the URL/UUID is correct."))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            logger.info(f"Snapshot UUID: {snapshot_uuid}")

            # ── Fetch snapshot components ───────────────────────────────────
            self.root.after(0, lambda: self._update_progress(
                20, "⬇️ Fetching snapshot components..."))

            def _snap_progress(current, total, message):
                if total > 0:
                    pct = 20 + int((current / total) * 25)
                    self.root.after(0, lambda p=pct, m=message:
                                    self._update_progress(p, m))

            snapshot_data = rtc_conn.fetch_snapshot_components(
                snapshot_uuid,
                snapshot_name="RTC Snapshot",
                progress_callback=_snap_progress,
            )

            if isinstance(snapshot_data, dict):
                components    = snapshot_data.get('components', [])
                snapshot_name = snapshot_data.get('name', 'RTC Snapshot')
            else:
                components    = snapshot_data if isinstance(snapshot_data, list) else []
                snapshot_name = 'RTC Snapshot'

            if not components:
                self.root.after(0, lambda: messagebox.showerror(
                    'Fetch Error',
                    f"No components found in snapshot.\n"
                    "Possible reasons:\n"
                    "• Invalid snapshot UUID\n"
                    "• Permission denied\n"
                    "• Network error"))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            logger.info(f"✓ Fetched {len(components)} components")

            # ── Component selection dialog ──────────────────────────────────
            self.root.after(0, lambda: self._update_progress(
                48, "📋 Opening component selection dialog..."))

            from src.gui.dialogs import show_hybrid_component_selection_dialog
            import threading as _threading

            dialog_event = _threading.Event()
            selection_result = {'canceled': True, 'selected': []}

            def _show_dialog():
                nonlocal selection_result
                try:
                    selection_result = show_hybrid_component_selection_dialog(
                        components, local_root_folder)
                except Exception as _de:
                    logger.error(f"Component selection dialog error: {_de}", exc_info=True)
                    selection_result = {'canceled': True, 'selected': []}
                finally:
                    dialog_event.set()

            self.root.after(0, _show_dialog)
            if not dialog_event.wait(timeout=600):
                self.root.after(0, lambda: messagebox.showwarning(
                    'Timeout', 'Component selection timed out.'))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            if selection_result['canceled']:
                logger.info("User cancelled component selection.")
                self.root.after(0, lambda: messagebox.showinfo(
                    'Cancelled', 'Comparison cancelled.'))
                self.root.after(0, lambda: self.compare_btn.config(state='normal'))
                self.root.after(0, lambda: self._update_progress(0, "Ready"))
                return

            selected = selection_result['selected']   # list of enriched comp dicts
            logger.info(f"User selected {len(selected)} component(s).")

            # ── Per-component download + compare ───────────────────────────
            from src.utils.comparison_engine import compare_folders as _compare_folders
            import tempfile, shutil

            # Timestamped output directory
            from datetime import datetime as _dt
            timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
            base_results_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(
                    os.path.abspath(__file__)))),
                "Snapshot_Comparison_Results")
            output_dir = os.path.join(
                base_results_dir, f"hybrid_{timestamp}")
            os.makedirs(output_dir, exist_ok=True)

            all_comp_results = []   # list of dicts: {name, local_folder, result}
            total_selected   = len(selected)

            for comp_idx, comp_entry in enumerate(selected, 1):
                comp_name    = comp_entry.get('name', 'Unknown')
                local_folder = comp_entry.get('local_folder', '')
                baseline_uuid = (comp_entry.get('baseline_uuid', '')
                                 or comp_entry.get('uuid', ''))

                step_pct = 50 + int((comp_idx / total_selected) * 40)

                # ── Validate local folder ───────────────────────────────
                if not os.path.isdir(local_folder):
                    resolved = self._resolve_local_folder_for_component(
                        local_root_folder, comp_name)
                    if os.path.isdir(resolved):
                        local_folder = resolved
                    else:
                        logger.warning(
                            f"[{comp_name}] Local folder not found: {local_folder}\n"
                            f"  Resolved path: {resolved}")
                        all_comp_results.append({
                            'name': comp_name,
                            'local_folder': local_folder,
                            'result': None,
                            'error': (
                                f"Local folder not found.\n"
                                f"Expected: {resolved}\n"
                                f"Check that the root folder contains the correct "
                                f"sub-directory structure for '{comp_name}'."
                            ),
                        })
                        continue

                self.root.after(0, lambda p=step_pct, n=comp_name:
                                self._update_progress(
                                    p,
                                    f"[{comp_idx}/{total_selected}] "
                                    f"Downloading online files for: {n}"))

                # ── Fetch online file list for this component ───────────
                if not baseline_uuid:
                    logger.warning(f"[{comp_name}] No baseline UUID – skipping download")
                    all_comp_results.append({
                        'name': comp_name,
                        'local_folder': local_folder,
                        'result': None,
                        'error': "No baseline UUID available for this component.",
                    })
                    continue

                logger.info(f"[{comp_name}] Fetching baseline structure "
                            f"(uuid={baseline_uuid[:16]}...)")
                online_structure = rtc_conn.fetch_baseline_folder_structure(
                    baseline_uuid, comp_name)
                online_files = online_structure.get('files', [])

                logger.info(f"[{comp_name}] Online file list: {len(online_files)} file(s)")

                # ── Create temp folder and download files ───────────────
                temp_comp_dir = tempfile.mkdtemp(prefix=f"hybrid_{comp_name[:20]}_")
                temp_dirs_to_cleanup.append(temp_comp_dir)

                if online_files:
                    downloaded = 0
                    total_online = len(online_files)
                    for _fi_idx, fi in enumerate(online_files, 1):
                        file_path = fi.get('path', fi.get('name', ''))
                        if not file_path:
                            continue
                        content = rtc_conn.fetch_file_content_from_baseline(
                            baseline_uuid, file_path, comp_name)
                        if content is not None:
                            dest = os.path.join(
                                temp_comp_dir,
                                file_path.replace('/', os.sep).replace('\\', os.sep))
                            os.makedirs(os.path.dirname(dest), exist_ok=True)
                            try:
                                with open(dest, 'w', encoding='utf-8',
                                          errors='replace') as fh:
                                    fh.write(content)
                                downloaded += 1
                            except Exception as _we:
                                logger.warning(f"  Write error {file_path}: {_we}")

                        # Update progress every 5 files or on last file
                        if _fi_idx % 5 == 0 or _fi_idx == total_online:
                            _ci = comp_idx
                            _ts = total_selected
                            _fi = _fi_idx
                            _to = total_online
                            _dl = downloaded
                            _cn = comp_name
                            self.root.after(0, lambda ci=_ci, ts=_ts, fi=_fi,
                                            to=_to, dl=_dl, cn=_cn:
                                self._update_progress(
                                    step_pct,
                                    f"[{ci}/{ts}] 📥 {cn}: "
                                    f"{fi}/{to} files fetched ({dl} downloaded)"
                                )
                            )
                    logger.info(f"[{comp_name}] Downloaded {downloaded}/{len(online_files)} files")
                else:
                    logger.warning(
                        f"[{comp_name}] No files returned from baseline structure fetch.\n"
                        "  Possible causes: SCM CLI not installed / LSCM_PATH not configured.\n"
                        "  Comparison will proceed with empty online side – all local files "
                        "will appear as 'Only in Project'.")

                # ── Compare temp (online) vs local folder ───────────────
                self.root.after(0, lambda p=step_pct + 2, n=comp_name:
                                self._update_progress(
                                    p,
                                    f"[{comp_idx}/{total_selected}] "
                                    f"Comparing: {n}"))

                comp_output_dir = os.path.join(output_dir, comp_name.replace('.', '_'))
                os.makedirs(comp_output_dir, exist_ok=True)

                try:
                    safe_comp_name = comp_name.replace('.', '_').replace('/', '_')
                    cmp_result = _compare_folders(
                        temp_comp_dir,
                        local_folder,
                        progress_callback=None,
                        custom_mappings=None,
                        rtc_info=None,
                        output_dir=comp_output_dir,
                        report_name=f"Migration_Analysis_Report_{safe_comp_name}",
                        component_name=self.component_name.get(),
                    )

                    # Override display names
                    if cmp_result.get('success'):
                        cmp_result['folder1_display'] = (
                            f"{comp_name} (RTC snapshot: {snapshot_name})")
                        cmp_result['folder2_display'] = (
                            f"{comp_name} (Local: {local_folder})")

                    # ── Generate per-component HTML diff report ─────────
                    comp_html_path = None
                    if cmp_result.get('success'):
                        try:
                            from src.utils.diff_utils import generate_hybrid_component_html
                            html_dir = os.path.join(comp_output_dir, 'html_diffs')
                            comp_html_path = generate_hybrid_component_html(
                                component_name=comp_name,
                                compare_results=cmp_result['results'],
                                temp_online_dir=temp_comp_dir,
                                local_folder_dir=local_folder,
                                output_dir=html_dir,
                                snap1_label=f"{comp_name} — RTC snapshot: {snapshot_name}",
                                snap2_label=f"{comp_name} — Local: {os.path.basename(local_folder)}",
                            )
                            if comp_html_path:
                                # Also update report_paths so results viewer can link to it
                                cmp_result.setdefault('report_paths', {})
                                cmp_result['report_paths']['html'] = comp_html_path
                                logger.info(f"[{comp_name}] HTML diff report: {comp_html_path}")
                            else:
                                logger.warning(f"[{comp_name}] HTML diff generation returned None")
                        except Exception as _he:
                            logger.warning(f"[{comp_name}] HTML diff generation error: {_he}",
                                           exc_info=True)

                    all_comp_results.append({
                        'name': comp_name,
                        'local_folder': local_folder,
                        'result': cmp_result,
                        'html_path': comp_html_path,
                    })

                except Exception as _ce:
                    logger.error(f"[{comp_name}] compare_folders error: {_ce}",
                                 exc_info=True)
                    all_comp_results.append({
                        'name': comp_name,
                        'local_folder': local_folder,
                        'result': None,
                        'error': str(_ce),
                    })

            # ── Generate master HTML summary report ────────────────────
            self.root.after(0, lambda: self._update_progress(
                92, "📊 Generating master HTML report..."))

            master_html_path = self._generate_hybrid_master_report(
                all_comp_results, snapshot_name, snapshot_url,
                local_root_folder, output_dir)

            self.root.after(0, lambda: self._update_progress(
                97, "✅ Opening results..."))

            # ── Display results ─────────────────────────────────────────
            def _show_results():
                import webbrowser

                # Open each per-component HTML diff report in the browser
                for cr in all_comp_results:
                    html_p = cr.get('html_path')
                    if html_p and os.path.isfile(html_p):
                        webbrowser.open(html_p)

                # Also show individual result viewer for each successful component
                shown = 0
                for cr in all_comp_results:
                    r = cr.get('result')
                    if r and r.get('success'):
                        try:
                            show_results_dialog(
                                self.root,
                                r['results'],
                                r['folder1_display'],
                                r['folder2_display'],
                                r.get('folder1', ''),
                                r.get('folder2', ''),
                                r.get('files1', {}),
                                r.get('files2', {}),
                                r.get('report_paths', {}),
                            )
                            shown += 1
                        except Exception as _ve:
                            logger.warning(f"Viewer error for {cr['name']}: {_ve}")

                # Open master report
                if master_html_path and os.path.isfile(master_html_path):
                    webbrowser.open(master_html_path)

                # Summary message
                success_count = sum(
                    1 for cr in all_comp_results
                    if cr.get('result') and cr['result'].get('success'))
                fail_count    = len(all_comp_results) - success_count
                summary = (
                    f"Online → Offline comparison complete!\n\n"
                    f"Components selected  : {total_selected}\n"
                    f"Components compared  : {success_count}\n"
                    f"Failed / skipped     : {fail_count}\n\n"
                    f"Reports saved to:\n{output_dir}"
                )
                if fail_count:
                    errors = "\n".join(
                        f"  • {cr['name']}: {cr.get('error', 'unknown')}"
                        for cr in all_comp_results if cr.get('error'))
                    summary += f"\n\nErrors:\n{errors}"
                messagebox.showinfo("Comparison Complete", summary)

            self.root.after(0, _show_results)

            self.root.after(0, lambda: self._update_progress(
                100,
                f"✅ Hybrid comparison done: "
                f"{sum(1 for c in all_comp_results if c.get('result') and c['result'].get('success'))} "
                f"component(s) compared"))

            logger.info("=" * 80)
            logger.info("HYBRID COMPARISON COMPLETED")
            logger.info("=" * 80)

            self.root.after(2000, lambda: self.compare_btn.config(state='normal'))

        except Exception as e:
            logger.error(f"Hybrid comparison error: {e}", exc_info=True)
            error_msg = (f"Hybrid comparison failed:\n\n"
                         f"{type(e).__name__}:\n{str(e)[:200]}")
            self.root.after(0, lambda: messagebox.showerror('Error', error_msg))
            self.root.after(0, lambda: self.compare_btn.config(state='normal'))
            self.root.after(0, lambda: self._update_progress(0, "Ready"))

        finally:
            for td in temp_dirs_to_cleanup:
                if td and os.path.exists(td):
                    try:
                        import shutil
                        shutil.rmtree(td)
                        logger.info(f"Cleaned up temp folder: {td}")
                    except Exception as _e:
                        logger.warning(f"Failed to clean up temp folder {td}: {_e}")

    def _generate_hybrid_master_report(
            self, all_comp_results, snapshot_name, snapshot_url,
            local_root_folder, output_dir):
        """
        Generate a self-contained HTML master report for the Online → Offline
        comparison that summarises every compared component with:
          • Status badges (found/missing local folder)
          • File-level change counts (added / modified / removed / identical)
          • Link to the per-component CSV/Excel/HTML diff reports

        Returns the absolute path to the generated HTML file, or None on error.
        """
        import html as _html
        from datetime import datetime as _dt

        try:
            html_path = os.path.join(output_dir, "Hybrid_Comparison_Report.html")

            # ── Statistics ─────────────────────────────────────────────
            total       = len(all_comp_results)
            succeeded   = sum(1 for c in all_comp_results
                              if c.get('result') and c['result'].get('success'))
            failed      = total - succeeded
            total_diff  = 0
            total_ident = 0
            total_added = 0
            total_remov = 0

            for cr in all_comp_results:
                r = cr.get('result') or {}
                if not r.get('success'):
                    continue
                for row in (r.get('results') or []):
                    status = row[4] if len(row) > 4 else ''
                    if status == 'Different':
                        total_diff  += 1
                    elif status == 'Identical':
                        total_ident += 1
                    elif 'Only in Platform' in status:
                        total_remov += 1
                    elif 'Only in Project' in status:
                        total_added += 1

            # ── Build component rows ────────────────────────────────────
            rows_html = ''
            for cr in all_comp_results:
                name         = _html.escape(cr['name'])
                local_folder = _html.escape(cr.get('local_folder', ''))
                r            = cr.get('result') or {}
                error_msg    = cr.get('error', '')

                if error_msg:
                    status_badge = (
                        '<span style="background:#ffebe9;color:#cf222e;'
                        'padding:2px 8px;border-radius:8px;font-size:12px;">'
                        '✘ Error</span>')
                    detail = _html.escape(error_msg)
                    diff_link = '—'
                elif not r.get('success'):
                    status_badge = (
                        '<span style="background:#fff3cd;color:#856404;'
                        'padding:2px 8px;border-radius:8px;font-size:12px;">'
                        '⚠ No result</span>')
                    detail = _html.escape(r.get('error', 'No comparison result'))
                    diff_link = '—'
                else:
                    results_list = r.get('results', [])
                    n_diff  = sum(1 for row in results_list
                                  if len(row) > 4 and row[4] == 'Different')
                    n_ident = sum(1 for row in results_list
                                  if len(row) > 4 and row[4] == 'Identical')
                    n_added = sum(1 for row in results_list
                                  if len(row) > 4 and 'Only in Project' in row[4])
                    n_remov = sum(1 for row in results_list
                                  if len(row) > 4 and 'Only in Platform' in row[4])

                    if n_diff == 0 and n_added == 0 and n_remov == 0:
                        status_badge = (
                            '<span style="background:#dafbe1;color:#1a7f37;'
                            'padding:2px 8px;border-radius:8px;font-size:12px;">'
                            '✔ Identical</span>')
                    else:
                        status_badge = (
                            '<span style="background:#fff8c5;color:#9a6700;'
                            'padding:2px 8px;border-radius:8px;font-size:12px;">'
                            '± Different</span>')

                    detail = (
                        f'<span style="color:#cf222e;">−{n_remov} removed  </span>'
                        f'<span style="color:#9a6700;">±{n_diff} modified  </span>'
                        f'<span style="color:#1a7f37;">+{n_added} added  </span>'
                        f'<span style="color:#57606a;">○{n_ident} identical</span>'
                    )

                    # Link to per-component HTML diff report
                    rp = r.get('report_paths') or {}
                    html_rep = (rp.get('html')
                                or cr.get('html_path')
                                or rp.get('csv', ''))
                    if html_rep and os.path.isfile(html_rep):
                        rel = os.path.relpath(html_rep, output_dir).replace('\\', '/')
                        diff_link = (
                            f'<a href="{_html.escape(rel)}" target="_blank">'
                            f'📄 View report</a>')
                    else:
                        diff_link = '—'

                rows_html += f'''
  <tr>
    <td style="padding:8px 12px;font-family:monospace;font-size:13px;">{name}</td>
    <td style="padding:8px 12px;font-size:12px;color:#555;">{local_folder}</td>
    <td style="padding:8px 12px;">{status_badge}</td>
    <td style="padding:8px 12px;font-size:13px;">{detail}</td>
    <td style="padding:8px 12px;font-size:13px;">{diff_link}</td>
  </tr>'''

            # ── Compose full HTML ───────────────────────────────────────
            html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Online → Offline Comparison — {_html.escape(snapshot_name)}</title>
  <style>
    body {{ font-family: "Segoe UI", Arial, sans-serif; margin: 0; background: #f0f4f8; color: #24292e; }}
    header {{ background: #003366; color: white; padding: 20px 30px; }}
    header h1 {{ margin: 0 0 6px; font-size: 22px; }}
    header p  {{ margin: 0; font-size: 13px; color: #b0c4de; }}
    .summary  {{ display: flex; gap: 16px; padding: 16px 30px; flex-wrap: wrap; }}
    .card {{ background: white; border-radius: 8px; padding: 14px 20px;
              border-left: 4px solid; min-width: 140px; box-shadow: 0 1px 4px rgba(0,0,0,.1); }}
    .card .num  {{ font-size: 28px; font-weight: 700; }}
    .card .lbl  {{ font-size: 12px; color: #57606a; margin-top: 2px; }}
    .c-blue  {{ border-color: #0969da; }}   .c-blue .num  {{ color: #0969da; }}
    .c-green {{ border-color: #1a7f37; }}   .c-green .num {{ color: #1a7f37; }}
    .c-amber {{ border-color: #9a6700; }}   .c-amber .num {{ color: #9a6700; }}
    .c-red   {{ border-color: #cf222e; }}   .c-red .num   {{ color: #cf222e; }}
    table {{ border-collapse: collapse; width: calc(100% - 60px);
              margin: 0 30px 30px; background: white; border-radius: 8px;
              overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,.1); }}
    thead tr {{ background: #003366; color: white; }}
    th {{ padding: 10px 12px; text-align: left; font-size: 13px; font-weight: 600; }}
    tbody tr:nth-child(even) {{ background: #f6f8fa; }}
    tbody tr:hover {{ background: #eaf3fb; }}
    td a {{ color: #0969da; text-decoration: none; }}
    td a:hover {{ text-decoration: underline; }}
    .ts {{ font-size: 11px; color: #888; padding: 0 30px 20px; }}
  </style>
</head>
<body>
<header>
  <h1>🔄 Online → Offline Comparison Report</h1>
  <p>Snapshot: <strong>{_html.escape(snapshot_name)}</strong> &nbsp;|&nbsp;
     Local root: <strong>{_html.escape(local_root_folder)}</strong></p>
</header>

<div class="summary">
  <div class="card c-blue">
    <div class="num">{total}</div>
    <div class="lbl">Components selected</div>
  </div>
  <div class="card c-green">
    <div class="num">{succeeded}</div>
    <div class="lbl">Successfully compared</div>
  </div>
  <div class="card c-amber">
    <div class="num">{total_diff}</div>
    <div class="lbl">Files modified</div>
  </div>
  <div class="card c-red">
    <div class="num">{failed}</div>
    <div class="lbl">Failed / skipped</div>
  </div>
</div>

<table>
  <thead>
    <tr>
      <th>Component</th>
      <th>Local Folder</th>
      <th>Status</th>
      <th>File Changes</th>
      <th>Report</th>
    </tr>
  </thead>
  <tbody>
{rows_html}
  </tbody>
</table>

<p class="ts">Generated: {_dt.now().strftime("%Y-%m-%d %H:%M:%S")} &nbsp;|&nbsp;
Snapshot URL: {_html.escape(snapshot_url)}</p>
</body>
</html>'''

            with open(html_path, 'w', encoding='utf-8') as fh:
                fh.write(html_content)

            logger.info(f"✓ Hybrid master report: {html_path}")
            return html_path

        except Exception as e:
            logger.error(f"Error generating hybrid master report: {e}", exc_info=True)
            return None

    def _update_progress(self, value, message):
        """Update progress bar and label"""
        self.progress_bar['value'] = value
        self.progress_label.config(text=message)
        self.root.update_idletasks()
    
    def _generate_file_diffs_for_comparison(self, comparison_results, rtc_conn, output_dir):
        """
        Generate HTML diffs for modified files in component comparisons (OPTIMIZED)
        
        Args:
            comparison_results: List of component comparison results
            rtc_conn: RTCConnection instance
            output_dir: Directory to save diff HTML files
        """
        try:
            import tempfile
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            total_diffs = 0
            total_files_to_diff = 0
            max_diffs_per_component = 50  # Process up to 50 files per component
            max_file_size_kb = 500  # Skip files larger than 500KB for diff (still compare)
            max_workers = 10  # Parallel diff generation
            
            # Create diffs subdirectory
            diffs_dir = os.path.join(output_dir, "file_diffs")
            os.makedirs(diffs_dir, exist_ok=True)
            logger.info(f"Created diffs directory: {diffs_dir}")
            
            # Count modified components
            modified_components = [r for r in comparison_results if r.get('file_comparison')]
            logger.info(f"Found {len(modified_components)} components with file comparison data")
            
            if not modified_components:
                logger.warning("No components have file comparison data - cannot generate diffs")
                return
            
            # Collect all file diff tasks
            all_tasks = []
            component_task_mapping = {}
            
            for comp_result in comparison_results:
                comp_name = comp_result.get('name', 'Unknown')
                file_comparison = comp_result.get('file_comparison')
                
                if not file_comparison:
                    continue
                
                modified_files = file_comparison.get('modified', [])
                if not modified_files:
                    continue
                
                logger.info(f"{comp_name}: Found {len(modified_files)} modified files")
                total_files_to_diff += len(modified_files)
                
                # Limit files to diff
                files_to_diff = modified_files[:max_diffs_per_component]
                if len(modified_files) > max_diffs_per_component:
                    logger.warning(f"{comp_name}: Limiting to {max_diffs_per_component} of {len(modified_files)} modified files")
                
                comp_result['file_diffs'] = []
                component_task_mapping[comp_name] = comp_result
                
                baseline1_uuid = comp_result.get('baseline1_uuid')
                baseline2_uuid = comp_result.get('baseline2_uuid')
                
                if not baseline1_uuid or not baseline2_uuid:
                    logger.warning(f"{comp_name}: Missing baseline UUIDs")
                    continue
                
                # Create tasks for parallel processing
                for file_path in files_to_diff:
                    all_tasks.append({
                        'comp_name': comp_name,
                        'file_path': file_path,
                        'baseline1_uuid': baseline1_uuid,
                        'baseline2_uuid': baseline2_uuid,
                        'comp_result': comp_result
                    })
            
            if not all_tasks:
                logger.warning("No files to generate diffs for")
                return
            
            logger.info(f"Starting parallel diff generation for {len(all_tasks)} files (workers={max_workers})...")
            
            def generate_single_diff(task):
                """Generate diff for a single file (runs in thread pool)"""
                comp_name = task['comp_name']
                file_path = task['file_path']
                baseline1_uuid = task['baseline1_uuid']
                baseline2_uuid = task['baseline2_uuid']
                
                try:
                    # Download both versions
                    content1 = rtc_conn.fetch_file_content_from_baseline(baseline1_uuid, file_path, comp_name)
                    if content1 is None:
                        return {'success': False, 'reason': 'download1_failed'}
                    
                    content2 = rtc_conn.fetch_file_content_from_baseline(baseline2_uuid, file_path, comp_name)
                    if content2 is None:
                        return {'success': False, 'reason': 'download2_failed'}
                    
                    # Check file size (skip very large files for HTML diff)
                    size1_kb = len(content1) / 1024
                    size2_kb = len(content2) / 1024
                    max_size = max(size1_kb, size2_kb)
                    
                    if max_size > max_file_size_kb:
                        logger.info(f"{comp_name}: Skipping HTML diff for {file_path} (size: {max_size:.1f}KB > {max_file_size_kb}KB)")
                        return {'success': False, 'reason': 'too_large', 'size': max_size}
                    
                    # Generate diff in memory (no temp files)
                    import difflib
                    lines1 = content1.splitlines(keepends=True)
                    lines2 = content2.splitlines(keepends=True)
                    
                    differ = difflib.HtmlDiff(wrapcolumn=120)
                    html_diff = differ.make_file(
                        lines1, lines2,
                        fromdesc=f"Baseline 1: {file_path}",
                        todesc=f"Baseline 2: {file_path}"
                    )
                    
                    # Count differences for summary
                    added_count = html_diff.count('class="diff_add"')
                    deleted_count = html_diff.count('class="diff_sub"')
                    changed_count = html_diff.count('class="diff_chg"')
                    total_diffs = added_count + deleted_count + changed_count
                    
                    # Add navigation summary at top of HTML
                    summary_html = f'''
<div style="background-color: #e8f4f8; border: 2px solid #1976d2; padding: 20px; margin: 20px; border-radius: 8px; font-family: Arial, sans-serif;">
    <h2 style="margin-top: 0; color: #1976d2;">📊 File Comparison Summary</h2>
    <p><strong>File:</strong> {file_path}</p>
    <p><strong>Component:</strong> {comp_name}</p>
    <table style="margin-top: 15px; border-collapse: collapse;">
        <tr>
            <td style="padding: 8px; background-color: palegreen; border: 1px solid #ccc;"><strong>Added Lines:</strong></td>
            <td style="padding: 8px; border: 1px solid #ccc;">{added_count}</td>
        </tr>
        <tr>
            <td style="padding: 8px; background-color: #ffaaaa; border: 1px solid #ccc;"><strong>Deleted Lines:</strong></td>
            <td style="padding: 8px; border: 1px solid #ccc;">{deleted_count}</td>
        </tr>
        <tr>
            <td style="padding: 8px; background-color: #ffff77; border: 1px solid #ccc;"><strong>Changed Lines:</strong></td>
            <td style="padding: 8px; border: 1px solid #ccc;">{changed_count}</td>
        </tr>
        <tr style="font-weight: bold; background-color: #f0f0f0;">
            <td style="padding: 8px; border: 1px solid #ccc;"><strong>Total Differences:</strong></td>
            <td style="padding: 8px; border: 1px solid #ccc;">{total_diffs}</td>
        </tr>
    </table>
    <p style="margin-top: 15px; font-size: 14px; color: #666;">
        💡 <strong>Tip:</strong> Click the "f" or "t" links in the first column to jump to the next difference.
        Differences are highlighted in <span style="background-color: palegreen; padding: 2px 4px;">green</span> (added),
        <span style="background-color: #ffaaaa; padding: 2px 4px;">red</span> (deleted), or
        <span style="background-color: #ffff77; padding: 2px 4px;">yellow</span> (changed).
    </p>
    {f'<p style="margin-top: 10px; padding: 10px; background-color: #fff3cd; border-left: 4px solid #ffc107;"><strong>⚠️ Warning:</strong> Files are identical - no differences found. The metadata may have changed but content is the same.</p>' if total_diffs == 0 else ''}
</div>
'''
                    
                    # Insert summary after <body> tag
                    html_diff = html_diff.replace('<body>', '<body>' + summary_html, 1)
                    
                    # Write to output
                    safe_filename = file_path.replace('/', '_').replace('\\', '_').replace(':', '_')
                    diff_path = os.path.join(diffs_dir, f"{safe_filename}_diff.html")
                    
                    with open(diff_path, 'w', encoding='utf-8') as f:
                        f.write(html_diff)
                    
                    logger.debug(f"{comp_name}: ✓ Generated diff for {file_path} ({total_diffs} differences)")
                    
                    return {
                        'success': True,
                        'comp_name': comp_name,
                        'file_path': file_path,
                        'diff_path': diff_path,
                        'size': max_size,
                        'diff_count': total_diffs
                    }
                    
                except Exception as e:
                    logger.warning(f"{comp_name}: ✗ Error generating diff for {file_path}: {e}", exc_info=True)
                    return {'success': False, 'comp_name': comp_name, 'file_path': file_path, 'reason': str(e)[:100]}
            
            # Process diffs in parallel
            completed_count = 0
            success_count = 0
            failed_count = 0
            failed_details = []  # Track failures for debugging
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_task = {executor.submit(generate_single_diff, task): task for task in all_tasks}
                
                for future in as_completed(future_to_task):
                    completed_count += 1
                    result = future.result()
                    
                    # Update progress every 5 files
                    if completed_count % 5 == 0 or completed_count == len(all_tasks):
                        progress_pct = 75 + (completed_count / len(all_tasks)) * 10  # 75-85% range
                        self.root.after(
                            0,
                            lambda p=progress_pct, c=completed_count, t=len(all_tasks):
                                self._update_progress(p, f"📄 Generating diffs: {c}/{t} files")
                        )
                    
                    if result['success']:
                        success_count += 1
                        comp_name = result['comp_name']
                        comp_result = component_task_mapping[comp_name]
                        
                        relative_diff_path = os.path.relpath(result['diff_path'], output_dir)
                        comp_result['file_diffs'].append({
                            'file_path': result['file_path'],
                            'diff_html': relative_diff_path.replace('\\', '/'),
                            'diff_html_abs': result['diff_path']
                        })
                    else:
                        failed_count += 1
                        # Track failure details for logging
                        comp_name = result.get('comp_name', 'Unknown')
                        file_path = result.get('file_path', 'Unknown')
                        reason = result.get('reason', 'Unknown')
                        failed_details.append(f"{comp_name}: {file_path} ({reason})")
            
            logger.info("=" * 80)
            logger.info(f"✓ FILE DIFF GENERATION COMPLETE (PARALLEL):")
            logger.info(f"  Total files processed: {len(all_tasks)}")
            logger.info(f"  Successfully generated diffs: {success_count}")
            logger.info(f"  Failed: {failed_count}")
            logger.info(f"  Diffs saved to: {diffs_dir}")
            
            if failed_count > 0 and failed_details:
                logger.warning(f"Failed diffs (showing first 10):")
                for detail in failed_details[:10]:
                    logger.warning(f"  - {detail}")
                if len(failed_details) > 10:
                    logger.warning(f"  ... and {len(failed_details) - 10} more")
            
            logger.info("=" * 80)
            
        except Exception as e:
            logger.error(f"Error generating file diffs: {e}", exc_info=True)

    def _transform_snapshot_results_for_viewer(self, comparison_results, rtc_conn=None, enable_changesets=False):
        """
        Transform snapshot comparison results into results_viewer format
        
        Match offline results format:
        [comp_name, metric1, metric2, line_status, status, html_link, purpose, changeset]
        
        For snapshots:
        - index[0] = component name
        - index[1] = snapshot1_uuid length (as numeric reference)
        - index[2] = snapshot2_uuid length (as numeric reference)
        - index[3] = line_status (baseline + UUID info)
        - index[4] = status (Modified|Different|Identical|Only in Snapshot 1|Only in Snapshot 2)
        - index[5] = html_link (N/A for snapshot components)
        - index[6] = purpose
        - index[7] = changeset_info
        
        Args:
            comparison_results: Results from compare_snapshots
            rtc_conn: RTCConnection object (optional, for changeset fetching)
            enable_changesets: Whether to fetch changeset information (default: False for performance)
        """
        viewer_results = []
        changeset_stats = {'total_modified': 0, 'changesets_found': 0, 'changesets_failed': 0}
        
        for result in comparison_results:
            comp_name = result['name']
            status = result['status']
            snap1 = result.get('snapshot1', {})
            snap2 = result.get('snapshot2', {})
            
            # Extract baseline UUIDs
            baseline1_uuid = snap1.get('baseline_uuid', 'N/A') if snap1 else 'N/A'
            baseline2_uuid = snap2.get('baseline_uuid', 'N/A') if snap2 else 'N/A'
            
            # Create numeric metrics (UUID string length for reference)
            metric1 = len(baseline1_uuid) if baseline1_uuid != 'N/A' else 0
            metric2 = len(baseline2_uuid) if baseline2_uuid != 'N/A' else 0
            
            # Create line status with baseline info
            line_status = f"Baseline: {baseline1_uuid[:8]}... → {baseline2_uuid[:8]}..."
            
            # Map status to results_viewer format
            # _compare_one_component already uses 'Identical' and 'Different' directly.
            if status == 'Added in Snapshot 2':
                status_type = 'Only in Snapshot 2'
            elif status == 'Removed in Snapshot 2':
                status_type = 'Only in Snapshot 1'
            else:
                status_type = status  # 'Identical' or 'Different' passed through unchanged
            
            # Generate purpose/description based on status
            if status == 'Different':
                # Show file change details in purpose
                file_comparison = result.get('file_comparison', {})
                if file_comparison:
                    # file_comparison values are ints (counts) from compare_folder_structures
                    added_count = file_comparison.get('added', 0)
                    modified_count = file_comparison.get('modified', 0)
                    removed_count = file_comparison.get('removed', 0)
                    purpose = f"Component baseline changed: {baseline1_uuid[:12]}... → {baseline2_uuid[:12]}... " \
                             f"(+{added_count} added, ~{modified_count} modified, -{removed_count} removed)"
                else:
                    purpose = f"Component baseline changed: {baseline1_uuid[:12]}... → {baseline2_uuid[:12]}..."
            elif status == 'Identical' or status == 'Unchanged':
                purpose = f"Component baseline unchanged: {baseline1_uuid[:12]}..."
            elif status == 'Added in Snapshot 2':
                purpose = f"Component added in snapshot 2: {baseline2_uuid[:12]}..."
            elif status == 'Removed in Snapshot 2':
                purpose = f"Component removed from snapshot 2 (was: {baseline1_uuid[:12]}...)"
            else:
                purpose = "Component comparison"
            
            # Check if file-level diffs were generated
            file_diffs = result.get('file_diffs', [])
            if file_diffs and len(file_diffs) > 0:
                # Use first diff file as the html_link
                html_link = file_diffs[0].get('diff_html', 'N/A')
            else:
                html_link = "Component-level (no file diff)"
            
            # Fetch changeset information if enabled and component was modified
            changeset_info = ""
            if enable_changesets and rtc_conn and status == 'Different':
                changeset_stats['total_modified'] += 1
                try:
                    if not settings.LSCM_PATH:
                        changeset_info = "LSCM not configured - Install RTC SCM CLI for changeset details"
                        changeset_stats['changesets_failed'] += 1
                        logger.warning(f"   ⚠ {comp_name}: LSCM not available")
                    else:
                        logger.info(f"📋 {comp_name}: Fetching changesets from baseline2 ({baseline2_uuid[:12]}...)...")
                        changesets = rtc_conn.fetch_baseline_changesets_scm(baseline2_uuid, comp_name)
                        if changesets:
                            changeset_count = len(changesets)
                            first_cs = changesets[0]
                            cs_uuid = first_cs.get('uuid', 'N/A')
                            cs_author = first_cs.get('author', '')
                            cs_comment = first_cs.get('comment', '')[:60]
                            changeset_info = (
                                f"Changesets: {changeset_count} | "
                                f"Latest: {cs_uuid[:12]}{'...' if len(cs_uuid) > 12 else ''} "
                                f"by {cs_author}"
                                + (f" — {cs_comment}" if cs_comment else "")
                            )
                            changeset_stats['changesets_found'] += 1
                            logger.info(f"   ✓ {comp_name}: Found {changeset_count} changeset(s)")
                        else:
                            changeset_info = "No changesets found via SCM CLI"
                            changeset_stats['changesets_failed'] += 1
                            logger.warning(f"   ⚠ {comp_name}: No changesets returned")
                except Exception as e:
                    logger.warning(f"{comp_name}: Failed to fetch changesets: {e}", exc_info=True)
                    changeset_info = f"Changeset fetch error: {str(e)[:50]}"
                    changeset_stats['changesets_failed'] += 1
            
            # Create result list matching offline format
            result_list = [
                comp_name,                              # index[0]: component name
                metric1,                                # index[1]: metric1 (baseline1 uuid length)
                metric2,                                # index[2]: metric2 (baseline2 uuid length)
                line_status,                            # index[3]: line_status
                status_type,                            # index[4]: status
                html_link,                              # index[5]: html_link (first file diff or component-level)
                purpose,                                # index[6]: purpose with baseline info
                changeset_info                          # index[7]: changeset_info (fetched if enabled)
            ]
            
            viewer_results.append(result_list)
        
        # Log changeset fetching statistics
        if enable_changesets:
            logger.info("=" * 80)
            logger.info("CHANGESET FETCHING SUMMARY:")
            logger.info(f"  Total modified components: {changeset_stats['total_modified']}")
            logger.info(f"  Components with changesets found: {changeset_stats['changesets_found']}")
            logger.info(f"  Components with no changesets: {changeset_stats['changesets_failed']}")
            if changeset_stats['changesets_found'] > 0:
                success_rate = (changeset_stats['changesets_found'] / changeset_stats['total_modified']) * 100 if changeset_stats['total_modified'] > 0 else 0
                logger.info(f"  Success rate: {success_rate:.1f}%")
            logger.info("=" * 80)
        
        logger.info(f"Transformed {len(viewer_results)} results for viewer (format: 8-element list)")
        return viewer_results
    
    def _export_snapshot_results_to_reports(self, viewer_results, csv_path, excel_path, snap1_name, snap2_name, comparison_results=None):
        """
        Export snapshot comparison results to CSV and Excel files with file-level details
        
        Args:
            viewer_results: List of 8-element result lists from viewer format
            csv_path: Path to save CSV file
            excel_path: Path to save Excel file
            snap1_name: Display name for snapshot 1
            snap2_name: Display name for snapshot 2
            comparison_results: Original comparison results with file diff info
        """
        try:
            import csv
            
            # Calculate summary statistics
            total_components = len(viewer_results)
            modified_count = sum(1 for r in viewer_results if r[4] == 'Different')
            unchanged_count = sum(1 for r in viewer_results if r[4] == 'Identical')
            added_count = sum(1 for r in viewer_results if 'Snapshot 2' in r[4])
            removed_count = sum(1 for r in viewer_results if 'Snapshot 1' in r[4])
            
            # Count file-level changes
            total_files_modified = 0
            total_files_added = 0
            total_files_removed = 0
            total_files_unchanged = 0
            
            if comparison_results:
                for comp in comparison_results:
                    file_comp = comp.get('file_comparison', {})
                    if file_comp:
                        # file_comparison values are ints (counts) from compare_folder_structures
                        total_files_modified += file_comp.get('modified', 0) if isinstance(file_comp.get('modified'), int) else len(file_comp.get('details', {}))
                        total_files_added += file_comp.get('added', 0) if isinstance(file_comp.get('added'), int) else 0
                        total_files_removed += file_comp.get('removed', 0) if isinstance(file_comp.get('removed'), int) else 0
                        total_files_unchanged += file_comp.get('unchanged', 0) if isinstance(file_comp.get('unchanged'), int) else 0
            
            # Write MAIN CSV report (component-level)
            logger.info(f"Exporting snapshot comparison to CSV: {csv_path}")
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write summary section
                writer.writerow(["SNAPSHOT COMPARISON SUMMARY"])
                writer.writerow(["Snapshot 1", snap1_name])
                writer.writerow(["Snapshot 2", snap2_name])
                writer.writerow(["Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow([])
                
                writer.writerow(["COMPONENT-LEVEL STATISTICS"])
                writer.writerow(["Total Components", total_components])
                writer.writerow(["Modified Components", modified_count])
                writer.writerow(["Unchanged Components", unchanged_count])
                writer.writerow(["Added Components", added_count])
                writer.writerow(["Removed Components", removed_count])
                writer.writerow([])
                
                if comparison_results:
                    writer.writerow(["FILE-LEVEL STATISTICS"])
                    writer.writerow(["Total Modified Files", total_files_modified])
                    writer.writerow(["Total Added Files", total_files_added])
                    writer.writerow(["Total Removed Files", total_files_removed])
                    writer.writerow(["Total Unchanged Files", total_files_unchanged])
                    writer.writerow([])
                
                # Write component headers
                writer.writerow(["COMPONENT-LEVEL DETAILS"])
                writer.writerow([
                    "Component Name",
                    f"Snapshot 1 Metric ({snap1_name})",
                    f"Snapshot 2 Metric ({snap2_name})",
                    "Baseline Comparison",
                    "Status",
                    "Details",
                    "Type",
                    "Changeset Info"  # Column 8: Changeset information
                ])
                # Write data rows
                for row in viewer_results:
                    try:
                        writer.writerow([str(cell) if cell else '' for cell in row])
                    except Exception as row_err:
                        logger.warning(f"Error writing row: {row_err}")
            
            logger.info(f"✓ CSV report exported: {csv_path}")
            
            # Write FILE-LEVEL CSV report
            if comparison_results:
                file_csv_path = csv_path.replace('.csv', '_FileDetails.csv')
                logger.info(f"Exporting file-level details to CSV: {file_csv_path}")
                
                with open(file_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow([
                        "Component Name",
                        "File Path",
                        "Change Type",
                        "HTML Diff Link"
                    ])
                    
                    for comp in comparison_results:
                        comp_name = comp.get('name', 'Unknown')
                        file_comp = comp.get('file_comparison', {})
                        file_diffs = comp.get('file_diffs', [])
                        
                        # Create lookup for diff links
                        diff_lookup = {fd['file_path']: fd['diff_html'] for fd in file_diffs}
                        
                        if file_comp:
                            # Use details dict: {file_path: 'modified'|'added'|'removed'|'unchanged'}
                            details = file_comp.get('details', {})
                            for file_path, change_type in sorted(details.items()):
                                if change_type == 'modified':
                                    diff_link = diff_lookup.get(file_path, 'N/A')
                                    writer.writerow([comp_name, file_path, 'Modified', diff_link])
                                elif change_type == 'added':
                                    writer.writerow([comp_name, file_path, 'Added', 'N/A'])
                                elif change_type == 'removed':
                                    writer.writerow([comp_name, file_path, 'Removed', 'N/A'])
                
                logger.info(f"✓ File-level CSV exported: {file_csv_path}")
            
            # Write HTML summary report
            html_path = csv_path.replace('.csv', '.html')
            logger.info(f"Exporting snapshot comparison to HTML: {html_path}")
            try:
                self._generate_snapshot_html_report(viewer_results, html_path, snap1_name, snap2_name, comparison_results)
                logger.info(f"✓ HTML report exported: {html_path}")
            except Exception as html_err:
                logger.error(f"Error generating HTML report: {html_err}")
            
            # Write Excel report with MULTIPLE SHEETS
            logger.info(f"Exporting snapshot comparison to Excel: {excel_path}")
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                wb = Workbook()
                
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
                
                # ========== SHEET 1: SUMMARY ==========
                ws_summary = wb.create_sheet("Summary", 0)
                
                # Title
                title_font = Font(bold=True, size=16, color="FFFFFF")
                title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                ws_summary['A1'] = "SNAPSHOT COMPARISON SUMMARY"
                ws_summary['A1'].font = title_font
                ws_summary['A1'].fill = title_fill
                ws_summary.merge_cells('A1:B1')
                
                # Snapshot info
                ws_summary['A3'] = "Snapshot 1:"
                ws_summary['B3'] = snap1_name
                ws_summary['A4'] = "Snapshot 2:"
                ws_summary['B4'] = snap2_name
                ws_summary['A5'] = "Comparison Date:"
                ws_summary['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Component statistics
                ws_summary['A7'] = "COMPONENT-LEVEL STATISTICS"
                ws_summary['A7'].font = Font(bold=True, size=12)
                ws_summary['A8'] = "Total Components"
                ws_summary['B8'] = total_components
                ws_summary['A9'] = "Modified Components"
                ws_summary['B9'] = modified_count
                ws_summary['A10'] = "Unchanged Components"
                ws_summary['B10'] = unchanged_count
                ws_summary['A11'] = "Added Components"
                ws_summary['B11'] = added_count
                ws_summary['A12'] = "Removed Components"
                ws_summary['B12'] = removed_count
                
                # File statistics
                if comparison_results:
                    ws_summary['A14'] = "FILE-LEVEL STATISTICS"
                    ws_summary['A14'].font = Font(bold=True, size=12)
                    ws_summary['A15'] = "Total Modified Files"
                    ws_summary['B15'] = total_files_modified
                    ws_summary['A16'] = "Total Added Files"
                    ws_summary['B16'] = total_files_added
                    ws_summary['A17'] = "Total Removed Files"
                    ws_summary['B17'] = total_files_removed
                    ws_summary['A18'] = "Total Unchanged Files"
                    ws_summary['B18'] = total_files_unchanged
                
                # Column widths
                ws_summary.column_dimensions['A'].width = 30
                ws_summary.column_dimensions['B'].width = 50
                
                # ========== SHEET 2: COMPONENT DETAILS ==========
                ws_components = wb.create_sheet("Component Details", 1)
                
                # Headers
                headers = [
                    "Component Name",
                    f"Snapshot 1 Metric",
                    f"Snapshot 2 Metric",
                    "Baseline Comparison",
                    "Status",
                    "Details",
                    "Type",
                    "Changeset Info"  # Column 8: Changeset information
                ]
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_components.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Data rows
                hyperlink_font = Font(color="0563C1", underline="single")
                for row_idx, row in enumerate(viewer_results, 2):
                    for col_idx, cell_value in enumerate(row, 1):
                        cell = ws_components.cell(row=row_idx, column=col_idx, value=str(cell_value) if cell_value else '')
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        
                        # Color code by status
                        if col_idx == 5:  # Status column
                            if cell_value == 'Different':
                                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                            elif cell_value == 'Identical':
                                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        
                        # col 6 (index 5) = HTML diff link — make clickable if file exists
                        if col_idx == 6 and cell_value and str(cell_value).lower().endswith('.html'):
                            html_abs = str(cell_value)
                            if os.path.isfile(html_abs):
                                cell.value = "Open Diff Report"
                                cell.hyperlink = 'file:///' + html_abs.replace('\\', '/')
                                cell.font = hyperlink_font
                
                # Auto-adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    ws_components.column_dimensions[chr(64 + col_idx)].width = 25
                
                # ========== SHEET 3: FILE-LEVEL DETAILS ==========
                if comparison_results:
                    ws_files = wb.create_sheet("File Details", 2)
                    
                    # Headers
                    file_headers = ["Component Name", "File Path", "Change Type", "HTML Diff Link"]
                    for col_idx, header in enumerate(file_headers, 1):
                        cell = ws_files.cell(row=1, column=col_idx, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Build lookup: comp_name → component HTML diff path (from viewer_results)
                    comp_html_lookup = {}
                    for vr in viewer_results:
                        html_val = vr[5] if len(vr) > 5 else ''
                        if html_val and str(html_val).lower().endswith('.html') and os.path.isfile(str(html_val)):
                            comp_html_lookup[vr[0]] = str(html_val)

                    # Data rows
                    file_row = 2
                    for comp in comparison_results:
                        comp_name = comp.get('name', 'Unknown')
                        file_comp = comp.get('file_comparison', {})
                        comp_html = comp_html_lookup.get(comp_name)  # per-component HTML report
                        
                        if file_comp:
                            # Use details dict: {file_path: 'modified'|'added'|'removed'|'unchanged'}
                            details = file_comp.get('details', {})
                            fill_map = {
                                'modified': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                                'added':    PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
                                'removed':  PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
                            }
                            label_map = {'modified': 'Modified', 'added': 'Added', 'removed': 'Removed'}
                            for file_path, change_type in sorted(details.items()):
                                if change_type not in label_map:
                                    continue
                                ws_files.cell(row=file_row, column=1, value=comp_name)
                                ws_files.cell(row=file_row, column=2, value=file_path)
                                cell_type = ws_files.cell(row=file_row, column=3, value=label_map[change_type])
                                cell_type.fill = fill_map[change_type]
                                # For modified files link to the component HTML diff report;
                                # added/removed files have no separate diff, so link the same report
                                if comp_html:
                                    link_cell = ws_files.cell(row=file_row, column=4, value="Open Diff Report")
                                    link_cell.hyperlink = 'file:///' + comp_html.replace('\\', '/')
                                    link_cell.font = Font(color="0563C1", underline="single")
                                else:
                                    ws_files.cell(row=file_row, column=4, value='N/A')
                                file_row += 1
                    
                    # Column widths
                    ws_files.column_dimensions['A'].width = 30
                    ws_files.column_dimensions['B'].width = 50
                    ws_files.column_dimensions['C'].width = 15
                    ws_files.column_dimensions['D'].width = 60
                    
                    logger.info(f"Added {file_row - 2} file-level entries to Excel")                
                
                wb.save(excel_path)
                logger.info(f"✓ Excel report exported with {len(wb.sheetnames)} sheets: {excel_path}")
            
            except ImportError:
                logger.warning("openpyxl not installed - skipping Excel export")
            except Exception as excel_err:
                logger.error(f"Error generating Excel report: {excel_err}")
        
        except Exception as e:
            logger.error(f"Error exporting snapshot results: {e}", exc_info=True)
    
    def _display_snapshot_results_with_viewer(self, viewer_results, comparison_metadata):
        """
        Display snapshot comparison results using enhanced results_viewer
        
        Args:
            viewer_results: List of results in viewer format (8-element lists)
            comparison_metadata: Dict with snapshot URLs and component data
        """
        try:
            # Extract metadata
            snap1_url = comparison_metadata['snap1_url']
            snap2_url = comparison_metadata['snap2_url']
            snap1_actual_name = comparison_metadata.get('snap1_name')
            snap2_actual_name = comparison_metadata.get('snap2_name')
            snap1_components = comparison_metadata['snap1_components']
            snap2_components = comparison_metadata['snap2_components']
            selected_components = comparison_metadata.get('selected_components', [])
            total_common_components = comparison_metadata.get('total_common_components', len(selected_components))
            output_dir = comparison_metadata.get('output_dir')  # Get output_dir from metadata
            
            # Create display names for snapshots - use actual names if available
            if snap1_actual_name:
                snap1_name = snap1_actual_name
            else:
                snap1_name = f"Snapshot 1: {snap1_url[:50]}..."
            
            if snap2_actual_name:
                snap2_name = snap2_actual_name
            else:
                snap2_name = f"Snapshot 2: {snap2_url[:50]}..."
            
            # Create component name mappings (like file dictionaries for offline mode)
            files1 = {}  # {component_name: component_uuid}
            files2 = {}  # {component_name: component_uuid}
            
            for comp in snap1_components:
                comp_name = comp.get('name', 'Unknown')
                files1[comp_name] = comp.get('uuid', '')
            
            for comp in snap2_components:
                comp_name = comp.get('name', 'Unknown')
                files2[comp_name] = comp.get('uuid', '')
            
            # Use the output directory already created during comparison
            # (output_dir was created earlier in run_snapshot_comparison_analysis)
            
            # Create report paths with absolute paths (consistent with offline mode)
            csv_report = os.path.join(output_dir, "Selected_Snapshot_Comparison.csv")
            excel_report = os.path.join(output_dir, "Selected_Snapshot_Comparison.xlsx")
            html_report = os.path.join(output_dir, "Selected_Snapshot_Comparison.html")
            
            report_paths = {
                'csv': csv_report,
                'output_dir': output_dir,
                'excel': excel_report,
                'html': html_report,
                'online_context': {
                    'mode': 'online_online',
                    'server_url': self.rtc_server_url.get(),
                    'username': self.rtc_username.get(),
                    'password': self.rtc_password.get(),
                    'snapshot1_components': snap1_components,
                    'snapshot2_components': snap2_components,
                    'selected_components': selected_components,
                }
            }
            
            logger.info(f"✓ Created output directory: {output_dir}")

            manifest_path = os.path.join(output_dir, "selected_components.txt")
            with open(manifest_path, 'w', encoding='utf-8') as manifest:
                manifest.write("Selected Online-Online Snapshot Components\n")
                manifest.write(f"Snapshot 1: {snap1_name}\n")
                manifest.write(f"Snapshot 2: {snap2_name}\n")
                manifest.write(f"Selected: {len(selected_components)} of {total_common_components} common components\n\n")
                for component_name in selected_components:
                    manifest.write(f"{component_name}\n")
            logger.info(f"✓ Selected component manifest exported: {manifest_path}")
            
            # ── Generate per-component HTML diff reports ─────────────────────
            # IMPORTANT: Must run BEFORE export so that viewer_results[idx][5]
            # holds the actual HTML path when the Excel/CSV is written.
            try:
                from src.utils.diff_utils import generate_snapshot_component_html
                full_comparison_results = comparison_metadata.get('comparison_results', [])
                html_dir = os.path.join(output_dir, "html_diffs")
                os.makedirs(html_dir, exist_ok=True)

                # Pre-fetched file contents and RTC server URL (from comparison thread)
                file_contents_by_comp = comparison_metadata.get('file_contents_by_component', {})
                rtc_srv_url           = comparison_metadata.get('rtc_server_url', '')

                # Build a quick lookup: component_name → viewer_results index
                name_to_idx = {vr[0]: i for i, vr in enumerate(viewer_results)}

                for comp_result in full_comparison_results:
                    if comp_result.get('status') != 'Different':
                        continue
                    cname  = comp_result['name']
                    b1uuid = comp_result.get('baseline1_uuid', '')
                    b2uuid = comp_result.get('baseline2_uuid', '')
                    fcmp   = comp_result.get('file_comparison')

                    # Per-component changeset metadata passed into HTML generator
                    csdata = comparison_metadata.get('changeset_by_component', {}).get(cname, {})

                    html_path = generate_snapshot_component_html(
                        component_name=cname,
                        baseline1_uuid=b1uuid,
                        baseline2_uuid=b2uuid,
                        file_comparison=fcmp,
                        output_dir=html_dir,
                        snap1_label=snap1_name,
                        snap2_label=snap2_name,
                        file_contents=file_contents_by_comp.get(cname),
                        changeset_data=csdata,
                        server_url=rtc_srv_url,
                    )
                    if html_path and cname in name_to_idx:
                        viewer_results[name_to_idx[cname]][5] = html_path

                logger.info(
                    f"✓ HTML diff reports generated for "
                    f"{sum(1 for vr in viewer_results if vr[5] not in ['N/A', '', None, 'Component-level (no file diff)'])} "
                    f"Different components in: {html_dir}"
                )
            except Exception as html_err:
                logger.warning(f"HTML diff generation failed (non-fatal): {html_err}")

            # Export snapshot comparison results to CSV, Excel, and HTML
            # (called after HTML diff generation so viewer_results[idx][5] holds actual paths)
            comparison_results = comparison_metadata.get('comparison_results', [])
            self._export_snapshot_results_to_reports(
                viewer_results,
                csv_report,
                excel_report,
                snap1_name,
                snap2_name,
                comparison_results  # Pass comparison results for file diff links
            )

            # ── Generate Beyond Compare master report ────────────────────────
            # One self-contained HTML with ALL components, dual tree view, inline
            # unified diffs, and changeset table — written AFTER per-component
            # HTML is generated so file_contents are already available.
            master_report_path = None
            try:
                from src.utils.diff_utils import generate_beyond_compare_master_report
                master_report_path = generate_beyond_compare_master_report(
                    comparison_results=comparison_results,
                    output_dir=output_dir,
                    snap1_label=snap1_name,
                    snap2_label=snap2_name,
                    file_contents_by_component=comparison_metadata.get('file_contents_by_component', {}),
                    changeset_by_component=comparison_metadata.get('changeset_by_component', {}),
                    server_url=comparison_metadata.get('rtc_server_url', ''),
                )
                if master_report_path:
                    report_paths['master_report'] = master_report_path
                    logger.info(f"✓ Beyond Compare master report: {master_report_path}")
                else:
                    logger.warning("Beyond Compare master report generation returned None")
            except Exception as mr_err:
                logger.warning(f"Beyond Compare master report failed (non-fatal): {mr_err}")

            # Log transformation details
            logger.info(f"Displaying {len(viewer_results)} snapshots results in viewer")
            if viewer_results:
                logger.info(f"Sample result: {viewer_results[0]} (type: {type(viewer_results[0])})")
                logger.info(f"Result length: {len(viewer_results[0]) if viewer_results else 'N/A'}")

            logger.info(f"Files1 count: {len(files1)}, Files2 count: {len(files2)}")
            logger.info(f"snap1_name: {snap1_name}, snap2_name: {snap2_name}")
            logger.info(f"Selected components exported: {len(selected_components)}/{total_common_components}")

            # Show results using enhanced viewer
            show_results_dialog(
                self.root,
                viewer_results,
                snap1_name,
                snap2_name,
                snap1_url,
                snap2_url,
                files1,
                files2,
                report_paths
            )

            # Notify user about master report
            if master_report_path and os.path.isfile(master_report_path):
                import webbrowser
                self.root.after(
                    800,
                    lambda p=master_report_path: (
                        messagebox.askquestion(
                            "📊 Beyond Compare Report Ready",
                            f"A complete Beyond Compare-style master report has been generated:\n\n"
                            f"{p}\n\n"
                            f"This report shows all {len(comparison_results)} components with:\n"
                            f"  • Left/Right dual folder tree view\n"
                            f"  • Inline file diffs (click any file)\n"
                            f"  • Changeset & baseline details\n\n"
                            f"Open it now?",
                        ) == 'yes' and webbrowser.open(p)
                    )
                )
            
            logger.info("✓ Snapshot comparison results displayed in enhanced viewer")
            
        except Exception as e:
            logger.error(f"Error displaying snapshot results in viewer: {e}", exc_info=True)
            # Fallback to simple display
            messagebox.showerror(
                "Error",
                f"Error displaying results in enhanced viewer:\n\n{str(e)[:200]}"
            )
    
    def _generate_snapshot_html_report(self, viewer_results, html_path, snap1_name, snap2_name, comparison_results=None):
        """Generate HTML summary report for snapshot comparison"""
        try:
            from datetime import datetime
            
            # Count statistics
            modified = sum(1 for r in viewer_results if r[4] == 'Different')
            added = sum(1 for r in viewer_results if 'Only in Snapshot 2' in r[4])
            removed = sum(1 for r in viewer_results if 'Only in Snapshot 1' in r[4])
            unchanged = sum(1 for r in viewer_results if r[4] == 'Identical')
            
            # Create a mapping from component name to comparison result for file diff links
            comp_results_map = {}
            if comparison_results:
                comp_results_map = {r['name']: r for r in comparison_results}
            
            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Snapshot Comparison Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .header {{ background-color: #003366; color: white; padding: 20px; border-radius: 5px; }}
        .header h1 {{ margin: 0; }}
        .header p {{ margin: 5px 0; opacity: 0.9; }}
        .summary {{ background-color: white; padding: 20px; margin: 20px 0; border-radius: 5px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .stats {{ display: flex; gap: 20px; margin: 20px 0; }}
        .stat-box {{ flex: 1; padding: 15px; border-radius: 5px; text-align: center; }}
        .stat-box h3 {{ margin: 0 0 10px 0; font-size: 32px; }}
        .stat-box p {{ margin: 0; color: #666; }}
        .modified {{ background-color: #fff3cd; border-left: 4px solid #ffc107; }}
        .added {{ background-color: #d4edda; border-left: 4px solid #28a745; }}
        .removed {{ background-color: #f8d7da; border-left: 4px solid #dc3545; }}
        .unchanged {{ background-color: #d1ecf1; border-left: 4px solid #17a2b8; }}
        table {{ width: 100%; border-collapse: collapse; background-color: white; border-radius: 5px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        th {{ background-color: #003366; color: white; padding: 12px; text-align: left; font-weight: bold; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        tr:hover {{ background-color: #f8f9fa; }}
        .status-Different {{ color: #ffc107; font-weight: bold; }}
        .status-Identical {{ color: #17a2b8; }}
        .status-added {{ color: #28a745; font-weight: bold; }}
        .status-removed {{ color: #dc3545; font-weight: bold; }}
        .file-diffs {{ margin-top: 10px; padding: 10px; background-color: #f8f9fa; border-radius: 3px; }}
        .file-diffs-header {{ font-weight: bold; margin-bottom: 5px; color: #666; font-size: 12px; }}
        .file-diff-link {{ display: inline-block; margin: 2px 5px; padding: 2px 8px; background-color: #007bff; color: white; text-decoration: none; border-radius: 3px; font-size: 11px; }}
        .file-diff-link:hover {{ background-color: #0056b3; }}
        .footer {{ text-align: center; margin-top: 30px; color: #666; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 Snapshot Comparison Report</h1>
        <p><strong>Snapshot 1:</strong> {snap1_name}</p>
        <p><strong>Snapshot 2:</strong> {snap2_name}</p>
        <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
    
    <div class="summary">
        <h2>Summary</h2>
        <div class="stats">
            <div class="stat-box modified">
                <h3>{modified}</h3>
                <p>Modified Components</p>
            </div>
            <div class="stat-box added">
                <h3>{added}</h3>
                <p>Added Components</p>
            </div>
            <div class="stat-box removed">
                <h3>{removed}</h3>
                <p>Removed Components</p>
            </div>
            <div class="stat-box unchanged">
                <h3>{unchanged}</h3>
                <p>Unchanged Components</p>
            </div>
        </div>
    </div>
    
    <table>
        <thead>
            <tr>
                <th>Component Name</th>
                <th>Baseline Comparison</th>
                <th>Status</th>
                <th>Description & File Diffs</th>
            </tr>
        </thead>
        <tbody>
"""
            
            for row in viewer_results:
                comp_name = row[0]
                line_status = row[3]
                status = row[4]
                purpose = row[6]
                
                status_class = status.replace(' ', '-').replace('Snapshot', '')
                
                # Check for file diffs
                file_diffs_html = ""
                if comp_name in comp_results_map:
                    comp_result = comp_results_map[comp_name]
                    file_diffs = comp_result.get('file_diffs', [])
                    if file_diffs:
                        file_diffs_html = '<div class="file-diffs"><div class="file-diffs-header">📄 File-Level Diffs:</div>'
                        for diff_info in file_diffs:
                            file_name = os.path.basename(diff_info['file_path'])
                            diff_link = diff_info['diff_html']
                            file_diffs_html += f'<a href="{diff_link}" class="file-diff-link" target="_blank">{file_name}</a>'
                        file_diffs_html += '</div>'
                
                html_content += f"""            <tr>
                <td><strong>{comp_name}</strong></td>
                <td><code>{line_status}</code></td>
                <td class="status-{status_class}">{status}</td>
                <td>{purpose}{file_diffs_html}</td>
            </tr>
"""
            
            html_content += f"""        </tbody>
    </table>
    
    <div class="footer">
        <p>Migration Analysis Tool - Bosch Engineering</p>
        <p>Online-Online Snapshot Comparison (Component + File Level)</p>
    </div>
</body>
</html>
"""
            
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
        except Exception as e:
            logger.error(f"Error generating HTML report: {e}", exc_info=True)
            raise
    
    def _display_snapshot_results(self, comparison_results):
        """Display snapshot comparison results"""
        # Create popup window for results
        result_window = tk.Toplevel(self.root)
        result_window.title("📊 Snapshot Comparison Results")
        result_window.geometry("900x600")
        
        # Header
        header = tk.Frame(result_window, bg='#003366')
        header.pack(fill=tk.X)
        tk.Label(
            header,
            text="📊 Snapshot Comparison Results",
            font=('Segoe UI', 12, 'bold'),
            bg='#003366',
            fg='white'
        ).pack(pady=10)
        
        # Statistics
        modified = sum(1 for r in comparison_results if r['status'] == 'Modified')
        added = sum(1 for r in comparison_results if 'Added' in r['status'])
        removed = sum(1 for r in comparison_results if 'Removed' in r['status'])
        unchanged = sum(1 for r in comparison_results if r['status'] == 'Unchanged')
        
        stats_frame = tk.Frame(result_window, bg='#EAF3FB')
        stats_frame.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Label(stats_frame, text=f"✏️ Modified: {modified}", bg='#EAF3FB', fg='#f39c12').pack(side=tk.LEFT, padx=10)
        tk.Label(stats_frame, text=f"➕ Added: {added}", bg='#EAF3FB', fg='#27ae60').pack(side=tk.LEFT, padx=10)
        tk.Label(stats_frame, text=f"➖ Removed: {removed}", bg='#EAF3FB', fg='#e74c3c').pack(side=tk.LEFT, padx=10)
        tk.Label(stats_frame, text=f"✓ Unchanged: {unchanged}", bg='#EAF3FB', fg='#95a5a6').pack(side=tk.LEFT, padx=10)
        
        # Tree view
        tree_frame = tk.Frame(result_window)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=('status', 'snapshot1', 'snapshot2'), show='tree headings', height=20)
        tree.pack(fill=tk.BOTH, expand=True)
        
        # Configure columns
        tree.column('#0', width=400, heading='Component Name')
        tree.column('status', width=150, heading='Status')
        tree.column('snapshot1', width=150, heading='Snapshot 1')
        tree.column('snapshot2', width=150, heading='Snapshot 2')
        
        # Configure tags for colors
        tree.tag_configure('modified', foreground='#f39c12')
        tree.tag_configure('added', foreground='#27ae60')
        tree.tag_configure('removed', foreground='#e74c3c')
        tree.tag_configure('unchanged', foreground='#95a5a6')
        
        # Populate tree
        for result in sorted(comparison_results, key=lambda r: (r['status'], r['name'])):
            comp_name = result['name']
            status = result['status']
            snap1 = result.get('snapshot1', {})
            snap2 = result.get('snapshot2', {})
            snap1_uuid = snap1.get('uuid', 'N/A')[:8] if snap1 else 'N/A'
            snap2_uuid = snap2.get('uuid', 'N/A')[:8] if snap2 else 'N/A'
            
            tag = status.lower().replace(' ', '_')
            tree.insert('', 'end', text=comp_name, values=(status, snap1_uuid, snap2_uuid), tags=(tag,))
        
        # Close button
        button_frame = tk.Frame(result_window, bg='#EAF3FB')
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        tk.Button(
            button_frame,
            text="Close",
            bg='#003366',
            fg='white',
            command=result_window.destroy
        ).pack()

    
    def run_folder_comparison(self, folder1, folder2, component_name="ALL", genmake_filter=None):
        """Run folder comparison in background thread"""
        # Reset progress
        self.progress_bar['value'] = 0
        self.progress_label.config(text="Starting comparison...")
        self.root.update()
        
        # Disable compare button during processing
        self.compare_btn.config(state='disabled')
        
        # Progress callback
        def update_progress(current, total, message):
            percentage = int((current / total) * 100) if total > 0 else 0
            self.progress_bar['value'] = percentage
            self.progress_label.config(text=message)
            self.root.update_idletasks()
        
        # RTC info (if enabled)
        rtc_info = None
        if self.rtc_enabled_var.get():
            rtc_info = {
                'enabled': True,
                'username': None,  # TODO: Add credential dialog
                'password': None,
                'repository_path': folder2,
                'workspace_name': None,
                'stream_name': None
            }
        
        # Run comparison in background thread
        def comparison_thread():
            try:
                result = compare_folders(
                    folder1,
                    folder2,
                    progress_callback=update_progress,
                    custom_mappings=None,  # TODO: Add file mapping dialog
                    rtc_info=rtc_info,
                    component_name=component_name,
                    genmake_filter=genmake_filter,
                )
                
                # Update UI on completion (must run in main thread)
                self.root.after(0, lambda: self.on_comparison_complete(result))
                
            except Exception as e:
                error_msg = f"Comparison failed: {str(e)}"
                self.root.after(0, lambda: self.on_comparison_error(error_msg))
        
        thread = threading.Thread(target=comparison_thread, daemon=True)
        thread.start()
    
    def on_comparison_complete(self, result):
        """Handle comparison completion"""
        self.compare_btn.config(state='normal')
        
        if result.get('success'):
            # Hide progress widgets temporarily
            self.progress_bar['value'] = 100
            self.progress_label.config(
                text=f"✅ Comparison complete! Showing results..."
            )
            self.root.update()
            
            # Show interactive results dialog
            try:
                show_results_dialog(
                    self.root,
                    result['results'],
                    result['folder1_display'],
                    result['folder2_display'],
                    result['folder1'],
                    result['folder2'],
                    result['files1'],
                    result['files2'],
                    result['report_paths']
                )
            except Exception as e:
                messagebox.showerror("Error", f"Error showing results: {str(e)}")
            
            # Clean up temp directories from ZIP extraction
            if result.get('temp_dirs'):
                cleanup_temp_dirs(result['temp_dirs'])
            
            # Show final success message
            output_dir = result['report_paths']['output_dir']
            total_files = len(result['results'])
            
            self.progress_label.config(
                text=f"✅ Done! {total_files} files compared. Reports saved to:\n{output_dir}"
            )
        else:
            error = result.get('error', 'Unknown error')
            messagebox.showerror("Comparison Failed", f"Error: {error}")
            self.progress_label.config(text=f"❌ Failed: {error}")
    
    def on_comparison_error(self, error_msg):
        """Handle comparison error"""
        self.compare_btn.config(state='normal')
        self.progress_label.config(text=f"❌ Error occurred")
        messagebox.showerror("Error", error_msg)
    
    def run(self):
        """Run the GUI application"""
        self.root.mainloop()


def launch_gui():
    """Launch the Migration Analysis Tool GUI"""
    app = MigrationAnalysisGUI()
    app.run()
