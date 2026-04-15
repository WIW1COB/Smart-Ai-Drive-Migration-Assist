import os
import difflib
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import re
import zipfile
import tempfile
import shutil
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from xml.dom import minidom
import subprocess
import json
import datetime

# ---------------------------------------------------------------------------
# RTC/ALM Configuration
# ---------------------------------------------------------------------------
RTC_SERVER_URL = "https://rb-alm-06-p.de.bosch.com/ccm"
CERT_PATH = os.path.join(os.path.dirname(__file__), "rb-alm-06-p-de-bosch-com-chain.pem")

# RTC Client Java Library path for workitem fetching
RTC_CLIENT_LIB_PATH = r"C:\Users\WIW1COB\Desktop\TOOL_Developed\Migration_Analysis_Report\NEW\NEW2\RTC-Client-plainJavaLib-6.0.6.1"

# RTC SCM CLI (lscm) Configuration - OPTIONAL for faster component fetching
# The tool will automatically fall back to REST API if lscm is not available or fails
# Using scm.exe directly instead of lscm.bat to bypass Java environment issues
# Example: r"C:\Program Files\IBM\RTC-SCM-CLI\scmtools\eclipse\scm.exe"
LSCM_PATH = r"C:\toolbase\lscm\7.0.3\jazz\scmtools\eclipse\scm.exe"  # BOSCH STEPS ALM SCM installation

# Global variables for RTC authentication
RTC_USERNAME = None
RTC_PASSWORD = None
RTC_ENABLED = False

# Global variables for RTC workspace/stream detection
RTC_WORKSPACE_NAME = None
RTC_STREAM_NAME = None
RTC_REPOSITORY_UUID = None

# Global variables for snapshot comparison
SNAPSHOT_MODE = False
SNAPSHOT1_URL = None
SNAPSHOT2_URL = None

# ---------------------------------------------------------------------------
# AI Configuration — set your API keys here (tool-level, no UI entry needed)
# ---------------------------------------------------------------------------
# Google Gemini Flash — FREE AI model used for AI Smart Merge
# Get a FREE key (no credit card) at: https://aistudio.google.com/app/apikey
# Sign in with any Google account → "Create API key" → paste below
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "AIzaSyD2TwpunJMKwCRsDh97-IqYsTWd__1W-UY")

# OpenAI GPT-4o-mini — used to enhance AI Suggest (optional; heuristics work without it)
# Get your key at: https://platform.openai.com/api-keys
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")

# ---------------------------------------------------------------------------
# Corporate Proxy Configuration (for Gemini API calls)
# ---------------------------------------------------------------------------
# Auto-detected Bosch proxy. Change PROXY_URL to "" to disable.
PROXY_URL    = os.environ.get("HTTPS_PROXY",
                   os.environ.get("HTTP_PROXY", "http://rb-proxy-in.bosch.com:8080"))
# Leave PROXY_USER / PROXY_PASS empty to be prompted once at runtime.
# Or fill in directly: PROXY_USER = "BOSCH\\WIW1COB", PROXY_PASS = "yourpassword"
PROXY_DOMAIN = os.environ.get("PROXY_DOMAIN", "BOSCH")
PROXY_USER   = os.environ.get("PROXY_USER",   "")  # e.g. "WIW1COB" (without domain)
PROXY_PASS   = os.environ.get("PROXY_PASS",   "")  # leave empty → prompted on first call

_proxy_cred_cache = {}   # in-memory cache so we only prompt once per session

# ---------------------------------------------------------------------------
# Utility: Sanitize text for Excel
# ---------------------------------------------------------------------------
def sanitize_for_excel(text):
    """Remove illegal characters that Excel cannot handle"""
    if not text:
        return text
    # Remove control characters (0x00-0x1F except tab, newline, carriage return)
    # and other problematic characters
    illegal_chars = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]')
    cleaned = illegal_chars.sub('', str(text))
    # Limit length to prevent Excel cell overflow
    if len(cleaned) > 32767:  # Excel cell character limit
        cleaned = cleaned[:32760] + "...[truncated]"
    return cleaned

# ---------------------------------------------------------------------------
# Utility: Count lines in file
# ---------------------------------------------------------------------------
def count_file_lines(file_path):
    """Count the number of lines in a file."""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return sum(1 for _ in f)
    except Exception:
        try:
            with open(file_path, 'rb') as f:
                return sum(1 for _ in f)
        except:
            return 0


# ---------------------------------------------------------------------------
# Utility: Determine line comparison status
# ---------------------------------------------------------------------------
def get_line_comparison_status(lines1, lines2, files_identical, text1_lines=None, text2_lines=None):
    """
    Determine the status based on line counts and file content.
    Returns a descriptive status string.
    """
    if lines1 == 0 and lines2 == 0:
        return "Both files empty"
    elif lines1 == 0:
        return f"{lines2} lines added in project"
    elif lines2 == 0:
        return f"{lines1} lines removed in project"
    elif lines1 == lines2:
        if files_identical:
            return "No change (same lines, same content)"
        else:
            # Check if only comments changed
            if text1_lines and text2_lines and is_only_comment_change(text1_lines, text2_lines):
                return "Comments update only"
            else:
                return "Same line count, but modifications occurred"
    elif lines2 > lines1:
        diff = lines2 - lines1
        # Check if only comments changed despite line count difference
        if text1_lines and text2_lines and is_only_comment_change(text1_lines, text2_lines):
            return f"Comments update only ({diff} line(s) added in project)"
        else:
            return f"{diff} line(s) added in project"
    else:  # lines2 < lines1
        diff = lines1 - lines2
        # Check if only comments changed despite line count difference
        if text1_lines and text2_lines and is_only_comment_change(text1_lines, text2_lines):
            return f"Comments update only ({diff} line(s) removed in project)"
        else:
            return f"{diff} line(s) removed in project"


# ---------------------------------------------------------------------------
# Utility: Remove comments from code
# ---------------------------------------------------------------------------
def remove_comments(text):
    """
    Remove C/C++ style comments (/* */ and //) from text.
    Returns text with comments removed.
    """
    # Remove multi-line comments /* ... */
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)
    # Remove single-line comments //
    text = re.sub(r'//.*?$', '', text, flags=re.MULTILINE)
    return text


# ---------------------------------------------------------------------------
# Utility: Check if only comments changed
# ---------------------------------------------------------------------------
def is_only_comment_change(text1_lines, text2_lines):
    """
    Check if the difference between two files is only in comments.
    Returns True if only comments changed, False otherwise.
    """
    # Join lines and remove comments from both
    text1 = ''.join(text1_lines)
    text2 = ''.join(text2_lines)
    
    text1_no_comments = remove_comments(text1)
    text2_no_comments = remove_comments(text2)
    
    # Normalize whitespace for comparison
    text1_normalized = ' '.join(text1_no_comments.split())
    text2_normalized = ' '.join(text2_no_comments.split())
    
    # If code without comments is the same, only comments changed
    return text1_normalized == text2_normalized


# ---------------------------------------------------------------------------
# Utility: Read file safely
# ---------------------------------------------------------------------------
def read_file_as_text(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.readlines()
    except Exception:
        with open(file_path, 'rb') as f:
            content = f.read().decode('latin-1', errors='ignore')
            return content.splitlines(keepends=True)

# ---------------------------------------------------------------------------
# XML comparison utilities
# ---------------------------------------------------------------------------
def normalize_xml(file_path):
    """Parse and normalize XML for comparison"""
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        # Pretty print XML for better comparison
        xml_str = ET.tostring(root, encoding='unicode')
        dom = minidom.parseString(xml_str)
        return dom.toprettyxml(indent="  ").splitlines(keepends=True)
    except Exception as e:
        # If XML parsing fails, treat as regular text file
        return read_file_as_text(file_path)

# ---------------------------------------------------------------------------
# ZIP extraction utilities
# ---------------------------------------------------------------------------
def extract_zip_to_temp(zip_path):
    """Extract ZIP file to temporary directory and return path"""
    try:
        temp_dir = tempfile.mkdtemp(prefix="migration_zip_")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        return temp_dir
    except Exception as e:
        print(f"Error extracting ZIP {zip_path}: {e}")
        return None

def prepare_folder_path(path):
    """Check if path is a ZIP file and extract it, otherwise return as-is
    Returns: (actual_path, is_temp_dir, original_path)
    """
    if not path:
        return None, False, None
    
    # Check if path is a ZIP file
    if os.path.isfile(path) and path.lower().endswith('.zip'):
        temp_dir = extract_zip_to_temp(path)
        if temp_dir:
            return temp_dir, True, path
        else:
            return None, False, path
    
    # Regular directory
    if os.path.isdir(path):
        return path, False, path
    
    return None, False, path

# ---------------------------------------------------------------------------
# Generate HTML diff
# ---------------------------------------------------------------------------
def generate_html_diff(file1, file2, file_name, output_dir):
    # Check if files are XML
    is_xml = file1.lower().endswith('.xml') and file2.lower().endswith('.xml')
    
    if is_xml:
        text1 = normalize_xml(file1)
        text2 = normalize_xml(file2)
    else:
        text1 = read_file_as_text(file1)
        text2 = read_file_as_text(file2)

    differ = difflib.HtmlDiff(wrapcolumn=120)
    html_diff = differ.make_file(
        text1, text2,
        fromdesc=f"{file1} (Migration Analysis)",
        todesc=f"{file2} (Migration Analysis)"
    )

    output_path = os.path.join(output_dir, f"{file_name.replace(os.sep,'_')}_diff.html")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_diff)
    return output_path, text1, text2

# ---------------------------------------------------------------------------
# Generate purpose of change
# ---------------------------------------------------------------------------
def generate_purpose_of_change(text1, text2):
    diff = list(difflib.ndiff(text1, text2))
    comments = []
    for line in diff:
        if line.startswith("+ "):
            line_text = line[2:].strip()
            # Only include printable text, skip binary/control characters
            if line_text and len(line_text) < 200:  # Skip very long lines
                # Check if line is mostly printable
                printable_ratio = sum(c.isprintable() or c in '\t\n\r' for c in line_text) / len(line_text)
                if printable_ratio > 0.7:  # At least 70% printable characters
                    comments.append(f"Added: {line_text}")
        elif line.startswith("- "):
            line_text = line[2:].strip()
            if line_text and len(line_text) < 200:
                printable_ratio = sum(c.isprintable() or c in '\t\n\r' for c in line_text) / len(line_text)
                if printable_ratio > 0.7:
                    comments.append(f"Removed: {line_text}")
    
    if not comments:
        return "No change detected."
    if len(comments) > 10:
        comments = comments[:10]
        comments.append("... (more differences omitted)")
    return " | ".join(comments)

# ---------------------------------------------------------------------------
# RTC Snapshot Comparison Functions
# ---------------------------------------------------------------------------
def fetch_snapshot_details(snapshot_url_or_id, username, password):
    """
    Fetch snapshot details from RTC using REST API.
    Returns snapshot information including name and UUID.
    Supports formats:
    - Direct UUID: _ojreQAAbEfG1br8X33nQcA
    - Web URL: https://rb-alm-06-p.de.bosch.com/ccm/web/projects/...&id=_i3S_vwAaEfG3rPS3zZLwKA&...
    - Resource URL: https://rb-alm-06-p.de.bosch.com/ccm/resource/itemOid/...
    """
    try:
        # Extract snapshot UUID from URL if needed
        snapshot_id = snapshot_url_or_id.strip()
        
        # Check if it's a web UI URL with id= parameter
        if "id=" in snapshot_id:
            # Extract UUID from id=_xxxxx format
            match = re.search(r'id=(_[a-zA-Z0-9_-]+)', snapshot_id)
            if match:
                snapshot_id = match.group(1)
                print(f"Extracted snapshot ID from web URL: {snapshot_id}")
        # Check if it's a resource URL format
        elif "/_" in snapshot_id:
            snapshot_id = snapshot_id.split("/_")[-1]
            if "?" in snapshot_id:
                snapshot_id = snapshot_id.split("?")[0]
            print(f"Extracted snapshot ID from resource URL: {snapshot_id}")
        # Otherwise assume it's a direct UUID
        else:
            print(f"Using direct snapshot UUID: {snapshot_id}")
        
        print(f"Fetching snapshot details for: {snapshot_id}")
        
        # Build REST API URL - use BaselineSet for snapshots
        api_url = f"{RTC_SERVER_URL}/resource/itemOid/com.ibm.team.scm.BaselineSet/{snapshot_id}"
        
        cert_param = []
        if os.path.exists(CERT_PATH):
            cert_param = ["--cacert", CERT_PATH]
        
        curl_command = [
            "curl",
            "-k",
            "-L",
            "-u", f"{username}:{password}",
            *cert_param,
            "-X", "GET",
            "-H", "Accept: application/json",
            api_url
        ]
        
        print("Fetching snapshot details (timeout: 180s, please wait...)")
        result = subprocess.run(curl_command, capture_output=True, text=True, timeout=180)
        
        if result.returncode != 0:
            print(f"Failed to fetch snapshot: {result.stderr}")
            return None
        
        if not result.stdout.strip():
            print("No data returned from snapshot query")
            return None
        
        snapshot_data = json.loads(result.stdout)
        
        # Extract snapshot name from various possible fields
        snapshot_name = snapshot_data.get('name', 
                        snapshot_data.get('dc:title', 
                        snapshot_data.get('title', 'Unknown')))
        
        print(f"✓ Snapshot fetched: {snapshot_name}")
        print(f"  Snapshot details: name={snapshot_name}, itemId={snapshot_data.get('itemId', 'N/A')}")
        
        return snapshot_data
        
    except Exception as e:
        print(f"Error fetching snapshot details: {e}")
        import traceback
        traceback.print_exc()
        return None

def fetch_components_using_lscm(snapshot_url_or_id, username, password):
    """
    Fetch components from a snapshot using lscm (RTC SCM command-line tool).
    Command: lscm list components <snapshot>
    Returns list of component dictionaries with name and uuid.
    """
    try:
        # Extract snapshot UUID if it's a URL
        snapshot_id = snapshot_url_or_id.strip()
        if "id=" in snapshot_id:
            match = re.search(r'id=(_[a-zA-Z0-9_-]+)', snapshot_id)
            if match:
                snapshot_id = match.group(1)
        elif "/_" in snapshot_id:
            snapshot_id = snapshot_id.split("/_")[-1]
            if "?" in snapshot_id:
                snapshot_id = snapshot_id.split("?")[0]
        
        print(f"Fetching components using scm for snapshot: {snapshot_id}")
        
        # Check if lscm is configured
        if not LSCM_PATH:
            # lscm disabled - use REST API directly
            return None
        
        scm_executable = LSCM_PATH
        
        # Check if scm.exe exists
        if not os.path.exists(scm_executable):
            print(f"scm executable not found at: {scm_executable}")
            return None
        
        print(f"✓ Using RTC SCM CLI (scm.exe) for component fetching")
        
        # Step 1: Login to RTC repository
        print(f"Logging in to {RTC_SERVER_URL}...")
        login_command = [
            scm_executable, "login",
            "-r", RTC_SERVER_URL,
            "-u", username,
            "-P", password
        ]
        
        # Set environment to disable proxy
        env = os.environ.copy()
        env['NO_PROXY'] = '*'
        env['no_proxy'] = '*'
        env['HTTP_PROXY'] = ''
        env['HTTPS_PROXY'] = ''
        env['http_proxy'] = ''
        env['https_proxy'] = ''
        
        try:
            login_result = subprocess.run(
                login_command,
                capture_output=True,
                text=True,
                timeout=60,
                cwd=os.path.dirname(scm_executable),
                env=env
            )
            
            if login_result.returncode != 0:
                print(f"Login failed: {login_result.stderr if login_result.stderr else 'Unknown error'}")
                return None
            
            print("✓ Login successful")
        except subprocess.TimeoutExpired:
            print("Login timeout")
            return None
        except Exception as e:
            print(f"Login error: {e}")
            return None
        
        # Step 2: List components in snapshot
        print(f"Listing components in snapshot {snapshot_id}...")
        list_command = [
            scm_executable, "list", "components",
            "-r", RTC_SERVER_URL,
            "-s", snapshot_id,
            "-j",  # JSON output format
            "--maximum", "10000"  # Maximum number of components to fetch (default is only 25)
        ]
        
        print(f"Running: scm list components -s {snapshot_id} -j --maximum 10000")
        
        # Execute scm list components command (use same env with no proxy)
        result = subprocess.run(
            list_command,
            capture_output=True,
            text=True,
            timeout=120,
            cwd=os.path.dirname(scm_executable),
            env=env
        )
        
        if result.returncode != 0:
            print(f"scm list components failed (exit code: {result.returncode})")
            if result.stderr:
                print(f"Error: {result.stderr}")
            if result.stdout:
                print(f"Output: {result.stdout}")
            print(f"Falling back to REST API method...")
            return None
        
        if not result.stdout.strip():
            print("No output from scm command")
            return None
        
        # Debug: Show raw output
        # print(f"\nRaw scm output:")
        # print(f"{'-'*80}")
        # print(result.stdout)
        # print(f"{'-'*80}\n")
        
        # Parse scm JSON output
        components = []
        try:
            # Parse JSON format from scm -j output
            output_data = json.loads(result.stdout)
            
            # Handle the actual scm JSON structure:
            # {
            #   "name": "snapshot_name",
            #   "uuid": "snapshot_uuid",
            #   "components": [
            #     {
            #       "name": "component_name",
            #       "uuid": "component_uuid",
            #       "baseline": { "name": "baseline_name", "uuid": "baseline_uuid" }
            #     }
            #   ]
            # }
            
            if isinstance(output_data, dict) and "components" in output_data:
                # New format with components array
                component_list = output_data.get("components", [])
                print(f"Found {len(component_list)} components in JSON response")
                
                for item in component_list:
                    comp_name = item.get("name", "")
                    comp_uuid = item.get("uuid", "")
                    
                    # Get baseline info if available
                    baseline_info = item.get("baseline", {})
                    baseline_uuid = baseline_info.get("uuid", "") if isinstance(baseline_info, dict) else ""
                    baseline_name = baseline_info.get("name", "") if isinstance(baseline_info, dict) else ""
                    
                    if comp_name or comp_uuid:
                        components.append({
                            "name": comp_name if comp_name else comp_uuid,
                            "uuid": comp_uuid,
                            "baseline_uuid": baseline_uuid,
                            "baseline_name": baseline_name,
                            "state_id": ""
                        })
                        print(f"  - {comp_name} ({comp_uuid[:20]}...)")
                        
            elif isinstance(output_data, list):
                # Legacy format - direct array of components
                for item in output_data:
                    comp_name = item.get("name", "")
                    comp_uuid = item.get("uuid", item.get("item-id", ""))
                    if comp_name or comp_uuid:
                        components.append({
                            "name": comp_name if comp_name else comp_uuid,
                            "uuid": comp_uuid,
                            "baseline_uuid": item.get("baseline-uuid", ""),
                            "baseline_name": "",
                            "state_id": ""
                        })
            else:
                print(f"Unexpected JSON structure: {type(output_data)}")
                print(f"Keys: {output_data.keys() if isinstance(output_data, dict) else 'N/A'}")
                
        except json.JSONDecodeError:
            # Fallback: Parse text output
            # Expected formats:
            # "(1234) Component_Name (_uuid123)"
            # "Component_Name (_uuid123)"
            # "Component_Name _uuid123"
            # or table format with headers
            lines = result.stdout.strip().split('\n')
            
            for line in lines:
                line = line.strip()
                
                # Skip empty lines, headers, separators
                if not line or line.startswith('-') or line.startswith('='):
                    continue
                if 'Component' in line and 'Name' in line:  # Header row
                    continue
                if line.startswith('Workspace:') or line.startswith('Stream:'):
                    continue
                
                # Try multiple parsing patterns
                # Pattern 1: "(1234) Component_Name (_uuid123)" or "Component_Name (_uuid123)"
                match = re.search(r'(?:\(\d+\)\s+)?([^(]+)\s+\((_[a-zA-Z0-9_-]+)\)', line)
                if match:
                    comp_name = match.group(1).strip()
                    comp_uuid = match.group(2).strip()
                    components.append({
                        "name": comp_name,
                        "uuid": comp_uuid,
                        "baseline_uuid": "",
                        "state_id": ""
                    })
                    continue
                
                # Pattern 2: Simple format "Component_Name _uuid123"
                parts = line.split()
                if len(parts) >= 2:
                    # Look for UUID pattern in parts
                    for i, part in enumerate(parts):
                        if part.startswith('_') and len(part) > 10:
                            comp_uuid = part
                            comp_name = ' '.join(parts[:i])
                            if comp_name:
                                components.append({
                                    "name": comp_name,
                                    "uuid": comp_uuid,
                                    "baseline_uuid": "",
                                    "state_id": ""
                                })
                            break
        
        if components:
            print(f"✓ Found {len(components)} components using scm CLI")
            print(f"\nComponents in snapshot:")
            print(f"{'-'*80}")
            for idx, comp in enumerate(components, 1):
                print(f"{idx:3d}. {comp['name']}")
                if comp['uuid']:
                    print(f"      UUID: {comp['uuid']}")
            print(f"{'-'*80}")
        else:
            print("No components found in scm output")
        
        return components
        
    except subprocess.TimeoutExpired:
        print("Timeout: scm command took too long (>120s)")
        print("Falling back to REST API method...")
        return None
    except Exception as e:
        print(f"Error using scm CLI: {e}")
        import traceback
        traceback.print_exc()
        print("Falling back to REST API method...")
        return None

def fetch_snapshot_components(snapshot_url_or_id, username, password, snapshot_details=None):
    """
    Fetch list of components in a snapshot.
    Returns list of component names and UUIDs.
    Supports formats:
    - Direct UUID: _ojreQAAbEfG1br8X33nQcA
    - Web URL: https://rb-alm-06-p.de.bosch.com/ccm/web/projects/...&id=_i3S_vwAaEfG3rPS3zZLwKA&...
    - Resource URL: https://rb-alm-06-p.de.bosch.com/ccm/resource/itemOid/...
    
    If snapshot_details is provided, it will extract components from it directly.
    Primary method: lscm list components command
    Fallback: REST API
    """
    try:
        # Method 1: Try lscm command first (fastest and most reliable)
        print("=" * 80)
        print("Method 1: Attempting to fetch components using lscm command...")
        print("=" * 80)
        lscm_components = fetch_components_using_lscm(snapshot_url_or_id, username, password)
        if lscm_components:
            return lscm_components
        
        print("\nlscm method failed, falling back to REST API...")
        print("=" * 80)
        
        # Method 2: If snapshot_details already provided, extract components from it
        if snapshot_details:
            print(f"Extracting components from snapshot details...")
            components = []
            
            # Check for baseline list in snapshot details
            if isinstance(snapshot_details, dict) and "com.ibm.team.scm.Baseline" in snapshot_details:
                baseline_list = snapshot_details["com.ibm.team.scm.Baseline"]
                print(f"Found {len(baseline_list) if isinstance(baseline_list, list) else 0} baselines in snapshot details")
                
                if isinstance(baseline_list, list):
                    print(f"Fetching component names for {len(baseline_list)} baselines...")
                    for idx, baseline_ref in enumerate(baseline_list):
                        if isinstance(baseline_ref, dict):
                            item_id = baseline_ref.get("itemId", "")
                            state_id = baseline_ref.get("stateId", "")
                            
                            # Fetch baseline details to get component name
                            baseline_url = f"{RTC_SERVER_URL}/resource/itemOid/com.ibm.team.scm.Baseline/{item_id}"
                            cert_param = []
                            if os.path.exists(CERT_PATH):
                                cert_param = ["--cacert", CERT_PATH]
                            
                            curl_baseline = [
                                "curl", "-k", "-L",
                                "-u", f"{username}:{password}",
                                *cert_param,
                                "-X", "GET",
                                "-H", "Accept: application/json",
                                baseline_url
                            ]
                            
                            try:
                                baseline_result = subprocess.run(curl_baseline, capture_output=True, text=True, timeout=30)
                                if baseline_result.returncode == 0 and baseline_result.stdout.strip():
                                    baseline_data = json.loads(baseline_result.stdout)
                                    
                                    # Extract component info from baseline
                                    comp_ref = baseline_data.get("com.ibm.team.scm.Component", {})
                                    comp_item_id = item_id
                                    comp_name = baseline_data.get("name", item_id)
                                    
                                    # If we have a component reference, fetch the component name
                                    if isinstance(comp_ref, dict) and comp_ref.get("itemId"):
                                        comp_item_id = comp_ref.get("itemId", item_id)
                                        
                                        # Try to get component name from the reference
                                        if "name" in comp_ref:
                                            comp_name = comp_ref.get("name")
                                        else:
                                            # Fetch component details to get the name
                                            comp_url = f"{RTC_SERVER_URL}/resource/itemOid/com.ibm.team.scm.Component/{comp_item_id}"
                                            curl_comp = [
                                                "curl", "-k", "-L",
                                                "-u", f"{username}:{password}",
                                                *cert_param,
                                                "-X", "GET",
                                                "-H", "Accept: application/json",
                                                comp_url
                                            ]
                                            try:
                                                comp_result = subprocess.run(curl_comp, capture_output=True, text=True, timeout=30)
                                                if comp_result.returncode == 0 and comp_result.stdout.strip():
                                                    comp_data = json.loads(comp_result.stdout)
                                                    comp_name = comp_data.get("name", comp_name)
                                            except:
                                                pass  # Keep baseline name as fallback
                                    
                                    components.append({
                                        "name": comp_name,
                                        "uuid": comp_item_id,
                                        "baseline_uuid": item_id,
                                        "state_id": state_id
                                    })
                                    
                                    if (idx + 1) % 10 == 0:
                                        print(f"  Processed {idx + 1}/{len(baseline_list)} baselines...")
                            except Exception as e:
                                print(f"  Warning: Failed to fetch baseline {item_id[:20]}: {e}")
                                components.append({
                                    "name": item_id,
                                    "uuid": item_id,
                                    "baseline_uuid": item_id,
                                    "state_id": state_id
                                })
                    
                    print(f"✓ Found {len(components)} components")
                    print(f"\nComponents in this snapshot:")
                    print(f"{'-'*80}")
                    for idx, comp in enumerate(components, 1):
                        print(f"{idx:3d}. {comp['name']}")
                        print(f"      Baseline ID: {comp['baseline_uuid']}")
                    print(f"{'-'*80}")
                    
                    return components
            
            print("Warning: Could not find baseline list in snapshot details, fetching separately...")
            # Extract snapshot UUID and fetch baseline list
            snapshot_uuid = snapshot_details.get("itemId", "")
            if snapshot_uuid:
                print(f"Using snapshot UUID from details: {snapshot_uuid}")
                snapshot_id = snapshot_uuid
            else:
                snapshot_id = snapshot_url_or_id.strip()
        else:
            snapshot_id = snapshot_url_or_id.strip()
        
        # Extract UUID if it's a URL
        
        # Check if it's a web UI URL with id= parameter
        if "id=" in snapshot_id:
            # Extract UUID from id=_xxxxx format
            match = re.search(r'id=(_[a-zA-Z0-9_-]+)', snapshot_id)
            if match:
                snapshot_id = match.group(1)
                print(f"Extracted snapshot ID from web URL: {snapshot_id}")
        # Check if it's a resource URL format
        elif "/_" in snapshot_id:
            snapshot_id = snapshot_id.split("/_")[-1]
            if "?" in snapshot_id:
                snapshot_id = snapshot_id.split("?")[0]
            print(f"Extracted snapshot ID from resource URL: {snapshot_id}")
        else:
            print(f"Using direct snapshot UUID: {snapshot_id}")
        
        print(f"Fetching components for snapshot: {snapshot_id}")
        
        # Query snapshot resource using BaselineSet type (correct RTC type for snapshots)
        api_url = f"{RTC_SERVER_URL}/resource/itemOid/com.ibm.team.scm.BaselineSet/{snapshot_id}"
        
        cert_param = []
        if os.path.exists(CERT_PATH):
            cert_param = ["--cacert", CERT_PATH]
        
        curl_command = [
            "curl",
            "-k",  # Skip SSL verification if certificate issues
            "-L",
            "-u", f"{username}:{password}",
            *cert_param,
            "-X", "GET",
            "-H", "Accept: application/json",
            api_url
        ]
        
        print("Fetching snapshot details (this may take a while for large snapshots)...")
        result = subprocess.run(curl_command, capture_output=True, text=True, timeout=180)
        
        if result.returncode != 0:
            print(f"Failed to fetch components from: {api_url}")
            print(f"Error: {result.stderr}")
            print(f"Response: {result.stdout}")
            
            # Try reportable REST API
            print("\nTrying reportable REST API endpoint...")
            alt_api_url = f"{RTC_SERVER_URL}/reportablerest/snapshot/{snapshot_id}"
            curl_command_alt = [
                "curl",
                "-k",
                "-L",
                "-u", f"{username}:{password}",
                *cert_param,
                "-X", "GET",
                "-H", "Accept: application/json",
                alt_api_url
            ]
            
            print("Fetching from alternative endpoint (this may take a while)...")
            result = subprocess.run(curl_command_alt, capture_output=True, text=True, timeout=180)
            if result.returncode != 0:
                print(f"Reportable REST API failed: {result.stderr}")
                return []
        
        if not result.stdout.strip():
            print("No components data returned (empty response)")
            return []
        
        print(f"API Response (first 1000 chars): {result.stdout[:1000]}")
        
        # Check if response is HTML (error page)
        if result.stdout.strip().startswith("<"):
            print("ERROR: Received HTML response instead of JSON (likely authentication or API error)")
            print(f"Full response: {result.stdout}")
            return []
        
        try:
            data = json.loads(result.stdout)
        except json.JSONDecodeError as e:
            print(f"ERROR: Failed to parse JSON response: {e}")
            print(f"Raw response: {result.stdout}")
            return []
        
        components = []
        
        # Debug: Show the top-level keys in the response
        if isinstance(data, dict):
            print(f"Response keys: {list(data.keys())}")
        
        # Parse components - RTC BaselineSet format
        # The response should contain "com.ibm.team.scm.Baseline" key with baseline references
        if isinstance(data, dict) and "com.ibm.team.scm.Baseline" in data:
            baseline_list = data["com.ibm.team.scm.Baseline"]
            print(f"Found 'com.ibm.team.scm.Baseline' key with {len(baseline_list) if isinstance(baseline_list, list) else 'unknown'} items")
            
            if isinstance(baseline_list, list):
                print(f"Fetching component details for {len(baseline_list)} baselines...")
                for idx, baseline_ref in enumerate(baseline_list):
                    if isinstance(baseline_ref, dict):
                        # Extract baseline UUID
                        item_id = baseline_ref.get("itemId", "")
                        state_id = baseline_ref.get("stateId", "")
                        
                        # Fetch the baseline details to get component information
                        baseline_url = f"{RTC_SERVER_URL}/resource/itemOid/com.ibm.team.scm.Baseline/{item_id}"
                        curl_baseline = [
                            "curl",
                            "-k",
                            "-L",
                            "-u", f"{username}:{password}",
                            *cert_param,
                            "-X", "GET",
                            "-H", "Accept: application/json",
                            baseline_url
                        ]
                        
                        try:
                            baseline_result = subprocess.run(curl_baseline, capture_output=True, text=True, timeout=30)
                            if baseline_result.returncode == 0 and baseline_result.stdout.strip():
                                baseline_data = json.loads(baseline_result.stdout)
                                
                                # Extract component info from baseline
                                # Try to get component reference first
                                comp_ref = baseline_data.get("com.ibm.team.scm.Component", {})
                                comp_item_id = ""
                                comp_name = baseline_data.get("name", "Unknown")
                                
                                # If we have a component reference, fetch the component name
                                if isinstance(comp_ref, dict) and comp_ref.get("itemId"):
                                    comp_item_id = comp_ref.get("itemId", "")
                                    
                                    # Try to get component name from the reference
                                    if "name" in comp_ref:
                                        comp_name = comp_ref.get("name")
                                    else:
                                        # Fetch component details to get the name
                                        comp_url = f"{RTC_SERVER_URL}/resource/itemOid/com.ibm.team.scm.Component/{comp_item_id}"
                                        curl_comp = [
                                            "curl", "-k", "-L",
                                            "-u", f"{username}:{password}",
                                            *cert_param,
                                            "-X", "GET",
                                            "-H", "Accept: application/json",
                                            comp_url
                                        ]
                                        try:
                                            comp_result = subprocess.run(curl_comp, capture_output=True, text=True, timeout=30)
                                            if comp_result.returncode == 0 and comp_result.stdout.strip():
                                                comp_data = json.loads(comp_result.stdout)
                                                comp_name = comp_data.get("name", comp_name)
                                        except:
                                            pass  # Keep baseline name as fallback
                                else:
                                    comp_item_id = item_id
                                
                                component = {
                                    "name": comp_name,
                                    "uuid": comp_item_id,
                                    "baseline_uuid": item_id,
                                    "state_id": state_id
                                }
                                components.append(component)
                                
                                if (idx + 1) % 10 == 0:
                                    print(f"  Processed {idx + 1}/{len(baseline_list)} baselines...")
                        except Exception as e:
                            print(f"  Warning: Failed to fetch baseline {item_id[:20]}: {e}")
                            # Use baseline UUID as fallback name
                            component = {
                                "name": item_id,
                                "uuid": item_id,
                                "baseline_uuid": item_id,
                                "state_id": state_id
                            }
                            components.append(component)
                
                print(f"Completed fetching component details")
                # Display all fetched components
                print(f"\nComponents found in snapshot:")
                for comp in components:
                    print(f"  - {comp['name']}")
            else:
                print(f"Warning: 'com.ibm.team.scm.Baseline' is not a list: {type(baseline_list)}")
        
        # Try alternative formats
        elif "baselines" in data:
            for baseline in data["baselines"]:
                component = {
                    "name": baseline.get("component", {}).get("name", baseline.get("name", "Unknown")),
                    "uuid": baseline.get("component", {}).get("itemId", baseline.get("itemId", "")),
                    "baseline_uuid": baseline.get("itemId", "")
                }
                components.append(component)
                print(f"  - Component: {component['name']}")
        
        elif isinstance(data, list):
            # Response might be a direct list
            for item in data:
                component = {
                    "name": item.get("name", item.get("component", {}).get("name", "Unknown")),
                    "uuid": item.get("itemId", item.get("uuid", "")),
                    "baseline_uuid": item.get("itemId", "")
                }
                components.append(component)
                print(f"  - Component: {component['name']}")
        
        else:
            # Show full response for debugging
            print(f"WARNING: Unexpected response structure")
            print(f"Response type: {type(data)}")
            if isinstance(data, dict):
                print(f"Response keys: {list(data.keys())}")
                print(f"Full response (first 2000 chars):")
                print(json.dumps(data, indent=2)[:2000])
        
        print(f"✓ Found {len(components)} components")
        return components
        
    except Exception as e:
        print(f"Error fetching snapshot components: {e}")
        import traceback
        traceback.print_exc()
        return []

# ---------------------------------------------------------------------------
# RTC Workspace Detection Functions
# ---------------------------------------------------------------------------
def detect_rtc_workspace_and_stream(folder_path, username=None, password=None):
    """
    Detect RTC workspace and stream information from a folder path.
    Returns a dict with workspace_name, stream_name, and repository_workspace_uuid.
    Returns None if the folder is not an RTC workspace.
    """
    if not folder_path or not os.path.exists(folder_path):
        print(f"Path does not exist: {folder_path}")
        return None
    
    # Handle ZIP files - cannot detect workspace from ZIP
    if folder_path.lower().endswith('.zip'):
        print(f"Workspace detection skipped: {folder_path} is a ZIP file")
        return None
    
    print(f"\nSearching for RTC workspace markers in: {folder_path}")
    
    # Check if this is an RTC workspace by looking for multiple possible markers
    workspace_root = folder_path
    is_rtc_workspace = False
    
    # Search upward for RTC workspace markers
    # Try: .jazz5, .metadata/.plugins/com.ibm.team.filesystem.client, .jazzShed
    current_path = folder_path
    max_depth = 10  # Limit upward search
    depth = 0
    
    while current_path and depth < max_depth:
        print(f"  Checking: {current_path}")
        
        # Check for .jazz5 folder (Eclipse-based RTC client)
        if os.path.exists(os.path.join(current_path, ".jazz5")):
            workspace_root = current_path
            is_rtc_workspace = True
            print(f"  ✓ Found .jazz5 folder at: {workspace_root}")
            break
        
        # Check for .metadata folder (Eclipse workspace)
        metadata_path = os.path.join(current_path, ".metadata", ".plugins", "com.ibm.team.filesystem.client")
        if os.path.exists(metadata_path):
            workspace_root = current_path
            is_rtc_workspace = True
            print(f"  ✓ Found Eclipse RTC metadata at: {workspace_root}")
            break
        
        # Check for .jazzShed folder (SCM CLI)
        if os.path.exists(os.path.join(current_path, ".jazzShed")):
            workspace_root = current_path
            is_rtc_workspace = True
            print(f"  ✓ Found .jazzShed folder at: {workspace_root}")
            break
        
        parent = os.path.dirname(current_path)
        if parent == current_path:  # Reached root
            break
        current_path = parent
        depth += 1
    
    if not is_rtc_workspace:
        print(f"  ✗ No RTC workspace markers found (.jazz5, .metadata, .jazzShed)")
        print(f"\n  Tip: Make sure you're comparing folders from an RTC workspace.")
        print(f"  The folder should be loaded from RTC and contain workspace metadata.\n")
        return None
    
    print(f"\n✓ RTC workspace root identified: {workspace_root}")
    
    # Get stream info from the workspace
    stream_info = get_stream_info_from_workspace(workspace_root, username, password)
    
    # If scm command failed, try reading from .jazz5 metadata
    if not stream_info:
        print(f"\n  Trying alternative method: reading from .jazz5 metadata...")
        stream_info = get_workspace_info_from_metadata(workspace_root)
    
    if stream_info:
        print(f"\n✓ RTC Workspace details:")
        print(f"  - Root: {workspace_root}")
        print(f"  - Workspace: {stream_info.get('workspace_name', 'Unknown')}")
        print(f"  - Stream: {stream_info.get('stream_name', 'Unknown')}")
        if stream_info.get('repository_workspace_uuid'):
            print(f"  - UUID: {stream_info.get('repository_workspace_uuid')}")
    else:
        print(f"\n⚠ Could not extract workspace details using 'scm status' command.")
        print(f"  This might mean:")
        print(f"  - SCM command-line tools are not installed")
        print(f"  - The workspace is not loaded/connected to RTC server")
        print(f"  - Authentication failed\n")
    
    return stream_info

def get_stream_info_from_workspace(workspace_path, username=None, password=None):
    """
    Extract workspace and stream information using 'scm status' command.
    Returns a dictionary with workspace_name, stream_name, and repository_workspace_uuid.
    """
    try:
        # First check if scm command is available
        try:
            check_scm = subprocess.run(["scm", "--version"], capture_output=True, timeout=10)
            if check_scm.returncode != 0:
                print(f"  ✗ 'scm' command not found or not working")
                print(f"    Please install RTC SCM Command Line Interface from your RTC server")
                return None
            print(f"  ✓ SCM command-line tool is available")
        except FileNotFoundError:
            print(f"  ✗ 'scm' command not found in system PATH")
            print(f"    Please install RTC SCM Command Line Interface")
            return None
        
        # Build scm status command
        scm_command = ["scm", "status", "-v"]
        
        # Add authentication if provided
        if username and password:
            scm_command.extend([
                "-r", RTC_SERVER_URL,
                "-u", username,
                "-P", password
            ])
            print(f"  Running: scm status -v -r {RTC_SERVER_URL} -u {username}")
        else:
            print(f"  Running: scm status -v (no authentication)")
        
        # Run the command in the workspace directory
        result = subprocess.run(
            scm_command,
            capture_output=True,
            text=True,
            timeout=60,
            cwd=workspace_path
        )
        
        if result.returncode != 0:
            print(f"  ✗ SCM status command failed (exit code: {result.returncode})")
            if result.stderr:
                print(f"    Error: {result.stderr.strip()}")
            if result.stdout:
                print(f"    Output: {result.stdout.strip()}")
            return None
        
        print(f"  ✓ SCM status command succeeded")
        
        # Parse the output to extract workspace and stream information
        output = result.stdout
        print(f"\n  SCM Status Output:")
        print(f"  {'-'*60}")
        for line in output.split('\n')[:15]:  # Show first 15 lines
            if line.strip():
                print(f"  {line}")
        print(f"  {'-'*60}\n")
        
        workspace_info = {
            'workspace_name': None,
            'stream_name': None,
            'repository_workspace_uuid': None
        }
        
        # Parse output lines
        for line in output.split('\n'):
            line = line.strip()
            
            # Extract workspace name
            if line.startswith("Workspace:"):
                match = re.search(r'Workspace:\s+(.+?)\s+\(([^)]+)\)', line)
                if match:
                    workspace_info['workspace_name'] = match.group(1).strip()
                    workspace_info['repository_workspace_uuid'] = match.group(2).strip()
                    print(f"  Extracted Workspace: {workspace_info['workspace_name']}")
                else:
                    workspace_name = line.replace("Workspace:", "").strip()
                    workspace_info['workspace_name'] = workspace_name
                    print(f"  Extracted Workspace (no UUID): {workspace_name}")
            
            # Extract stream/flow target
            elif "Stream:" in line or "Flow Target:" in line:
                match = re.search(r'(Stream|Flow Target):\s+(.+?)(?:\s+\(([^)]+)\))?$', line)
                if match:
                    workspace_info['stream_name'] = match.group(2).strip()
                    print(f"  Extracted Stream: {workspace_info['stream_name']}")
        
        if workspace_info['workspace_name']:
            print(f"\n  ✓ Successfully parsed workspace information")
        else:
            print(f"\n  ✗ Could not parse workspace name from scm status output")
        
        return workspace_info if workspace_info['workspace_name'] else None
        
    except subprocess.TimeoutExpired:
        print(f"  ✗ Timeout running scm status in {workspace_path}")
        return None
    except Exception as e:
        print(f"  ✗ Error getting stream info: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_workspace_info_from_metadata(workspace_root):
    """
    Extract workspace information from .jazz5 metadata files.
    This is a fallback when scm command is not available.
    """
    try:
        jazz5_path = os.path.join(workspace_root, ".jazz5")
        if not os.path.exists(jazz5_path):
            return None
        
        # Try to read properties file
        props_file = os.path.join(jazz5_path, "repository.properties")
        if os.path.exists(props_file):
            workspace_info = {
                'workspace_name': None,
                'stream_name': None,
                'repository_workspace_uuid': None
            }
            
            with open(props_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
                # Extract workspace name
                match = re.search(r'workspace[._-]?name\s*[=:]\s*(.+)', content, re.IGNORECASE)
                if match:
                    workspace_info['workspace_name'] = match.group(1).strip()
                
                # Extract UUID
                match = re.search(r'workspace[._-]?uuid\s*[=:]\s*([a-zA-Z0-9_-]+)', content, re.IGNORECASE)
                if match:
                    workspace_info['repository_workspace_uuid'] = match.group(1).strip()
                
                # Extract stream
                match = re.search(r'stream[._-]?name\s*[=:]\s*(.+)', content, re.IGNORECASE)
                if match:
                    workspace_info['stream_name'] = match.group(1).strip()
            
            if workspace_info['workspace_name']:
                print(f"  ✓ Extracted from metadata: {workspace_info['workspace_name']}")
                return workspace_info
        
        # Try to infer from directory structure
        workspace_name = os.path.basename(workspace_root)
        print(f"  ⚠ Using folder name as workspace: {workspace_name}")
        return {
            'workspace_name': workspace_name,
            'stream_name': 'Unknown',
            'repository_workspace_uuid': None
        }
        
    except Exception as e:
        print(f"  Error reading metadata: {e}")
        # Last resort: use folder name
        workspace_name = os.path.basename(workspace_root)
        return {
            'workspace_name': workspace_name,
            'stream_name': 'Unknown',
            'repository_workspace_uuid': None
        }

# ---------------------------------------------------------------------------
# RTC/ALM Integration Functions
# ---------------------------------------------------------------------------
def fetch_file_changesets_from_scm(file_path, repository_path, username=None, password=None, workspace_name=None, stream_name=None):
    """
    Fetch changesets/history for a specific file from RTC SCM using scm command-line tool.
    Returns a list of changeset information with UUID, comment, and date.
    """
    if not RTC_ENABLED or not username or not password:
        return []
    
    try:
        # Normalize file path to be relative to repository
        if os.path.isabs(file_path):
            try:
                rel_file_path = os.path.relpath(file_path, repository_path)
            except ValueError:
                rel_file_path = os.path.basename(file_path)
        else:
            rel_file_path = file_path
        
        # Convert to forward slashes for RTC
        rel_file_path = rel_file_path.replace("\\", "/")
        
        # Try using scm command-line tool first
        # scm history command shows the history of a file
        scm_command = [
            "scm",
            "history",
            rel_file_path,
            "-r", RTC_SERVER_URL,
            "-u", username,
            "-P", password,
            "-m", "10"  # Maximum 10 recent changesets
        ]
        
        # Add workspace context if available
        if workspace_name:
            scm_command.extend(["--workspace", workspace_name])
            print(f"Using workspace context: {workspace_name}")
        elif RTC_WORKSPACE_NAME:
            scm_command.extend(["--workspace", RTC_WORKSPACE_NAME])
            print(f"Using global workspace context: {RTC_WORKSPACE_NAME}")
        
        # Set working directory to repository path for scm command
        work_dir = repository_path if os.path.isdir(repository_path) else os.path.dirname(repository_path)
        
        result = subprocess.run(
            scm_command, 
            capture_output=True, 
            text=True, 
            timeout=60,
            cwd=work_dir
        )
        
        changesets = []
        
        if result.returncode == 0 and result.stdout:
            # Parse scm history output
            # Format typically includes change-set UUID, date, user, and comment
            lines = result.stdout.split('\n')
            current_changeset = {}
            
            for line in lines:
                line = line.strip()
                if not line:
                    if current_changeset and current_changeset.get("uuid"):
                        changesets.append(current_changeset)
                        current_changeset = {}
                    continue
                
                # Parse changeset information
                if line.startswith("Change sets:") or line.startswith("("):
                    # Extract UUID from format like: (1234) "comment" author date
                    match = re.search(r'\((\d+)\)', line)
                    if match:
                        cs_id = match.group(1)
                        current_changeset["uuid"] = cs_id
                        
                        # Extract comment
                        comment_match = re.search(r'"([^"]*)"', line)
                        if comment_match:
                            current_changeset["comment"] = comment_match.group(1)
                        
                        # Build changeset URL
                        current_changeset["url"] = f"{RTC_SERVER_URL}/resource/itemName/com.ibm.team.scm.ChangeSet/{cs_id}"
            
            # Add last changeset if exists
            if current_changeset and current_changeset.get("uuid"):
                changesets.append(current_changeset)
        
        # Fallback to REST API if scm command fails
        if not changesets:
            changesets = fetch_file_changesets_from_rest_api(rel_file_path, username, password)
        
        return changesets
    except subprocess.TimeoutExpired:
        print(f"Timeout fetching file history for {file_path}")
        return []
    except Exception as e:
        print(f"Error fetching file history: {e}")
        return []

def fetch_file_changesets_from_rest_api(rel_file_path, username=None, password=None):
    """
    Fallback method using REST API to fetch file changesets.
    """
    try:
        cert_param = []
        if os.path.exists(CERT_PATH):
            cert_param = ["--cacert", CERT_PATH]
        
        # Use OSLC SCM API to query changesets
        history_url = f"{RTC_SERVER_URL}/service/com.ibm.team.filesystem.service.rest.IFileContentService/history"
        
        curl_command = [
            "curl",
            "-L",
            "-u", f"{username}:{password}",
            *cert_param,
            "-X", "GET",
            "-H", "Accept: application/json",
            f"{history_url}?file={rel_file_path}"
        ]
        
        result = subprocess.run(curl_command, capture_output=True, text=True, timeout=30)
        
        if result.returncode != 0 or not result.stdout.strip():
            return []
        
        response_json = json.loads(result.stdout)
        changesets = []
        
        if "changes" in response_json:
            for change in response_json["changes"]:
                changeset_info = {
                    "uuid": change.get("changeSetUuid", ""),
                    "comment": change.get("comment", ""),
                    "date": change.get("creationDate", ""),
                    "url": f"{RTC_SERVER_URL}/resource/itemName/com.ibm.team.scm.ChangeSet/{change.get('changeSetUuid', '')}"
                }
                changesets.append(changeset_info)
        
        return changesets
    except Exception as e:
        print(f"REST API fallback error: {e}")
        return []

def fetch_workitems_using_java_client(changeset_uuid, username=None, password=None):
    """
    Fetch workitem IDs using RTC Java Client Library.
    Returns a list of workitem IDs.
    """
    if not RTC_ENABLED or not username or not password:
        return []
    
    if not os.path.exists(RTC_CLIENT_LIB_PATH):
        print(f"Warning: RTC Java Client Library not found at: {RTC_CLIENT_LIB_PATH}")
        return []
    
    try:
        # Build classpath from all JAR files in the library directory
        jar_files = []
        for root, dirs, files in os.walk(RTC_CLIENT_LIB_PATH):
            for file in files:
                if file.endswith('.jar'):
                    jar_files.append(os.path.join(root, file))
        
        if not jar_files:
            print(f"Warning: No JAR files found in: {RTC_CLIENT_LIB_PATH}")
            return []
        
        # Add current script directory to classpath for WorkItemFetcher.class
        script_dir = os.path.dirname(os.path.abspath(__file__))
        jar_files.insert(0, script_dir)
        
        classpath = os.pathsep.join(jar_files)
        
        # Create a simple Java command to fetch workitems
        java_command = [
            "java",
            "-cp", classpath,
            "com.ibm.team.rtc.client.WorkItemFetcher",
            RTC_SERVER_URL,
            username,
            password,
            changeset_uuid
        ]
        
        result = subprocess.run(java_command, capture_output=True, text=True, timeout=60)
        
        if result.returncode != 0:
            print(f"Java client failed: {result.stderr}")
            return []
        
        # Parse output (assuming it returns workitem IDs one per line)
        workitem_ids = []
        for line in result.stdout.strip().split('\n'):
            line = line.strip()
            if line.isdigit():
                workitem_ids.append(line)
        
        return workitem_ids
    except FileNotFoundError:
        print("Error: Java runtime not found. Please ensure Java is installed and in PATH.")
        return []
    except subprocess.TimeoutExpired:
        print(f"Timeout fetching workitems for changeset: {changeset_uuid}")
        return []
    except Exception as e:
        print(f"Error using Java client: {e}")
        return []

def fetch_workitems_from_changeset(changeset_url, username=None, password=None):
    """
    Fetch workitem IDs associated with a changeset from RTC/ALM.
    Uses REST API first, falls back to Java client if available.
    Returns a list of workitem IDs.
    """
    if not RTC_ENABLED or not username or not password:
        return []
    
    try:
        # Check if certificate file exists
        cert_param = []
        if os.path.exists(CERT_PATH):
            cert_param = ["--cacert", CERT_PATH]
        
        curl_command = [
            "curl",
            "-L",  # Follow redirects
            "-u", f"{username}:{password}",
            *cert_param,
            "-X", "GET",
            changeset_url,
            "-H", "Accept: application/json"
        ]
        
        result = subprocess.run(curl_command, capture_output=True, text=True, timeout=30)
        
        if result.returncode != 0:
            print(f"Failed to fetch changeset details: {result.stderr}")
            return []
        
        if not result.stdout.strip():
            return []
        
        # Parse JSON response to extract workitem IDs
        response_json = json.loads(result.stdout)
        workitem_ids = []
        
        # Extract workitem references from changeset
        # The structure may vary - adjust based on actual RTC response
        if "rtc_cm:com.ibm.team.workitem.linktype.relatedworkitem.related" in response_json:
            for item in response_json["rtc_cm:com.ibm.team.workitem.linktype.relatedworkitem.related"]:
                resource_url = item.get("rdf:resource", "")
                if resource_url:
                    # Extract workitem ID from URL
                    parts = resource_url.split("/")
                    if parts and parts[-1].isdigit():
                        workitem_ids.append(parts[-1])
        
        if workitem_ids:
            return workitem_ids
        
        # If REST API didn't return workitems, try Java client as fallback
        print("No workitems found via REST API, trying Java client...")
        changeset_uuid = changeset_url.split("/")[-1] if "/" in changeset_url else changeset_url
        return fetch_workitems_using_java_client(changeset_uuid, username, password)
        
    except subprocess.TimeoutExpired:
        print(f"Timeout fetching changeset: {changeset_url}")
        # Try Java client as fallback
        changeset_uuid = changeset_url.split("/")[-1] if "/" in changeset_url else changeset_url
        return fetch_workitems_using_java_client(changeset_uuid, username, password)
    except json.JSONDecodeError as e:
        print(f"Failed to parse changeset JSON: {e}")
        # Try Java client as fallback
        changeset_uuid = changeset_url.split("/")[-1] if "/" in changeset_url else changeset_url
        return fetch_workitems_using_java_client(changeset_uuid, username, password)
    except Exception as e:
        print(f"Error fetching workitems from changeset: {e}")
        return []

def get_workitems_for_file(file_path, repository_path, username=None, password=None, workspace_name=None, stream_name=None):
    """
    Get changeset and workitem information for a file from RTC SCM history.
    Returns a dictionary with changeset URL and associated workitem IDs.
    
    Parameters:
        file_path: Path to the file
        repository_path: Path to the repository root
        username: RTC username
        password: RTC password
        workspace_name: Optional workspace name for context
        stream_name: Optional stream name for context
    """
    if not RTC_ENABLED:
        return {"changeset_url": "", "workitem_ids": []}
    
    try:
        # Fetch changesets from SCM for this file with workspace context
        changesets = fetch_file_changesets_from_scm(file_path, repository_path, username, password, 
                                                    workspace_name=workspace_name, stream_name=stream_name)
        
        if not changesets:
            return {"changeset_url": "", "workitem_ids": []}
        
        # Use the most recent changeset
        latest_changeset = changesets[0]
        changeset_url = latest_changeset.get("url", "")
        
        # Fetch workitems associated with this changeset
        workitem_ids = fetch_workitems_from_changeset(changeset_url, username, password)
        
        return {
            "changeset_url": changeset_url,
            "changeset_comment": latest_changeset.get("comment", ""),
            "workitem_ids": workitem_ids
        }
    except Exception as e:
        print(f"Error getting workitems for {file_path}: {e}")
        return {"changeset_url": "", "workitem_ids": []}

# ---------------------------------------------------------------------------
# Create Overview Sheet
# ---------------------------------------------------------------------------
def create_overview_sheet(wb, results, folder1, folder2):
    """
    Create an overview sheet with summary statistics and comparison info.
    """
    from datetime import datetime
    from openpyxl.styles import Border, Side
    
    ws = wb.active
    ws.title = "Overview"
    
    # Define styles
    title_font = Font(size=14, bold=True, color="1F4E78")
    subtitle_font = Font(size=11, bold=True, color="003366")
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Calculate statistics
    # Results structure: [File Path, Lines1, Lines2, Line Status, Status, HTML Link, Purpose]
    # Status is at index 4, Lines1 at index 1, Lines2 at index 2
    total_files = len(results)
    identical_count = sum(1 for r in results if r[4] == "Identical")
    different_count = sum(1 for r in results if r[4] == "Different")
    comments_only_count = sum(1 for r in results if r[4] == "Comments update only")
    only_folder1_count = sum(1 for r in results if r[4] == "Only in Platform")
    only_folder2_count = sum(1 for r in results if r[4] == "Only in Project")
    
    # Calculate line statistics
    def get_lines(r):
        """Get line counts safely"""
        try:
            lines1 = int(r[1]) if r[1] and str(r[1]).isdigit() else 0
            lines2 = int(r[2]) if r[2] and str(r[2]).isdigit() else 0
            return lines1, lines2
        except:
            return 0, 0
    
    # Total lines
    total_lines_folder1 = sum(get_lines(r)[0] for r in results)
    total_lines_folder2 = sum(get_lines(r)[1] for r in results)
    total_lines_combined = total_lines_folder1 + total_lines_folder2
    
    # Lines by category
    identical_lines = sum(get_lines(r)[0] + get_lines(r)[1] for r in results if r[4] == "Identical")
    different_lines = sum(get_lines(r)[0] + get_lines(r)[1] for r in results if r[4] == "Different")
    comments_only_lines = sum(get_lines(r)[0] + get_lines(r)[1] for r in results if r[4] == "Comments update only")
    only_folder1_lines = sum(get_lines(r)[0] for r in results if r[4] == "Only in Platform")
    only_folder2_lines = sum(get_lines(r)[1] for r in results if r[4] == "Only in Project")
    
    # Determine complexity level based on percentage of different files
    def get_complexity(diff_pct):
        """Determine complexity level based on percentage of differences"""
        if diff_pct <= 10:
            return "Minor"
        elif diff_pct <= 30:
            return "Medium"
        else:
            return "Major"
    
    # Title
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "Migration Analysis Report - Overview"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Comparison Information
    ws.append([])
    ws.append(["Comparison Information:"])
    ws['A3'].font = subtitle_font
    
    ws.append(["Platform (Baseline):", folder1])
    ws.append(["Project:", folder2])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    
    # Summary Statistics (following template format)
    ws.append(["Summary Statistics:"])
    ws['A8'].font = subtitle_font
    
    # Header Row with 9 columns (E-F merged for "No of lines(LOC)")
    ws.append([
        "Category", 
        "No of files", 
        "% of files", 
        "% of files\n(Platform vs Project)",
        "No of lines(LOC)",
        None,  # Merged with E
        "% of LOC",
        "% of LOC\n(Platform vs Project)",
        "Complexity Level"
    ])
    
    header_row = 9
    # Format header cells
    for col in range(1, 10):
        cell = ws.cell(header_row, col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30
    
    # Merge E1:F1 for "No of lines(LOC)" header
    ws.merge_cells(f'E{header_row}:F{header_row}')
    
    # Determine individual complexity levels
    different_pct_val = (different_count/total_files*100) if total_files > 0 else 0
    complexity_identical = "Minor"
    complexity_platform_only = "Minor"
    complexity_comments = "Minor" if comments_only_count > 0 else "Minor"
    complexity_different = get_complexity(different_pct_val)
    complexity_project_only = "Minor"
    
    # Check if all complexities are the same
    all_complexities = [complexity_identical, complexity_platform_only, complexity_comments, 
                        complexity_different, complexity_project_only]
    all_same_complexity = len(set(all_complexities)) == 1
    overall_complexity = all_complexities[0] if all_same_complexity else get_complexity(different_pct_val)
    
    # Row 2: Total (with Excel formulas)
    row = header_row + 1
    ws.append([
        "Total",
        f"=B{row+1}+B{row+2}+B{row+3}+B{row+4}+B{row+6}",  # Sum of files
        f"=C{row+1}+C{row+2}+C{row+3}+C{row+4}+C{row+6}",  # Sum of file %
        f"=SUM(D{row+1}+D{row+4})",  # Platform vs Project %
        None,
        f"=F{row+1}+F{row+3}+F{row+4}+F{row+5}+F{row+6}",  # Sum of lines
        f"=G{row+1}+G{row+3}+G{row+4}+G{row+6}",  # Sum of line %
        f"=H{row+1}+H{row+4}",  # Platform vs Project line %
        overall_complexity if all_same_complexity else None  # Complexity (merged if all same)
    ])
    ws.cell(row, 1).font = Font(bold=True, size=11)
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
    
    # Row 3: Identical (Files with no differences)
    row += 1
    row_identical = row
    ws.append([
        "Identical (Files with no differences)",
        identical_count,
        f"=B{row}/$B${header_row+1}",
        f"=SUM(C{row}:C{row+2})",
        None,
        identical_lines,
        f"=F{row}/F${header_row+1}",
        f"=SUM(G{row}:G{row+2})",
        complexity_identical if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
    
    # Row 4: Files exist only in platform
    row += 1
    ws.append([
        "Files exist only in platform",
        only_folder1_count,
        f"=B{row}/$B${header_row+1}",
        None,
        None,
        "Not Applicable" if only_folder1_count > 0 else 0,
        "NA" if only_folder1_count > 0 else 0,
        None,
        complexity_platform_only if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FF9BC2E6", end_color="FF9BC2E6", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FF9BC2E6", end_color="FF9BC2E6", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
    
    # Row 5: Comments update only
    row += 1
    ws.append([
        "Comments update only",
        comments_only_count,
        f"=B{row}/$B${header_row+1}",
        None,
        None,
        comments_only_lines if comments_only_count > 0 else 0,
        f"=F{row}/F${header_row+1}" if comments_only_count > 0 else 0,
        None,
        complexity_comments if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
    
    # Merge D cells for identical, platform, comments (rows 3-5)
    ws.merge_cells(f'D{row_identical}:D{row}')
    ws.merge_cells(f'H{row_identical}:H{row}')
    
    # Row 6-7: Files with code differences in Project (with Added/Removed sub-rows)
    row += 1
    row_diff_start = row
    # Calculate added and removed lines for different files
    added_lines_diff = sum(get_lines(r)[1] for r in results if r[4] == "Different")
    removed_lines_diff = -sum(get_lines(r)[0] for r in results if r[4] == "Different")
    
    ws.append([
        "Files with code differences in Project",
        different_count,
        f"=B{row}/B${header_row+1}",
        f"=SUM(C{row}:C{row+2})",
        "Added",
        added_lines_diff,
        f"=F{row}/F${header_row+1}",
        f"=SUM(G{row}:G{row+2})",
        complexity_different if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
    
    # Row 7: Sub-row for Removed lines
    row += 1
    ws.append([
        None,
        None,
        None,
        None,
        "Removed",
        removed_lines_diff,
        f"=F{row}/F${header_row+1}",
        None,
        None
    ])
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
    
    # Row 8: Files exist only in Project
    row += 1
    ws.append([
        "Files exist only in Project",
        only_folder2_count,
        f"=B{row}/B${header_row+1}",
        None,
        "Added",
        only_folder2_lines,
        f"=F{row}/F${header_row+1}",
        None,
        complexity_project_only if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
    
    # Merge cells for "Files with code differences" (A, B, C, D columns across rows 6-7)
    ws.merge_cells(f'A{row_diff_start}:A{row_diff_start+1}')
    ws.merge_cells(f'B{row_diff_start}:B{row_diff_start+1}')
    ws.merge_cells(f'C{row_diff_start}:C{row_diff_start+1}')
    ws.merge_cells(f'D{row_diff_start}:D{row}')
    ws.merge_cells(f'H{row_diff_start}:H{row}')
    
    # If all complexities are the same, merge the Complexity Level column (I) for all data rows
    if all_same_complexity:
        ws.merge_cells(f'I{header_row+1}:I{row}')
        ws.cell(header_row+1, 9).font = Font(bold=True, size=12)
    else:
        # Merge complexity for sub-rows that don't have individual complexity
        ws.merge_cells(f'I{row_diff_start}:I{row_diff_start+1}')
    
    # Apply percentage formatting to percentage columns (C, D, G, H)
    from openpyxl.styles import numbers
    for row_num in range(header_row+1, row+1):
        # Column C: % of files
        ws.cell(row_num, 3).number_format = numbers.FORMAT_PERCENTAGE_00
        # Column D: % of files (Platform vs Project)
        if ws.cell(row_num, 4).value and ws.cell(row_num, 4).value != "":
            ws.cell(row_num, 4).number_format = numbers.FORMAT_PERCENTAGE_00
        # Column G: % of LOC
        if ws.cell(row_num, 7).value not in ["NA", None, ""]:
            ws.cell(row_num, 7).number_format = numbers.FORMAT_PERCENTAGE_00
        # Column H: % of LOC (Platform vs Project)
        if ws.cell(row_num, 8).value and ws.cell(row_num, 8).value != "":
            ws.cell(row_num, 8).number_format = numbers.FORMAT_PERCENTAGE_00
    
    row = ws.max_row
    
    
    ws.append([])
    
    # Color Legend
    legend_row_start = ws.max_row + 1
    ws.append(["Color Coding Legend:"])
    ws.cell(legend_row_start, 1).font = subtitle_font
    ws.append([])
    
    # List of color descriptions
    row = legend_row_start + 1
    ws.append(["Identical (Files with no differences) -- Green color"])
    ws.cell(row, 1).font = Font(size=10)
    
    row += 1
    ws.append(["Files exist only in platform -- Blue color"])
    ws.cell(row, 1).font = Font(size=10)
    
    row += 1
    ws.append(["Comments update only -- Yellow color"])
    ws.cell(row, 1).font = Font(size=10)
    
    row += 1
    ws.append(["Files with code differences in Project -- Red color"])
    ws.cell(row, 1).font = Font(size=10)
    
    row += 1
    ws.append(["Files exist only in Project -- Orange color"])
    ws.cell(row, 1).font = Font(size=10)
    
    # Set column widths to match template + complexity column
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 13.5
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 16
    
    return ws


# ---------------------------------------------------------------------------
# Fetch Changesets for a Baseline (RTC REST API)
# ---------------------------------------------------------------------------
def fetch_baseline_changesets(baseline_uuid, username, password):
    """
    Fetch all changesets for a given baseline UUID using RTC REST API.
    Returns list of changeset dictionaries with uuid, comment, author, date, workitems with owner names.
    """
    try:
        import requests
        from requests.auth import HTTPBasicAuth
        
        # RTC API endpoint for baseline changesets
        url = f"{RTC_SERVER_URL}/oslc/baselines/{baseline_uuid}/changesets"
        
        headers = {
            'Accept': 'application/json',
            'OSLC-Core-Version': '2.0'
        }
        
        response = requests.get(
            url,
            auth=HTTPBasicAuth(username, password),
            headers=headers,
            verify=False,
            timeout=60
        )
        
        if response.status_code == 200:
            data = response.json()
            changesets = []
            
            # Parse changesets from response
            if isinstance(data, dict) and 'oslc_cm:changeRequests' in data:
                cs_list = data.get('oslc_cm:changeRequests', [])
            elif isinstance(data, list):
                cs_list = data
            else:
                cs_list = []
            
            for cs in cs_list:
                cs_uuid = cs.get('dcterms:identifier', cs.get('rdf:resource', '').split('/')[-1])
                cs_comment = cs.get('dcterms:title', cs.get('dcterms:description', ''))
                cs_author = cs.get('dcterms:creator', {}).get('foaf:name', 'Unknown') if isinstance(cs.get('dcterms:creator'), dict) else str(cs.get('dcterms:creator', 'Unknown'))
                cs_date = cs.get('dcterms:modified', cs.get('dcterms:created', ''))
                
                # Extract work items with owner names
                workitems = []
                wi_links = cs.get('rtc_cm:com.ibm.team.filesystem.workitems', [])
                if not isinstance(wi_links, list):
                    wi_links = [wi_links]
                
                for wi in wi_links:
                    if isinstance(wi, dict):
                        wi_id = wi.get('oslc_cm:id', wi.get('dcterms:identifier', ''))
                    else:
                        wi_id = str(wi).split('/')[-1]
                    
                    if wi_id:
                        # Fetch work item details to get owner name
                        wi_details = fetch_workitem_details(wi_id, username, password)
                        workitems.append(wi_details)
                
                changesets.append({
                    'uuid': cs_uuid,
                    'comment': cs_comment[:500] if cs_comment else '',  # Limit comment length
                    'author': cs_author,
                    'date': cs_date,
                    'workitems': workitems  # List of dicts with 'id' and 'owner'
                })
            
            return changesets
        else:
            print(f"Failed to fetch changesets for baseline {baseline_uuid[:20]}...: HTTP {response.status_code}")
            return []
            
    except Exception as e:
        print(f"Error fetching changesets for baseline: {e}")
        return []


def fetch_workitems_for_changeset(changeset_uuid, username, password):
    """
    Fetch work items associated with a changeset using lscm.
    Returns list of dicts with work item ID and owner name.
    """
    try:
        if not LSCM_PATH or not os.path.exists(LSCM_PATH):
            print(f"            DEBUG: LSCM path not available")
            return []
        
        scm_executable = LSCM_PATH
        
        # Use scm list workitems command
        list_command = [
            scm_executable, "list", "workitems",
            "-r", RTC_SERVER_URL,
            "-c", changeset_uuid,  # Changeset UUID
            "-j"  # JSON output
        ]
        
        env = os.environ.copy()
        env['NO_PROXY'] = '*'
        env['no_proxy'] = '*'
        env['HTTP_PROXY'] = ''
        env['HTTPS_PROXY'] = ''
        env['http_proxy'] = ''
        env['https_proxy'] = ''
        
        print(f"            DEBUG: Running scm list workitems for changeset {changeset_uuid[:16]}...")
        
        result = subprocess.run(
            list_command,
            capture_output=True,
            text=True,
            timeout=30,
            cwd=os.path.dirname(scm_executable),
            env=env
        )
        
        print(f"            DEBUG: Return code: {result.returncode}")
        
        if result.returncode != 0:
            print(f"            DEBUG: Error output: {result.stderr[:200] if result.stderr else 'None'}")
            return []
        
        if not result.stdout.strip():
            print(f"            DEBUG: Empty output from scm list workitems")
            return []
        
        print(f"            DEBUG: Raw output (first 300 chars): {result.stdout[:300]}")
        
        # Parse work items
        workitems = []
        try:
            data = json.loads(result.stdout)
            
            print(f"            DEBUG: Parsed JSON type: {type(data)}")
            if isinstance(data, dict):
                print(f"            DEBUG: JSON keys: {list(data.keys())}")
            
            wi_list = []
            if isinstance(data, dict):
                # Try different possible keys
                for key in ['workitems', 'work-items', 'items', 'workItems']:
                    if key in data:
                        wi_list = data[key]
                        print(f"            DEBUG: Found work items under key '{key}': {len(wi_list)}")
                        break
            elif isinstance(data, list):
                wi_list = data
                print(f"            DEBUG: Data is a list with {len(wi_list)} items")
            
            for idx, wi in enumerate(wi_list):
                print(f"            DEBUG: Work item {idx+1}: {wi}")
                
                wi_id = wi.get('id', wi.get('workitem-id', wi.get('number', '')))
                wi_owner = wi.get('owner', wi.get('owned-by', wi.get('ownedBy', {})))
                
                # Extract owner name
                if isinstance(wi_owner, dict):
                    owner_name = wi_owner.get('name', wi_owner.get('userId', wi_owner.get('userName', 'Unknown')))
                else:
                    owner_name = str(wi_owner) if wi_owner else 'Unknown'
                
                if wi_id:
                    workitems.append({
                        'id': str(wi_id),
                        'owner': owner_name
                    })
                    print(f"            DEBUG: Extracted WI #{wi_id} - Owner: {owner_name}")
        
        except json.JSONDecodeError as e:
            print(f"            DEBUG: JSON decode error: {e}")
            pass
        
        print(f"            DEBUG: Returning {len(workitems)} work items")
        return workitems
        
    except Exception as e:
        print(f"            DEBUG: Exception in fetch_workitems_for_changeset: {e}")
        return []


def fetch_workitem_details(workitem_id, username, password):
    """
    Fetch work item details including owner name using RTC REST API.
    Returns dict with work item info.
    """
    try:
        import requests
        from requests.auth import HTTPBasicAuth
        
        # RTC Work Item API
        url = f"{RTC_SERVER_URL}/oslc/workitems/{workitem_id}"
        
        headers = {
            'Accept': 'application/json',
            'OSLC-Core-Version': '2.0'
        }
        
        response = requests.get(
            url,
            auth=HTTPBasicAuth(username, password),
            headers=headers,
            verify=False,
            timeout=15
        )
        
        if response.status_code == 200:
            data = response.json()
            
            # Extract owner information
            owner = data.get('rtc_cm:ownedBy', data.get('dcterms:creator', {}))
            if isinstance(owner, dict):
                owner_name = owner.get('foaf:name', owner.get('dcterms:title', 'Unknown'))
            else:
                owner_name = str(owner) if owner else 'Unknown'
            
            return {
                'id': workitem_id,
                'owner': owner_name,
                'summary': data.get('dcterms:title', ''),
                'status': data.get('rtc_cm:status', '')
            }
        else:
            return {'id': workitem_id, 'owner': 'Unknown'}
            
    except Exception as e:
        return {'id': workitem_id, 'owner': 'Unknown'}


def fetch_changeset_files(changeset_uuid, baseline_uuid, username, password):
    """
    Fetch list of files changed in a changeset.
    Returns list of file path strings.
    """
    try:
        import requests
        from requests.auth import HTTPBasicAuth
        
        # RTC API endpoint for changeset changes
        url = f"{RTC_SERVER_URL}/service/com.ibm.team.filesystem.service.rest.IFileContentService/changes"
        
        params = {
            'changeSetUuid': changeset_uuid,
            'baselineUuid': baseline_uuid
        }
        
        headers = {
            'Accept': 'application/json'
        }
        
        response = requests.get(
            url,
            params=params,
            auth=HTTPBasicAuth(username, password),
            headers=headers,
            verify=False,
            timeout=30
        )
        
        if response.status_code == 200:
            data = response.json()
            files = []
            
            changes = data.get('changes', [])
            for change in changes:
                file_path = change.get('path', change.get('name', ''))
                if file_path:
                    files.append(file_path)
            
            return files
        else:
            return []
            
    except Exception as e:
        print(f"Error fetching files for changeset: {e}")
        return []


def compare_baseline_changesets(baseline1_uuid, baseline2_uuid, username, password):
    """
    Compare changesets between two baselines using scm compare command.
    Returns dict with 'only_in_1', 'only_in_2', 'common' changesets.
    """
    try:
        if not LSCM_PATH or not os.path.exists(LSCM_PATH):
            print("  lscm not available, cannot compare changesets")
            return {'only_in_1': [], 'only_in_2': [], 'common': []}
        
        scm_executable = LSCM_PATH
        
        # Use scm compare command to compare two baselines
        # scm compare baseline <baseline1> baseline <baseline2> -r <repo> -j
        compare_command = [
            scm_executable, "compare",
            "baseline", baseline1_uuid,
            "baseline", baseline2_uuid,
            "-r", RTC_SERVER_URL,
            "-j"  # JSON output
        ]
        
        # Set environment to disable proxy
        env = os.environ.copy()
        env['NO_PROXY'] = '*'
        env['no_proxy'] = '*'
        env['HTTP_PROXY'] = ''
        env['HTTPS_PROXY'] = ''
        env['http_proxy'] = ''
        env['https_proxy'] = ''
        
        print(f"  Comparing baselines: {baseline1_uuid[:20]}... vs {baseline2_uuid[:20]}...")
        
        result = subprocess.run(
            compare_command,
            capture_output=True,
            text=True,
            timeout=120,
            cwd=os.path.dirname(scm_executable),
            env=env
        )
        
        if result.returncode != 0:
            print(f"  Failed to compare baselines: {result.stderr if result.stderr else 'Unknown error'}")
            return {'only_in_1': [], 'only_in_2': [], 'common': []}
        
        if not result.stdout.strip():
            print("  No differences found")
            return {'only_in_1': [], 'only_in_2': [], 'common': []}
        
        # Debug: Print raw output
        print(f"  Raw compare output (first 500 chars): {result.stdout[:500]}")
        
        # Parse JSON output
        changesets_only_1 = []
        changesets_only_2 = []
        
        try:
            data = json.loads(result.stdout)
            
            # Debug: Print parsed structure
            print(f"  Parsed JSON type: {type(data)}")
            if isinstance(data, dict):
                print(f"  JSON keys: {list(data.keys())}")
            
            # Parse the comparison results
            # Structure: { "direction": [ { "components": [ { "changesets": [...] } ] } ] }
            if isinstance(data, dict):
                # Look for the direction array which contains components with changesets
                directions = data.get('direction', [])
                
                if not directions:
                    print(f"  No 'direction' key found in response")
                    return {'only_in_1': [], 'only_in_2': [], 'common': []}
                
                print(f"  Found {len(directions)} direction entries")
                
                # Process each direction entry
                for dir_idx, dir_entry in enumerate(directions):
                    dir_type = dir_entry.get('type', dir_entry.get('name', 'unknown'))
                    print(f"\n  Direction Entry {dir_idx+1}: type='{dir_type}'")
                    
                    components = dir_entry.get('components', [])
                    print(f"    Found {len(components)} components in this direction")
                    
                    # Determine if this is incoming or outgoing direction
                    is_incoming = dir_type.lower() in ['incoming', '-->', 'in', 'i', 'right']
                    is_outgoing = dir_type.lower() in ['outgoing', '<--', 'out', 'o', 'left']
                    
                    # Process each component
                    for comp_idx, component in enumerate(components):
                        comp_name = component.get('name', f'Component_{comp_idx+1}')
                        changesets = component.get('changesets', [])
                        
                        print(f"      Component '{comp_name}': {len(changesets)} changesets")
                        
                        # Process each changeset
                        for cs_idx, changeset in enumerate(changesets):
                            cs_uuid = changeset.get('uuid', changeset.get('item-id', ''))
                            
                            if not cs_uuid:
                                print(f"        WARNING: Changeset {cs_idx+1} has no UUID, skipping")
                                continue
                            
                            # Extract author information
                            cs_author = 'Unknown'
                            author_field = changeset.get('author', {})
                            if isinstance(author_field, dict):
                                cs_author = author_field.get('userName', 
                                           author_field.get('userId', 
                                           author_field.get('name', 'Unknown')))
                            elif isinstance(author_field, str):
                                cs_author = author_field
                            
                            cs_comment = changeset.get('comment', '')
                            cs_date = changeset.get('modified', changeset.get('created', changeset.get('modifiedDate', '')))
                            
                            print(f"        [{cs_idx+1}/{len(changesets)}] {cs_uuid[:16]}... by {cs_author}")
                            
                            # First, try to extract work items directly from the changeset JSON
                            workitems = []
                            
                            # Check for work items in the changeset data itself
                            wi_from_json = changeset.get('workitems', changeset.get('work-items', changeset.get('workItems', [])))
                            
                            if wi_from_json:
                                print(f"          Found {len(wi_from_json)} work items in changeset JSON")
                                for wi in wi_from_json:
                                    if isinstance(wi, dict):
                                        wi_id = wi.get('id', wi.get('number', wi.get('workitem-id', '')))
                                        wi_owner = wi.get('owner', wi.get('ownedBy', {}))
                                        
                                        if isinstance(wi_owner, dict):
                                            owner_name = wi_owner.get('userName', wi_owner.get('userId', wi_owner.get('name', 'Unknown')))
                                        else:
                                            owner_name = str(wi_owner) if wi_owner else 'Unknown'
                                        
                                        if wi_id:
                                            workitems.append({'id': str(wi_id), 'owner': owner_name})
                                    elif isinstance(wi, (str, int)):
                                        workitems.append({'id': str(wi), 'owner': 'Unknown'})
                            
                            # If no work items found in JSON, try fetching via scm command
                            if not workitems:
                                print(f"          No work items in JSON, fetching via scm command...")
                                workitems = fetch_workitems_for_changeset(cs_uuid, username, password)
                            
                            if workitems:
                                print(f"          Total work items: {len(workitems)}")
                                for wi in workitems:
                                    print(f"            - WI #{wi.get('id')} - Owner: {wi.get('owner')}")
                            else:
                                print(f"          No work items found for this changeset")
                            
                            cs_data = {
                                'uuid': cs_uuid,
                                'comment': cs_comment[:500] if cs_comment else '',
                                'author': cs_author,
                                'date': cs_date,
                                'workitems': workitems
                            }
                            
                            # Add to appropriate list based on direction
                            if is_outgoing:
                                print(f"          -> BASELINE 1 ONLY (outgoing/removed)")
                                changesets_only_1.append(cs_data)
                            elif is_incoming:
                                print(f"          -> BASELINE 2 ONLY (incoming/new)")
                                changesets_only_2.append(cs_data)
                            else:
                                # If direction type unclear, default to baseline 2 (new)
                                print(f"          -> DEFAULT to BASELINE 2 (direction unclear: '{dir_type}')")
                                changesets_only_2.append(cs_data)
            
            print(f"  Result: {len(changesets_only_1)} changesets only in B1, {len(changesets_only_2)} only in B2")
            
            return {
                'only_in_1': changesets_only_1,
                'only_in_2': changesets_only_2,
                'common': []
            }
            
        except json.JSONDecodeError as e:
            print(f"  Failed to parse comparison JSON: {e}")
            return {'only_in_1': [], 'only_in_2': [], 'common': []}
        
    except Exception as e:
        print(f"  Error comparing baselines: {e}")
        return {'only_in_1': [], 'only_in_2': [], 'common': []}


# ---------------------------------------------------------------------------
# Generate Excel Report for Snapshot Component Comparison
# ---------------------------------------------------------------------------
def generate_snapshot_comparison_excel(selected_components, only_in_1, only_in_2, 
                                       components1, components2, snap1_name, snap2_name):
    """
    Generate Excel report for snapshot component comparison with changeset details.
    Shows component-level differences with baseline information, changesets, work items, and file diffs.
    """
    # Ask user where to save the report
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"Snapshot_Comparison_{timestamp}.xlsx"
    
    excel_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_filename,
        title="Save Snapshot Comparison Report"
    )
    
    if not excel_path:
        messagebox.showinfo("Cancelled", "Report generation cancelled.")
        return
    
    # Create HTML diff output directory
    html_output_dir = os.path.join(os.path.dirname(excel_path), f"Snapshot_Diffs_{timestamp}")
    os.makedirs(html_output_dir, exist_ok=True)
    
    # Show progress dialog
    progress_win = tk.Toplevel(root)
    progress_win.title("Generating Report")
    progress_win.geometry("600x180")
    progress_win.configure(bg="white")
    progress_win.grab_set()
    
    tk.Label(progress_win, text="📊 Generating Snapshot Comparison Report...", 
             font=("Segoe UI", 12, "bold"), bg="white").pack(pady=15)
    
    progress_label = tk.Label(progress_win, text="Preparing data...", 
                              font=("Segoe UI", 10), bg="white", fg="gray")
    progress_label.pack(pady=10)
    
    progress_bar = ttk.Progressbar(progress_win, mode='determinate', length=400)
    progress_bar.pack(pady=10)
    
    progress_win.update()
    
    try:
        # Create workbook
        wb = Workbook()
        
        # Create component dictionaries for lookup
        comp1_dict = {c['name']: c for c in components1}
        comp2_dict = {c['name']: c for c in components2}
        
        # ===== Sheet 1: Summary =====
        progress_label.config(text="Creating summary sheet...")
        progress_win.update()
        
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        # Title
        ws_summary['A1'] = "Snapshot Component Comparison Report"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        
        # Snapshot info
        ws_summary['A3'] = "Snapshot 1:"
        ws_summary['B3'] = snap1_name
        ws_summary['A4'] = "Snapshot 2:"
        ws_summary['B4'] = snap2_name
        ws_summary['A5'] = "Generated:"
        ws_summary['B5'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for cell in ['A3', 'A4', 'A5']:
            ws_summary[cell].font = Font(bold=True)
        
        # Summary statistics
        ws_summary['A7'] = "Component Summary"
        ws_summary['A7'].font = Font(bold=True, size=12)
        ws_summary['A8'] = "Common Components (Selected):"
        ws_summary['B8'] = len(selected_components)
        ws_summary['A9'] = "Components only in Snapshot 1:"
        ws_summary['B9'] = len(only_in_1)
        ws_summary['A10'] = "Components only in Snapshot 2:"
        ws_summary['B10'] = len(only_in_2)
        ws_summary['A11'] = "Total Components in Snapshot 1:"
        ws_summary['B11'] = len(components1)
        ws_summary['A12'] = "Total Components in Snapshot 2:"
        ws_summary['B12'] = len(components2)
        
        for cell in ['A8', 'A9', 'A10', 'A11', 'A12']:
            ws_summary[cell].font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 50
        
        # ===== Sheet 2: Common Components (Selected) with Changeset Details =====
        progress_label.config(text="Creating common components sheet...")
        progress_bar['value'] = 20
        progress_win.update()
        
        ws_common = wb.create_sheet("Common Components Details")
        
        # Create detailed sheet with changeset information
        ws_detailed = wb.create_sheet("Changeset & File Details")
        detail_headers = ["Component Name", "Changeset UUID", "Comment", "Changeset Author", "Date", 
                         "Work Item IDs", "Work Item Owners", "Changed Files", "HTML Diff Links"]
        ws_detailed.append(detail_headers)
        
        # Header formatting for detailed sheet
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detailed.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Simple summary for common components
        summary_headers = ["Component Name", "Baseline Status", "Changesets Only in S1", 
                          "Changesets Only in S2", "Total Changed Files"]
        ws_common.append(summary_headers)
        
        for col, header in enumerate(summary_headers, 1):
            cell = ws_common.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Status colors
        same_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        diff_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        cs_fill = PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid")  # Pink
        
        total_components = len(selected_components)
        
        for idx, comp_name in enumerate(selected_components, 1):
            progress_label.config(text=f"Analyzing component {idx}/{total_components}: {comp_name[:40]}...")
            progress_bar['value'] = 20 + (idx / total_components * 50)
            progress_win.update()
            
            comp1 = comp1_dict.get(comp_name, {})
            comp2 = comp2_dict.get(comp_name, {})
            
            baseline1_uuid = comp1.get('baseline_uuid', '')
            baseline2_uuid = comp2.get('baseline_uuid', '')
            
            baseline_status = "Same Baseline" if baseline1_uuid == baseline2_uuid else "Different Baseline"
            
            # If baselines are different, compare changesets
            changesets_only_1 = []
            changesets_only_2 = []
            total_files = 0
            
            if baseline_status == "Different Baseline" and baseline1_uuid and baseline2_uuid:
                print(f"\n{'='*80}")
                print(f"COMPARING CHANGESETS FOR COMPONENT: {comp_name}")
                print(f"{'='*80}")
                print(f"  Baseline 1 UUID: {baseline1_uuid}")
                print(f"  Baseline 2 UUID: {baseline2_uuid}")
                
                cs_comparison = compare_baseline_changesets(baseline1_uuid, baseline2_uuid, 
                                                            RTC_USERNAME, RTC_PASSWORD)
                
                changesets_only_1 = cs_comparison.get('only_in_1', [])
                changesets_only_2 = cs_comparison.get('only_in_2', [])
                
                print(f"\n  RESULT:")
                print(f"    Changesets only in Baseline 1: {len(changesets_only_1)}")
                print(f"    Changesets only in Baseline 2: {len(changesets_only_2)}")
                
                # If no changesets found but baselines are different, add a note row
                if len(changesets_only_1) == 0 and len(changesets_only_2) == 0:
                    print(f"  WARNING: No changesets found, but baselines are different!")
                    print(f"  This might indicate a comparison issue or empty difference.")
                    
                    # Add informational row
                    detail_row = [
                        comp_name,
                        "N/A",
                        "Different baselines but no changesets detected by scm compare",
                        "N/A",
                        "N/A",
                        "N/A",
                        "N/A",
                        "Check baseline comparison manually",
                        "N/A"
                    ]
                    ws_detailed.append(detail_row)
                    
                    row_num = ws_detailed.max_row
                    warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    for col in range(1, 10):
                        ws_detailed.cell(row=row_num, column=col).fill = warning_fill
                
                # Process changesets new in Snapshot 2
                if changesets_only_2:
                    print(f"\n  PROCESSING CHANGESETS NEW IN SNAPSHOT 2:")
                
                for idx, cs in enumerate(changesets_only_2, 1):
                    print(f"\n    [{idx}/{len(changesets_only_2)}] Changeset: {cs.get('uuid', 'NO-UUID')[:20]}...")
                    print(f"        Author: {cs.get('author', 'Unknown')}")
                    print(f"        Date: {cs.get('date', 'Unknown')}")
                    print(f"        Comment: {cs.get('comment', 'No comment')[:80]}...")
                    
                    # Extract work item IDs and owners
                    workitem_ids = []
                    workitem_owners = []
                    
                    workitems_list = cs.get('workitems', [])
                    print(f"        Work items: {len(workitems_list)} found")
                    
                    if workitems_list:
                        print(f"        Work item details:")
                        for wi in workitems_list:
                            if isinstance(wi, dict):
                                wi_id = wi.get('id', '')
                                wi_owner = wi.get('owner', 'Unknown')
                                if wi_id:
                                    workitem_ids.append(str(wi_id))
                                    workitem_owners.append(wi_owner)
                                    print(f"          ✓ Work Item #{wi_id} - Owner: {wi_owner}")
                            else:
                                wi_str = str(wi)
                                workitem_ids.append(wi_str)
                                workitem_owners.append('Unknown')
                                print(f"          ✓ Work Item: {wi_str} - Owner: Unknown")
                    else:
                        print(f"        ⚠ No work items associated with this changeset")
                    
                    workitem_ids_str = ", ".join(workitem_ids) if workitem_ids else "No Work Items"
                    workitem_owners_str = ", ".join(workitem_owners) if workitem_owners else "N/A"
                    
                    # Add work items to comment for visibility
                    comment_with_wi = cs.get('comment', 'No comment')[:150]
                    if workitem_ids:
                        comment_with_wi = f"[WI: {workitem_ids_str}] {comment_with_wi}"
                    
                    # Fetch files changed in this changeset
                    print(f"        Fetching changed files...")
                    changed_files = fetch_changeset_files(cs['uuid'], baseline2_uuid, 
                                                          RTC_USERNAME, RTC_PASSWORD)
                    total_files += len(changed_files)
                    
                    print(f"        Changed files: {len(changed_files)}")
                    
                    files_str = "\n".join(changed_files[:10]) if changed_files else "No files information"
                    if len(changed_files) > 10:
                        files_str += f"\n... and {len(changed_files) - 10} more files"
                    
                    # Create HTML diff links
                    html_links = []
                    for file_path in changed_files[:5]:
                        html_filename = f"{comp_name}_{cs['uuid'][:8]}_{os.path.basename(file_path)}_diff.html"
                        html_file_path = os.path.join(html_output_dir, html_filename)
                        html_links.append(f"file:///{html_file_path.replace(os.sep, '/')}")
                    
                    html_links_str = "\n".join(html_links) if html_links else "N/A"
                    
                    detail_row = [
                        comp_name,
                        cs.get('uuid', 'NO-UUID')[:20] + "..." if len(cs.get('uuid', '')) > 20 else cs.get('uuid', 'NO-UUID'),
                        comment_with_wi,
                        cs.get('author', 'Unknown'),
                        str(cs.get('date', 'Unknown')),
                        workitem_ids_str,
                        workitem_owners_str,
                        files_str,
                        html_links_str
                    ]
                    
                    print(f"        ✓ Excel Row Data:")
                    print(f"           Component: {comp_name}")
                    print(f"           Work Items: {workitem_ids_str}")
                    print(f"           Owners: {workitem_owners_str}")
                    print(f"        ✓ Adding to Excel sheet (row {ws_detailed.max_row + 1})")
                    ws_detailed.append(detail_row)
                    
                    # Color changeset rows
                    row_num = ws_detailed.max_row
                    for col in range(1, 10):
                        ws_detailed.cell(row=row_num, column=col).fill = cs_fill
                
                # Process changesets removed in Snapshot 2
                if changesets_only_1:
                    print(f"\n  PROCESSING CHANGESETS REMOVED FROM SNAPSHOT 2:")
                
                for idx, cs in enumerate(changesets_only_1, 1):
                    print(f"\n    [{idx}/{len(changesets_only_1)}] [REMOVED] Changeset: {cs.get('uuid', 'NO-UUID')[:20]}...")
                    print(f"        Author: {cs.get('author', 'Unknown')}")
                    
                    # Extract work item IDs and owners
                    workitem_ids = []
                    workitem_owners = []
                    
                    workitems_list = cs.get('workitems', [])
                    print(f"        Work items: {len(workitems_list)} found")
                    
                    if workitems_list:
                        print(f"        Work item details:")
                        for wi in workitems_list:
                            if isinstance(wi, dict):
                                wi_id = wi.get('id', '')
                                wi_owner = wi.get('owner', 'Unknown')
                                if wi_id:
                                    workitem_ids.append(str(wi_id))
                                    workitem_owners.append(wi_owner)
                                    print(f"          ✓ Work Item #{wi_id} - Owner: {wi_owner}")
                            else:
                                workitem_ids.append(str(wi))
                                workitem_owners.append('Unknown')
                                print(f"          ✓ Work Item: {str(wi)} - Owner: Unknown")
                    
                    workitem_ids_str = ", ".join(workitem_ids) if workitem_ids else "No Work Items"
                    workitem_owners_str = ", ".join(workitem_owners) if workitem_owners else "N/A"
                    
                    # Add work items to comment for visibility
                    comment_text = cs.get('comment', 'No comment')[:130]
                    if workitem_ids:
                        comment_text = f"[WI: {workitem_ids_str}] {comment_text}"
                    
                    files_str = "[Changeset removed in Snapshot 2]"
                    
                    detail_row = [
                        comp_name,
                        cs.get('uuid', 'NO-UUID')[:20] + "..." if len(cs.get('uuid', '')) > 20 else cs.get('uuid', 'NO-UUID'),
                        "[REMOVED] " + comment_text,
                        cs.get('author', 'Unknown'),
                        str(cs.get('date', 'Unknown')),
                        workitem_ids_str,
                        workitem_owners_str,
                        files_str,
                        "N/A"
                    ]
                    
                    print(f"        ✓ Excel Row Data:")
                    print(f"           Component: {comp_name}")
                    print(f"           Work Items: {workitem_ids_str}")
                    print(f"           Owners: {workitem_owners_str}")
                    print(f"        ✓ Adding to Excel sheet (row {ws_detailed.max_row + 1})")
                    ws_detailed.append(detail_row)
                    
                    # Color removed changeset rows differently (light gray)
                    removed_cs_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    row_num = ws_detailed.max_row
                    for col in range(1, 10):
                        ws_detailed.cell(row=row_num, column=col).fill = removed_cs_fill
                
                print(f"\n  ✓ Finished processing component: {comp_name}")
                print(f"{'='*80}\n")
            
            # Add summary row to common components sheet
            summary_row = [
                comp_name,
                baseline_status,
                len(changesets_only_1),
                len(changesets_only_2),
                total_files
            ]
            ws_common.append(summary_row)
            
            # Apply status coloring
            row_num = ws_common.max_row
            status_cell = ws_common.cell(row=row_num, column=2)
            if baseline_status == "Same Baseline":
                status_cell.fill = same_fill
            else:
                status_cell.fill = diff_fill
        
        # Column widths for common components
        ws_common.column_dimensions['A'].width = 45
        ws_common.column_dimensions['B'].width = 20
        ws_common.column_dimensions['C'].width = 25
        ws_common.column_dimensions['D'].width = 25
        ws_common.column_dimensions['E'].width = 20
        
        # Column widths for detailed sheet
        ws_detailed.column_dimensions['A'].width = 40  # Component Name
        ws_detailed.column_dimensions['B'].width = 20  # Changeset UUID
        ws_detailed.column_dimensions['C'].width = 50  # Comment
        ws_detailed.column_dimensions['D'].width = 25  # Changeset Author
        ws_detailed.column_dimensions['E'].width = 20  # Date
        ws_detailed.column_dimensions['F'].width = 30  # Work Item IDs
        ws_detailed.column_dimensions['G'].width = 40  # Work Item Owners
        ws_detailed.column_dimensions['H'].width = 50  # Changed Files
        ws_detailed.column_dimensions['I'].width = 60  # HTML Diff Links
        
        # ===== Sheet 3: Only in Snapshot 1 =====
        progress_label.config(text="Creating Snapshot 1 unique components sheet...")
        progress_bar['value'] = 80
        progress_win.update()
        
        ws_only1 = wb.create_sheet("Only in Snapshot 1")
        headers_only = ["Component Name", "UUID", "Baseline Name", "Baseline UUID"]
        ws_only1.append(headers_only)
        
        for col, header in enumerate(headers_only, 1):
            cell = ws_only1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        removed_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")  # Blue
        
        for comp_name in only_in_1:
            comp = comp1_dict.get(comp_name, {})
            row = [
                comp_name,
                comp.get('uuid', 'N/A'),
                comp.get('baseline_name', 'N/A'),
                comp.get('baseline_uuid', 'N/A')
            ]
            ws_only1.append(row)
            
            # Color the entire row
            row_num = ws_only1.max_row
            for col in range(1, 5):
                ws_only1.cell(row=row_num, column=col).fill = removed_fill
        
        ws_only1.column_dimensions['A'].width = 40
        ws_only1.column_dimensions['B'].width = 35
        ws_only1.column_dimensions['C'].width = 35
        ws_only1.column_dimensions['D'].width = 35
        
        # ===== Sheet 4: Only in Snapshot 2 =====
        progress_label.config(text="Creating Snapshot 2 unique components sheet...")
        progress_bar['value'] = 90
        progress_win.update()
        
        ws_only2 = wb.create_sheet("Only in Snapshot 2")
        ws_only2.append(headers_only)
        
        for col, header in enumerate(headers_only, 1):
            cell = ws_only2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        added_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # Orange
        
        for comp_name in only_in_2:
            comp = comp2_dict.get(comp_name, {})
            row = [
                comp_name,
                comp.get('uuid', 'N/A'),
                comp.get('baseline_name', 'N/A'),
                comp.get('baseline_uuid', 'N/A')
            ]
            ws_only2.append(row)
            
            # Color the entire row
            row_num = ws_only2.max_row
            for col in range(1, 5):
                ws_only2.cell(row=row_num, column=col).fill = added_fill
        
        ws_only2.column_dimensions['A'].width = 40
        ws_only2.column_dimensions['B'].width = 35
        ws_only2.column_dimensions['C'].width = 35
        ws_only2.column_dimensions['D'].width = 35
        
        # Save workbook
        progress_label.config(text="Saving Excel file...")
        progress_bar['value'] = 95
        progress_win.update()
        
        wb.save(excel_path)
        
        progress_bar['value'] = 100
        progress_win.update()
        progress_win.destroy()
        
        # Success message with option to open
        result = messagebox.askyesno(
            "Report Generated",
            f"Snapshot comparison report generated successfully!\\n\\n"
            f"Excel Report: {os.path.basename(excel_path)}\\n"
            f"HTML Diffs Folder: {os.path.basename(html_output_dir)}\\n\\n"
            f"Common Components: {len(selected_components)}\\n"
            f"Only in Snapshot 1: {len(only_in_1)}\\n"
            f"Only in Snapshot 2: {len(only_in_2)}\\n\\n"
            f"Would you like to open the report now?"
        )
        
        if result:
            os.startfile(excel_path)
            
    except Exception as e:
        progress_win.destroy()
        messagebox.showerror("Error", f"Failed to generate report:\\n{str(e)}")


# ---------------------------------------------------------------------------
# Excel writer - Optimized for speed
# ---------------------------------------------------------------------------
def write_excel_report(results, excel_report_path, folder1="", folder2=""):
    wb = Workbook()
    
    # Create Overview sheet first
    create_overview_sheet(wb, results, folder1, folder2)
    
    # Create detailed report sheet
    ws = wb.create_sheet("Migration Analysis Report")
    wb.active = ws

    # Updated headers with new columns B, C, D
    folder1_short = folder1[-50:] if len(folder1) > 50 else folder1
    folder2_short = folder2[-50:] if len(folder2) > 50 else folder2
    
    headers = [
        "File Path", 
        f"Lines in Platform\n({folder1_short})", 
        f"Lines in Project\n({folder2_short})", 
        "Line Comparison Status",
        "Status", 
        "HTML Diff Report", 
        "Purpose of Change (Migration Analysis)",
        "ChangeSet & WorkItem from RTC History"
    ]
    
    # Pre-define all styles once
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    different_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    identical_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    comments_only_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    only_folder1_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")  # Blue
    only_folder2_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # Orange

    # Write all data at once (fastest method) - sanitize for Excel compatibility
    ws.append(headers)
    for row_data in results:
        # Sanitize each cell value to remove illegal Excel characters
        sanitized_row = [sanitize_for_excel(cell) if isinstance(cell, str) else cell for cell in row_data]
        # Try to extract workitem_id(s) from row_data if available (customize this logic as needed)
        workitem_ids = []
        if isinstance(row_data, dict):
            if 'workitem_id' in row_data:
                ids = row_data['workitem_id']
                if isinstance(ids, list):
                    workitem_ids.extend(ids)
                else:
                    workitem_ids.append(ids)
            elif 'workitem_ids' in row_data:
                ids = row_data['workitem_ids']
                if isinstance(ids, list):
                    workitem_ids.extend(ids)
                else:
                    workitem_ids.append(ids)
        elif isinstance(row_data, (list, tuple)):
            # Extract all numeric strings that could be workitem IDs
            for item in row_data:
                if isinstance(item, str) and item.isdigit() and len(item) >= 4:  # Assuming workitem IDs are at least 4 digits
                    workitem_ids.append(item)
        
        # Build the ALM links if workitem_ids are found
        if workitem_ids:
            # Create a list of hyperlinks for all workitems
            alm_links = []
            for wid in workitem_ids:
                link = f"https://rb-alm-06-p.de.bosch.com/ccm/resource/itemName/com.ibm.team.workitem.WorkItem/{wid}"
                alm_links.append(f"WorkItem {wid}: {link}")
            alm_link = "\n".join(alm_links)
        else:
            alm_link = ""
        sanitized_row.append(alm_link)
        ws.append(sanitized_row)

    # Apply header formatting (only once per column)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Batch apply status coloring using ranges
    for row_idx in range(2, len(results) + 2):
        status_cell = ws.cell(row=row_idx, column=5)  # Status column is at E
        
        if status_cell.value == "Different":
            status_cell.fill = different_fill
        elif status_cell.value == "Identical":
            status_cell.fill = identical_fill
        elif status_cell.value == "Comments update only":
            status_cell.fill = comments_only_fill
        elif status_cell.value == "Only in Platform":
            # Color entire row blue for files only in Platform
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = only_folder1_fill
        elif status_cell.value == "Only in Project":
            # Color entire row orange for files only in Project
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = only_folder2_fill

    # Set fixed column widths (no calculation needed for speed)
    ws.column_dimensions['A'].width = 50  # File Path
    ws.column_dimensions['B'].width = 20  # Lines in Platform
    ws.column_dimensions['C'].width = 20  # Lines in Project
    ws.column_dimensions['D'].width = 40  # Line Comparison Status
    ws.column_dimensions['E'].width = 15  # Status
    ws.column_dimensions['F'].width = 25  # HTML Diff Report
    ws.column_dimensions['G'].width = 80  # Purpose of Change
    ws.column_dimensions['H'].width = 80  # ChangeSet & WorkItem from RTC History

    wb.save(excel_report_path)

# ---------------------------------------------------------------------------
# File Mapping Preview and Confirmation Dialog
# ---------------------------------------------------------------------------
def show_file_mapping_dialog(folder1, folder2):
    """
    Show a preview dialog of file mappings with ability to manually map files.
    Returns: (confirmed, custom_mappings)
        confirmed: True if user confirmed, False if cancelled
        custom_mappings: dict of {file1_path: file2_path} for custom mappings
    """
    dialog = tk.Toplevel()
    dialog.title("File Mapping Preview - Confirm Comparison")
    dialog.geometry("1400x800")
    dialog.configure(bg="#f0f4f7")
    dialog.grab_set()  # Modal dialog
    
    # Result variables
    result = {'confirmed': False, 'mappings': {}}
    
    # Prepare folders (extract ZIP if needed)
    temp_dirs_to_cleanup = []
    folder1_actual, is_temp1, orig1 = prepare_folder_path(folder1)
    folder2_actual, is_temp2, orig2 = prepare_folder_path(folder2)
    
    if is_temp1:
        temp_dirs_to_cleanup.append(folder1_actual)
    if is_temp2:
        temp_dirs_to_cleanup.append(folder2_actual)
    
    if not folder1_actual or not folder2_actual:
        messagebox.showerror("Error", "Could not prepare folders for mapping preview.")
        for temp_dir in temp_dirs_to_cleanup:
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
        dialog.destroy()
        return False, {}
    
    # Title
    title_frame = tk.Frame(dialog, bg="#003366", height=60)
    title_frame.pack(fill="x")
    title_frame.pack_propagate(False)
    
    tk.Label(
        title_frame,
        text="File Mapping Preview - Review and Confirm",
        font=("Segoe UI", 14, "bold"),
        bg="#003366",
        fg="white"
    ).pack(pady=15)
    
    # Info frame
    info_frame = tk.Frame(dialog, bg="#f0f4f7")
    info_frame.pack(fill="x", padx=20, pady=10)
    
    # Show original paths (ZIP names if applicable)
    folder1_display = f"{folder1}" + (" [ZIP - Extracted]" if is_temp1 else "")
    folder2_display = f"{folder2}" + (" [ZIP - Extracted]" if is_temp2 else "")
    
    tk.Label(
        info_frame,
        text=f"Platform (Baseline): {folder1_display}",
        font=("Segoe UI", 9, "bold"),
        bg="#f0f4f7",
        fg="#003366"
    ).pack(anchor="w")
    
    tk.Label(
        info_frame,
        text=f"Project (Comparison): {folder2_display}",
        font=("Segoe UI", 9, "bold"),
        bg="#f0f4f7",
        fg="#003366"
    ).pack(anchor="w")
    
    tk.Label(
        info_frame,
        text="Review the file mappings below. You can manually map files by selecting rows and clicking 'Map Selected'.",
        font=("Segoe UI", 9),
        bg="#f0f4f7",
        fg="#666666"
    ).pack(anchor="w", pady=(5, 0))
    
    # Create frame for treeview
    tree_frame = tk.Frame(dialog, bg="#ffffff")
    tree_frame.pack(fill="both", expand=True, padx=20, pady=10)
    
    # Scrollbars
    vsb = tk.Scrollbar(tree_frame, orient="vertical")
    hsb = tk.Scrollbar(tree_frame, orient="horizontal")
    
    # Treeview with columns
    columns = ("Platform File", "Project File", "Status", "Action")
    tree = ttk.Treeview(tree_frame, columns=columns, show="headings", 
                        yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=25)
    
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)
    
    # Configure columns
    tree.heading("Platform File", text="Platform File")
    tree.heading("Project File", text="Project File")
    tree.heading("Status", text="Status")
    tree.heading("Action", text="Action")
    
    tree.column("Platform File", width=500, anchor="w")
    tree.column("Project File", width=500, anchor="w")
    tree.column("Status", width=200, anchor="center")
    tree.column("Action", width=150, anchor="center")
    
    # Pack treeview and scrollbars
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)
    
    # Collect file mappings from extracted/actual folders
    files1 = {}
    files2 = {}
    
    for root_dir, dirs, files in os.walk(folder1_actual):
        for file in files:
            full_path = os.path.join(root_dir, file)
            rel_path = os.path.relpath(full_path, folder1_actual)
            files1[rel_path] = full_path
    
    for root_dir, dirs, files in os.walk(folder2_actual):
        for file in files:
            full_path = os.path.join(root_dir, file)
            rel_path = os.path.relpath(full_path, folder2_actual)
            files2[rel_path] = full_path
    
    # Populate treeview with mappings
    all_files = sorted(set(files1.keys()) | set(files2.keys()))
    
    for rel_path in all_files:
        file1 = files1.get(rel_path, "")
        file2 = files2.get(rel_path, "")
        
        if file1 and file2:
            status = "Will Compare"
            tag = "compare"
        elif file1 and not file2:
            status = "Only in Platform"
            tag = "only1"
        elif file2 and not file1:
            status = "Only in Project"
            tag = "only2"
        else:
            continue
        
        tree.insert("", "end", values=(
            rel_path if file1 else "[Not in Platform]",
            rel_path if file2 else "[Not in Project]",
            status,
            "Auto-mapped"
        ), tags=(tag,))
    
    # Configure tags for colors
    tree.tag_configure("compare", background="#E8F5E9")
    tree.tag_configure("only1", background="#E3F2FD")
    tree.tag_configure("only2", background="#FFF3E0")
    tree.tag_configure("custom", background="#F3E5F5")
    
    # Statistics label
    stats_frame = tk.Frame(dialog, bg="#f0f4f7")
    stats_frame.pack(fill="x", padx=20, pady=5)
    
    stats_label = tk.Label(
        stats_frame,
        text=f"Total Files: {len(all_files)} | Will Compare: {len([f for f in all_files if f in files1 and f in files2])} | "
             f"Only in Platform: {len([f for f in all_files if f in files1 and f not in files2])} | "
             f"Only in Project: {len([f for f in all_files if f in files2 and f not in files1])}",
        font=("Segoe UI", 9, "bold"),
        bg="#f0f4f7",
        fg="#003366"
    )
    stats_label.pack()
    
    # Button frame
    button_frame = tk.Frame(dialog, bg="#f0f4f7")
    button_frame.pack(fill="x", padx=20, pady=15)
    
    # Manual mapping function
    def manual_mapping():
        """Allow user to manually map selected files"""
        mapping_window = tk.Toplevel(dialog)
        mapping_window.title("Manual File Mapping")
        mapping_window.geometry("900x600")
        mapping_window.configure(bg="#f0f4f7")
        mapping_window.grab_set()
        
        tk.Label(
            mapping_window,
            text="Select files to map together",
            font=("Segoe UI", 12, "bold"),
            bg="#f0f4f7"
        ).pack(pady=10)
        
        # Instructions
        instruction_text = (
            "📋 How to map files:\n"
            "1. Click to select a file from Folder 1 (LEFT)\n"
            "2. Click to select a file from Folder 2 (RIGHT)\n"
            "3. Click 'Create Mapping' button to map them together\n"
            "   OR double-click on Folder 2 file to auto-map with selected Folder 1 file"
        )
        tk.Label(
            mapping_window,
            text=instruction_text,
            font=("Segoe UI", 9),
            bg="#FFF9C4",
            fg="#000",
            justify="left",
            relief="solid",
            borderwidth=1,
            padx=10,
            pady=10
        ).pack(padx=20, pady=(0, 10), fill="x")
        
        # Two listboxes side by side
        list_frame = tk.Frame(mapping_window, bg="#f0f4f7")
        list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Selection indicator labels
        selected_file1 = tk.StringVar(value="No file selected")
        selected_file2 = tk.StringVar(value="No file selected")
        
        # Folder 1 section
        folder1_frame = tk.Frame(list_frame, bg="#f0f4f7")
        folder1_frame.grid(row=0, column=0, sticky="nsew", padx=5)
        
        tk.Label(folder1_frame, text="📁 Folder 1 Files:", font=("Segoe UI", 10, "bold"), bg="#f0f4f7").pack(anchor="w")
        tk.Label(
            folder1_frame,
            textvariable=selected_file1,
            font=("Segoe UI", 8, "italic"),
            bg="#E3F2FD",
            fg="#1565C0",
            anchor="w",
            wraplength=450,
            relief="sunken",
            padx=5,
            pady=3
        ).pack(fill="x", pady=(3, 5))
        
        list1_scroll = tk.Scrollbar(folder1_frame)
        list1 = tk.Listbox(
            folder1_frame,
            width=55,
            height=20,
            yscrollcommand=list1_scroll.set,
            selectmode=tk.SINGLE,
            font=("Consolas", 9),
            bg="#FAFAFA"
        )
        list1_scroll.config(command=list1.yview)
        list1.pack(side="left", fill="both", expand=True)
        list1_scroll.pack(side="right", fill="y")
        
        # Arrow indicator in middle
        arrow_frame = tk.Frame(list_frame, bg="#f0f4f7", width=60)
        arrow_frame.grid(row=0, column=1, padx=10)
        tk.Label(
            arrow_frame,
            text="➜\nMAP\n➜",
            font=("Segoe UI", 14, "bold"),
            bg="#f0f4f7",
            fg="#FF6F00"
        ).pack(expand=True)
        
        # Folder 2 section
        folder2_frame = tk.Frame(list_frame, bg="#f0f4f7")
        folder2_frame.grid(row=0, column=2, sticky="nsew", padx=5)
        
        tk.Label(folder2_frame, text="📁 Folder 2 Files:", font=("Segoe UI", 10, "bold"), bg="#f0f4f7").pack(anchor="w")
        tk.Label(
            folder2_frame,
            textvariable=selected_file2,
            font=("Segoe UI", 8, "italic"),
            bg="#E8F5E9",
            fg="#2E7D32",
            anchor="w",
            wraplength=450,
            relief="sunken",
            padx=5,
            pady=3
        ).pack(fill="x", pady=(3, 5))
        
        list2_scroll = tk.Scrollbar(folder2_frame)
        list2 = tk.Listbox(
            folder2_frame,
            width=55,
            height=20,
            yscrollcommand=list2_scroll.set,
            selectmode=tk.SINGLE,
            font=("Consolas", 9),
            bg="#FAFAFA"
        )
        list2_scroll.config(command=list2.yview)
        list2.pack(side="left", fill="both", expand=True)
        list2_scroll.pack(side="right", fill="y")
        
        # Make columns expand
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_columnconfigure(2, weight=1)
        list_frame.grid_rowconfigure(0, weight=1)
        
        # Populate listboxes
        for rel_path in sorted(files1.keys()):
            list1.insert(tk.END, rel_path)
        
        for rel_path in sorted(files2.keys()):
            list2.insert(tk.END, rel_path)
        
        # Store persistent selections
        persistent_selection = {'list1_index': None, 'list2_index': None}
        
        # Selection handlers with persistent storage
        def on_list1_select(event):
            sel = list1.curselection()
            if sel:
                persistent_selection['list1_index'] = sel[0]
                selected_file1.set(f"✓ Selected: {list1.get(sel[0])}")
                # Highlight selected item with color
                list1.itemconfig(sel[0], bg="#BBDEFB", fg="#000")
                # Remove highlight from previously selected items
                for i in range(list1.size()):
                    if i != sel[0]:
                        list1.itemconfig(i, bg="white", fg="black")
        
        def on_list2_select(event):
            sel = list2.curselection()
            if sel:
                persistent_selection['list2_index'] = sel[0]
                selected_file2.set(f"✓ Selected: {list2.get(sel[0])}")
                # Highlight selected item with color
                list2.itemconfig(sel[0], bg="#C8E6C9", fg="#000")
                # Remove highlight from previously selected items
                for i in range(list2.size()):
                    if i != sel[0]:
                        list2.itemconfig(i, bg="white", fg="black")
        
        def on_list2_double_click(event):
            """Double-click on list2 to auto-map with selected list1 file"""
            idx1 = persistent_selection['list1_index']
            idx2 = persistent_selection['list2_index']
            
            if idx1 is not None and idx2 is not None:
                file1_rel = list1.get(idx1)
                file2_rel = list2.get(idx2)
                
                # Add to custom mappings
                result['mappings'][file1_rel] = file2_rel
                
                # Add to treeview
                tree.insert("", "end", values=(
                    file1_rel,
                    file2_rel,
                    "Will Compare",
                    "Custom Mapping"
                ), tags=("custom",))
                
                messagebox.showinfo("Mapping Created", f"✓ Successfully Mapped:\n\n{file1_rel}\n     ↓\n{file2_rel}")
                mapping_window.destroy()
            elif idx1 is None:
                messagebox.showwarning("No Selection", "⚠ Please select a file from Platform first")
        
        # Bind events
        list1.bind('<<ListboxSelect>>', on_list1_select)
        list2.bind('<<ListboxSelect>>', on_list2_select)
        list2.bind('<Double-Button-1>', on_list2_double_click)
        
        def create_mapping():
            idx1 = persistent_selection['list1_index']
            idx2 = persistent_selection['list2_index']
            
            if idx1 is None or idx2 is None:
                messagebox.showwarning("Selection Required", "⚠ Please select one file from EACH list")
                return
            
            file1_rel = list1.get(idx1)
            file2_rel = list2.get(idx2)
            
            # Add to custom mappings
            result['mappings'][file1_rel] = file2_rel
            
            # Add to treeview
            tree.insert("", "end", values=(
                file1_rel,
                file2_rel,
                "Will Compare",
                "Custom Mapping"
            ), tags=("custom",))
            
            messagebox.showinfo("Mapping Created", f"✓ Successfully Mapped:\n\n{file1_rel}\n     ↓\n{file2_rel}")
            mapping_window.destroy()
        
        # Button frame
        button_frame = tk.Frame(mapping_window, bg="#f0f4f7")
        button_frame.pack(pady=15)
        
        tk.Button(
            button_frame,
            text="✓ Create Mapping",
            command=create_mapping,
            bg="#007B3E",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            width=20,
            height=2
        ).pack(side="left", padx=10)
        
        tk.Button(
            button_frame,
            text="✗ Cancel",
            command=mapping_window.destroy,
            bg="#C62828",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            width=15,
            height=2
        ).pack(side="left", padx=10)
    
    tk.Button(
        button_frame,
        text="Manual File Mapping",
        command=manual_mapping,
        bg="#FF8C00",
        fg="white",
        font=("Segoe UI", 10, "bold"),
        width=20
    ).pack(side="left", padx=5)
    
    def confirm_and_proceed():
        result['confirmed'] = True
        # Cleanup temp directories for mapping preview
        for temp_dir in temp_dirs_to_cleanup:
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
        dialog.destroy()
    
    def cancel_comparison():
        result['confirmed'] = False
        # Cleanup temp directories for mapping preview
        for temp_dir in temp_dirs_to_cleanup:
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
        dialog.destroy()
    
    tk.Button(
        button_frame,
        text="✓ Confirm & Generate Report",
        command=confirm_and_proceed,
        bg="#007B3E",
        fg="white",
        font=("Segoe UI", 11, "bold"),
        width=25
    ).pack(side="right", padx=5)
    
    tk.Button(
        button_frame,
        text="✗ Cancel",
        command=cancel_comparison,
        bg="#C62828",
        fg="white",
        font=("Segoe UI", 11, "bold"),
        width=15
    ).pack(side="right", padx=5)
    
    # Wait for dialog to close
    dialog.wait_window()
    
    return result['confirmed'], result['mappings']


# ---------------------------------------------------------------------------
# Process single file comparison (for parallel execution)
# ---------------------------------------------------------------------------
def process_file_comparison(args):
    """Process a single file comparison - designed for parallel execution"""
    rel_path, path1, path2, output_dir, is_custom_mapping = args
    
    try:
        # Handle ZIP files
        temp_dirs = []
        original_path1, original_path2 = path1, path2
        
        if path1 and path1.lower().endswith('.zip'):
            temp_dir1 = extract_zip_to_temp(path1)
            if temp_dir1:
                temp_dirs.append(temp_dir1)
                path1 = temp_dir1
        
        if path2 and path2.lower().endswith('.zip'):
            temp_dir2 = extract_zip_to_temp(path2)
            if temp_dir2:
                temp_dirs.append(temp_dir2)
                path2 = temp_dir2
        
        # Handle case where one path doesn't exist
        if path1 and not path2:
            lines1 = count_file_lines(path1) if os.path.isfile(path1) else 0
            result = [
                rel_path + (" (Custom Mapping)" if is_custom_mapping else ""),
                lines1,
                0,
                f"File only exists in platform ({lines1} lines)",
                "Only in Platform",
                "",
                "File exists only in Platform" + (" (Custom Mapping)" if is_custom_mapping else ""),
                ""  # Workitem column placeholder
            ]
        elif path2 and not path1:
            lines2 = count_file_lines(path2) if os.path.isfile(path2) else 0
            result = [
                rel_path + (" (Custom Mapping)" if is_custom_mapping else ""),
                0,
                lines2,
                f"File only exists in project ({lines2} lines)",
                "Only in Project",
                "",
                "File exists only in Project" + (" (Custom Mapping)" if is_custom_mapping else ""),
                ""  # Workitem column placeholder
            ]
        else:
            # Both exist - compare them
            lines1 = count_file_lines(path1) if os.path.isfile(path1) else 0
            lines2 = count_file_lines(path2) if os.path.isfile(path2) else 0
            
            # Check if XML files
            is_xml = path1.lower().endswith('.xml') and path2.lower().endswith('.xml')
            
            if is_xml:
                text1 = normalize_xml(path1)
                text2 = normalize_xml(path2)
            else:
                text1 = read_file_as_text(path1)
                text2 = read_file_as_text(path2)
            
            files_identical = (text1 == text2)
            line_status = get_line_comparison_status(lines1, lines2, files_identical, text1, text2)
            
            if files_identical:
                result = [
                    rel_path + (" (Custom Mapping)" if is_custom_mapping else ""),
                    lines1,
                    lines2,
                    line_status,
                    "Identical",
                    "",
                    "No difference found." + (" (Custom Mapping)" if is_custom_mapping else ""),
                    ""  # Workitem column placeholder
                ]
            else:
                # Use original paths for diff generation
                html_path, t1, t2 = generate_html_diff(
                    original_path1 if original_path1 else path1,
                    original_path2 if original_path2 else path2,
                    rel_path.replace(" ← → ", "_vs_"),
                    output_dir
                )
                purpose = generate_purpose_of_change(t1, t2)
                html_link = f'=HYPERLINK("{html_path}","View Diff")'
                
                if is_only_comment_change(text1, text2):
                    status = "Comments update only"
                else:
                    status = "Different"
                
                result = [
                    rel_path + (" (Custom Mapping)" if is_custom_mapping else ""),
                    lines1,
                    lines2,
                    line_status,
                    status,
                    html_link,
                    purpose + (" (Custom Mapping)" if is_custom_mapping else ""),
                    ""  # Workitem column - will be populated below
                ]
        
        # Fetch changeset and workitem information from RTC if enabled
        workitem_data = ""
        if RTC_ENABLED:
            print(f"\n[RTC] Processing file: {rel_path}")
            print(f"[RTC] RTC_ENABLED: {RTC_ENABLED}")
            print(f"[RTC] RTC_USERNAME: {RTC_USERNAME}")
            print(f"[RTC] RTC_WORKSPACE_NAME: {RTC_WORKSPACE_NAME}")
            print(f"[RTC] RTC_STREAM_NAME: {RTC_STREAM_NAME}")
            
            try:
                # Get the repository path and file to check
                file_to_check = path2 if path2 else path1  # Prefer path2 (newer version)
                repo_path = path2 if path2 else path1
                
                print(f"[RTC] File to check: {file_to_check}")
                
                # Get the repository root (parent directory)
                if repo_path:
                    # Try to get repository root
                    repo_root = os.path.dirname(repo_path)
                    while repo_root and not os.path.exists(os.path.join(repo_root, ".jazz5")):
                        parent = os.path.dirname(repo_root)
                        if parent == repo_root:
                            break
                        repo_root = parent
                    
                    print(f"[RTC] Repository root: {repo_root}")
                    
                    if file_to_check and os.path.isfile(file_to_check):
                        print(f"[RTC] Calling get_workitems_for_file...")
                        changeset_info = get_workitems_for_file(file_to_check, repo_root, RTC_USERNAME, RTC_PASSWORD,
                                                               workspace_name=RTC_WORKSPACE_NAME, stream_name=RTC_STREAM_NAME)
                        
                        print(f"[RTC] Changeset info returned: {changeset_info}")
                        
                        if changeset_info:
                            changeset_url = changeset_info.get("changeset_url", "")
                            changeset_comment = changeset_info.get("changeset_comment", "")
                            workitem_ids = changeset_info.get("workitem_ids", [])
                            
                            output_lines = []
                            
                            # Always show workspace context
                            if RTC_WORKSPACE_NAME:
                                output_lines.append(f"Workspace: {RTC_WORKSPACE_NAME}")
                            if RTC_STREAM_NAME:
                                output_lines.append(f"Stream: {RTC_STREAM_NAME}")
                            
                            # Add changeset information
                            if changeset_url:
                                output_lines.append(f"ChangeSet: {changeset_url}")
                                if changeset_comment:
                                    output_lines.append(f"Comment: {changeset_comment}")
                            
                            # Add workitem information
                            if workitem_ids:
                                for wid in workitem_ids:
                                    wi_link = f"{RTC_SERVER_URL}/resource/itemName/com.ibm.team.workitem.WorkItem/{wid}"
                                    output_lines.append(f"WorkItem {wid}: {wi_link}")
                            
                            if not output_lines:
                                output_lines.append("No changeset/workitem data available")
                            
                            workitem_data = "\n".join(output_lines)
                            print(f"[RTC] Final workitem_data: {workitem_data}")
                        else:
                            # Show at least workspace info
                            output_lines = []
                            if RTC_WORKSPACE_NAME:
                                output_lines.append(f"Workspace: {RTC_WORKSPACE_NAME}")
                            if RTC_STREAM_NAME:
                                output_lines.append(f"Stream: {RTC_STREAM_NAME}")
                            output_lines.append("(Changeset history unavailable - SCM tools not installed)")
                            workitem_data = "\n".join(output_lines)
                            print(f"[RTC] Using fallback data: {workitem_data}")
            except Exception as e:
                print(f"[RTC] Error fetching changeset and workitems for {rel_path}: {e}")
                import traceback
                traceback.print_exc()
                # Show error in column
                workitem_data = f"Error: {str(e)}"
        # RTC disabled – no per-file log needed (reduces noise when processing many files)
        
        # Update the workitem column in result
        result[7] = workitem_data
        
        # Cleanup temp directories
        for temp_dir in temp_dirs:
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
        
        return result
    except Exception as e:
        print(f"Error processing {rel_path}: {e}")
        return [
            rel_path,
            0,
            0,
            "Error during comparison",
            "Error",
            "",
            f"Error: {str(e)}",
            ""  # Workitem column placeholder
        ]

# ---------------------------------------------------------------------------
# AI-powered analysis engine (heuristic + optional OpenAI)
# ---------------------------------------------------------------------------
_AI_IMPORTANT_EXTS = {
    ".c", ".h", ".cpp", ".hpp", ".cs", ".py", ".java",
    ".xml", ".arxml", ".cfg", ".mk", ".mak", ".cmake",
    ".json", ".yaml", ".yml", ".ini", ".properties",
    ".bat", ".cmd", ".sh",
}
_AI_GENERATED_MARKERS  = ["generated", "auto-generated", "do not edit", "autogenerated"]
_AI_DEPRECATED_MARKERS = ["deprecated", "obsolete", "legacy", "unused", "todo: remove"]
_AI_IMPORTANT_MARKERS  = ["copyright", "license", "version", "author", "@brief", "@file",
                           "important", "critical", "required"]


def _read_head(path, max_lines=60):
    """Read up to max_lines of text from a file for heuristic analysis."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            return "".join(fh.readline() for _ in range(max_lines)).lower()
    except Exception:
        return ""


def _score_file(path, lines):
    """
    Return a numeric score for a file based on heuristics.
    Higher = more likely the version to keep.
    """
    if not path or not os.path.isfile(path):
        return 0
    score = 0
    ext = os.path.splitext(path)[1].lower()
    head = _read_head(path)

    # Prefer source-code / config files over temp/log files
    if ext in _AI_IMPORTANT_EXTS:
        score += 20
    elif ext in {".log", ".tmp", ".bak", ".orig", ".swp"}:
        score -= 20

    # Auto-generated files are lower priority (the generator will recreate them)
    if any(m in head for m in _AI_GENERATED_MARKERS):
        score -= 15

    # Deprecated / legacy markers reduce priority
    if any(m in head for m in _AI_DEPRECATED_MARKERS):
        score -= 25

    # Important content markers increase priority
    if any(m in head for m in _AI_IMPORTANT_MARKERS):
        score += 15

    # Larger files usually carry more content / work done
    score += min(lines, 2000) // 40

    # File size on disk (catch binary files)
    try:
        score += min(os.path.getsize(path), 200_000) // 10_000
    except Exception:
        pass

    return score


def ai_analyze_file(path1, path2, status, lines1, lines2, openai_key=""):
    """
    Heuristic + optional AI analysis of which file version to keep.
    Returns a detailed recommendation string.
    """
    ext = os.path.splitext(path1 or path2 or "")[1].lower()
    fname = os.path.basename(path1 or path2 or "")

    # ── Heuristic analysis ────────────────────────────────────────────────────
    lines = []
    verdict = ""
    confidence = ""

    if status == "Only in Platform":
        head = _read_head(path1)
        is_generated  = any(m in head for m in _AI_GENERATED_MARKERS)
        is_deprecated = any(m in head for m in _AI_DEPRECATED_MARKERS)
        is_important  = any(m in head for m in _AI_IMPORTANT_MARKERS) or ext in _AI_IMPORTANT_EXTS

        if is_deprecated:
            verdict = "DO NOT COPY — file appears deprecated/obsolete"
            confidence = "High"
            lines += [
                f"\u2716 File '{fname}' contains deprecation markers.",
                "  It is likely intentionally absent from Project.",
                "  Recommendation: LEAVE IT only in Platform, do not migrate.",
            ]
        elif is_generated:
            verdict = "LOW PRIORITY — auto-generated file"
            confidence = "Medium"
            lines += [
                f"\u2139 File '{fname}' appears to be auto-generated.",
                "  The build system will recreate it in Project if needed.",
                "  Recommendation: Skip migration unless you confirmed it is hand-edited.",
            ]
        elif is_important:
            verdict = "COPY Platform \u2192 Project — likely needed in target"
            confidence = "High"
            lines += [
                f"\u2714 File '{fname}' ({ext}) contains important markers (copyright, version, @brief ...).",
                f"  It has {lines1} lines and looks like meaningful source code/config.",
                "  Recommendation: Copy this file to Project — it was probably missed during migration.",
            ]
        else:
            verdict = "REVIEW BEFORE COPYING"
            confidence = "Low"
            lines += [
                f"\u26a0 File '{fname}' exists only in Platform ({lines1} lines, type: {ext or 'unknown'}).",
                "  Could not determine automatically whether it is needed in Project.",
                "  Recommendation: Open the file, review its purpose, then decide.",
            ]

    elif status == "Only in Project":
        head = _read_head(path2)
        is_generated  = any(m in head for m in _AI_GENERATED_MARKERS)
        is_important  = any(m in head for m in _AI_IMPORTANT_MARKERS) or ext in _AI_IMPORTANT_EXTS

        if is_generated:
            verdict = "LOW PRIORITY — auto-generated in target"
            confidence = "Medium"
            lines += [
                f"\u2139 File '{fname}' was auto-generated by the new build/tool in Project.",
                "  Platform does not need it — the generator was not present there.",
                "  Recommendation: Keep it in Folder 2 only; no action for Folder 1.",
            ]
        elif is_important:
            verdict = "KEEP IN F2 — new meaningful addition"
            confidence = "High"
            lines += [
                f"\u2714 File '{fname}' ({ext}) is a NEW file in Folder 2 ({lines2} lines).",
                "  It contains important content markers and looks like hand-crafted work.",
                "  Recommendation: Keep it in Folder 2. Consider back-porting to Folder 1 if needed.",
            ]
        else:
            verdict = "REVIEW — new file of unknown purpose"
            confidence = "Low"
            lines += [
                f"\u26a0 New file '{fname}' ({lines2} lines, {ext or 'unknown'}) exists only in Project.",
                "  Cannot determine automatically whether it is hand-crafted or auto-generated.",
                "  Recommendation: Open and review before deciding.",
            ]

    elif status in ("Modified", "Different", "Comments update only"):
        s1 = _score_file(path1, lines1)
        s2 = _score_file(path2, lines2)
        delta = lines2 - lines1
        head1 = _read_head(path1)
        head2 = _read_head(path2)
        dep1  = any(m in head1 for m in _AI_DEPRECATED_MARKERS)
        dep2  = any(m in head2 for m in _AI_DEPRECATED_MARKERS)

        if dep1 and not dep2:
            verdict = "PREFER F2 — baseline has deprecation markers"
            confidence = "High"
            lines += [
                "\u2714 Folder 1 version contains deprecation/legacy markers; Folder 2 does not.",
                "  Recommendation: Trust Folder 2 (target) as the canonical version.",
            ]
        elif dep2 and not dep1:
            verdict = "PREFER F1 — target has deprecation markers"
            confidence = "High"
            lines += [
                "\u2714 Folder 2 version contains deprecation markers; Folder 1 does not.",
                "  Recommendation: Investigate why the target version was deprecated.",
            ]
        elif status == "Comments update only":
            verdict = "PREFER F2 — only comments differ, keep the newer description"
            confidence = "High"
            lines += [
                "\u2714 Only comments changed — code logic is identical.",
                "  Folder 2 has updated comments/documentation.",
                "  Recommendation: Keep Folder 2 version (more up-to-date comments).",
            ]
        elif delta > 50:
            verdict = f"PREFER F2 — significantly larger (+{delta} lines)"
            confidence = "Medium"
            lines += [
                f"\u2714 Folder 2 version is {delta} lines larger ({lines1} \u2192 {lines2}).",
                "  This suggests new features or expanded implementation were added.",
                "  Recommendation: Use Folder 2; cherry-pick any Folder 1-only logic if needed.",
            ]
        elif delta < -50:
            verdict = f"PREFER F1 — Folder 2 is shorter by {-delta} lines (possible truncation?)"
            confidence = "Medium"
            lines += [
                f"\u26a0 Folder 2 version is {-delta} lines SHORTER ({lines1} \u2192 {lines2}).",
                "  Code may have been accidentally removed or the file was partially migrated.",
                "  Recommendation: Review the diff carefully before accepting Folder 2.",
            ]
        elif s2 >= s1:
            verdict = "PREFER F2 (target) — heuristic score favors target version"
            confidence = "Medium"
            lines += [
                f"\u2714 Both versions are similar in size ({lines1} vs {lines2} lines).",
                "  Heuristic analysis slightly favors Folder 2 (target/newer).",
                "  Recommendation: Keep Folder 2. View the diff to verify nothing was lost.",
            ]
        else:
            verdict = "PREFER F1 (baseline) — heuristic score favors baseline version"
            confidence = "Medium"
            lines += [
                f"\u26a0 Both versions similar in size ({lines1} vs {lines2} lines).",
                "  Heuristic analysis slightly favors Folder 1 (baseline).",
                "  Recommendation: Manually review the diff to confirm.",
            ]
    else:
        return "AI analysis not applicable for status: " + status

    heuristic_result = "\n".join(lines)

    # ── Optional OpenAI enhancement ───────────────────────────────────────────
    openai_result = ""
    if openai_key and openai_key.strip().startswith("sk-"):
        try:
            import urllib.request, json as _json
            # Build a compact diff summary to send (max ~800 chars each side)
            def _snippet(p, max_c=800):
                if not p or not os.path.isfile(p):
                    return "(not available)"
                try:
                    with open(p, "r", encoding="utf-8", errors="replace") as fh:
                        content = fh.read(max_c)
                    return content + ("..." if os.path.getsize(p) > max_c else "")
                except Exception:
                    return "(unreadable)"

            snip1 = _snippet(path1)
            snip2 = _snippet(path2)
            prompt = (
                f"You are a software migration expert. Compare these two versions of a file "
                f"named '{fname}' and decide which version to keep.\n"
                f"Comparison status: {status}\n"
                f"Lines in F1 (baseline): {lines1}\n"
                f"Lines in F2 (target): {lines2}\n\n"
                f"F1 content snippet:\n```\n{snip1}\n```\n\n"
                f"F2 content snippet:\n```\n{snip2}\n```\n\n"
                "In 3-5 sentences: which version should be kept and why?"
            )
            payload = _json.dumps({
                "model": "gpt-4o-mini",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 300,
                "temperature": 0.3,
            }).encode()
            req = urllib.request.Request(
                "https://api.openai.com/v1/chat/completions",
                data=payload,
                headers={
                    "Authorization": f"Bearer {openai_key.strip()}",
                    "Content-Type": "application/json",
                },
            )
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = _json.loads(resp.read())
            openai_result = data["choices"][0]["message"]["content"].strip()
        except Exception as ex:
            openai_result = f"(OpenAI call failed: {ex})"

    # ── Compose final output ──────────────────────────────────────────────────
    out = []
    out.append(f"\u2728 AI RECOMMENDATION  [Confidence: {confidence}]")
    out.append(f"   \u27a4 {verdict}")
    out.append("")
    out.append("\u25b6 Heuristic Analysis:")
    out.extend(["  " + l for l in heuristic_result.splitlines()])
    if openai_result:
        out.append("")
        out.append("\u25b6 GPT-4o-mini Analysis:")
        for l in openai_result.splitlines():
            out.append("  " + l)
    return "\n".join(out)


# ---------------------------------------------------------------------------
# Google Gemini Flash AI merge engine  (FREE — https://aistudio.google.com/app/apikey)
# ---------------------------------------------------------------------------
_GEMINI_MODEL   = "gemini-2.0-flash"
_GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
_DEP_SCAN_EXTS  = {".c", ".h", ".cpp", ".hpp", ".cs", ".py", ".java",
                   ".xml", ".arxml", ".cfg", ".mk", ".mak",
                   ".json", ".yaml", ".yml", ".properties"}


def _gemini_call(api_key, system_prompt, user_prompt, max_tokens=8192):
    """Send a message to Google Gemini Flash and return the text response.
    Automatically routes through the corporate Bosch NTLM proxy.
    Password is requested once via a popup and cached for the session.
    """
    import json as _json, urllib.parse, getpass

    url = f"{_GEMINI_API_URL}?key={urllib.parse.quote(api_key.strip())}"
    headers  = {"content-type": "application/json"}
    payload  = _json.dumps({
        "system_instruction": {"parts": [{"text": system_prompt}]},
        "contents":           [{"parts": [{"text": user_prompt}]}],
        "generationConfig":   {"maxOutputTokens": max_tokens, "temperature": 0.3},
    })

    import requests as _req

    def _build_session(user, password):
        """Return a requests.Session configured with NTLM proxy auth."""
        sess = _req.Session()
        if PROXY_URL and PROXY_URL.strip():
            _pu = PROXY_URL.strip().rstrip("/")
            sess.proxies = {"http": _pu, "https": _pu}
            if user and password:
                try:
                    from requests_ntlm import HttpNtlmAuth
                    domain_user = f"{PROXY_DOMAIN}\\{user}" if "\\" not in user else user
                    sess.auth = HttpNtlmAuth(domain_user, password)
                except ImportError:
                    pass  # no NTLM library — try basic proxy URL embedding
        return sess

    def _get_credentials():
        """Return (user, password) from cache, constants, or a one-time Tkinter dialog."""
        if _proxy_cred_cache:
            return _proxy_cred_cache.get("user", ""), _proxy_cred_cache.get("pass", "")

        _u = PROXY_USER.strip() if PROXY_USER else getpass.getuser()
        _p = PROXY_PASS.strip() if PROXY_PASS else ""

        if not _p:
            # Show a one-time password popup
            import tkinter as _tk, tkinter.simpledialog as _sd
            _root_hidden = _tk.Tk(); _root_hidden.withdraw()
            _p = _sd.askstring(
                "Proxy Authentication",
                f"Enter your Bosch Windows password for:\n"
                f"  Proxy : {PROXY_URL}\n"
                f"  User  : {PROXY_DOMAIN}\\{_u}\n",
                show="*", parent=_root_hidden
            ) or ""
            _root_hidden.destroy()

        _proxy_cred_cache["user"] = _u
        _proxy_cred_cache["pass"] = _p
        return _u, _p

    user, password = _get_credentials()
    sess = _build_session(user, password)

    resp = sess.post(url, data=payload.encode(), headers=headers, timeout=90, verify=True)
    resp.raise_for_status()
    data = resp.json()
    return data["candidates"][0]["content"]["parts"][0]["text"]


def _read_file_safe(path, max_chars=8000):
    """Read a file as text, truncating if very large."""
    if not path or not os.path.isfile(path):
        return ""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            content = fh.read(max_chars)
        if os.path.getsize(path) > max_chars:
            content += f"\n... [truncated - file is {os.path.getsize(path)} bytes total]"
        return content
    except Exception:
        return ""


def _scan_related_files(fname, target_files_dict, max_files=8):
    """
    Find files in the TARGET folder that reference fname.
    Returns list of (rel_path, abs_path, lineno, match_line) tuples.
    """
    base        = os.path.basename(fname)
    base_no_ext = os.path.splitext(base)[0]
    patterns = [
        f'include "{base}"',
        f"include '{base}'",
        f"include <{base}>",
        f'import "{base_no_ext}"',
        f"import '{base_no_ext}'",
        f"from {base_no_ext} import",
    ]
    found = []
    for rel, abs_path in sorted(target_files_dict.items()):
        if os.path.splitext(rel)[1].lower() not in _DEP_SCAN_EXTS:
            continue
        try:
            with open(abs_path, "r", encoding="utf-8", errors="replace") as fh:
                for lineno, line in enumerate(fh, 1):
                    ll = line.lower()
                    if any(p.lower() in ll for p in patterns):
                        found.append((rel, abs_path, lineno, line.strip()))
                        break
        except Exception:
            pass
        if len(found) >= max_files:
            break
    return found


def ai_merge_with_gemini(
    path1, path2, status, rel_path,
    files1, files2, folder1_root, folder2_root,
    api_key
):
    """
    Use Google Gemini Flash to intelligently merge two file versions.
    Returns (merged_content, dependency_report, warnings) tuple.
    """
    if not api_key or not api_key.strip():
        raise ValueError("Gemini API key is required for AI Merge.")

    fname = os.path.basename(path1 or path2 or rel_path)
    ext   = os.path.splitext(fname)[1].lower()
    lang  = {
        ".c": "C", ".h": "C", ".cpp": "C++", ".hpp": "C++",
        ".cs": "C#", ".py": "Python", ".java": "Java",
        ".xml": "XML", ".arxml": "ARXML", ".cfg": "Config",
        ".mk": "Makefile", ".mak": "Makefile",
        ".json": "JSON", ".yaml": "YAML", ".yml": "YAML",
        ".properties": "Properties",
    }.get(ext, "text")

    content1 = _read_file_safe(path1)
    content2 = _read_file_safe(path2)

    # Dependency scan on the TARGET side
    if status == "Only in Platform":
        related = _scan_related_files(fname, files2)
        target_label = "Project (target)"
        source_label = "Platform (baseline)"
    elif status == "Only in Project":
        related = _scan_related_files(fname, files1)
        target_label = "Platform (baseline)"
        source_label = "Project (target)"
    else:
        related = _scan_related_files(fname, files2)
        target_label = "Folder 2 (target)"
        source_label = "Folder 1 (baseline)"

    dep_snippets = []
    for rel, abs_path, lineno, match_line in related:
        snippet = _read_file_safe(abs_path, max_chars=1200)
        dep_snippets.append(
            f"--- Dependent: {rel} (reference at line {lineno}: {match_line}) ---\n{snippet}\n"
        )
    dep_context = "\n".join(dep_snippets) if dep_snippets else "(no direct dependencies found in target folder)"

    # Build Gemini prompt
    system_prompt = (
        "You are an expert software engineer specialising in code migration and merging. "
        "Your task is to produce a SINGLE merged file that:\n"
        "1. Combines both versions without losing any logic.\n"
        "2. Resolves all conflicts intelligently.\n"
        "3. Does NOT break any of the dependent files shown (preserve signatures, types, macros).\n"
        "4. Keeps all copyright headers, version comments, and important annotations.\n"
        "5. Preserves the original language syntax and style.\n"
        "\nOUTPUT FORMAT - return EXACTLY two sections separated by ===DEPENDENCY_REPORT===\n"
        "Section 1: Complete merged file (raw file only, no markdown fences, no explanation).\n"
        "Section 2: Bullet-point dependency impact report (max 15 bullets) covering:\n"
        "  - Which dependent files were checked\n"
        "  - Whether the merge could break any of them\n"
        "  - Warnings and recommended follow-up actions\n"
    )

    if status in ("Only in Platform", "Only in Project"):
        src_content = content1 if status == "Only in Platform" else content2
        user_prompt = (
            f"File: {fname} ({lang})\nStatus: {status}\n"
            f"This file exists ONLY in {source_label} and must be integrated into {target_label}.\n\n"
            f"=== FILE CONTENT from {source_label} ===\n{src_content}\n\n"
            f"=== DEPENDENT FILES already in {target_label} ===\n{dep_context}\n\n"
            "Instructions:\n"
            "- Adapt the file if needed to integrate cleanly into the target context.\n"
            "- Adjust include paths / namespaces ONLY if required by the dependent files.\n"
            "- Keep ALL logic intact. Do not simplify or remove code.\n"
            "- Produce the adapted file content then the dependency report.\n"
        )
    else:
        user_prompt = (
            f"File: {fname} ({lang})\nStatus: {status}\n\n"
            f"=== VERSION 1 - {source_label} ({len(content1.splitlines())} lines) ===\n{content1}\n\n"
            f"=== VERSION 2 - {target_label} ({len(content2.splitlines())} lines) ===\n{content2}\n\n"
            f"=== DEPENDENT FILES in target folder ===\n{dep_context}\n\n"
            "Instructions:\n"
            "- Produce a single merged file combining both versions.\n"
            "- Resolve conflicts by taking the most complete/correct implementation.\n"
            "- Do NOT add conflict markers (<<<<, ====, >>>>) in the output.\n"
            "- Preserve all unique code from both versions.\n"
            "- Ensure the merged file is syntactically correct.\n"
            "- Produce merged content then the dependency report.\n"
        )

    raw = _gemini_call(api_key, system_prompt, user_prompt, max_tokens=8192)

    # Parse response
    sep = "===DEPENDENCY_REPORT==="
    if sep in raw:
        parts = raw.split(sep, 1)
        merged_content    = parts[0].strip()
        dependency_report = parts[1].strip()
    else:
        merged_content    = raw.strip()
        dependency_report = "(Gemini did not return a separate dependency report)"

    # Strip markdown code fences if Gemini added them
    for fence in (f"```{lang.lower()}", f"```{ext.lstrip('.')}", "```"):
        if merged_content.startswith(fence):
            merged_content = merged_content[len(fence):].lstrip()
            if merged_content.endswith("```"):
                merged_content = merged_content[:-3].rstrip()
            break

    warnings = [
        line.strip("- \u2022\t ")
        for line in dependency_report.splitlines()
        if any(w in line.lower() for w in ("warn", "break", "conflict", "caution", "risk", "alert"))
    ]

    return merged_content, dependency_report, warnings


# ---------------------------------------------------------------------------
# Interactive comparison results dialog
# ---------------------------------------------------------------------------
def show_comparison_results_dialog(results, folder1_display, folder2_display,
                                   folder1_root, folder2_root, files1, files2):
    """
    Show an interactive results viewer after folder comparison.
    Allows the user to:
      - See all files with their comparison status and a smart suggestion
      - Open/edit either version of a file directly in the dialog
      - Save edits back to disk
      - Copy a file from one side to the other
      - Open the HTML diff report for modified files
    """
    dialog = tk.Toplevel()
    dialog.title("Comparison Results — Interactive Review & Edit")
    dialog.geometry("1600x920")
    dialog.configure(bg="#f0f4f7")
    dialog.grab_set()

    # ── Top title bar ─────────────────────────────────────────────────────────
    top = tk.Frame(dialog, bg="#003366", height=55)
    top.pack(fill="x")
    top.pack_propagate(False)
    tk.Label(
        top,
        text="Comparison Results — Review, Edit & Apply Suggestions",
        font=("Segoe UI", 13, "bold"),
        bg="#003366",
        fg="white"
    ).pack(pady=14)

    # ── AI status bar (keys embedded — no user entry needed) ─────────────────
    ai_bar = tk.Frame(dialog, bg="#1B5E20", height=28)
    ai_bar.pack(fill="x")
    ai_bar.pack_propagate(False)
    _gemini_ready = bool(GEMINI_API_KEY and not GEMINI_API_KEY.startswith("YOUR-"))
    _openai_ready = bool(OPENAI_API_KEY and OPENAI_API_KEY.strip().startswith("sk-"))
    _gemini_status = f"Gemini Flash ({_GEMINI_MODEL})  \u2714 READY" if _gemini_ready else "Gemini Flash  \u2718 Key not set \u2014 edit GEMINI_API_KEY in script (free at aistudio.google.com)"
    _openai_status = "OpenAI GPT-4o-mini  \u2714 READY" if _openai_ready else "OpenAI  (offline \u2014 heuristics only)"
    tk.Label(
        ai_bar,
        text=f"  \U0001f916 AI Enabled  |  {_gemini_status}  |  {_openai_status}",
        font=("Segoe UI", 8, "bold"),
        bg="#1B5E20" if _gemini_ready else "#B71C1C",
        fg="white"
    ).pack(side="left", padx=12, pady=5)

    # ── Stats bar ─────────────────────────────────────────────────────────────
    status_counts = {}
    for r in results:
        st = r[4] if len(r) > 4 else "Error"
        status_counts[st] = status_counts.get(st, 0) + 1

    only1 = status_counts.get("Only in Platform", 0)
    only2 = status_counts.get("Only in Project", 0)
    modified = (status_counts.get("Modified", 0) +
                status_counts.get("Different", 0) +
                status_counts.get("Comments update only", 0))
    identical = status_counts.get("Identical", 0)

    stats_bar = tk.Frame(dialog, bg="#E3F2FD", height=32)
    stats_bar.pack(fill="x")
    stats_bar.pack_propagate(False)
    stats_text = (
        f"Total: {len(results)}   |   "
        f"Only in Platform (Baseline): {only1}   |   "
        f"Only in Project (Target): {only2}   |   "
        f"Modified/Different: {modified}   |   "
        f"Identical: {identical}"
    )
    tk.Label(stats_bar, text=stats_text, font=("Segoe UI", 9, "bold"),
             bg="#E3F2FD", fg="#003366").pack(side="left", padx=15, pady=6)

    # ── Main paned window ─────────────────────────────────────────────────────
    main_frame = tk.Frame(dialog, bg="#f0f4f7")
    main_frame.pack(fill="both", expand=True, padx=8, pady=5)

    pw = tk.PanedWindow(main_frame, orient="horizontal", bg="#b0bec5",
                        sashwidth=5, sashrelief="raised")
    pw.pack(fill="both", expand=True)

    # ── LEFT PANE: File list ──────────────────────────────────────────────────
    left_frame = tk.Frame(pw, bg="#ffffff")
    pw.add(left_frame, minsize=550)

    tk.Label(
        left_frame,
        text=" Compared Files",
        font=("Segoe UI", 10, "bold"),
        bg="#003366",
        fg="white",
        anchor="w"
    ).pack(fill="x")

    # Filter bar
    filter_frame = tk.Frame(left_frame, bg="#ffffff")
    filter_frame.pack(fill="x", pady=2, padx=4)
    tk.Label(filter_frame, text="Filter:", font=("Segoe UI", 8),
             bg="#ffffff").pack(side="left")
    filter_var = tk.StringVar()
    filter_entry = tk.Entry(filter_frame, textvariable=filter_var,
                            font=("Segoe UI", 9), width=22)
    filter_entry.pack(side="left", padx=4)

    status_filter_var = tk.StringVar(value="All")
    status_options = ["All", "Only in Platform", "Only in Project",
                      "Modified", "Different", "Comments update only",
                      "Identical", "Error"]
    status_dd = ttk.Combobox(filter_frame, textvariable=status_filter_var,
                              values=status_options, state="readonly", width=20,
                              font=("Segoe UI", 9))
    status_dd.pack(side="left", padx=4)

    tree_frame = tk.Frame(left_frame, bg="#ffffff")
    tree_frame.pack(fill="both", expand=True)

    vsb = tk.Scrollbar(tree_frame, orient="vertical")
    hsb = tk.Scrollbar(tree_frame, orient="horizontal")
    cols = ("File Path", "Status", "Lines F1", "Lines F2")
    tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                        yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)

    tree.heading("File Path", text="File Path")
    tree.column("File Path", width=350, anchor="w")
    tree.heading("Status", text="Status")
    tree.column("Status", width=180, anchor="center")
    tree.heading("Lines F1", text="Lines F1")
    tree.column("Lines F1", width=65, anchor="center")
    tree.heading("Lines F2", text="Lines F2")
    tree.column("Lines F2", width=65, anchor="center")

    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)

    tree.tag_configure("only1",    background="#BBDEFB", foreground="#0D47A1")
    tree.tag_configure("only2",    background="#FFE0B2", foreground="#BF360C")
    tree.tag_configure("modified", background="#FFF9C4", foreground="#827717")
    tree.tag_configure("identical",background="#C8E6C9", foreground="#1B5E20")
    tree.tag_configure("comments", background="#E8F5E9", foreground="#2E7D32")
    tree.tag_configure("error",    background="#FFCDD2", foreground="#B71C1C")

    STATUS_TAG = {
        "Only in Platform":     "only1",
        "Only in Project":     "only2",
        "Modified":             "modified",
        "Different":            "modified",
        "Identical":            "identical",
        "Comments update only": "comments",
        "Error":                "error",
    }

    def populate_tree(filter_text="", filter_status="All"):
        for item in tree.get_children():
            tree.delete(item)
        ft = filter_text.lower()
        for r in results:
            rel_path = str(r[0]) if len(r) > 0 else ""
            status   = str(r[4]) if len(r) > 4 else "Error"
            lines1   = r[1] if len(r) > 1 else 0
            lines2   = r[2] if len(r) > 2 else 0
            if ft and ft not in rel_path.lower():
                continue
            if filter_status != "All" and status != filter_status:
                continue
            tag = STATUS_TAG.get(status, "")
            tree.insert("", "end",
                        values=(rel_path, status, lines1, lines2),
                        tags=(tag,))

    populate_tree()

    def on_filter_change(*_):
        populate_tree(filter_var.get(), status_filter_var.get())

    filter_var.trace_add("write", on_filter_change)
    status_dd.bind("<<ComboboxSelected>>", on_filter_change)

    # ── RIGHT PANE: Suggestion + editor ──────────────────────────────────────
    right_frame = tk.Frame(pw, bg="#f8f9fa")
    pw.add(right_frame, minsize=700)

    # Suggestion panel
    sug_outer = tk.LabelFrame(
        right_frame,
        text="  Suggestion  ",
        font=("Segoe UI", 9, "bold"),
        bg="#f8f9fa",
        fg="#003366",
        bd=2,
        relief="groove"
    )
    sug_outer.pack(fill="x", padx=8, pady=(6, 3))

    suggestion_text = tk.Text(
        sug_outer, height=5, wrap="word",
        font=("Segoe UI", 9), bg="#FFFDE7", fg="#333333",
        relief="flat", state="disabled"
    )
    suggestion_text.pack(fill="x", padx=5, pady=5)

    # Action buttons row
    btn_frame = tk.Frame(right_frame, bg="#f8f9fa")
    btn_frame.pack(fill="x", padx=8, pady=3)

    def make_btn(parent, label, color, cmd):
        return tk.Button(
            parent, text=label, bg=color, fg="white",
            font=("Segoe UI", 9, "bold"), relief="flat",
            padx=8, pady=4, cursor="hand2", command=cmd
        )

    make_btn(btn_frame, "Open File (F1)", "#1565C0",
             lambda: open_file(1)).pack(side="left", padx=3)
    make_btn(btn_frame, "Open File (F2)", "#BF360C",
             lambda: open_file(2)).pack(side="left", padx=3)
    make_btn(btn_frame, "Copy F1 → F2",  "#2E7D32",
             lambda: copy_file(1, 2)).pack(side="left", padx=3)
    make_btn(btn_frame, "Copy F2 → F1",  "#6A1B9A",
             lambda: copy_file(2, 1)).pack(side="left", padx=3)
    make_btn(btn_frame, "View HTML Diff","#00695C",
             lambda: open_diff()).pack(side="left", padx=3)
    make_btn(btn_frame, "\U0001f916 AI Suggest: Which to Keep", "#4527A0",
             lambda: run_ai_suggest()).pack(side="left", padx=6)
    make_btn(btn_frame, "\U0001f500 AI Smart Merge (Gemini Flash)", "#006064",
             lambda: run_ai_merge()).pack(side="left", padx=6)

    # File content editor (tabbed)
    editor_label_frame = tk.Frame(right_frame, bg="#f8f9fa")
    editor_label_frame.pack(fill="x", padx=8, pady=(6, 0))
    tk.Label(
        editor_label_frame,
        text="File Content Editor  (select a row to load files — edit then Save)",
        font=("Segoe UI", 9, "bold"),
        bg="#f8f9fa", fg="#003366", anchor="w"
    ).pack(side="left")

    editor_tabs = ttk.Notebook(right_frame)
    editor_tabs.pack(fill="both", expand=True, padx=8, pady=3)

    tab_f1 = tk.Frame(editor_tabs, bg="#1e1e1e")
    tab_f2 = tk.Frame(editor_tabs, bg="#1e1e1e")
    editor_tabs.add(tab_f1, text="  Folder 1 File  ")
    editor_tabs.add(tab_f2, text="  Folder 2 File  ")

    def make_editor_pane(parent):
        frm = tk.Frame(parent, bg="#1e1e1e")
        frm.pack(fill="both", expand=True)
        xsb = tk.Scrollbar(frm, orient="horizontal")
        ysb = tk.Scrollbar(frm, orient="vertical")
        xsb.pack(side="bottom", fill="x")
        ysb.pack(side="right",  fill="y")
        ed = tk.Text(
            frm, wrap="none",
            font=("Consolas", 9),
            bg="#1e1e1e", fg="#d4d4d4",
            insertbackground="white",
            selectbackground="#264f78",
            relief="flat",
            yscrollcommand=ysb.set,
            xscrollcommand=xsb.set
        )
        ed.pack(fill="both", expand=True)
        ysb.config(command=ed.yview)
        xsb.config(command=ed.xview)
        return ed

    editor_f1 = make_editor_pane(tab_f1)
    editor_f2 = make_editor_pane(tab_f2)

    # Save buttons
    save_frame = tk.Frame(right_frame, bg="#f8f9fa")
    save_frame.pack(fill="x", padx=8, pady=3)
    make_btn(save_frame, "Save Folder 1 File", "#1565C0",
             lambda: save_content(1)).pack(side="left", padx=3)
    make_btn(save_frame, "Save Folder 2 File", "#BF360C",
             lambda: save_content(2)).pack(side="left", padx=3)

    # ── State ─────────────────────────────────────────────────────────────────
    current = {"f1": None, "f2": None, "diff_html": None}

    SUGGESTIONS = {
        "Only in Platform": (
            "This file exists ONLY in Platform (baseline).\n\n"
            "Possible reasons:\n"
            "  \u2022 The file was not yet migrated to Folder 2\n"
            "  \u2022 The file was intentionally removed / deprecated in the target\n"
            "  \u2022 The file was renamed or moved (search Folder 2 for a similar name)\n\n"
            "Suggested actions:\n"
            "  \u2714 Click 'Copy F1 \u2192 F2' to copy this file into Folder 2\n"
            "  \u2714 Open the file in the editor below to review content before deciding\n"
            "  \u2714 If deprecated, document the reason in a migration log"
        ),
        "Only in Project": (
            "This file exists ONLY in Project (target / comparison).\n\n"
            "Possible reasons:\n"
            "  \u2022 A newly added file introduced during migration\n"
            "  \u2022 Renamed from an existing Folder 1 file\n"
            "  \u2022 Auto-generated by the new version / build system\n\n"
            "Suggested actions:\n"
            "  \u2714 Click 'Copy F2 \u2190 F1' if this file should also exist in Folder 1\n"
            "  \u2714 Open the file in the editor below to review its content\n"
            "  \u2714 Document this as a new addition in your migration report"
        ),
        "Modified": (
            "This file exists in BOTH folders but has been MODIFIED.\n\n"
            "Suggested actions:\n"
            "  \u2714 Click 'View HTML Diff' to see the exact line-by-line changes\n"
            "  \u2714 Review both versions side-by-side in the editor tabs below\n"
            "  \u2714 Edit either version if corrections are needed, then click Save"
        ),
        "Different": (
            "This file exists in BOTH folders but has DIFFERENCES.\n\n"
            "Suggested actions:\n"
            "  \u2714 Click 'View HTML Diff' to see the exact line-by-line changes\n"
            "  \u2714 Review both versions side-by-side in the editor tabs below\n"
            "  \u2714 Edit either version if corrections are needed, then click Save"
        ),
        "Comments update only": (
            "Only COMMENTS were changed in this file (code logic is identical).\n\n"
            "Low-impact change — verify whether the comment update was intentional.\n"
            "  \u2714 Open the diff to review which comments changed\n"
            "  \u2714 No code action usually required unless the comments carry important info"
        ),
        "Identical": (
            "This file is IDENTICAL in both folders.\n\n"
            "No action needed. The file has not changed between F1 and F2."
        ),
        "Error": (
            "An ERROR occurred during comparison of this file.\n\n"
            "Check the console / terminal output for the detailed error message.\n"
            "  \u2714 The file may be binary, locked, or have an unsupported encoding"
        ),
    }

    def run_ai_suggest():
        """Run AI analysis on the currently selected file and show recommendation."""
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("No Selection", "Please select a file row first.", parent=dialog)
            return
        vals = tree.item(sel[0])["values"]
        if not vals:
            return
        rel_path_raw = str(vals[0])
        status       = str(vals[1]) if len(vals) > 1 else ""
        lines1_val   = int(vals[2]) if len(vals) > 2 and str(vals[2]).isdigit() else 0
        lines2_val   = int(vals[3]) if len(vals) > 3 and str(vals[3]).isdigit() else 0
        p1, p2 = _resolve_paths(rel_path_raw, status)

        if status in ("Identical", "Error"):
            messagebox.showinfo(
                "AI Suggest",
                "AI suggestion is most useful for:\n"
                "  \u2022 Only in Platform  \u2022 Only in Project  \u2022 Modified / Different\n\n"
                f"Current status is '{status}' — no recommendation needed.",
                parent=dialog
            )
            return

        # Show loading indicator
        suggestion_text.config(state="normal", bg="#EDE7F6")
        suggestion_text.delete("1.0", "end")
        suggestion_text.insert("1.0", "\u23f3 Analysing... please wait")
        suggestion_text.config(state="disabled")
        dialog.update()

        key = OPENAI_API_KEY
        try:
            result_text = ai_analyze_file(p1, p2, status, lines1_val, lines2_val, openai_key=key)
        except Exception as ex:
            result_text = f"Error during AI analysis: {ex}"

        suggestion_text.config(state="normal", bg="#EDE7F6")
        suggestion_text.delete("1.0", "end")
        suggestion_text.insert("1.0", result_text)
        suggestion_text.config(state="disabled")

        # Auto-size the suggestion panel to 9 rows for better readability
        suggestion_text.config(height=9)

    def run_ai_merge():
        """Invoke Gemini Flash to merge the selected file and show confirm-before-save preview."""
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("No Selection", "Please select a file row first.", parent=dialog)
            return
        vals = tree.item(sel[0])["values"]
        if not vals:
            return
        rel_path_raw = str(vals[0])
        status       = str(vals[1]) if len(vals) > 1 else ""
        p1, p2 = _resolve_paths(rel_path_raw, status)

        if status == "Identical":
            messagebox.showinfo("AI Merge", "Files are identical — no merge needed.", parent=dialog)
            return
        if status == "Error":
            messagebox.showinfo("AI Merge", "Cannot merge a file with comparison errors.", parent=dialog)
            return

        if not GEMINI_API_KEY or GEMINI_API_KEY.startswith("YOUR-"):
            messagebox.showwarning(
                "Gemini API Key Not Configured",
                "The Google Gemini API key is not set.\n\n"
                "1. Go to: https://aistudio.google.com/app/apikey\n"
                "2. Sign in with your Google account (free, no credit card)\n"
                "3. Click 'Create API key' and copy it\n"
                "4. Open Migration_V2.py and set near the top:\n\n"
                "   GEMINI_API_KEY = \"AIza...\"\n",
                parent=dialog
            )
            return

        # Show loading indicator in suggestion panel
        suggestion_text.config(state="normal", bg="#E0F7FA", height=6)
        suggestion_text.delete("1.0", "end")
        suggestion_text.insert(
            "1.0",
            f"Gemini Flash is merging '{os.path.basename(p1 or p2 or rel_path_raw)}'\n"
            "  Scanning dependencies ...\n"
            "  Calling Gemini API (may take 10-30 seconds) ..."
        )
        suggestion_text.config(state="disabled")
        dialog.update()

        try:
            merged, dep_report, warnings = ai_merge_with_gemini(
                p1, p2, status, rel_path_raw,
                files1, files2, folder1_root, folder2_root,
                api_key=GEMINI_API_KEY
            )
        except Exception as ex:
            suggestion_text.config(state="normal", bg="#FFCDD2", height=5)
            suggestion_text.delete("1.0", "end")
            suggestion_text.insert("1.0", f"Gemini merge failed:\n{ex}")
            suggestion_text.config(state="disabled")
            return

        # Update suggestion panel summary
        suggestion_text.config(state="normal", bg="#E0F7FA", height=8)
        suggestion_text.delete("1.0", "end")
        suggestion_text.insert("1.0",
            "Gemini merge complete! Review the preview window, edit if needed, then Save.\n\n"
            "Dependency Impact Report:\n" + dep_report[:1000]
        )
        suggestion_text.config(state="disabled")

        # ── Merge preview window ──────────────────────────────────────────────
        preview = tk.Toplevel(dialog)
        preview.title(
            f"AI Merge Preview  -  {os.path.basename(p1 or p2 or rel_path_raw)}"
        )
        preview.geometry("1440x920")
        preview.configure(bg="#003333")
        preview.grab_set()

        # Title bar
        pt = tk.Frame(preview, bg="#006064", height=52)
        pt.pack(fill="x")
        pt.pack_propagate(False)
        tk.Label(
            pt,
            text=f"Gemini Flash ({_GEMINI_MODEL}) Merge Preview  -  "
                 f"{os.path.basename(p1 or p2 or rel_path_raw)}   (Status: {status})",
            font=("Segoe UI", 12, "bold"), bg="#006064", fg="white"
        ).pack(pady=14)

        # Warning banner if Gemini flagged risks
        if warnings:
            wb = tk.Frame(preview, bg="#E65100", height=28)
            wb.pack(fill="x")
            wb.pack_propagate(False)
            wtext = "  WARNINGS: " + "  |  ".join(warnings[:5])
            tk.Label(wb, text=wtext, font=("Segoe UI", 8, "bold"),
                     bg="#E65100", fg="white").pack(side="left", padx=10, pady=5)

        # Three-tab editor: Original F1 | Original F2 | AI Merged (editable) | Dep Report
        nb = ttk.Notebook(preview)
        nb.pack(fill="both", expand=True, padx=8, pady=6)

        def _make_pe(parent, content, bg="#1e1e1e", editable=False):
            frm = tk.Frame(parent, bg=bg)
            frm.pack(fill="both", expand=True)
            ysb = tk.Scrollbar(frm, orient="vertical")
            xsb = tk.Scrollbar(frm, orient="horizontal")
            ysb.pack(side="right", fill="y")
            xsb.pack(side="bottom", fill="x")
            ed = tk.Text(
                frm, wrap="none",
                font=("Consolas", 9),
                bg=bg, fg="#d4d4d4",
                insertbackground="white",
                selectbackground="#37474f",
                relief="flat",
                yscrollcommand=ysb.set,
                xscrollcommand=xsb.set,
            )
            ed.pack(fill="both", expand=True)
            ysb.config(command=ed.yview)
            xsb.config(command=ed.xview)
            ed.insert("1.0", content if content else "(not available)")
            if not editable:
                ed.config(state="disabled")
            return ed

        c1 = _read_file_safe(p1) or "(file not in Folder 1)"
        c2 = _read_file_safe(p2) or "(file not in Folder 2)"

        tab_f1  = tk.Frame(nb, bg="#1e1e1e")
        tab_f2  = tk.Frame(nb, bg="#1e1e1e")
        tab_mrg = tk.Frame(nb, bg="#0d2137")
        tab_dep = tk.Frame(nb, bg="#1a1a2e")

        nb.add(tab_f1,  text=f"  Original F1 ({len(c1.splitlines())} lines)  ")
        nb.add(tab_f2,  text=f"  Original F2 ({len(c2.splitlines())} lines)  ")
        nb.add(tab_mrg, text=f"  AI Merged ({len(merged.splitlines())} lines) - EDIT HERE  ")
        nb.add(tab_dep, text="  Dependency Report  ")
        nb.select(tab_mrg)

        _make_pe(tab_f1, c1, editable=False)
        _make_pe(tab_f2, c2, editable=False)
        merged_ed = _make_pe(tab_mrg, merged, bg="#0d2137", editable=True)

        dep_txt = tk.Text(tab_dep, wrap="word", font=("Segoe UI", 9),
                          bg="#003333", fg="#B0BEC5", relief="flat")
        dep_txt.pack(fill="both", expand=True, padx=5, pady=5)
        dep_txt.insert("1.0",
            f"Gemini Flash ({_GEMINI_MODEL})  -  Dependency Impact Analysis\n"
            f"{'=' * 55}\n\n"
            f"File merged:  {rel_path_raw}\n"
            f"Status:       {status}\n"
            f"Dependencies scanned: {len(dep_report.splitlines())} lines of analysis\n\n"
            f"{dep_report}"
        )
        dep_txt.config(state="disabled")

        # Guidance
        guide = tk.Frame(preview, bg="#263238", height=28)
        guide.pack(fill="x")
        guide.pack_propagate(False)
        tk.Label(
            guide,
            text="  Review the 'AI Merged' tab. You can freely edit it before saving. "
                 "Choose which folder to write the merged result into.",
            font=("Segoe UI", 8), bg="#263238", fg="#80CBC4"
        ).pack(side="left", padx=10, pady=6)

        # Bottom action buttons
        bot = tk.Frame(preview, bg="#37474F", height=52)
        bot.pack(fill="x", side="bottom")
        bot.pack_propagate(False)

        def _get_merged():
            return merged_ed.get("1.0", "end-1c")

        def _save_merged(to_folder_num):
            content_out = _get_merged()
            if not content_out.strip():
                messagebox.showwarning("Empty", "Merged content is empty — please review.", parent=preview)
                return
            # Determine target path
            if to_folder_num == 1:
                tpath = p1
                if not tpath:
                    base = rel_path_raw.replace(" (Custom Mapping)", "").strip()
                    tpath = os.path.join(folder1_root, base)
            else:
                tpath = p2
                if not tpath:
                    base = rel_path_raw.replace(" (Custom Mapping)", "").strip()
                    tpath = os.path.join(folder2_root, base)
            if os.path.exists(tpath):
                if not messagebox.askyesno("Overwrite?",
                        f"File already exists:\n{tpath}\n\nOverwrite with merged version?",
                        parent=preview):
                    return
            try:
                os.makedirs(os.path.dirname(tpath), exist_ok=True)
                with open(tpath, "w", encoding="utf-8") as fh:
                    fh.write(content_out)
                messagebox.showinfo("Saved",
                    f"Merged file saved successfully:\n{tpath}", parent=preview)
                # Refresh parent dialog editors
                if to_folder_num == 1:
                    _load_editor(editor_f1, tpath, "[File not available in Folder 1]")
                    current["f1"] = tpath
                else:
                    _load_editor(editor_f2, tpath, "[File not available in Folder 2]")
                    current["f2"] = tpath
                preview.destroy()
            except Exception as ex:
                messagebox.showerror("Save Error", f"Could not save merged file:\n{ex}", parent=preview)

        def _pbtn(parent, label, bg, cmd):
            tk.Button(parent, text=label, bg=bg, fg="white",
                      font=("Segoe UI", 9, "bold"), relief="flat",
                      padx=12, pady=10, cursor="hand2", command=cmd
                      ).pack(side="left", padx=6, pady=8)

        _pbtn(bot, "Save Merged -> Folder 1 (Baseline)", "#1565C0", lambda: _save_merged(1))
        _pbtn(bot, "Save Merged -> Folder 2 (Target)",   "#2E7D32", lambda: _save_merged(2))
        _pbtn(bot, "Cancel - Discard Merge",             "#c62828", preview.destroy)
        tk.Label(bot,
            text="  Tip: The 'AI Merged' tab is fully editable — adjust before saving.",
            font=("Segoe UI", 8, "italic"), bg="#37474F", fg="#BDBDBD"
        ).pack(side="right", padx=14)

        preview.wait_window()

    def _resolve_paths(rel_path_raw, status):
        """Return (path1, path2) from rel_path raw string and status."""
        base = rel_path_raw.replace(" (Custom Mapping)", "").strip()
        p1 = files1.get(base)
        p2 = files2.get(base)
        # Handle custom mapping "file1 ← → file2"
        if p1 is None and p2 is None and " \u2190 \u2192 " in base:
            parts = base.split(" \u2190 \u2192 ")
            p1 = files1.get(parts[0].strip())
            p2 = files2.get(parts[1].strip()) if len(parts) > 1 else None
        # If still None, fall back using folder roots
        if p1 is None and status != "Only in Project" and folder1_root:
            candidate = os.path.join(folder1_root, base)
            if os.path.isfile(candidate):
                p1 = candidate
        if p2 is None and status != "Only in Platform" and folder2_root:
            candidate = os.path.join(folder2_root, base)
            if os.path.isfile(candidate):
                p2 = candidate
        return p1, p2

    def _load_editor(editor, path, placeholder):
        editor.config(state="normal")
        editor.delete("1.0", "end")
        if path and os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8", errors="replace") as fh:
                    editor.insert("1.0", fh.read())
            except Exception as ex:
                editor.insert("1.0", f"[Could not read file: {ex}]")
        else:
            editor.insert("1.0", placeholder)

    def on_select(event=None):
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0])["values"]
        if not vals:
            return
        rel_path_raw = str(vals[0])
        status       = str(vals[1]) if len(vals) > 1 else ""

        p1, p2 = _resolve_paths(rel_path_raw, status)
        current["f1"] = p1
        current["f2"] = p2

        # Suggestion (reset to static text and default colour when a new row is selected)
        sug = SUGGESTIONS.get(status, f"Status: {status}")
        suggestion_text.config(state="normal", bg="#FFFDE7", height=5)
        suggestion_text.delete("1.0", "end")
        suggestion_text.insert("1.0", sug)
        suggestion_text.config(state="disabled")

        # Update tab labels
        f1_name = os.path.basename(p1) if p1 else "N/A"
        f2_name = os.path.basename(p2) if p2 else "N/A"
        editor_tabs.tab(0, text=f"  F1: {f1_name}  ")
        editor_tabs.tab(1, text=f"  F2: {f2_name}  ")

        # Load editors
        _load_editor(editor_f1, p1, "[File not available in Folder 1]")
        _load_editor(editor_f2, p2, "[File not available in Folder 2]")

        # Store diff HTML path
        idx = tree.index(sel[0])
        current["diff_html"] = None
        if idx < len(results):
            raw_link = results[idx][5] if len(results[idx]) > 5 else ""
            if isinstance(raw_link, str) and raw_link.startswith('=HYPERLINK("'):
                html_path = raw_link.split('"')[1]
                if os.path.isfile(html_path):
                    current["diff_html"] = html_path

    tree.bind("<<TreeviewSelect>>", on_select)

    def open_file(folder_num):
        path = current[f"f{folder_num}"]
        if path and os.path.isfile(path):
            try:
                os.startfile(path)
            except Exception as ex:
                messagebox.showerror("Error", f"Could not open file:\n{ex}", parent=dialog)
        else:
            messagebox.showinfo("Not Available",
                                f"File is not available in Folder {folder_num}.",
                                parent=dialog)

    def copy_file(src_num, dst_num):
        src = current[f"f{src_num}"]
        if not src or not os.path.isfile(src):
            messagebox.showwarning("No Source",
                                   f"No source file loaded from Folder {src_num}.",
                                   parent=dialog)
            return

        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0])["values"]
        rel_path_raw = str(vals[0]).replace(" (Custom Mapping)", "").strip()
        base_rel = rel_path_raw.split(" \u2190 \u2192 ")[0].strip() if " \u2190 \u2192 " in rel_path_raw else rel_path_raw

        dst_root = folder2_root if dst_num == 2 else folder1_root
        if not dst_root:
            messagebox.showwarning("Error",
                                   f"Cannot determine root path for Folder {dst_num}.",
                                   parent=dialog)
            return

        dest = os.path.join(dst_root, base_rel)
        if os.path.exists(dest):
            if not messagebox.askyesno("Overwrite?",
                                       f"File already exists at:\n{dest}\n\nOverwrite?",
                                       parent=dialog):
                return
        try:
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            shutil.copy2(src, dest)
            messagebox.showinfo("Copied", f"File copied to:\n{dest}", parent=dialog)
            # Refresh the other folder's editor
            if dst_num == 2:
                _load_editor(editor_f2, dest, "[File not available in Folder 2]")
                current["f2"] = dest
            else:
                _load_editor(editor_f1, dest, "[File not available in Folder 1]")
                current["f1"] = dest
        except Exception as ex:
            messagebox.showerror("Copy Failed", f"Could not copy file:\n{ex}", parent=dialog)

    def open_diff():
        html = current.get("diff_html")
        if html and os.path.isfile(html):
            try:
                os.startfile(html)
            except Exception as ex:
                messagebox.showerror("Error", f"Could not open diff:\n{ex}", parent=dialog)
        else:
            messagebox.showinfo("No Diff",
                                "No HTML diff report available for this file.\n"
                                "(Identical files and files present in only one folder have no diff.)",
                                parent=dialog)

    def save_content(folder_num):
        path = current.get(f"f{folder_num}")
        if not path:
            messagebox.showwarning("No File",
                                   f"No file is loaded for Folder {folder_num}.",
                                   parent=dialog)
            return
        editor = editor_f1 if folder_num == 1 else editor_f2
        content = editor.get("1.0", "end-1c")
        try:
            with open(path, "w", encoding="utf-8") as fh:
                fh.write(content)
            messagebox.showinfo("Saved", f"File saved successfully:\n{path}", parent=dialog)
        except Exception as ex:
            messagebox.showerror("Save Error", f"Could not save file:\n{ex}", parent=dialog)

    # Export function
    def export_results(format_type):
        """Export comparison results to Excel or CSV"""
        output_dir = os.path.join(os.getcwd(), "Migration_Analysis_Reports")
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        try:
            if format_type == "excel":
                file_path = os.path.join(output_dir, f"Comparison_Results_{timestamp}.xlsx")
                write_excel_report(results, file_path, folder1_display, folder2_display)
                messagebox.showinfo("Export Successful", 
                                   f"Excel report saved successfully:\n{file_path}", 
                                   parent=dialog)
            elif format_type == "csv":
                file_path = os.path.join(output_dir, f"Comparison_Results_{timestamp}.csv")
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow([
                        "File Path", 
                        f"Lines in Platform ({folder1_display})", 
                        f"Lines in Project ({folder2_display})", 
                        "Line Comparison Status",
                        "Status", 
                        "HTML Diff Report", 
                        "Purpose of Change (Migration Analysis)",
                        "ChangeSet & WorkItem from RTC History"
                    ])
                    writer.writerows(results)
                messagebox.showinfo("Export Successful", 
                                   f"CSV report saved successfully:\n{file_path}", 
                                   parent=dialog)
        except Exception as ex:
            messagebox.showerror("Export Failed", 
                               f"Could not export results:\n{ex}", 
                               parent=dialog)

    # ── Bottom bar ─────────────────────────────────────────────────────────────
    bottom = tk.Frame(dialog, bg="#CFD8DC", height=42)
    bottom.pack(fill="x", side="bottom")
    bottom.pack_propagate(False)
    
    # Close button (right side)
    tk.Button(
        bottom, text="Close", font=("Segoe UI", 9, "bold"),
        bg="#c62828", fg="white", relief="flat",
        padx=20, pady=5, cursor="hand2",
        command=dialog.destroy
    ).pack(side="right", padx=12, pady=6)
    
    # Export buttons (right side, before Close)
    tk.Button(
        bottom, text="💾 Export to Excel", font=("Segoe UI", 9, "bold"),
        bg="#1976D2", fg="white", relief="flat",
        padx=15, pady=5, cursor="hand2",
        command=lambda: export_results("excel")
    ).pack(side="right", padx=5, pady=6)
    
    tk.Button(
        bottom, text="💾 Export to CSV", font=("Segoe UI", 9, "bold"),
        bg="#388E3C", fg="white", relief="flat",
        padx=15, pady=5, cursor="hand2",
        command=lambda: export_results("csv")
    ).pack(side="right", padx=5, pady=6)
    
    # Info label (left side)
    tk.Label(
        bottom,
        text="Select a file row to load its content and see a smart suggestion.",
        font=("Segoe UI", 9),
        bg="#CFD8DC", fg="#37474F"
    ).pack(side="left", padx=12, pady=10)

    dialog.wait_window()


# ---------------------------------------------------------------------------
# Folder comparison with parallel processing
# ---------------------------------------------------------------------------
def compare_folders(folder1, folder2, progress_label, progress_bar, custom_mappings=None):
    output_dir = os.path.join(os.getcwd(), "Migration_Analysis_Reports")
    os.makedirs(output_dir, exist_ok=True)

    csv_report_path = os.path.join(output_dir, "Migration_Analysis_Report.csv")
    excel_report_path = os.path.join(output_dir, "Migration_Analysis_Report.xlsx")

    # Handle ZIP folder inputs - extract if needed
    temp_dirs_to_cleanup = []
    original_folder1 = folder1
    original_folder2 = folder2
    
    progress_label.config(text="Preparing folders (extracting if ZIP)...")
    progress_label.update()
    
    folder1_actual, is_temp1, orig1 = prepare_folder_path(folder1)
    folder2_actual, is_temp2, orig2 = prepare_folder_path(folder2)
    
    if is_temp1:
        temp_dirs_to_cleanup.append(folder1_actual)
        progress_label.config(text=f"Extracted Folder 1 ZIP: {os.path.basename(orig1)}")
        progress_label.update()
    
    if is_temp2:
        temp_dirs_to_cleanup.append(folder2_actual)
        progress_label.config(text=f"Extracted Folder 2 ZIP: {os.path.basename(orig2)}")
        progress_label.update()
    
    if not folder1_actual or not folder2_actual:
        messagebox.showerror("Error", "Could not prepare folders for comparison. Check if paths are valid.")
        # Cleanup any temp dirs created
        for temp_dir in temp_dirs_to_cleanup:
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
        return
    
    # Use actual folders for comparison
    folder1 = folder1_actual
    folder2 = folder2_actual

    results = []
    files1 = {os.path.relpath(os.path.join(dp, f), folder1): os.path.join(dp, f)
              for dp, _, fnames in os.walk(folder1) for f in fnames}
    files2 = {os.path.relpath(os.path.join(dp, f), folder2): os.path.join(dp, f)
              for dp, _, fnames in os.walk(folder2) for f in fnames}

    all_files = sorted(set(files1.keys()) | set(files2.keys()))
    
    # Track which files have been processed via custom mappings
    processed_files = set()
    
    # Prepare comparison tasks
    comparison_tasks = []
    
    # Add custom mappings first
    if custom_mappings:
        for file1_rel, file2_rel in custom_mappings.items():
            path1 = files1.get(file1_rel)
            path2 = files2.get(file2_rel)
            
            if not path1 or not path2:
                continue
            
            processed_files.add(file1_rel)
            processed_files.add(file2_rel)
            
            rel_path_combined = f"{file1_rel} ← → {file2_rel}"
            comparison_tasks.append((rel_path_combined, path1, path2, output_dir, True))
    
    # Add regular file comparisons
    for rel_path in all_files:
        if rel_path in processed_files:
            continue
        
        path1 = files1.get(rel_path)
        path2 = files2.get(rel_path)
        comparison_tasks.append((rel_path, path1, path2, output_dir, False))
    
    total_files = len(comparison_tasks)
    completed = 0
    
    # Use ThreadPoolExecutor for parallel processing
    # Adjust max_workers based on CPU cores (default is min(32, cpu_count + 4))
    max_workers = min(8, os.cpu_count() or 4)  # Limit to 8 to avoid overwhelming the system
    
    progress_label.config(text=f"Starting parallel comparison with {max_workers} workers...")
    progress_label.update()
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_task = {executor.submit(process_file_comparison, task): task for task in comparison_tasks}
        
        # Process completed tasks as they finish
        for future in as_completed(future_to_task):
            task = future_to_task[future]
            try:
                result = future.result()
                results.append(result)
                completed += 1
                
                progress = int((completed / total_files) * 100)
                progress_bar['value'] = progress
                progress_label.config(text=f"Compared: {task[0][:50]}... ({completed}/{total_files} - {progress}%)")
                progress_label.update()
            except Exception as e:
                print(f"Error processing {task[0]}: {e}")
                results.append([
                    task[0],
                    0,
                    0,
                    "Error during comparison",
                    "Error",
                    "",
                    f"Error: {str(e)}",
                    ""  # Workitem column placeholder
                ])
                completed += 1

    # Use original folder names in reports (show ZIP names if applicable)
    folder1_display = original_folder1
    folder2_display = original_folder2
    
    with open(csv_report_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([
            "File Path", 
            f"Lines in Platform ({folder1_display})", 
            f"Lines in Project ({folder2_display})", 
            "Line Comparison Status",
            "Status", 
            "HTML Diff Report", 
            "Purpose of Change (Migration Analysis)",
            "ChangeSet & WorkItem from RTC History"
        ])
        writer.writerows(results)

    write_excel_report(results, excel_report_path, folder1_display, folder2_display)

    progress_bar['value'] = 100
    progress_label.config(
        text=f"✅ Comparison complete!\nReports saved at:\n{output_dir}\n(Tag: Migration Analysis)"
    )

    # Show the interactive results dialog (modal – blocks until the user closes it).
    # Temp dirs are cleaned up AFTER the dialog so extracted ZIP files remain
    # accessible for the open/copy actions inside the dialog.
    show_comparison_results_dialog(
        results, folder1_display, folder2_display,
        folder1, folder2, files1, files2
    )

    # Cleanup temporary directories from ZIP extraction
    for temp_dir in temp_dirs_to_cleanup:
        try:
            shutil.rmtree(temp_dir)
            print(f"Cleaned up temp directory: {temp_dir}")
        except Exception as e:
            print(f"Warning: Could not cleanup temp directory {temp_dir}: {e}")

# ---------------------------------------------------------------------------
# GUI Setup
# ---------------------------------------------------------------------------
root = tk.Tk()
root.title("Migration Analysis Report Generator")
root.geometry("780x540")
root.config(bg="#EAF3FB")

# Bosch-style header
header_frame = tk.Frame(root, bg="#EAF3FB")
header_frame.pack(fill="x")

top_strip = tk.Label(header_frame, bg="#E60000", height=1)
top_strip.pack(fill="x")

title_frame = tk.Frame(header_frame, bg="#003366")
title_frame.pack(fill="x")

title_label = tk.Label(title_frame, text="Migration Analysis Report Generator",
                       font=("Segoe UI", 18, "bold"), bg="#003366", fg="white")
title_label.pack(padx=10, pady=10)

# Comparison Mode Selection
mode_frame = tk.Frame(root, bg="#EAF3FB")
mode_frame.pack(pady=(20, 10))

tk.Label(mode_frame, text="Comparison Mode:", bg="#EAF3FB", font=("Segoe UI", 11, "bold")).pack(side="left", padx=5)

comparison_mode = tk.StringVar(value="folder")

tk.Radiobutton(mode_frame, text="Folder/ZIP Comparison", variable=comparison_mode, value="folder", 
               bg="#EAF3FB", font=("Segoe UI", 10), command=lambda: toggle_input_mode()).pack(side="left", padx=10)
tk.Radiobutton(mode_frame, text="RTC Snapshot Comparison", variable=comparison_mode, value="snapshot", 
               bg="#EAF3FB", font=("Segoe UI", 10), command=lambda: toggle_input_mode()).pack(side="left", padx=10)

# Folder input frame (default visible)
folder_input_frame = tk.Frame(root, bg="#EAF3FB")
folder_input_frame.pack(fill="x")

tk.Label(folder_input_frame, text="Select Platform Folder (or ZIP file):", bg="#EAF3FB", font=("Segoe UI", 11)).pack(pady=(5, 5))
folder1_entry = tk.Entry(folder_input_frame, width=85)
folder1_entry.pack()
tk.Button(folder_input_frame, text="Browse", bg="#007B3E", fg="white",
          command=lambda: browse_folder(folder1_entry)).pack(pady=5)

tk.Label(folder_input_frame, text="Select Project Folder (or ZIP file):", bg="#EAF3FB", font=("Segoe UI", 11)).pack(pady=(10, 5))
folder2_entry = tk.Entry(folder_input_frame, width=85)
folder2_entry.pack()
tk.Button(folder_input_frame, text="Browse", bg="#007B3E", fg="white",
          command=lambda: browse_folder(folder2_entry)).pack(pady=5)

# Snapshot input frame (initially hidden)
snapshot_input_frame = tk.Frame(root, bg="#EAF3FB")

tk.Label(snapshot_input_frame, text="Platform Snapshot URL/ID:", bg="#EAF3FB", font=("Segoe UI", 11)).pack(pady=(5, 5))
snapshot1_entry = tk.Entry(snapshot_input_frame, width=85)
snapshot1_entry.pack()
tk.Label(snapshot_input_frame, text="From RTC web: Copy snapshot URL (with id=_xxxxx) or just the UUID like _ojreQAAbEfG1br8X33nQcA", 
         bg="#EAF3FB", font=("Segoe UI", 8), fg="gray").pack()

tk.Label(snapshot_input_frame, text="Project Snapshot URL/ID:", bg="#EAF3FB", font=("Segoe UI", 11)).pack(pady=(10, 5))
snapshot2_entry = tk.Entry(snapshot_input_frame, width=85)
snapshot2_entry.pack()
tk.Label(snapshot_input_frame, text="From RTC web: Copy snapshot URL (with id=_xxxxx) or just the UUID like _i3S_vwAaEfG3rPS3zZLwKA", 
         bg="#EAF3FB", font=("Segoe UI", 8), fg="gray").pack()

def toggle_input_mode():
    """Toggle between folder and snapshot input modes"""
    if comparison_mode.get() == "folder":
        snapshot_input_frame.pack_forget()
        folder_input_frame.pack(fill="x", before=rtc_frame)
    else:
        folder_input_frame.pack_forget()
        snapshot_input_frame.pack(fill="x", before=rtc_frame)

# Folder inputs (kept for reference but now in folder_input_frame above)

def browse_folder(entry_field):
    # Ask user to choose between folder or ZIP file
    choice = messagebox.askyesnocancel(
        "Select Input Type",
        "Do you want to select a FOLDER?\n\nYes = Select Folder\nNo = Select ZIP File\nCancel = Abort"
    )
    
    if choice is None:  # Cancel
        return
    elif choice:  # Yes - Select Folder
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            entry_field.delete(0, tk.END)
            entry_field.insert(0, folder_selected)
    else:  # No - Select ZIP File
        zip_selected = filedialog.askopenfilename(
            title="Select ZIP File",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        if zip_selected:
            entry_field.delete(0, tk.END)
            entry_field.insert(0, zip_selected)

# RTC/ALM Integration Section
rtc_frame = tk.Frame(root, bg="#EAF3FB")
rtc_frame.pack(pady=10)

rtc_enabled_var = tk.BooleanVar(value=False)
rtc_checkbox = tk.Checkbutton(rtc_frame, text="Enable RTC/ALM WorkItem Integration", 
                               variable=rtc_enabled_var, bg="#EAF3FB", 
                               font=("Segoe UI", 10))
rtc_checkbox.grid(row=0, column=0, sticky="w", padx=5)

def show_component_selection_dialog(components1, components2):
    """
    Show dialog to select components to compare from two snapshots.
    Shows: 1) Common components (selectable for comparison)
           2) Components only in Snapshot 1
           3) Components only in Snapshot 2
    Returns dict with selected components and metadata.
    """
    dialog = tk.Toplevel(root)
    dialog.title("Component Comparison - Select Components")
    dialog.geometry("900x650")
    dialog.configure(bg="#f0f4f7")
    dialog.grab_set()
    
    result = {'selected_components': [], 'only_in_1': [], 'only_in_2': []}
    
    # Header
    header_frame = tk.Frame(dialog, bg="#003366")
    header_frame.pack(fill="x")
    tk.Label(header_frame, text="Component Analysis & Selection", 
             font=("Segoe UI", 16, "bold"), bg="#003366", fg="white").pack(pady=12)
    
    # Create notebook for tabs
    notebook = ttk.Notebook(dialog)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Tab 1: Common Components (Selectable)
    common_tab = tk.Frame(notebook, bg="white")
    notebook.add(common_tab, text="📊 Common Components")
    
    # Tab 2: Only in Snapshot 1
    snap1_only_tab = tk.Frame(notebook, bg="white")
    notebook.add(snap1_only_tab, text="➕ Only in Snapshot 1")
    
    # Tab 3: Only in Snapshot 2
    snap2_only_tab = tk.Frame(notebook, bg="white")
    notebook.add(snap2_only_tab, text="➕ Only in Snapshot 2")
    
    # Find component differences
    comp1_dict = {c['name']: c for c in components1}
    comp2_dict = {c['name']: c for c in components2}
    
    comp1_names = set(comp1_dict.keys())
    comp2_names = set(comp2_dict.keys())
    
    common_components = sorted(comp1_names.intersection(comp2_names))
    only_in_1 = sorted(comp1_names - comp2_names)
    only_in_2 = sorted(comp2_names - comp1_names)
    
    result['only_in_1'] = only_in_1
    result['only_in_2'] = only_in_2
    
    # ===== TAB 1: Common Components =====
    tk.Label(common_tab, text=f"Components present in both snapshots: {len(common_components)}", 
             font=("Segoe UI", 12, "bold"), bg="white", fg="#003366").pack(anchor="w", padx=15, pady=10)
    
    tk.Label(common_tab, text="✓ Select components to compare their changesets and files:", 
             font=("Segoe UI", 10), bg="white", fg="gray").pack(anchor="w", padx=15, pady=(0,10))
    
    # Scrollable frame for common components
    common_canvas_frame = tk.Frame(common_tab, bg="white")
    common_canvas_frame.pack(fill="both", expand=True, padx=15, pady=(0,10))
    
    common_canvas = tk.Canvas(common_canvas_frame, bg="white", highlightthickness=0)
    common_scrollbar = tk.Scrollbar(common_canvas_frame, orient="vertical", command=common_canvas.yview)
    common_scrollable = tk.Frame(common_canvas, bg="white")
    
    common_scrollable.bind("<Configure>", lambda e: common_canvas.configure(scrollregion=common_canvas.bbox("all")))
    common_canvas.create_window((0, 0), window=common_scrollable, anchor="nw")
    common_canvas.configure(yscrollcommand=common_scrollbar.set)
    
    common_canvas.pack(side="left", fill="both", expand=True)
    common_scrollbar.pack(side="right", fill="y")
    
    # Component checkboxes with additional info
    component_vars = {}
    for idx, comp_name in enumerate(common_components):
        comp1 = comp1_dict[comp_name]
        comp2 = comp2_dict[comp_name]
        
        frame = tk.Frame(common_scrollable, bg="white" if idx % 2 == 0 else "#f8f8f8", relief="flat")
        frame.pack(fill="x", padx=5, pady=2)
        
        var = tk.BooleanVar(value=True)
        component_vars[comp_name] = var
        
        cb = tk.Checkbutton(frame, text=comp_name, variable=var, bg=frame['bg'], 
                           font=("Segoe UI", 10), anchor="w")
        cb.pack(side="left", padx=10, fill="x", expand=True)
        
        # Show baseline UUIDs if different
        if comp1.get('baseline_uuid') != comp2.get('baseline_uuid'):
            tk.Label(frame, text="⚠ Different baselines", fg="orange", bg=frame['bg'], 
                    font=("Segoe UI", 8)).pack(side="right", padx=10)
    
    # Select/Deselect all buttons
    btn_frame_common = tk.Frame(common_tab, bg="white")
    btn_frame_common.pack(fill="x", padx=15, pady=5)
    
    def select_all_common():
        for var in component_vars.values():
            var.set(True)
    
    def deselect_all_common():
        for var in component_vars.values():
            var.set(False)
    
    tk.Button(btn_frame_common, text="✓ Select All", command=select_all_common, 
             bg="#007B3E", fg="white", width=12).pack(side="left", padx=5)
    tk.Button(btn_frame_common, text="✗ Deselect All", command=deselect_all_common, 
             bg="#666666", fg="white", width=12).pack(side="left", padx=5)
    
    # ===== TAB 2: Only in Snapshot 1 =====
    tk.Label(snap1_only_tab, text=f"Components only in Snapshot 1: {len(only_in_1)}", 
             font=("Segoe UI", 12, "bold"), bg="white", fg="#E60000").pack(anchor="w", padx=15, pady=10)
    
    if only_in_1:
        tk.Label(snap1_only_tab, text="These components will be marked as 'Removed' or 'Not in Snapshot 2':", 
                font=("Segoe UI", 10), bg="white", fg="gray").pack(anchor="w", padx=15, pady=(0,10))
        
        snap1_canvas_frame = tk.Frame(snap1_only_tab, bg="white")
        snap1_canvas_frame.pack(fill="both", expand=True, padx=15, pady=(0,10))
        
        snap1_canvas = tk.Canvas(snap1_canvas_frame, bg="white", highlightthickness=0)
        snap1_scrollbar = tk.Scrollbar(snap1_canvas_frame, orient="vertical", command=snap1_canvas.yview)
        snap1_scrollable = tk.Frame(snap1_canvas, bg="white")
        
        snap1_scrollable.bind("<Configure>", lambda e: snap1_canvas.configure(scrollregion=snap1_canvas.bbox("all")))
        snap1_canvas.create_window((0, 0), window=snap1_scrollable, anchor="nw")
        snap1_canvas.configure(yscrollcommand=snap1_scrollbar.set)
        
        snap1_canvas.pack(side="left", fill="both", expand=True)
        snap1_scrollbar.pack(side="right", fill="y")
        
        for idx, comp_name in enumerate(only_in_1):
            comp = comp1_dict[comp_name]
            frame = tk.Frame(snap1_scrollable, bg="white" if idx % 2 == 0 else "#fff5f5", relief="flat", bd=1)
            frame.pack(fill="x", padx=5, pady=2)
            
            tk.Label(frame, text=f"➖ {comp_name}", bg=frame['bg'], 
                    font=("Segoe UI", 10), anchor="w").pack(side="left", padx=10, fill="x", expand=True)
            
            tk.Label(frame, text=f"UUID: {comp.get('uuid', 'N/A')[:20]}...", bg=frame['bg'], 
                    font=("Segoe UI", 8), fg="gray").pack(side="right", padx=10)
    else:
        tk.Label(snap1_only_tab, text="✓ No unique components in Snapshot 1", 
                font=("Segoe UI", 11), bg="white", fg="green").pack(pady=50)
    
    # ===== TAB 3: Only in Snapshot 2 =====
    tk.Label(snap2_only_tab, text=f"Components only in Snapshot 2: {len(only_in_2)}", 
             font=("Segoe UI", 12, "bold"), bg="white", fg="#007B3E").pack(anchor="w", padx=15, pady=10)
    
    if only_in_2:
        tk.Label(snap2_only_tab, text="These components will be marked as 'Added' or 'New in Snapshot 2':", 
                font=("Segoe UI", 10), bg="white", fg="gray").pack(anchor="w", padx=15, pady=(0,10))
        
        snap2_canvas_frame = tk.Frame(snap2_only_tab, bg="white")
        snap2_canvas_frame.pack(fill="both", expand=True, padx=15, pady=(0,10))
        
        snap2_canvas = tk.Canvas(snap2_canvas_frame, bg="white", highlightthickness=0)
        snap2_scrollbar = tk.Scrollbar(snap2_canvas_frame, orient="vertical", command=snap2_canvas.yview)
        snap2_scrollable = tk.Frame(snap2_canvas, bg="white")
        
        snap2_scrollable.bind("<Configure>", lambda e: snap2_canvas.configure(scrollregion=snap2_canvas.bbox("all")))
        snap2_canvas.create_window((0, 0), window=snap2_scrollable, anchor="nw")
        snap2_canvas.configure(yscrollcommand=snap2_scrollbar.set)
        
        snap2_canvas.pack(side="left", fill="both", expand=True)
        snap2_scrollbar.pack(side="right", fill="y")
        
        for idx, comp_name in enumerate(only_in_2):
            comp = comp2_dict[comp_name]
            frame = tk.Frame(snap2_scrollable, bg="white" if idx % 2 == 0 else "#f5fff5", relief="flat", bd=1)
            frame.pack(fill="x", padx=5, pady=2)
            
            tk.Label(frame, text=f"➕ {comp_name}", bg=frame['bg'], 
                    font=("Segoe UI", 10), anchor="w").pack(side="left", padx=10, fill="x", expand=True)
            
            tk.Label(frame, text=f"UUID: {comp.get('uuid', 'N/A')[:20]}...", bg=frame['bg'], 
                    font=("Segoe UI", 8), fg="gray").pack(side="right", padx=10)
    else:
        tk.Label(snap2_only_tab, text="✓ No unique components in Snapshot 2", 
                font=("Segoe UI", 11), bg="white", fg="green").pack(pady=50)
    
    # Bottom buttons
    bottom_frame = tk.Frame(dialog, bg="#f0f4f7")
    bottom_frame.pack(fill="x", pady=15)
    
    summary_text = f"Common: {len(common_components)} | Only in S1: {len(only_in_1)} | Only in S2: {len(only_in_2)}"
    tk.Label(bottom_frame, text=summary_text, font=("Segoe UI", 10), 
            bg="#f0f4f7", fg="#003366").pack(pady=5)
    
    def on_confirm():
        selected = [name for name, var in component_vars.items() if var.get()]
        result['selected_components'] = selected
        
        if not selected:
            messagebox.showwarning("No Selection", "Please select at least one common component to compare.")
            return
        
        dialog.destroy()
    
    def on_cancel():
        result['selected_components'] = []
        dialog.destroy()
    
    btn_frame = tk.Frame(bottom_frame, bg="#f0f4f7")
    btn_frame.pack()
    
    tk.Button(btn_frame, text=f"Compare Selected Components ({len(common_components)} available)", 
              bg="#003366", fg="white", width=40, height=2, font=("Segoe UI", 10, "bold"),
              command=on_confirm).pack(side="left", padx=10)
    tk.Button(btn_frame, text="Cancel", bg="#666666", fg="white", 
              width=15, height=2, command=on_cancel).pack(side="left", padx=10)
    
    dialog.wait_window()
    return result

def show_rtc_credentials_dialog():
    """Show dialog to get RTC credentials"""
    global RTC_USERNAME, RTC_PASSWORD, RTC_ENABLED
    
    if not rtc_enabled_var.get():
        RTC_ENABLED = False
        return True
    
    dialog = tk.Toplevel(root)
    dialog.title("RTC/ALM Credentials")
    dialog.geometry("400x200")
    dialog.configure(bg="#f0f4f7")
    dialog.grab_set()  # Modal dialog
    
    result = {'confirmed': False}
    
    tk.Label(dialog, text="Enter RTC/ALM Credentials", 
             font=("Segoe UI", 12, "bold"), bg="#f0f4f7").pack(pady=10)
    
    # Username
    tk.Label(dialog, text="Username:", bg="#f0f4f7").pack(anchor="w", padx=20)
    username_entry = tk.Entry(dialog, width=40)
    username_entry.pack(padx=20, pady=5)
    
    # Password
    tk.Label(dialog, text="Password:", bg="#f0f4f7").pack(anchor="w", padx=20)
    password_entry = tk.Entry(dialog, width=40, show="*")
    password_entry.pack(padx=20, pady=5)
    
    def on_confirm():
        global RTC_USERNAME, RTC_PASSWORD, RTC_ENABLED
        RTC_USERNAME = username_entry.get().strip()
        RTC_PASSWORD = password_entry.get().strip()
        
        if not RTC_USERNAME or not RTC_PASSWORD:
            messagebox.showwarning("Warning", "Please enter both username and password")
            return
        
        RTC_ENABLED = True
        result['confirmed'] = True
        dialog.destroy()
    
    def on_cancel():
        global RTC_ENABLED
        RTC_ENABLED = False
        rtc_enabled_var.set(False)
        result['confirmed'] = True
        dialog.destroy()
    
    btn_frame = tk.Frame(dialog, bg="#f0f4f7")
    btn_frame.pack(pady=20)
    
    tk.Button(btn_frame, text="OK", bg="#003366", fg="white", width=10,
              command=on_confirm).pack(side="left", padx=10)
    tk.Button(btn_frame, text="Cancel", bg="#666666", fg="white", width=10,
              command=on_cancel).pack(side="left", padx=10)
    
    dialog.wait_window()
    return result['confirmed']

tk.Button(root, text="Start Comparison", bg="#003366", fg="white",
          font=("Segoe UI", 12, "bold"), width=25,
          command=lambda: start_comparison()).pack(pady=20)

def start_comparison():
    global RTC_WORKSPACE_NAME, RTC_STREAM_NAME, RTC_REPOSITORY_UUID, SNAPSHOT_MODE, SNAPSHOT1_URL, SNAPSHOT2_URL
    
    # Check which mode is selected
    mode = comparison_mode.get()
    
    if mode == "snapshot":
        # Snapshot comparison mode
        SNAPSHOT_MODE = True
        snapshot1 = snapshot1_entry.get().strip()
        snapshot2 = snapshot2_entry.get().strip()
        
        if not snapshot1 or not snapshot2:
            messagebox.showerror("Error", "Please provide both Snapshot 1 and Snapshot 2 URLs/IDs.")
            return
        
        SNAPSHOT1_URL = snapshot1
        SNAPSHOT2_URL = snapshot2
        
        # RTC credentials are required for snapshot mode
        if not show_rtc_credentials_dialog():
            return
        
        # Fetch snapshot details and show component selection
        print("\\n" + "="*80)
        print("Fetching Snapshot Information")
        print("="*80)
        
        snap1_details = fetch_snapshot_details(snapshot1, RTC_USERNAME, RTC_PASSWORD)
        snap2_details = fetch_snapshot_details(snapshot2, RTC_USERNAME, RTC_PASSWORD)
        
        if not snap1_details or not snap2_details:
            messagebox.showerror("Error", "Failed to fetch snapshot details. Check URLs and credentials.")
            return
        
        # Display snapshot names
        snap1_name = snap1_details.get('name', 'Unknown')
        snap2_name = snap2_details.get('name', 'Unknown')
        print(f"\\n{'='*80}")
        print(f"SNAPSHOT 1: {snap1_name}")
        print(f"{'='*80}")
        
        # Fetch components from both snapshots (pass the already-fetched details)
        components1 = fetch_snapshot_components(snapshot1, RTC_USERNAME, RTC_PASSWORD, snap1_details)
        
        print(f"\\n{'='*80}")
        print(f"SNAPSHOT 2: {snap2_name}")
        print(f"{'='*80}")
        
        components2 = fetch_snapshot_components(snapshot2, RTC_USERNAME, RTC_PASSWORD, snap2_details)
        
        if not components1 or not components2:
            messagebox.showerror("Error", "Failed to fetch components from snapshots.")
            return
        
        # Show component selection dialog
        selection_result = show_component_selection_dialog(components1, components2)
        
        if not selection_result or not selection_result.get('selected_components'):
            messagebox.showinfo("Cancelled", "No components selected for comparison.")
            return
        
        selected_components = selection_result['selected_components']
        only_in_1 = selection_result['only_in_1']
        only_in_2 = selection_result['only_in_2']
        
        # Generate Excel report
        generate_snapshot_comparison_excel(
            selected_components, only_in_1, only_in_2,
            components1, components2, 
            snap1_name, snap2_name
        )
        return
    
    else:
        # Folder comparison mode (existing logic)
        SNAPSHOT_MODE = False
        folder1 = folder1_entry.get().strip()
        folder2 = folder2_entry.get().strip()
        
        # Validate that paths exist (can be folders or ZIP files)
        valid1 = os.path.isdir(folder1) or (os.path.isfile(folder1) and folder1.lower().endswith('.zip'))
        valid2 = os.path.isdir(folder2) or (os.path.isfile(folder2) and folder2.lower().endswith('.zip'))
    
    if not valid1 or not valid2:
        messagebox.showerror("Error", "Please select valid folders or ZIP files for comparison.")
        return
    
    # Show RTC credentials dialog if enabled
    if rtc_enabled_var.get():
        if not show_rtc_credentials_dialog():
            return
        
        # Detect RTC workspace and stream from the folder paths
        print("\n" + "="*80)
        print("RTC Workspace Detection")
        print("="*80)
        
        # Try to detect workspace from folder2 (newer version) first
        print(f"\nChecking Folder 2: {folder2}")
        workspace_info = detect_rtc_workspace_and_stream(folder2, RTC_USERNAME, RTC_PASSWORD)
        
        if not workspace_info:
            # Fallback to folder1
            print(f"\nChecking Folder 1: {folder1}")
            workspace_info = detect_rtc_workspace_and_stream(folder1, RTC_USERNAME, RTC_PASSWORD)
        
        if workspace_info:
            # Store detected information globally
            RTC_WORKSPACE_NAME = workspace_info.get('workspace_name')
            RTC_STREAM_NAME = workspace_info.get('stream_name')
            RTC_REPOSITORY_UUID = workspace_info.get('repository_workspace_uuid')
            
            print("\n✓ RTC Workspace/Stream Information Successfully Detected!")
            print(f"  Workspace: {RTC_WORKSPACE_NAME}")
            print(f"  Stream: {RTC_STREAM_NAME}")
            if RTC_REPOSITORY_UUID:
                print(f"  UUID: {RTC_REPOSITORY_UUID}")
            print("="*80 + "\n")
            
            # Show confirmation dialog
            msg = f"RTC Workspace Detected:\n\n"
            msg += f"Workspace: {RTC_WORKSPACE_NAME}\n"
            if RTC_STREAM_NAME:
                msg += f"Stream: {RTC_STREAM_NAME}\n"
            msg += "\nThis workspace context will be used for changeset queries.\n\n"
            msg += "Continue with comparison?"
            
            if not messagebox.askyesno("RTC Workspace Detected", msg):
                print("Comparison cancelled by user after workspace detection.")
                return
        else:
            print("\n⚠ Warning: No RTC workspace detected from the provided folders.")
            print("Changeset queries will proceed without workspace context.")
            print("This may result in less accurate or missing changeset information.")
            print("="*80 + "\n")
            
            # Ask user if they want to continue without workspace context
            msg = "No RTC workspace was detected from the provided folders.\n\n"
            msg += "Changeset queries will proceed without workspace context,\n"
            msg += "which may result in incomplete changeset information.\n\n"
            msg += "Do you want to continue anyway?"
            
            if not messagebox.askyesno("No Workspace Detected", msg):
                print("Comparison cancelled - no RTC workspace detected.")
                return
    
    # Show file mapping preview and get confirmation
    confirmed, custom_mappings = show_file_mapping_dialog(folder1, folder2)
    if not confirmed:
        messagebox.showinfo("Cancelled", "Comparison cancelled by user.")
        return
    
    # Pass custom mappings to comparison
    compare_folders(folder1, folder2, progress_label, progress_bar, custom_mappings)

# Progress bar
progress_bar = ttk.Progressbar(root, orient="horizontal", length=600, mode="determinate")
progress_bar.pack(pady=5)
progress_label = tk.Label(root, text="", bg="#EAF3FB", fg="#333333", wraplength=650, justify="center")
progress_label.pack(pady=10)

# Footer
footer = tk.Label(root,
                  text="Developed under WIW1COB | Bosch Internal | Tag: Migration Analysis",
                  bg="#003366", fg="white", font=("Segoe UI", 9))
footer.pack(fill="x", side="bottom")

root.mainloop()

