"""
Interface List Report Generator
Generates HTML and Excel reports for interface analysis
"""

import os
import logging
from datetime import datetime
from typing import Dict, List
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class InterfaceListReportGenerator:
    """Generate HTML and Excel reports for interface analysis"""
    
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or os.path.join(os.getcwd(), 'Interface_Analysis_Results')
        os.makedirs(self.output_dir, exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def generate_single_workspace_report(self, analysis, workspace_name: str = "Workspace"):
        """Generate report for single workspace analysis"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Generate HTML report
        html_path = os.path.join(
            self.output_dir,
            f'Interface_Analysis_{workspace_name}_{timestamp}.html'
        )
        self._generate_single_html(analysis, html_path, workspace_name)
        
        # Generate Excel report with HTML path reference
        excel_path = os.path.join(
            self.output_dir,
            f'Interface_Analysis_{workspace_name}_{timestamp}.xlsx'
        )
        self._generate_single_excel(analysis, excel_path, workspace_name, html_path)
        
        self.logger.info(f"Reports generated:\n  HTML: {html_path}\n  Excel: {excel_path}")
        return html_path, excel_path
    
    def generate_comparison_report(self, comparison_data, platform_name: str = "Platform", 
                                   project_name: str = "Project"):
        """Generate report for workspace comparison"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Generate HTML report
        html_path = os.path.join(
            self.output_dir,
            f'Interface_Comparison_{timestamp}.html'
        )
        self._generate_comparison_html(comparison_data, html_path, platform_name, project_name)
        
        # Generate Excel report
        excel_path = os.path.join(
            self.output_dir,
            f'Interface_Comparison_{timestamp}.xlsx'
        )
        self._generate_comparison_excel(comparison_data, excel_path, platform_name, project_name)
        
        self.logger.info(f"Comparison reports generated:\n  HTML: {html_path}\n  Excel: {excel_path}")
        return html_path, excel_path
    
    def _generate_single_html(self, analysis, output_path: str, workspace_name: str):
        """Generate HTML report for single workspace"""
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Interface Analysis Report - {workspace_name}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .header {{
            background: linear-gradient(135deg, #003366 0%, #005599 100%);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        .header h1 {{
            margin: 0;
            font-size: 2.5em;
        }}
        .header .subtitle {{
            font-size: 1.1em;
            opacity: 0.9;
            margin-top: 10px;
        }}
        .summary {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-top: 20px;
        }}
        .summary-item {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            border-left: 4px solid #003366;
        }}
        .summary-item h3 {{
            margin: 0 0 10px 0;
            color: #003366;
            font-size: 0.9em;
            text-transform: uppercase;
        }}
        .summary-item .value {{
            font-size: 2em;
            font-weight: bold;
            color: #005599;
        }}
        .section {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .section h2 {{
            color: #003366;
            border-bottom: 3px solid #005599;
            padding-bottom: 10px;
            margin-top: 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #003366;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .type-badge {{
            display: inline-block;
            padding: 4px 12px;
            border-radius: 12px;
            font-size: 0.85em;
            font-weight: 600;
            background-color: #e3f2fd;
            color: #1565c0;
        }}
        .status-enabled {{
            color: #28a745;
            font-weight: bold;
        }}
        .status-disabled {{
            color: #dc3545;
            font-weight: bold;
        }}
        .status-conditional {{
            color: #ffc107;
            font-weight: bold;
        }}
        code {{
            background-color: #f4f4f4;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
        }}
        .file-path {{
            color: #666;
            font-size: 0.9em;
            font-family: 'Courier New', monospace;
        }}
        .file-path a {{
            color: #0563C1;
            text-decoration: none;
        }}
        .file-path a:hover {{
            text-decoration: underline;
            color: #004080;
        }}
        .dependency-list {{
            list-style: none;
            padding-left: 0;
        }}
        .dependency-list li {{
            padding: 8px;
            margin: 4px 0;
            background: #f8f9fa;
            border-radius: 4px;
        }}
        .flow-info {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            margin: 10px 0;
        }}
        .complexity-badge {{
            background: #ffc107;
            color: #000;
            padding: 4px 12px;
            border-radius: 12px;
            font-weight: bold;
            font-size: 0.9em;
        }}
        .chart-container {{
            margin: 20px 0;
            padding: 20px;
            background: #f8f9fa;
            border-radius: 8px;
        }}
        .summary-item a {{
            text-decoration: none;
            color: inherit;
            display: block;
            transition: transform 0.2s;
        }}
        .summary-item a:hover {{
            transform: scale(1.05);
            cursor: pointer;
        }}
        html {{
            scroll-behavior: smooth;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🔍 Interface Analysis Report</h1>
        <div class="subtitle">
            Workspace: {workspace_name}<br>
            Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}<br>
            Path: {analysis.workspace_path}
        </div>
    </div>
    
    <div class="summary">
        <h2>📊 Analysis Summary</h2>
        <p style="color: #666; font-size: 0.9em; margin-top: 5px;">Click on any section below to navigate</p>
        <div class="summary-grid">
            <div class="summary-item">
                <h3>Total Files</h3>
                <div class="value">{analysis.total_files}</div>
            </div>
            <div class="summary-item">
                <h3>Analyzed Files</h3>
                <div class="value">{analysis.analyzed_files}</div>
            </div>
            <div class="summary-item">
                <a href="#interfaces-section">
                    <h3>Interfaces Found</h3>
                    <div class="value">{len(analysis.interfaces)}</div>
                </a>
            </div>
            <div class="summary-item">
                <a href="#switches-section">
                    <h3>Switches Found</h3>
                    <div class="value">{len(analysis.switches)}</div>
                </a>
            </div>
            <div class="summary-item">
                <a href="#datatypes-section">
                    <h3>Data Types Used</h3>
                    <div class="value">{len(analysis.data_types_used)}</div>
                </a>
            </div>
        </div>
    </div>
"""
        
        # Interfaces section
        html_content += self._generate_interfaces_html(analysis.interfaces, analysis.workspace_path)
        
        # Switches section
        html_content += self._generate_switches_html(analysis.switches, analysis.workspace_path)
        
        # Data Types section
        html_content += self._generate_datatypes_html(analysis.data_types_used)
        
        # Dependencies section
        html_content += self._generate_dependencies_html(analysis.dependencies, analysis.workspace_path)
        
        # Flow Information section
        html_content += self._generate_flow_html(analysis.flow_info, analysis.workspace_path)
        
        html_content += """
</body>
</html>
"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _generate_interfaces_html(self, interfaces: List, base_path: str) -> str:
        """Generate HTML for interfaces section"""
        html = """
    <div class="section" id="interfaces-section">
        <h2>📋 Interfaces Catalog</h2>
        <table>
            <thead>
                <tr>
                    <th>Interface Name</th>
                    <th>Type</th>
                    <th>Data Type</th>
                    <th>File</th>
                    <th>Line</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for interface in interfaces:
            full_path = interface.file_path.replace('\\', '/')  # Normalize path separators
            # Create unique anchor ID for this interface
            anchor_id = f"interface-{interface.interface_name}-{interface.line_number}".replace(' ', '-').replace('/', '-')
            # Create vscode link to open file at line
            file_link = f"vscode://file/{interface.file_path.replace(chr(92), '/')}:{interface.line_number}"
            html += f"""
                <tr id="{anchor_id}">
                    <td><code>{interface.interface_name}</code></td>
                    <td><span class="type-badge">{interface.interface_type}</span></td>
                    <td><code>{interface.data_type}</code></td>
                    <td class="file-path"><a href="{file_link}" title="Open in editor">{full_path}</a></td>
                    <td>{interface.line_number}</td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
    </div>
"""
        return html
    
    def _generate_switches_html(self, switches: List, base_path: str) -> str:
        """Generate HTML for switches section"""
        html = """
    <div class="section" id="switches-section">
        <h2>⚙️ Preprocessor Switches</h2>
        <p>Compilation switches and their current status:</p>
        <table>
            <thead>
                <tr>
                    <th>Switch Name</th>
                    <th>Type</th>
                    <th>Status</th>
                    <th>Condition</th>
                    <th>File</th>
                    <th>Line</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for switch in switches:
            full_path = switch.file_path.replace('\\', '/')  # Normalize path separators
            status_class = f"status-{switch.status}"
            # Create unique anchor ID for this switch
            anchor_id = f"switch-{switch.switch_name}-{switch.line_number}".replace(' ', '-').replace('/', '-')
            # Create vscode link to open file at line
            file_link = f"vscode://file/{switch.file_path.replace(chr(92), '/')}:{switch.line_number}"
            html += f"""
                <tr id="{anchor_id}">
                    <td><code>{switch.switch_name}</code></td>
                    <td>{switch.switch_type}</td>
                    <td class="{status_class}">{switch.status.upper()}</td>
                    <td><code>{switch.condition}</code></td>
                    <td class="file-path"><a href="{file_link}" title="Open in editor">{full_path}</a></td>
                    <td>{switch.line_number}</td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
    </div>
"""
        return html
    
    def _generate_datatypes_html(self, data_types: Dict) -> str:
        """Generate HTML for data types section"""
        html = """
    <div class="section" id="datatypes-section">
        <h2>📊 Data Type Usage Statistics</h2>
        <div class="chart-container">
            <table>
                <thead>
                    <tr>
                        <th>Data Type</th>
                        <th>Usage Count</th>
                        <th>Percentage</th>
                    </tr>
                </thead>
                <tbody>
"""
        
        total = sum(data_types.values())
        for dtype, count in sorted(data_types.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total * 100) if total > 0 else 0
            html += f"""
                    <tr>
                        <td><code>{dtype}</code></td>
                        <td>{count}</td>
                        <td>{percentage:.1f}%</td>
                    </tr>
"""
        
        html += """
                </tbody>
            </table>
        </div>
    </div>
"""
        return html
    
    def _generate_dependencies_html(self, dependencies: Dict, base_path: str) -> str:
        """Generate HTML for dependencies section"""
        html = """
    <div class="section" id="dependencies-section">
        <h2>🔗 File Dependencies</h2>
        <p>Shows what each file includes and what includes it:</p>
"""
        
        for file_path, dep_info in list(dependencies.items())[:20]:  # Limit to first 20
            full_path = file_path.replace('\\', '/')  # Normalize path separators
            html += f"""
        <div class="flow-info">
            <h3>{full_path}</h3>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                <div>
                    <h4>Includes:</h4>
                    <ul class="dependency-list">
"""
            for inc in dep_info.includes[:10]:
                html += f"                        <li><code>{inc}</code></li>\n"
            
            html += """
                    </ul>
                </div>
                <div>
                    <h4>Included By:</h4>
                    <ul class="dependency-list">
"""
            for inc_by in dep_info.included_by[:10]:
                full_inc = inc_by.replace('\\', '/')  # Normalize path separators
                html += f"                        <li class=\"file-path\">{full_inc}</li>\n"
            
            html += """
                    </ul>
                </div>
            </div>
            <div>
                <h4>💡 Impact Analysis:</h4>
                <p>If this file is modified, <strong>{}</strong> file(s) may be affected.</p>
            </div>
        </div>
""".format(len(dep_info.included_by))
        
        html += """
    </div>
"""
        return html
    
    def _generate_flow_html(self, flow_info: Dict, base_path: str) -> str:
        """Generate HTML for flow information section"""
        html = """
    <div class="section">
        <h2>🔄 Control Flow Analysis</h2>
        <p>Analysis of code complexity and control structures:</p>
"""
        
        for file_path, flow in list(flow_info.items())[:15]:  # Limit to first 15
            full_path = file_path.replace('\\', '/')  # Normalize path separators
            html += f"""
        <div class="flow-info">
            <h3>{full_path}</h3>
            <p><span class="complexity-badge">Complexity: {flow.complexity}</span></p>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                <div>
                    <h4>Functions ({len(flow.functions)}):</h4>
                    <ul class="dependency-list">
"""
            for func in flow.functions[:10]:
                html += f"                        <li><code>{func}()</code></li>\n"
            
            html += """
                    </ul>
                </div>
                <div>
                    <h4>Control Structures:</h4>
                    <ul class="dependency-list">
"""
            ctrl_counts = {}
            for ctrl in flow.control_structures:
                ctrl_counts[ctrl['type']] = ctrl_counts.get(ctrl['type'], 0) + 1
            
            for ctrl_type, count in ctrl_counts.items():
                html += f"                        <li><strong>{ctrl_type}</strong>: {count}</li>\n"
            
            html += """
                    </ul>
                </div>
            </div>
        </div>
"""
        
        html += """
    </div>
"""
        return html
    
    def _generate_single_excel(self, analysis, output_path: str, workspace_name: str, html_path: str = None):
        """Generate Excel report for single workspace"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        self._create_summary_sheet(wb, analysis, workspace_name)
        
        # Interfaces sheet
        self._create_interfaces_sheet(wb, analysis, html_path)
        
        # Switches sheet
        self._create_switches_sheet(wb, analysis, html_path)
        
        # Data Types sheet
        self._create_datatypes_sheet(wb, analysis)
        
        # Dependencies sheet
        self._create_dependencies_sheet(wb, analysis)
        
        # Save workbook
        wb.save(output_path)
    
    def _create_summary_sheet(self, wb, analysis, workspace_name: str):
        """Create summary sheet in Excel"""
        ws = wb.create_sheet("Summary")
        
        # Header
        ws['A1'] = "Interface Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Workspace: {workspace_name}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Path: {analysis.workspace_path}"
        
        # Statistics
        ws['A6'] = "Metric"
        ws['B6'] = "Value"
        ws['A6'].font = Font(bold=True)
        ws['B6'].font = Font(bold=True)
        
        metrics = [
            ("Total Files", analysis.total_files),
            ("Analyzed Files", analysis.analyzed_files),
            ("Interfaces Found", len(analysis.interfaces)),
            ("Switches Found", len(analysis.switches)),
            ("Data Types Used", len(analysis.data_types_used)),
            ("Dependencies Tracked", len(analysis.dependencies)),
        ]
        
        row = 7
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_interfaces_sheet(self, wb, analysis, html_path: str = None):
        """Create interfaces sheet in Excel"""
        ws = wb.create_sheet("Interfaces")
        
        # Headers
        headers = ["Interface Name", "Type", "Data Type", "File", "Line", "Declaration", "Definition", "View in HTML", "Open File"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        row = 2
        for interface in analysis.interfaces:
            full_path = interface.file_path.replace('\\', '/')  # Normalize path separators
            ws.cell(row, 1, interface.interface_name)
            ws.cell(row, 2, interface.interface_type)
            ws.cell(row, 3, interface.data_type)
            ws.cell(row, 4, full_path)
            ws.cell(row, 5, interface.line_number)
            ws.cell(row, 6, "Yes" if interface.is_declaration else "No")
            ws.cell(row, 7, "Yes" if interface.is_definition else "No")
            
            # Add hyperlink to HTML report
            if html_path:
                anchor_id = f"interface-{interface.interface_name}-{interface.line_number}".replace(' ', '-').replace('/', '-')
                html_url = f"file:///{html_path.replace(chr(92), '/')}#{anchor_id}"
                link_cell = ws.cell(row, 8, "🔗 View")
                link_cell.hyperlink = html_url
                link_cell.font = Font(color="0563C1", underline="single")
                link_cell.alignment = Alignment(horizontal="center")
            
            # Add hyperlink to open source file at specific line
            file_url = f"vscode://file/{interface.file_path.replace(chr(92), '/')}:{interface.line_number}"
            file_link_cell = ws.cell(row, 9, "📂 Open")
            file_link_cell.hyperlink = file_url
            file_link_cell.font = Font(color="0563C1", underline="single")
            file_link_cell.alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Auto-adjust columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_switches_sheet(self, wb, analysis, html_path: str = None):
        """Create switches sheet in Excel"""
        ws = wb.create_sheet("Switches")
        
        # Headers
        headers = ["Switch Name", "Type", "Status", "Condition", "File", "Line", "View in HTML", "Open File"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        row = 2
        for switch in analysis.switches:
            full_path = switch.file_path.replace('\\', '/')  # Normalize path separators
            ws.cell(row, 1, switch.switch_name)
            ws.cell(row, 2, switch.switch_type)
            
            status_cell = ws.cell(row, 3, switch.status.upper())
            if switch.status == 'enabled':
                status_cell.font = Font(color="008000", bold=True)
            elif switch.status == 'disabled':
                status_cell.font = Font(color="FF0000", bold=True)
            else:
                status_cell.font = Font(color="FFA500", bold=True)
            
            ws.cell(row, 4, switch.condition)
            ws.cell(row, 5, full_path)
            ws.cell(row, 6, switch.line_number)
            
            # Add hyperlink to HTML report
            if html_path:
                anchor_id = f"switch-{switch.switch_name}-{switch.line_number}".replace(' ', '-').replace('/', '-')
                html_url = f"file:///{html_path.replace(chr(92), '/')}#{anchor_id}"
                link_cell = ws.cell(row, 7, "🔗 View")
                link_cell.hyperlink = html_url
                link_cell.font = Font(color="0563C1", underline="single")
                link_cell.alignment = Alignment(horizontal="center")
            
            # Add hyperlink to open source file at specific line
            file_url = f"vscode://file/{switch.file_path.replace(chr(92), '/')}:{switch.line_number}"
            file_link_cell = ws.cell(row, 8, "📂 Open")
            file_link_cell.hyperlink = file_url
            file_link_cell.font = Font(color="0563C1", underline="single")
            file_link_cell.alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Auto-adjust columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_datatypes_sheet(self, wb, analysis):
        """Create data types sheet in Excel"""
        ws = wb.create_sheet("Data Types")
        
        # Headers
        headers = ["Data Type", "Usage Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        row = 2
        total = sum(analysis.data_types_used.values())
        for dtype, count in sorted(analysis.data_types_used.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total * 100) if total > 0 else 0
            ws.cell(row, 1, dtype)
            ws.cell(row, 2, count)
            ws.cell(row, 3, f"{percentage:.1f}%")
            row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_dependencies_sheet(self, wb, analysis):
        """Create dependencies sheet in Excel"""
        ws = wb.create_sheet("Dependencies")
        
        # Headers
        headers = ["File", "Includes", "Included By", "Defines Functions", "Impact Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        row = 2
        for file_path, dep_info in analysis.dependencies.items():
            full_path = file_path.replace('\\', '/')  # Normalize path separators
            ws.cell(row, 1, full_path)
            ws.cell(row, 2, ", ".join(dep_info.includes[:5]))
            ws.cell(row, 3, len(dep_info.included_by))
            ws.cell(row, 4, ", ".join(dep_info.defines_functions[:5]))
            ws.cell(row, 5, len(dep_info.included_by))  # Impact = number of files that include this
            row += 1
        
        # Auto-adjust columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 30
    
    def _generate_comparison_html(self, comparison_data, output_path: str, 
                                   platform_name: str, project_name: str):
        """Generate HTML for workspace comparison"""
        platform = comparison_data['platform']
        project = comparison_data['project']
        differences = comparison_data['differences']
        
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Interface Comparison Report - {platform_name} vs {project_name}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .header {{
            background: linear-gradient(135deg, #003366 0%, #005599 100%);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        .header h1 {{
            margin: 0;
            font-size: 2.5em;
        }}
        .header .subtitle {{
            font-size: 1.1em;
            opacity: 0.9;
            margin-top: 10px;
        }}
        .comparison-summary {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
            margin-bottom: 30px;
        }}
        .workspace-card {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .workspace-card h2 {{
            color: #003366;
            margin-top: 0;
        }}
        .stat-item {{
            padding: 10px 0;
            border-bottom: 1px solid #eee;
        }}
        .stat-item:last-child {{
            border-bottom: none;
        }}
        .section {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .section h2 {{
            color: #003366;
            border-bottom: 3px solid #005599;
            padding-bottom: 10px;
            margin-top: 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #003366;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .diff-added {{
            background-color: #d4edda;
            border-left: 4px solid #28a745;
        }}
        .diff-removed {{
            background-color: #f8d7da;
            border-left: 4px solid #dc3545;
        }}
        .diff-modified {{
            background-color: #fff3cd;
            border-left: 4px solid #ffc107;
        }}
        code {{
            background-color: #f4f4f4;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
        }}
        .file-path {{
            color: #666;
            font-size: 0.9em;
            font-family: 'Courier New', monospace;
        }}
        .file-path a {{
            color: #0563C1;
            text-decoration: none;
        }}
        .file-path a:hover {{
            text-decoration: underline;
            color: #004080;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🔄 Interface Comparison Report</h1>
        <div class="subtitle">
            Comparing: {platform_name} ↔️ {project_name}<br>
            Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        </div>
    </div>
    
    <div class="comparison-summary">
        <div class="workspace-card">
            <h2>📁 {platform_name}</h2>
            <div class="stat-item">Files: <strong>{platform.analyzed_files}</strong></div>
            <div class="stat-item">Interfaces: <strong>{len(platform.interfaces)}</strong></div>
            <div class="stat-item">Switches: <strong>{len(platform.switches)}</strong></div>
            <div class="stat-item">Path: <code>{platform.workspace_path}</code></div>
        </div>
        <div class="workspace-card">
            <h2>📁 {project_name}</h2>
            <div class="stat-item">Files: <strong>{project.analyzed_files}</strong></div>
            <div class="stat-item">Interfaces: <strong>{len(project.interfaces)}</strong></div>
            <div class="stat-item">Switches: <strong>{len(project.switches)}</strong></div>
            <div class="stat-item">Path: <code>{project.workspace_path}</code></div>
        </div>
    </div>
"""
        
        # Differences sections
        html_content += self._generate_interface_differences_html(differences, platform, project)
        html_content += self._generate_switch_differences_html(differences, platform, project)
        
        html_content += """
</body>
</html>
"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _generate_interface_differences_html(self, differences, platform, project) -> str:
        """Generate HTML for interface differences"""
        html = """
    <div class="section">
        <h2>📋 Interface Differences</h2>
"""
        
        # Only in platform
        if differences['interfaces']['only_in_platform']:
            html += f"""
        <h3>Only in {platform.workspace_path} ({len(differences['interfaces']['only_in_platform'])} items)</h3>
        <table>
            <thead>
                <tr>
                    <th>Interface Name</th>
                    <th>Type</th>
                    <th>Data Type</th>
                    <th>File</th>
                </tr>
            </thead>
            <tbody>
"""
            for interface in differences['interfaces']['only_in_platform'][:50]:
                full_path = interface.file_path.replace('\\', '/')  # Normalize path separators
                file_link = f"vscode://file/{interface.file_path.replace(chr(92), '/')}:{interface.line_number}"
                html += f"""
                <tr class="diff-removed">
                    <td><code>{interface.interface_name}</code></td>
                    <td>{interface.interface_type}</td>
                    <td><code>{interface.data_type}</code></td>
                    <td class="file-path"><a href="{file_link}" title="Open in editor">{full_path}</a></td>
                </tr>
"""
            html += """
            </tbody>
        </table>
"""
        
        # Only in project
        if differences['interfaces']['only_in_project']:
            html += f"""
        <h3>Only in {project.workspace_path} ({len(differences['interfaces']['only_in_project'])} items)</h3>
        <table>
            <thead>
                <tr>
                    <th>Interface Name</th>
                    <th>Type</th>
                    <th>Data Type</th>
                    <th>File</th>
                </tr>
            </thead>
            <tbody>
"""
            for interface in differences['interfaces']['only_in_project'][:50]:
                full_path = interface.file_path.replace('\\', '/')  # Normalize path separators
                file_link = f"vscode://file/{interface.file_path.replace(chr(92), '/')}:{interface.line_number}"
                html += f"""
                <tr class="diff-added">
                    <td><code>{interface.interface_name}</code></td>
                    <td>{interface.interface_type}</td>
                    <td><code>{interface.data_type}</code></td>
                    <td class="file-path"><a href="{file_link}" title="Open in editor">{full_path}</a></td>
                </tr>
"""
            html += """
            </tbody>
        </table>
"""
        
        # Modified
        if differences['interfaces']['modified']:
            html += f"""
        <h3>Modified Interfaces ({len(differences['interfaces']['modified'])} items)</h3>
        <table>
            <thead>
                <tr>
                    <th>Interface Name</th>
                    <th>Platform Type</th>
                    <th>Project Type</th>
                    <th>File</th>
                </tr>
            </thead>
            <tbody>
"""
            for diff in differences['interfaces']['modified'][:50]:
                platform_int = diff['platform']
                project_int = diff['project']
                full_path = platform_int.file_path.replace('\\', '/')  # Normalize path separators
                file_link = f"vscode://file/{platform_int.file_path.replace(chr(92), '/')}:{platform_int.line_number}"
                html += f"""
                <tr class="diff-modified">
                    <td><code>{platform_int.interface_name}</code></td>
                    <td><code>{platform_int.data_type}</code></td>
                    <td><code>{project_int.data_type}</code></td>
                    <td class="file-path"><a href="{file_link}" title="Open in editor">{full_path}</a></td>
                </tr>
"""
            html += """
            </tbody>
        </table>
"""
        
        html += """
    </div>
"""
        return html
    
    def _generate_switch_differences_html(self, differences, platform, project) -> str:
        """Generate HTML for switch differences"""
        html = """
    <div class="section">
        <h2>⚙️ Switch Status Differences</h2>
"""
        
        if differences['switches']['status_changed']:
            html += f"""
        <h3>Changed Switch Status ({len(differences['switches']['status_changed'])} items)</h3>
        <table>
            <thead>
                <tr>
                    <th>Switch Name</th>
                    <th>Platform Status</th>
                    <th>Project Status</th>
                    <th>File</th>
                </tr>
            </thead>
            <tbody>
"""
            for diff in differences['switches']['status_changed']:
                platform_sw = diff['platform']
                project_sw = diff['project']
                full_path = platform_sw.file_path.replace('\\', '/')  # Normalize path separators
                html += f"""
                <tr class="diff-modified">
                    <td><code>{platform_sw.switch_name}</code></td>
                    <td><strong>{platform_sw.status}</strong></td>
                    <td><strong>{project_sw.status}</strong></td>
                    <td class="file-path">{full_path}</td>
                </tr>
"""
            html += """
            </tbody>
        </table>
"""
        else:
            html += "<p>No switch status changes detected.</p>"
        
        html += """
    </div>
"""
        return html
    
    def _generate_comparison_excel(self, comparison_data, output_path: str, 
                                    platform_name: str, project_name: str):
        """Generate Excel for workspace comparison"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Comparison summary sheet
        ws = wb.create_sheet("Comparison Summary")
        ws['A1'] = "Workspace Comparison Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"{platform_name} vs {project_name}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Statistics table
        ws['A5'] = "Metric"
        ws['B5'] = platform_name
        ws['C5'] = project_name
        for col in ['A5', 'B5', 'C5']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            ws[col].font = Font(color="FFFFFF", bold=True)
        
        metrics = [
            ("Files Analyzed", comparison_data['platform'].analyzed_files, comparison_data['project'].analyzed_files),
            ("Interfaces", len(comparison_data['platform'].interfaces), len(comparison_data['project'].interfaces)),
            ("Switches", len(comparison_data['platform'].switches), len(comparison_data['project'].switches)),
        ]
        
        row = 6
        for metric, platform_val, project_val in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = platform_val
            ws[f'C{row}'] = project_val
            row += 1
        
        # Difference sheets
        self._create_interface_diff_sheet(wb, comparison_data, platform_name, project_name)
        self._create_switch_diff_sheet(wb, comparison_data, platform_name, project_name)
        
        wb.save(output_path)
    
    def _create_interface_diff_sheet(self, wb, comparison_data, platform_name: str, project_name: str):
        """Create interface differences sheet"""
        ws = wb.create_sheet("Interface Differences")
        
        headers = ["Interface Name", "Type", "Data Type", "Status", "File", "Line", "Open File"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        
        # Only in platform
        for interface in comparison_data['differences']['interfaces']['only_in_platform']:
            full_path = interface.file_path.replace('\\', '/')  # Normalize path separators
            ws.cell(row, 1, interface.interface_name)
            ws.cell(row, 2, interface.interface_type)
            ws.cell(row, 3, interface.data_type)
            status_cell = ws.cell(row, 4, f"Only in {platform_name}")
            status_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            ws.cell(row, 5, full_path)
            ws.cell(row, 6, interface.line_number)
            
            # Add hyperlink to open source file
            file_url = f"vscode://file/{interface.file_path.replace(chr(92), '/')}:{interface.line_number}"
            file_link_cell = ws.cell(row, 7, "📂 Open")
            file_link_cell.hyperlink = file_url
            file_link_cell.font = Font(color="0563C1", underline="single")
            file_link_cell.alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Only in project
        for interface in comparison_data['differences']['interfaces']['only_in_project']:
            full_path = interface.file_path.replace('\\', '/')  # Normalize path separators
            ws.cell(row, 1, interface.interface_name)
            ws.cell(row, 2, interface.interface_type)
            ws.cell(row, 3, interface.data_type)
            status_cell = ws.cell(row, 4, f"Only in {project_name}")
            status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            ws.cell(row, 5, full_path)
            ws.cell(row, 6, interface.line_number)
            
            # Add hyperlink to open source file
            file_url = f"vscode://file/{interface.file_path.replace(chr(92), '/')}:{interface.line_number}"
            file_link_cell = ws.cell(row, 7, "📂 Open")
            file_link_cell.hyperlink = file_url
            file_link_cell.font = Font(color="0563C1", underline="single")
            file_link_cell.alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Modified
        for diff in comparison_data['differences']['interfaces']['modified']:
            platform_int = diff['platform']
            project_int = diff['project']
            full_path = platform_int.file_path.replace('\\', '/')  # Normalize path separators
            ws.cell(row, 1, platform_int.interface_name)
            ws.cell(row, 2, platform_int.interface_type)
            ws.cell(row, 3, f"{platform_int.data_type} → {project_int.data_type}")
            status_cell = ws.cell(row, 4, "Modified")
            status_cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            ws.cell(row, 5, full_path)
            ws.cell(row, 6, platform_int.line_number)
            
            # Add hyperlink to open source file (platform version)
            file_url = f"vscode://file/{platform_int.file_path.replace(chr(92), '/')}:{platform_int.line_number}"
            file_link_cell = ws.cell(row, 7, "📂 Open")
            file_link_cell.hyperlink = file_url
            file_link_cell.font = Font(color="0563C1", underline="single")
            file_link_cell.alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 30
    
    def _create_switch_diff_sheet(self, wb, comparison_data, platform_name: str, project_name: str):
        """Create switch differences sheet"""
        ws = wb.create_sheet("Switch Differences")
        
        headers = ["Switch Name", "Platform Status", "Project Status", "File", "Line", "Open File"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        for diff in comparison_data['differences']['switches']['status_changed']:
            platform_sw = diff['platform']
            project_sw = diff['project']
            full_path = platform_sw.file_path.replace('\\', '/')  # Normalize path separators
            
            ws.cell(row, 1, platform_sw.switch_name)
            ws.cell(row, 2, platform_sw.status)
            ws.cell(row, 3, project_sw.status)
            ws.cell(row, 4, full_path)
            ws.cell(row, 5, platform_sw.line_number)
            
            # Color code based on change
            for col in range(1, 6):
                ws.cell(row, col).fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            
            # Add hyperlink to open source file (platform version)
            file_url = f"vscode://file/{platform_sw.file_path.replace(chr(92), '/')}:{platform_sw.line_number}"
            file_link_cell = ws.cell(row, 6, "📂 Open")
            file_link_cell.hyperlink = file_url
            file_link_cell.font = Font(color="0563C1", underline="single")
            file_link_cell.alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Auto-adjust columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 30
