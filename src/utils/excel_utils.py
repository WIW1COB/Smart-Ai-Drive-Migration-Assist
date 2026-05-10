"""Excel utility functions for Migration Analysis Tool"""
 
import os
import csv
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles import numbers
 
from .file_utils import sanitize_for_excel
 
 
def create_overview_sheet(wb, results, folder1, folder2):
    """
    Create an overview sheet with summary statistics and comparison info.
    Contains detailed analysis with color coding and complexity assessment.
    """
    from datetime import datetime
    from openpyxl.styles import Border, Side
   
    ws = wb.active
    ws.title = "Overview"
   
    # Define styles
    title_font = Font(size=14, bold=True, color="1F4E78")
    subtitle_font = Font(size=11, bold=True, color="003366")
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
   
    # Calculate statistics
    # Results structure: [File Path, Lines1, Lines2, Line Status, Status, HTML Link, Purpose]
    # Status is at index 4, Lines1 at index 1, Lines2 at index 2
    total_files = len(results)
    identical_count = sum(1 for r in results if r[4] == "Identical")
    different_count = sum(1 for r in results if r[4] == "Different")
    comments_only_count = sum(1 for r in results if r[4] == "Comments update only")
    only_folder1_count = sum(1 for r in results if r[4] == "Only in Platform")
    only_folder2_count = sum(1 for r in results if r[4] == "Only in Project")
   
    # Calculate line statistics
    def get_lines(r):
        """Get line counts safely"""
        try:
            lines1 = int(r[1]) if r[1] and str(r[1]).isdigit() else 0
            lines2 = int(r[2]) if r[2] and str(r[2]).isdigit() else 0
            return lines1, lines2
        except:
            return 0, 0
   
    # Total lines
    total_lines_folder1 = sum(get_lines(r)[0] for r in results)
    total_lines_folder2 = sum(get_lines(r)[1] for r in results)
    total_lines_combined = total_lines_folder1 + total_lines_folder2
   
    # Lines by category
    identical_lines = sum(get_lines(r)[0] + get_lines(r)[1] for r in results if r[4] == "Identical")
    different_lines = sum(get_lines(r)[0] + get_lines(r)[1] for r in results if r[4] == "Different")
    comments_only_lines = sum(get_lines(r)[0] + get_lines(r)[1] for r in results if r[4] == "Comments update only")
    only_folder1_lines = sum(get_lines(r)[0] for r in results if r[4] == "Only in Platform")
    only_folder2_lines = sum(get_lines(r)[1] for r in results if r[4] == "Only in Project")
   
    # Determine complexity level based on percentage of different files
    def get_complexity(diff_pct):
        """Determine complexity level based on percentage of differences"""
        if diff_pct <= 10:
            return "Minor"
        elif diff_pct <= 30:
            return "Medium"
        else:
            return "Major"
   
    # Title
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "Migration Analysis Report - Overview"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
   
    # Comparison Information
    ws.append([])
    ws.append(["Comparison Information:"])
    ws['A3'].font = subtitle_font
   
    ws.append(["Platform (Baseline):", folder1])
    ws.append(["Project:", folder2])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
   
    # Summary Statistics (following template format)
    ws.append(["Summary Statistics:"])
    ws['A8'].font = subtitle_font
   
    # Header Row with 9 columns (E-F merged for "No of lines(LOC)")
    ws.append([
        "Category",
        "No of files",
        "% of files",
        "% of files\n(Platform vs Project)",
        "No of lines(LOC)",
        None,  # Merged with E
        "% of LOC",
        "% of LOC\n(Platform vs Project)",
        "Complexity Level"
    ])
   
    header_row = 9
    # Format header cells
    for col in range(1, 10):
        cell = ws.cell(header_row, col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30
   
    # Merge E1:F1 for "No of lines(LOC)" header
    ws.merge_cells(f'E{header_row}:F{header_row}')
   
    # Determine individual complexity levels
    different_pct_val = (different_count/total_files*100) if total_files > 0 else 0
    complexity_identical = "Minor"
    complexity_platform_only = "Minor"
    complexity_comments = "Minor" if comments_only_count > 0 else "Minor"
    complexity_different = get_complexity(different_pct_val)
    complexity_project_only = "Minor"
   
    # Check if all complexities are the same
    all_complexities = [complexity_identical, complexity_platform_only, complexity_comments,
                        complexity_different, complexity_project_only]
    all_same_complexity = len(set(all_complexities)) == 1
    overall_complexity = all_complexities[0] if all_same_complexity else get_complexity(different_pct_val)
   
    # Row 2: Total (with Excel formulas)
    row = header_row + 1
    ws.append([
        "Total",
        f"=B{row+1}+B{row+2}+B{row+3}+B{row+4}+B{row+6}",  # Sum of files
        f"=C{row+1}+C{row+2}+C{row+3}+C{row+4}+C{row+6}",  # Sum of file %
        f"=SUM(D{row+1}+D{row+4})",  # Platform vs Project %
        None,
        f"=F{row+1}+F{row+3}+F{row+4}+F{row+5}+F{row+6}",  # Sum of lines
        f"=G{row+1}+G{row+3}+G{row+4}+G{row+6}",  # Sum of line %
        f"=H{row+1}+H{row+4}",  # Platform vs Project line %
        overall_complexity if all_same_complexity else None  # Complexity (merged if all same)
    ])
    ws.cell(row, 1).font = Font(bold=True, size=11)
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
   
    # Row 3: Identical (Files with no differences)
    row += 1
    row_identical = row
    ws.append([
        "Identical (Files with no differences)",
        identical_count,
        f"=B{row}/$B${header_row+1}",
        f"=SUM(C{row}:C{row+2})",
        None,
        identical_lines,
        f"=F{row}/F${header_row+1}",
        f"=SUM(G{row}:G{row+2})",
        complexity_identical if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
   
    # Row 4: Files exist only in platform
    row += 1
    ws.append([
        "Files exist only in platform",
        only_folder1_count,
        f"=B{row}/$B${header_row+1}",
        None,
        None,
        "Not Applicable" if only_folder1_count > 0 else 0,
        "NA" if only_folder1_count > 0 else 0,
        None,
        complexity_platform_only if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FF9BC2E6", end_color="FF9BC2E6", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FF9BC2E6", end_color="FF9BC2E6", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
   
    # Row 5: Comments update only
    row += 1
    ws.append([
        "Comments update only",
        comments_only_count,
        f"=B{row}/$B${header_row+1}",
        None,
        None,
        comments_only_lines if comments_only_count > 0 else 0,
        f"=F{row}/F${header_row+1}" if comments_only_count > 0 else 0,
        None,
        complexity_comments if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
   
    # Merge D cells for identical, platform, comments (rows 3-5)
    ws.merge_cells(f'D{row_identical}:D{row}')
    ws.merge_cells(f'H{row_identical}:H{row}')
   
    # Row 6-7: Files with code differences in Project (with Added/Removed sub-rows)
    row += 1
    row_diff_start = row
    # Calculate added and removed lines for different files
    added_lines_diff = sum(get_lines(r)[1] for r in results if r[4] == "Different")
    removed_lines_diff = -sum(get_lines(r)[0] for r in results if r[4] == "Different")
   
    ws.append([
        "Files with code differences in Project",
        different_count,
        f"=B{row}/B${header_row+1}",
        f"=SUM(C{row}:C{row+2})",
        "Added",
        added_lines_diff,
        f"=F{row}/F${header_row+1}",
        f"=SUM(G{row}:G{row+2})",
        complexity_different if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
   
    # Row 7: Sub-row for Removed lines
    row += 1
    ws.append([
        None,
        None,
        None,
        None,
        "Removed",
        removed_lines_diff,
        f"=F{row}/F${header_row+1}",
        None,
        None
    ])
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
   
    # Row 8: Files exist only in Project
    row += 1
    ws.append([
        "Files exist only in Project",
        only_folder2_count,
        f"=B{row}/B${header_row+1}",
        None,
        "Added",
        only_folder2_lines,
        f"=F{row}/F${header_row+1}",
        None,
        complexity_project_only if not all_same_complexity else None
    ])
    ws.cell(row, 1).fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    ws.cell(row, 2).fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for col in range(1, 10):
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, col).border = thin_border
   
    # Merge cells for "Files with code differences" (A, B, C, D columns across rows 6-7)
    ws.merge_cells(f'A{row_diff_start}:A{row_diff_start+1}')
    ws.merge_cells(f'B{row_diff_start}:B{row_diff_start+1}')
    ws.merge_cells(f'C{row_diff_start}:C{row_diff_start+1}')
    ws.merge_cells(f'D{row_diff_start}:D{row}')
    ws.merge_cells(f'H{row_diff_start}:H{row}')
   
    # If all complexities are the same, merge the Complexity Level column (I) for all data rows
    if all_same_complexity:
        ws.merge_cells(f'I{header_row+1}:I{row}')
        ws.cell(header_row+1, 9).font = Font(bold=True, size=12)
    else:
        # Merge complexity for sub-rows that don't have individual complexity
        ws.merge_cells(f'I{row_diff_start}:I{row_diff_start+1}')
   
    # Apply percentage formatting to percentage columns (C, D, G, H)
    from openpyxl.styles import numbers
    for row_num in range(header_row+1, row+1):
        # Column C: % of files
        ws.cell(row_num, 3).number_format = numbers.FORMAT_PERCENTAGE_00
        # Column D: % of files (Platform vs Project)
        if ws.cell(row_num, 4).value and ws.cell(row_num, 4).value != "":
            ws.cell(row_num, 4).number_format = numbers.FORMAT_PERCENTAGE_00
        # Column G: % of LOC
        if ws.cell(row_num, 7).value not in ["NA", None, ""]:
            ws.cell(row_num, 7).number_format = numbers.FORMAT_PERCENTAGE_00
        # Column H: % of LOC (Platform vs Project)
        if ws.cell(row_num, 8).value and ws.cell(row_num, 8).value != "":
            ws.cell(row_num, 8).number_format = numbers.FORMAT_PERCENTAGE_00
   
    row = ws.max_row
   
   
    ws.append([])
   
    # Color Legend
    legend_row_start = ws.max_row + 1
    ws.append(["Color Coding Legend:"])
    ws.cell(legend_row_start, 1).font = subtitle_font
    ws.append([])
   
    # List of color descriptions
    row = legend_row_start + 1
    ws.append(["Identical (Files with no differences) -- Green color"])
    ws.cell(row, 1).font = Font(size=10)
   
    row += 1
    ws.append(["Files exist only in platform -- Blue color"])
    ws.cell(row, 1).font = Font(size=10)
   
    row += 1
    ws.append(["Comments update only -- Yellow color"])
    ws.cell(row, 1).font = Font(size=10)
   
    row += 1
    ws.append(["Files with code differences in Project -- Red color"])
    ws.cell(row, 1).font = Font(size=10)
   
    row += 1
    ws.append(["Files exist only in Project -- Orange color"])
    ws.cell(row, 1).font = Font(size=10)
   
    # Set column widths to match template + complexity column
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 13.5
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 16
   
    return ws
    ws['A3'].font = subtitle_font
    ws.append(["Platform (Baseline):", folder1])
    ws.append(["Project:", folder2])
    ws.append(["Generated:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
   
    # Summary Statistics
    ws.append(["Summary Statistics:"])
    ws.append(["Identical Files:", identical_count])
    ws.append(["Different Files:", different_count])
    ws.append(["Comment Changes Only:", comments_only_count])
    ws.append(["Only in Platform:", only_folder1_count])
    ws.append(["Only in Project:", only_folder2_count])
    ws.append(["Total Files:", total_files])
   
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
   
    return ws
 
 
def write_excel_report(results, output_path, folder1_name, folder2_name):
    """Write comparison results to Excel file with formatting"""
    wb = Workbook()
   
    # Create overview sheet
    create_overview_sheet(wb, results, folder1_name, folder2_name)
   
    # Create detailed results sheet
    ws_details = wb.create_sheet(title="Detailed Results")
   
    # Header row
    headers = [
        "File Path",
        f"Lines in Platform ({folder1_name})",
        f"Lines in Project ({folder2_name})",
        "Line Comparison Status",
        "Status",
        "HTML Diff Report",
        "Purpose of Change",
        "ChangeSet & WorkItem from RTC History"
    ]
    ws_details.append(headers)
   
    # Style header row
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
   
    for col_num, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
   
    # Data rows with color coding
    for row_idx, row_data in enumerate(results, 2):
        ws_details.append([sanitize_for_excel(str(cell)) for cell in row_data])
       
        # Add clickable hyperlink for HTML Diff Report (column F, index 6)
        if len(row_data) > 5 and row_data[5]:
            html_path = str(row_data[5])
            # Check if it's a valid path (not empty or "N/A")
            if html_path and html_path.lower() not in ['n/a', 'na', '', 'none']:
                # Verify the file exists
                if os.path.exists(html_path):
                    cell_f = ws_details.cell(row=row_idx, column=6)
                    # Create hyperlink to HTML file
                    cell_f.hyperlink = html_path
                    cell_f.value = "View Diff"
                    # Style the hyperlink
                    cell_f.font = Font(color="0000FF", underline="single")
                    cell_f.alignment = Alignment(horizontal="center", vertical="center")
       
        # Apply color based on status
        status = row_data[4] if len(row_data) > 4 else ""
        fill_color = None
       
        if status == "Identical":
            fill_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "Different":
            fill_color = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        elif status == "Comments update only":
            fill_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif status == "Only in Platform":
            fill_color = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        elif status == "Only in Project":
            fill_color = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
       
        if fill_color:
            for col_num in range(1, len(headers) + 1):
                ws_details.cell(row=row_idx, column=col_num).fill = fill_color
   
    # Set column widths
    ws_details.column_dimensions['A'].width = 50
    ws_details.column_dimensions['B'].width = 15
    ws_details.column_dimensions['C'].width = 15
    ws_details.column_dimensions['D'].width = 30
    ws_details.column_dimensions['E'].width = 20
    ws_details.column_dimensions['F'].width = 40
    ws_details.column_dimensions['G'].width = 50
    ws_details.column_dimensions['H'].width = 40
   
    # Save workbook
    wb.save(output_path)
    print(f"Excel report saved: {output_path}")
 