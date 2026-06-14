from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_PATH = Path(r"c:\Users\WIW1COB\Documents\Mig_inp.xlsx")
OUTPUT_PATH = Path(r"c:\Users\WIW1COB\Documents\Mig_OEM_Planning_Report.xlsx")

TITLE_FILL = PatternFill("solid", fgColor="0F4C81")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="B4C7E7")
ACCENT_FILL = PatternFill("solid", fgColor="E2F0D9")
RISK_FILL = PatternFill("solid", fgColor="FCE4D6")
GANTT_FILL = PatternFill("solid", fgColor="5B9BD5")
GANTT_REVIEW_FILL = PatternFill("solid", fgColor="A9D18E")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


@dataclass
class InputRow:
    serial_no: int
    topic: str
    feature_coverage: str
    estimated_hours_raw: str
    status: str

    @property
    def feature_lines(self) -> list[str]:
        return [line.strip() for line in self.feature_coverage.splitlines() if line.strip()]

    @property
    def effort_values(self) -> list[float]:
        values: list[float] = []
        for item in str(self.estimated_hours_raw).splitlines():
            item = item.strip()
            if not item:
                continue
            try:
                values.append(float(item))
            except ValueError:
                continue
        if not values:
            try:
                values.append(float(self.estimated_hours_raw))
            except Exception:
                pass
        return values

    @property
    def base_hours(self) -> float:
        values = self.effort_values
        if values:
            return sum(values)
        try:
            return float(self.estimated_hours_raw)
        except Exception:
            return 0.0


def style_range(ws, cell_range: str, fill: PatternFill, bold: bool = False, font_color: str = "000000") -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=bold, color=font_color)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_cell(cell, fill: PatternFill | None = None, bold: bool = False, font_color: str = "000000", size: int = 11) -> None:
    if fill is not None:
        cell.fill = fill
    cell.font = Font(bold=bold, color=font_color, size=size)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def read_input_rows(input_path: Path) -> list[InputRow]:
    workbook = load_workbook(input_path, data_only=False)
    sheet = workbook.active
    rows: list[InputRow] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        serial_no, topic, feature_coverage, estimated_hours, status = values[:5]
        rows.append(
            InputRow(
                serial_no=int(serial_no),
                topic=str(topic or "").strip(),
                feature_coverage=str(feature_coverage or "").strip(),
                estimated_hours_raw=str(estimated_hours or "").strip(),
                status=str(status or "").strip(),
            )
        )
    return rows


def infer_owner(topic: str) -> str:
    topic_lower = topic.lower()
    if "component" in topic_lower:
        return "Tool Engineer / Component SME"
    if "req" in topic_lower or "requirement" in topic_lower:
        return "Requirements Engineer"
    if "code" in topic_lower:
        return "Software Engineer / Architecture Reviewer"
    if "migration" in topic_lower:
        return "Migration Lead"
    return "Migration Analyst"


def infer_deliverable(topic: str) -> str:
    topic_lower = topic.lower()
    if "migration" in topic_lower:
        return "Baseline migration analysis pack"
    if "component" in topic_lower:
        return "Component-specific optimization baseline"
    if "newfile" in topic_lower or "customer" in topic_lower:
        return "Customer delta and new-file assessment"
    if "req" in topic_lower:
        return "Requirement rationalization note"
    if "code" in topic_lower:
        return "Code modification justification report"
    return "OEM analysis artifact"


def infer_dependency(index: int) -> str:
    if index == 1:
        return "RTC access, sample snapshots/workspaces, local project baseline"
    if index == 2:
        return "Completion of baseline comparison and OEM component list"
    if index == 3:
        return "Component optimization completed; DOORS access available"
    if index == 4:
        return "Customer feature delta list and platform SME review window"
    return "Requirement mapping completed; OEM SPOC and platform SPOC review"


def build_tool_coverage() -> list[tuple[str, str, str, str, str]]:
    return [
        ("Offline to Offline comparison", "Implemented", "High", "High", "Folder/ZIP comparison with parallel processing and report generation"),
        ("Online to Online RTC snapshot comparison", "Implemented", "High", "High", "Snapshot fetch, component comparison, baseline-aware diff logic"),
        ("Online to Offline hybrid mode", "Implemented", "Medium", "High", "RTC snapshot download to temp workspace and local hierarchy comparison"),
        ("Component-aware filtering", "Implemented", "Medium", "High", "Supports component-level filtering and extension-based inclusion rules"),
        ("HTML diff reporting", "Implemented", "High", "High", "Per-file and master comparison reports for technical review"),
        ("Excel and CSV reporting", "Implemented", "High", "High", "Structured output for engineering and program reporting"),
        ("Comment-only change detection", "Implemented", "High", "Medium", "Separates cosmetic deltas from functional changes"),
        ("Changeset enrichment", "Conditional", "Medium", "High", "Available when RTC SCM CLI is installed/configured"),
        ("Interface analysis", "Implemented", "Medium", "High", "Detects functions, variables, types, switches, and file-level interface deltas"),
        ("Dependency and switch analysis", "Implemented", "Medium", "Medium", "Useful for impact and migration complexity assessment"),
        ("AI-assisted recommendation", "Optional", "Low to Medium", "Medium", "Heuristic and AI-backed analysis depending on key/network availability"),
        ("Standalone executable deployment", "Implemented", "Medium", "Medium", "Portable distribution path available; SCM bundling optional"),
    ]


def build_phase_rows() -> list[tuple[str, str, str, int, str, str]]:
    return [
        (
            "Phase 1 - Mobilization and Baseline",
            "Confirm tool setup, RTC access, local workspaces, input component list, and reporting conventions",
            "Input 1",
            40,
            "3 to 5 working days",
            "Tool runs on target environment and baseline comparison reports are reproducible",
        ),
        (
            "Phase 2 - Core Migration Analysis",
            "Run stream/workspace/snapshot comparisons and generate technical evidence pack",
            "Input 1",
            160,
            "10 to 12 working days",
            "Snapshot or folder delta pack reviewed internally and shared for triage",
        ),
        (
            "Phase 3 - Component Optimization",
            "Tune include/exclude rules and OEM-specific filters for DEM, DCOM, NET, and customer conventions",
            "Input 2",
            60,
            "5 to 7 working days",
            "Component-focused comparison outputs produce low-noise actionable deltas",
        ),
        (
            "Phase 4 - Customer Delta and Requirement Analysis",
            "Identify new files, map deltas to DOORS/requirements, and classify reusable vs OEM-specific needs",
            "Inputs 3 and 4",
            81,
            "6 to 8 working days",
            "New files and requirement-driven deltas are categorized and justified",
        ),
        (
            "Phase 5 - Code Rationalization and Governance",
            "Validate code-level modifications, remove non-required carry-over changes, and close with OEM/platform reviews",
            "Input 5",
            30,
            "3 to 4 working days",
            "Final recommendation set is agreed with OEM SPOC and platform SPOC",
        ),
    ]


def create_summary_sheet(ws, rows: list[InputRow]) -> None:
    base_hours = sum(row.base_hours for row in rows)
    contingency_hours = round(base_hours * 0.15, 1)
    governance_hours = round(base_hours * 0.10, 1)
    total_hours = round(base_hours + contingency_hours + governance_hours, 1)
    person_days = round(total_hours / 8.0, 1)

    ws.title = "Executive Summary"
    ws.merge_cells("A1:F2")
    ws["A1"] = "Migration Analysis Tool - OEM Planning and Estimation Report"
    style_cell(ws["A1"], TITLE_FILL, bold=True, font_color="FFFFFF", size=16)

    ws.merge_cells("A4:F4")
    ws["A4"] = (
        "Purpose: establish a one-OEM delivery plan using the current Migration Analysis Tool coverage, "
        "the provided backlog inputs, and an execution-oriented estimation model."
    )
    style_cell(ws["A4"], SECTION_FILL, bold=False)

    summary_rows = [
        ("OEM scope baseline", "One OEM migration readiness and delta assessment"),
        ("Primary objective", "Reduce manual migration analysis effort through structured comparison, interface impact review, and report-driven triage"),
        ("Tool operating modes", "Offline to Offline, Online to Online, Online to Offline (Hybrid)"),
        ("Base engineering effort", base_hours),
        ("Contingency reserve (15%)", contingency_hours),
        ("Governance and review reserve (10%)", governance_hours),
        ("Recommended total effort", total_hours),
        ("Equivalent effort in person-days", person_days),
        ("Indicative duration with 2 engineers", f"{round(total_hours / 16.0, 1)} working days"),
        ("Indicative duration with 3 engineers", f"{round(total_hours / 24.0, 1)} working days"),
    ]

    start_row = 6
    for index, (label, value) in enumerate(summary_rows, start=start_row):
        ws[f"A{index}"] = label
        ws[f"B{index}"] = value
        style_cell(ws[f"A{index}"], HEADER_FILL, bold=True)
        style_cell(ws[f"B{index}"])
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=6)

    risk_note_row = start_row + len(summary_rows) + 2
    ws.merge_cells(start_row=risk_note_row, start_column=1, end_row=risk_note_row, end_column=6)
    ws.cell(risk_note_row, 1).value = (
        "Management note: the recommended total includes review loops with OEM and platform SPOCs, "
        "analysis reruns caused by evolving customer requirements, and tool tuning for component-specific filtering."
    )
    style_cell(ws.cell(risk_note_row, 1), ACCENT_FILL, bold=False)

    set_widths(ws, {"A": 28, "B": 24, "C": 20, "D": 20, "E": 20, "F": 20})


def create_coverage_sheet(ws) -> None:
    ws.title = "Tool Coverage"
    headers = ["Capability", "Implementation Status", "Technical Readiness", "OEM Value", "Remarks"]
    ws.append(headers)
    style_range(ws, "A1:E1", HEADER_FILL, bold=True)

    for row in build_tool_coverage():
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            style_cell(cell)

    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 34, "B": 18, "C": 20, "D": 14, "E": 58})


def create_wbs_sheet(ws, rows: list[InputRow]) -> None:
    ws.title = "Detailed WBS"
    headers = [
        "WP",
        "Topic",
        "Activity / Scope Item",
        "Base Hours",
        "Owner Role",
        "Dependency",
        "Deliverable",
        "Current Status",
    ]
    ws.append(headers)
    style_range(ws, "A1:H1", HEADER_FILL, bold=True)

    for row in rows:
        features = row.feature_lines or [row.feature_coverage]
        efforts = row.effort_values
        for idx, feature in enumerate(features, start=1):
            effort = efforts[idx - 1] if idx - 1 < len(efforts) else "Review / unallocated"
            ws.append(
                [
                    f"WP-{row.serial_no}.{idx}",
                    row.topic,
                    feature,
                    effort,
                    infer_owner(row.topic),
                    infer_dependency(row.serial_no),
                    infer_deliverable(row.topic),
                    row.status,
                ]
            )

    total_row = ws.max_row + 1
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"D{total_row}"] = f"=SUM(D2:D{total_row - 1})"
    for col in range(1, 9):
        style_cell(ws.cell(total_row, col), ACCENT_FILL, bold=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=8):
        for cell in row:
            style_cell(cell)

    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 12, "B": 34, "C": 75, "D": 12, "E": 24, "F": 34, "G": 30, "H": 16})


def create_phase_plan_sheet(ws, rows: list[InputRow]) -> None:
    ws.title = "Phase Plan"
    headers = ["Phase", "Objective", "Mapped Inputs", "Estimated Hours", "Indicative Duration", "Exit Criteria"]
    ws.append(headers)
    style_range(ws, "A1:F1", HEADER_FILL, bold=True)

    phase_rows = build_phase_rows()

    for row in phase_rows:
        ws.append(row)

    ws.append(["Management Reserve", "Cross-phase review cycles, planning overhead, escalations, and rework reserve", "All inputs", 93, "Integrated across phases", "No open delivery-critical blockers"]) 

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        fill = ACCENT_FILL if row[0].row == ws.max_row else None
        for cell in row:
            style_cell(cell, fill=fill, bold=(row[0].row == ws.max_row))

    set_widths(ws, {"A": 28, "B": 45, "C": 16, "D": 14, "E": 20, "F": 42})


def create_gantt_sheet(ws) -> None:
    ws.title = "Gantt Timeline"
    ws.merge_cells("A1:N2")
    ws["A1"] = "One-OEM Delivery Timeline - Gantt View"
    style_cell(ws["A1"], TITLE_FILL, bold=True, font_color="FFFFFF", size=16)

    ws.merge_cells("A4:N4")
    ws["A4"] = (
        "Indicative week-based schedule assuming a focused delivery team. "
        "Core execution phases are sequenced, while governance reserve spans the full program window."
    )
    style_cell(ws["A4"], SECTION_FILL)

    headers = ["Phase", "Hours", "Owner", "Start Week", "Duration (Weeks)"] + [f"W{k}" for k in range(1, 10)]
    header_row = 6
    for col, value in enumerate(headers, start=1):
        ws.cell(header_row, col).value = value
    style_range(ws, f"A{header_row}:N{header_row}", HEADER_FILL, bold=True)

    schedule_rows = [
        ("Phase 1 - Mobilization and Baseline", 40, "Migration Lead", 1, 1),
        ("Phase 2 - Core Migration Analysis", 160, "Migration Analysts + Tool Engineer", 2, 3),
        ("Phase 3 - Component Optimization", 60, "Tool Engineer / Component SME", 5, 2),
        ("Phase 4 - Customer Delta and Requirement Analysis", 81, "Requirements Engineer + Migration Analyst", 7, 2),
        ("Phase 5 - Code Rationalization and Governance", 30, "Software Engineer / Architecture Reviewer", 9, 1),
        ("Management Reserve and Reviews", 93, "Program Lead + OEM / Platform SPOCs", 1, 9),
    ]

    for row_idx, phase_data in enumerate(schedule_rows, start=header_row + 1):
        phase, hours, owner, start_week, duration = phase_data
        ws.cell(row_idx, 1).value = phase
        ws.cell(row_idx, 2).value = hours
        ws.cell(row_idx, 3).value = owner
        ws.cell(row_idx, 4).value = start_week
        ws.cell(row_idx, 5).value = duration

        bar_fill = GANTT_REVIEW_FILL if "Reserve" in phase else GANTT_FILL
        for week in range(1, 10):
            cell = ws.cell(row_idx, 5 + week)
            if start_week <= week < start_week + duration:
                cell.value = "X"
                style_cell(cell, fill=bar_fill, bold=True, font_color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                style_cell(cell)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 6):
            style_cell(ws.cell(row_idx, col))

    legend_row = header_row + len(schedule_rows) + 2
    ws.cell(legend_row, 1).value = "Legend"
    style_cell(ws.cell(legend_row, 1), ACCENT_FILL, bold=True)
    ws.cell(legend_row, 2).value = "Core execution phase"
    style_cell(ws.cell(legend_row, 2), GANTT_FILL, bold=True, font_color="FFFFFF")
    ws.cell(legend_row, 3).value = "Review / reserve / governance"
    style_cell(ws.cell(legend_row, 3), GANTT_REVIEW_FILL, bold=True)

    ws.freeze_panes = "F7"
    set_widths(
        ws,
        {
            "A": 34,
            "B": 10,
            "C": 32,
            "D": 12,
            "E": 16,
            "F": 6,
            "G": 6,
            "H": 6,
            "I": 6,
            "J": 6,
            "K": 6,
            "L": 6,
            "M": 6,
            "N": 6,
        },
    )


def create_risk_sheet(ws) -> None:
    ws.title = "Risks and Dependencies"
    headers = ["Category", "Risk / Dependency", "Impact", "Mitigation"]
    ws.append(headers)
    style_range(ws, "A1:D1", HEADER_FILL, bold=True)

    entries = [
        ("RTC connectivity", "Snapshot fetch and changeset enrichment depend on valid credentials, server access, and optional SCM CLI availability", "High", "Validate credentials early, pre-check connectivity, and keep offline fallback path ready"),
        ("OEM-specific filtering", "Noise level increases if include/exclude rules are not tuned for the target component set", "High", "Run component optimization workshops with OEM SPOC before large-scale analysis"),
        ("Requirement traceability", "DOORS or equivalent requirement mapping may be incomplete or inconsistent", "Medium", "Define a traceability convention and review missing links weekly"),
        ("Tooling dependencies", "AI-assisted features and changeset enrichment are optional and environment-dependent", "Medium", "Treat them as accelerators, not core delivery gates"),
        ("Review bandwidth", "Platform SPOC and OEM SPOC review cycles can delay closure of ambiguous findings", "Medium", "Book recurring governance slots and maintain decision log"),
        ("Input quality", "Workspace, snapshot, or folder baselines may not be aligned or may include obsolete branches", "High", "Freeze comparison baselines before execution and version the run inputs"),
    ]

    for entry in entries:
        ws.append(entry)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            style_cell(cell, fill=RISK_FILL if cell.column == 3 else None)

    set_widths(ws, {"A": 18, "B": 60, "C": 12, "D": 46})


def create_traceability_sheet(ws, rows: list[InputRow]) -> None:
    ws.title = "Input Traceability"
    headers = ["S.NO", "Topic", "Feature coverage", "Estimated Hours", "Status", "Interpreted Base Hours"]
    ws.append(headers)
    style_range(ws, "A1:F1", HEADER_FILL, bold=True)

    for row in rows:
        ws.append([
            row.serial_no,
            row.topic,
            row.feature_coverage,
            row.estimated_hours_raw,
            row.status,
            row.base_hours,
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            style_cell(cell)

    set_widths(ws, {"A": 10, "B": 34, "C": 80, "D": 18, "E": 18, "F": 18})


def generate_report(input_path: Path = INPUT_PATH, output_path: Path = OUTPUT_PATH) -> Path:
    rows = read_input_rows(input_path)
    workbook = Workbook()

    summary_sheet = workbook.active
    create_summary_sheet(summary_sheet, rows)
    create_coverage_sheet(workbook.create_sheet())
    create_wbs_sheet(workbook.create_sheet(), rows)
    create_phase_plan_sheet(workbook.create_sheet(), rows)
    create_gantt_sheet(workbook.create_sheet())
    create_risk_sheet(workbook.create_sheet())
    create_traceability_sheet(workbook.create_sheet(), rows)

    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    generated_path = generate_report()
    print(f"Generated report: {generated_path}")